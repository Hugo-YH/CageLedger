import re
import secrets
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel as openpyxl_from_excel
except ImportError:
    openpyxl = None
    openpyxl_from_excel = None

from server_app.domains.iacuc import normalize_iacuc_number
from server_app.domains.reimbursement.service import coerce_money as coerce_reimbursement_money
from server_app.domains.reimbursement.service import reimbursement_business_key
from server_app.shared.text import clean_text


def ensure_excel_import_supported():
    if not openpyxl:
        raise ValueError("当前运行环境缺少 Excel 导入依赖 openpyxl")


def month_key(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    if isinstance(value, int | float) and openpyxl_from_excel:
        try:
            parsed = openpyxl_from_excel(value)
        except Exception:
            parsed = None
        if isinstance(parsed, datetime):
            return parsed.strftime("%Y-%m")
        if isinstance(parsed, date):
            return parsed.strftime("%Y-%m")
    text = clean_text(value)
    if not text:
        return ""
    compact = re.search(r"(?<!\d)(20\d{2})(\d{2})(?!\d)", text)
    if compact:
        return f"{compact.group(1)}-{compact.group(2)}"
    dotted = re.search(r"(20\d{2})[./-](\d{1,2})", text)
    if dotted:
        return f"{dotted.group(1)}-{int(dotted.group(2)):02d}"
    chinese = re.search(r"(20\d{2})年\s*(\d{1,2})月", text)
    if chinese:
        return f"{chinese.group(1)}-{int(chinese.group(2)):02d}"
    return ""


def split_multiline_values(value):
    text = str(value or "").replace("\r", "\n")
    parts = [clean_text(part) for part in re.split(r"[\n、/]", text)]
    seen = set()
    values = []
    for part in parts:
        if not part or part in seen:
            continue
        seen.add(part)
        values.append(part)
    return values


def distinct_text_list(values):
    seen = set()
    result = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def find_header_row(sheet, required_labels, max_scan_rows=12):
    normalized_required = [clean_text(label).lower() for label in required_labels]
    for row_index in range(1, min(sheet.max_row, max_scan_rows) + 1):
        labels = [clean_text(sheet.cell(row_index, column).value).lower() for column in range(1, sheet.max_column + 1)]
        if all(any(required in label for label in labels if label) for required in normalized_required):
            return row_index, labels
    raise ValueError(f"工作表 {sheet.title} 缺少必需表头")


def find_header_column(labels, *patterns):
    candidates = [clean_text(pattern).lower() for pattern in patterns]
    for index, label in enumerate(labels, start=1):
        if not label:
            continue
        for candidate in candidates:
            if candidate and candidate in label:
                return index
    return 0


def monthly_sheet_month(sheet, header_row):
    for row_index in range(1, min(header_row, 3) + 1):
        for column_index in range(1, min(sheet.max_column, 8) + 1):
            key = month_key(sheet.cell(row_index, column_index).value)
            if key:
                return key
    key = month_key(sheet.title)
    if key:
        return key
    raise ValueError(f"无法识别工作表 {sheet.title} 的月份")


def next_imported_record_id(prefix="reim-import"):
    return f"{prefix}-{secrets.token_hex(8)}"


def build_monthly_import_groups(workbook):
    groups = {}
    for sheet in workbook.worksheets:
        try:
            header_row, labels = find_header_row(sheet, ["项目负责人", "IACUC", "应缴纳"])
        except ValueError:
            continue
        month = monthly_sheet_month(sheet, header_row)
        pi_col = find_header_column(labels, "项目负责人")
        facility_col = find_header_column(labels, "设施")
        iacuc_col = find_header_column(labels, "iacuc")
        funding_col = find_header_column(labels, "伦理对应的经费")
        species_col = find_header_column(labels, "动物品系")
        current_amount_col = find_header_column(labels, "产生的饲养费")
        support_amount_col = find_header_column(labels, "单位支持")
        payable_amount_col = find_header_column(labels, "应缴纳")
        fund_book_col = find_header_column(labels, "经费本号")
        reimbursement_form_col = find_header_column(labels, "单号")
        notes_col = find_header_column(labels, "备注")
        room_col = find_header_column(labels, "房间号")
        start_col = find_header_column(labels, "实验开始")
        end_col = find_header_column(labels, "实验结束")
        if not (pi_col and iacuc_col and current_amount_col and payable_amount_col):
            continue
        empty_rows = 0
        for row_index in range(header_row + 1, min(sheet.max_row, 1200) + 1):
            pi_name = clean_text(sheet.cell(row_index, pi_col).value)
            iacuc = normalize_iacuc_number(sheet.cell(row_index, iacuc_col).value)
            amount = coerce_reimbursement_money(sheet.cell(row_index, current_amount_col).value)
            if not pi_name and not iacuc and amount <= 0:
                empty_rows += 1
                if empty_rows >= 20:
                    break
                continue
            empty_rows = 0
            if not pi_name:
                continue
            business_key = reimbursement_business_key(month, pi_name)
            entry = groups.setdefault(
                business_key,
                {
                    "month": month,
                    "pi": pi_name,
                    "source": "imported",
                    "workflowId": "",
                    "workflowStatus": "",
                    "details": [],
                    "currentMonthAmount": 0.0,
                    "supportAmount": 0.0,
                    "payableAmount": 0.0,
                    "fundBookNos": [],
                    "reimbursementFormNos": [],
                    "notes": [],
                    "species": [],
                    "funding": [],
                },
            )
            support_amount = (
                coerce_reimbursement_money(sheet.cell(row_index, support_amount_col).value)
                if support_amount_col
                else 0.0
            )
            payable_amount = coerce_reimbursement_money(sheet.cell(row_index, payable_amount_col).value)
            note_parts = [clean_text(sheet.cell(row_index, notes_col).value)] if notes_col else []
            start_month = month_key(sheet.cell(row_index, start_col).value) if start_col else ""
            end_month = month_key(sheet.cell(row_index, end_col).value) if end_col else ""
            if start_month:
                note_parts.append(f"实验开始：{start_month}")
            if end_month:
                note_parts.append(f"实验结束：{end_month}")
            funding = clean_text(sheet.cell(row_index, funding_col).value) if funding_col else ""
            species = clean_text(sheet.cell(row_index, species_col).value) if species_col else ""
            rooms = split_multiline_values(sheet.cell(row_index, room_col).value) if room_col else []
            facility = clean_text(sheet.cell(row_index, facility_col).value) if facility_col else ""
            entry["currentMonthAmount"] += amount
            entry["supportAmount"] += support_amount
            entry["payableAmount"] += payable_amount
            entry["fundBookNos"].append(clean_text(sheet.cell(row_index, fund_book_col).value) if fund_book_col else "")
            entry["reimbursementFormNos"].append(
                clean_text(sheet.cell(row_index, reimbursement_form_col).value) if reimbursement_form_col else ""
            )
            entry["notes"].extend([part for part in note_parts if part])
            if species:
                entry["species"].append(species)
            if funding:
                entry["funding"].append(funding)
            entry["details"].append(
                {
                    "iacuc": iacuc,
                    "facility": facility,
                    "funding": funding,
                    "species": species,
                    "project": "",
                    "owner": "",
                    "amount": coerce_reimbursement_money(amount),
                    "supportAmount": coerce_reimbursement_money(support_amount),
                    "payableAmount": coerce_reimbursement_money(payable_amount),
                    "roomNames": rooms,
                    "statementVersionId": "",
                }
            )
    return groups


def arrears_summary_columns(sheet):
    month_columns = []
    max_columns = min(sheet.max_column, 120)
    for column_index in range(6, max_columns + 1):
        label = sheet.cell(1, column_index).value
        month = month_key(label)
        if month:
            month_columns.append((column_index, month))
    return month_columns
