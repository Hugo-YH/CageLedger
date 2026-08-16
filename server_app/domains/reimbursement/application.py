"""Reimbursement record import, accumulation, and detail application services."""

import io

try:
    import openpyxl
except ImportError:
    openpyxl = None

from server_app.cache import cache_get, cache_key, cache_set, invalidate_data_cache_prefixes
from server_app.domains.administration import audit_event, merge_audit_logs, write_audit_events
from server_app.domains.reimbursement.facade import (
    get_reimbursement_record_by_key,
    list_reimbursement_record_summaries_for_pi,
    list_reimbursement_records_for_pi,
    upsert_reimbursement_record,
)
from server_app.domains.reimbursement.importer import (
    arrears_summary_columns,
    build_monthly_import_groups,
    distinct_text_list,
    ensure_excel_import_supported,
    next_imported_record_id,
)
from server_app.domains.reimbursement.projector import build_reimbursement_record_payload, join_distinct_text
from server_app.domains.workflow.facade import (
    get_billing_workflow_detail,
    list_billing_workflow_events,
    list_billing_workflow_versions,
)
from server_app.services.reimbursement import (
    REIMBURSEMENT_STATUS_COMPLETED,
    REIMBURSEMENT_STATUS_PENDING,
    infer_import_status,
    normalize_reimbursement_status,
    reimbursement_business_key,
)
from server_app.services.reimbursement import (
    coerce_money as coerce_reimbursement_money,
)
from server_app.shared import clean_text, now_iso


def upsert_reimbursement_record_from_statement(
    conn, workflow, statement, lines, detail_context_by_iacuc, source="workflow"
):
    business_key = reimbursement_business_key(statement.get("month", ""), statement.get("pi", ""))
    existing = get_reimbursement_record_by_key(conn, business_key) or {}
    payload = build_reimbursement_record_payload(existing, workflow, statement, lines, detail_context_by_iacuc, source)
    upsert_reimbursement_record(conn, payload)
    return payload


def recalculate_reimbursement_accumulations(conn, pi_name):
    records = list_reimbursement_records_for_pi(conn, clean_text(pi_name))
    records = sorted(
        records,
        key=lambda item: (
            clean_text(item.get("month", "")),
            clean_text(item.get("latestEventAt", "")),
            clean_text(item.get("id", "")),
        ),
    )
    accumulated_payable = 0.0
    accumulated_paid = 0.0
    for record in records:
        payable_amount = coerce_reimbursement_money(record.get("payableAmount", 0))
        paid_amount = coerce_reimbursement_money(record.get("paidAmount", 0))
        record["unpaidAmount"] = coerce_reimbursement_money(max(payable_amount - paid_amount, 0))
        accumulated_payable += payable_amount
        accumulated_paid += paid_amount
        record["accumulatedPayable"] = coerce_reimbursement_money(accumulated_payable)
        record["accumulatedPaid"] = coerce_reimbursement_money(accumulated_paid)
        record["accumulatedUnpaid"] = coerce_reimbursement_money(accumulated_payable - accumulated_paid)
        record["updatedAt"] = now_iso()
        upsert_reimbursement_record(conn, record)


def recalculate_all_reimbursement_accumulations(conn):
    rows = conn.execute("SELECT DISTINCT pi FROM reimbursement_records WHERE TRIM(COALESCE(pi, '')) != ''").fetchall()
    for row in rows:
        recalculate_reimbursement_accumulations(conn, row["pi"])


def import_monthly_reimbursement_workbook(conn, file_body, actor):
    ensure_excel_import_supported()
    workbook = openpyxl.load_workbook(io.BytesIO(file_body), data_only=True)
    groups = build_monthly_import_groups(workbook)
    if not groups:
        raise ValueError("未识别到可导入的月度汇总数据")
    saved = []
    audits = []
    imported_months = set()
    imported_pis = set()
    for entry in groups.values():
        existing = get_reimbursement_record_by_key(conn, reimbursement_business_key(entry["month"], entry["pi"])) or {}
        fund_book_no = "；".join(distinct_text_list(entry["fundBookNos"]))
        reimbursement_form_no = "；".join(distinct_text_list(entry["reimbursementFormNos"]))
        notes = "；".join(distinct_text_list(entry["notes"]))
        payload = {
            "id": existing.get("id") or next_imported_record_id(),
            "businessKey": reimbursement_business_key(entry["month"], entry["pi"]),
            "month": entry["month"],
            "pi": entry["pi"],
            "workflowId": existing.get("workflowId", ""),
            "workflowStatus": existing.get("workflowStatus", ""),
            "reimbursementStatus": normalize_reimbursement_status(
                existing.get("reimbursementStatus") or infer_import_status(fund_book_no, reimbursement_form_no, notes)
            ),
            "currentMonthAmount": coerce_reimbursement_money(entry["currentMonthAmount"]),
            "supportAmount": coerce_reimbursement_money(entry["supportAmount"]),
            "payableAmount": coerce_reimbursement_money(entry["payableAmount"]),
            "paidAmount": coerce_reimbursement_money(existing.get("paidAmount", 0)),
            "unpaidAmount": 0,
            "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0)),
            "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0)),
            "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0)),
            "fundBookNo": fund_book_no or clean_text(existing.get("fundBookNo", "")),
            "reimbursementFormNo": reimbursement_form_no or clean_text(existing.get("reimbursementFormNo", "")),
            "approvedBudget": existing.get("approvedBudget", ""),
            "notes": notes or clean_text(existing.get("notes", "")),
            "completedAt": clean_text(existing.get("completedAt", "")),
            "source": "imported",
            "latestEventAt": now_iso(),
            "updatedAt": now_iso(),
            "details": entry["details"],
            "iacucs": distinct_text_list(detail.get("iacuc", "") for detail in entry["details"]),
            "statementVersionId": existing.get("statementVersionId", ""),
            "documentNumber": existing.get("documentNumber", ""),
            "billingUnit": existing.get("billingUnit", ""),
            "project": clean_text(existing.get("project", "")),
            "owner": clean_text(existing.get("owner", "")),
            "funding": "；".join(distinct_text_list(entry["funding"])) or clean_text(existing.get("funding", "")),
        }
        payload["unpaidAmount"] = coerce_reimbursement_money(max(payload["payableAmount"] - payload["paidAmount"], 0))
        if (
            payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED
            and payload["paidAmount"] + 1e-9 < payload["payableAmount"]
        ):
            payload["reimbursementStatus"] = infer_import_status(
                payload["fundBookNo"], payload["reimbursementFormNo"], payload["notes"]
            )
            payload["completedAt"] = ""
        if payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED and not payload["completedAt"]:
            payload["completedAt"] = now_iso()
        upsert_reimbursement_record(conn, payload)
        saved.append(payload)
        imported_months.add(payload["month"])
        imported_pis.add(payload["pi"])
    for pi_name in imported_pis:
        recalculate_reimbursement_accumulations(conn, pi_name)
    event_time = now_iso()
    audit = audit_event(
        actor,
        "reimbursement.import_monthly",
        "reimbursement_record",
        f"monthly:{len(saved)}",
        f"{actor['displayName']} 导入月度报销台账 {len(saved)} 条",
        [],
        event_time,
        None,
        {"months": sorted(imported_months), "count": len(saved)},
    )
    write_audit_events(conn, [audit])
    audits.append(audit)
    invalidate_data_cache_prefixes("reimbursement_records::", "billing_workflows::")
    return {
        "items": saved,
        "auditLogs": merge_audit_logs([], audits),
        "count": len(saved),
        "months": sorted(imported_months),
    }


def import_arrears_reimbursement_workbook(conn, file_body, actor):
    ensure_excel_import_supported()
    workbook = openpyxl.load_workbook(io.BytesIO(file_body), data_only=True)
    if not workbook.worksheets:
        raise ValueError("欠缴工作簿为空")
    sheet = workbook.worksheets[0]
    month_columns = arrears_summary_columns(sheet)
    if not month_columns:
        raise ValueError("未识别到欠缴月份列")
    saved = []
    imported_pis = set()
    empty_rows = 0
    for row_index in range(2, 1200):
        pi_name = clean_text(sheet.cell(row_index, 2).value)
        total_amount = coerce_reimbursement_money(sheet.cell(row_index, 5).value)
        if not pi_name and total_amount <= 0:
            empty_rows += 1
            if empty_rows >= 20:
                break
            continue
        empty_rows = 0
        if not pi_name:
            continue
        species = clean_text(sheet.cell(row_index, 3).value)
        funding = clean_text(sheet.cell(row_index, 4).value)
        for column_index, month in month_columns:
            monthly_unpaid = coerce_reimbursement_money(sheet.cell(row_index, column_index).value)
            if monthly_unpaid <= 0:
                continue
            business_key = reimbursement_business_key(month, pi_name)
            existing = get_reimbursement_record_by_key(conn, business_key) or {}
            if existing and existing.get("source") != "imported":
                continue
            details = existing.get("details") or [
                {
                    "iacuc": "",
                    "facility": "",
                    "funding": funding,
                    "species": species,
                    "project": "",
                    "owner": "",
                    "amount": monthly_unpaid,
                    "supportAmount": 0,
                    "payableAmount": monthly_unpaid,
                    "roomNames": [],
                    "statementVersionId": "",
                }
            ]
            payload = {
                "id": existing.get("id") or next_imported_record_id("reim-arrears"),
                "businessKey": business_key,
                "month": month,
                "pi": pi_name,
                "workflowId": existing.get("workflowId", ""),
                "workflowStatus": existing.get("workflowStatus", ""),
                "reimbursementStatus": normalize_reimbursement_status(
                    existing.get("reimbursementStatus") or REIMBURSEMENT_STATUS_PENDING
                ),
                "currentMonthAmount": coerce_reimbursement_money(
                    existing.get("currentMonthAmount", monthly_unpaid) or monthly_unpaid
                ),
                "supportAmount": coerce_reimbursement_money(existing.get("supportAmount", 0)),
                "payableAmount": coerce_reimbursement_money(
                    existing.get("payableAmount", monthly_unpaid) or monthly_unpaid
                ),
                "paidAmount": coerce_reimbursement_money(existing.get("paidAmount", 0)),
                "unpaidAmount": 0,
                "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0)),
                "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0)),
                "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0)),
                "fundBookNo": clean_text(existing.get("fundBookNo", "")),
                "reimbursementFormNo": clean_text(existing.get("reimbursementFormNo", "")),
                "approvedBudget": existing.get("approvedBudget", ""),
                "notes": join_distinct_text(existing.get("notes", ""), f"欠缴汇算导入：累计欠缴 {total_amount:.2f} 元"),
                "completedAt": clean_text(existing.get("completedAt", "")),
                "source": "imported",
                "latestEventAt": now_iso(),
                "updatedAt": now_iso(),
                "details": details,
                "iacucs": distinct_text_list(detail.get("iacuc", "") for detail in details),
                "statementVersionId": existing.get("statementVersionId", ""),
                "documentNumber": existing.get("documentNumber", ""),
                "billingUnit": existing.get("billingUnit", ""),
                "project": clean_text(existing.get("project", "")),
                "owner": clean_text(existing.get("owner", "")),
                "funding": funding or clean_text(existing.get("funding", "")),
            }
            payload["unpaidAmount"] = coerce_reimbursement_money(
                max(payload["payableAmount"] - payload["paidAmount"], 0)
            )
            upsert_reimbursement_record(conn, payload)
            saved.append(payload)
            imported_pis.add(pi_name)
    if not saved:
        raise ValueError("未识别到可导入的欠缴记录")
    for pi_name in imported_pis:
        recalculate_reimbursement_accumulations(conn, pi_name)
    event_time = now_iso()
    audit = audit_event(
        actor,
        "reimbursement.import_arrears",
        "reimbursement_record",
        f"arrears:{len(saved)}",
        f"{actor['displayName']} 导入欠缴汇算 {len(saved)} 条",
        [],
        event_time,
        None,
        {"count": len(saved)},
    )
    write_audit_events(conn, [audit])
    invalidate_data_cache_prefixes("reimbursement_records::", "billing_workflows::")
    return {"items": saved, "auditLogs": merge_audit_logs([], [audit]), "count": len(saved)}


def reimbursement_detail_payload(conn, record):
    record_id = clean_text(record.get("id", ""))
    workflow_id = clean_text(record.get("workflowId", ""))
    cache_key_value = cache_key(
        "reimbursement_records::detail",
        record_id=record_id,
        workflow_id=workflow_id,
        updated_at=clean_text(record.get("updatedAt", "")),
        latest_event_at=clean_text(record.get("latestEventAt", "")),
    )
    cached = cache_get(cache_key_value)
    if cached is not None:
        return cached
    workflow = get_billing_workflow_detail(conn, record.get("workflowId", "")) if record.get("workflowId") else None
    workflow_versions = list_billing_workflow_versions(conn, workflow["id"]) if workflow else []
    workflow_events = list_billing_workflow_events(conn, workflow["id"]) if workflow else []
    history = list_reimbursement_record_summaries_for_pi(conn, record.get("pi", ""))
    payload = {
        "item": record,
        "workflow": workflow if workflow else None,
        "workflowVersions": workflow_versions,
        "workflowEvents": workflow_events,
        "history": history,
    }
    return cache_set(cache_key_value, payload)
