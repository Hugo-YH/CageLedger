"""项目负责人结算单 Excel 导出，版式与打印/PDF 汇总表保持一致。

表格结构与 PDF 相同：标题、信息区、汇总块 + 按 IACUC 分列（笼数/数量、减免、阶梯、金额）、
单项合计、本月待缴纳总计、自定义收费明细与签字区。

每日金额在列不含自定义收费时写为公式 = (笼数-减免-阶梯) × 基础价 + 阶梯 × 超额价；
含自定义收费的列金额为引擎计算值（与 PDF 数字一致），自定义部分在附表按 数量 × 单价 公式单独核对。
合计与总计全部使用 SUM 公式，打开文件时强制重算。
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
_CUSTOM_FILL = PatternFill("solid", fgColor="FFF2CC")
_TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True)
_HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True)
_BODY_FONT = Font(name="Microsoft YaHei", size=10)
_MONEY_FORMAT = "#,##0.00"
_INTEGER_FORMAT = "0"

_UNIT_LABELS = {"cage_day": "笼/天", "animal_day": "只/天", "mixed": "混合"}
_SPECIES_ORDER = {"mouse": 0, "rat": 1, "guinea_pig": 2, "rabbit": 3, "monkey": 4, "pig": 5, "dog": 6}
_SPECIES_LABELS = {
    "mouse": "小鼠",
    "rat": "大鼠",
    "guinea_pig": "豚鼠",
    "rabbit": "兔",
    "monkey": "猴",
    "pig": "猪",
    "dog": "犬",
}


def _as_int(value):
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _as_money(value):
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _column_key(item):
    return "|".join(
        [
            str(item.get("iacuc") or ""),
            str(item.get("species") or ""),
            str(item.get("billingItem") or ""),
            str(item.get("billingUnit") or ""),
            f"{_as_money(item.get('statementUnitPrice', item.get('unitPrice'))):.2f}",
            f"{_as_money(item.get('statementOverageUnitPrice', item.get('overageUnitPrice'))):.2f}",
            "1" if item.get("statementTiered", item.get("tiered")) else "0",
            "1" if item.get("statementFreeAllowance", item.get("freeAllowance")) else "0",
            "1" if item.get("statementFullExemption", item.get("fullExemption")) else "0",
        ]
    )


def _statement_columns(statement, lines):
    iacuc_order = {iacuc: index for index, iacuc in enumerate(statement.get("iacucs") or [])}
    columns = []
    seen = set()
    for line in lines:
        for item in line.get("iacucBreakdown") or []:
            key = _column_key(item)
            if key in seen:
                continue
            seen.add(key)
            unit = str(item.get("billingUnit") or "cage_day")
            columns.append(
                {
                    "key": key,
                    "iacuc": str(item.get("iacuc") or ""),
                    "species": str(item.get("species") or ""),
                    "unit": unit,
                    "basePrice": _as_money(item.get("statementUnitPrice", item.get("unitPrice"))),
                    "overagePrice": _as_money(item.get("statementOverageUnitPrice", item.get("overageUnitPrice"))),
                    "tiered": bool(item["tiered"]) if "tiered" in item else bool(item.get("statementTiered")),
                    "freeAllowance": bool(item.get("statementFreeAllowance", item.get("freeAllowance"))),
                    "fullExemption": bool(item.get("statementFullExemption", item.get("fullExemption"))),
                    "hasCustom": False,
                }
            )
    for line in lines:
        for item in line.get("iacucBreakdown") or []:
            if not item.get("customBilling"):
                continue
            key = _column_key(item)
            for column in columns:
                if column["key"] == key:
                    column["hasCustom"] = True
    columns.sort(
        key=lambda column: (
            _SPECIES_ORDER.get(column["species"], len(_SPECIES_ORDER)),
            iacuc_order.get(column["iacuc"], 10**9),
            column["iacuc"],
            column["unit"],
            column["basePrice"],
        )
    )
    return columns


def _custom_range(item):
    start = str(item.get("customBillingStartDate") or "").strip()
    end = str(item.get("customBillingEndDate") or "").strip()
    if start and end:
        return f"{start} ~ {end}"
    return start or end or ""


def _statement_rows(lines, columns):
    rows = []
    for line in lines:
        per_column = {
            column["key"]: {"count": 0, "free": 0, "tier1": 0, "tier2": 0, "amount": 0.0} for column in columns
        }
        custom = []
        has_activity = False
        for item in line.get("iacucBreakdown") or []:
            key = _column_key(item)
            if key in per_column:
                cell = per_column[key]
                cell["count"] += _as_int(item.get("cageCount") or item.get("animalCount"))
                cell["free"] += _as_int(item.get("freeCages"))
                cell["tier1"] += _as_int(item.get("tier1BillableCages"))
                cell["tier2"] += _as_int(item.get("tier2BillableCages"))
                payable = item.get("payableAmount")
                cell["amount"] += _as_money(payable if payable is not None else item.get("amount"))
            if item.get("customBilling"):
                quantity = _as_int(item.get("cageCount") or item.get("animalCount"))
                custom.append(
                    {
                        "date": line.get("date", ""),
                        "iacuc": str(item.get("iacuc") or ""),
                        "range": _custom_range(item),
                        "quantity": quantity,
                        "unitPrice": _as_money(item.get("unitPrice")),
                        "note": str(item.get("customBillingNote") or ""),
                    }
                )
                has_activity = has_activity or quantity > 0
            elif key in per_column:
                cell = per_column[key]
                if cell["count"] or cell["free"] or cell["tier1"] or cell["tier2"] or cell["amount"]:
                    has_activity = True
        if has_activity:
            rows.append({"date": line.get("date", ""), "perColumn": per_column, "custom": custom})
    return rows


def _sanitize_sheet_name(name):
    cleaned = "".join("_" if character in "[]:*?/\\" else character for character in name)
    return (cleaned or "结算单")[:31]


def _write_custom_appendix(worksheet, custom_rows, header_row):
    """写自定义收费明细表，返回 (合计单元格引用, 下一可用行)。"""
    if not custom_rows:
        return None, header_row
    custom_first_row = header_row + 2
    custom_last_row = custom_first_row + len(custom_rows) - 1
    worksheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=7)
    custom_title = worksheet.cell(header_row, 1, "自定义收费明细（金额 = 数量 × 单价，已计入每日金额）")
    custom_title.font = _HEADER_FONT
    custom_title.fill = _CUSTOM_FILL
    custom_title.border = _BORDER
    custom_headers = ["日期", "IACUC", "收费区间", "数量", "单价", "金额", "备注"]
    for index, label in enumerate(custom_headers):
        cell = worksheet.cell(header_row + 1, index + 1, label)
        cell.font = _HEADER_FONT
        cell.fill = _CUSTOM_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for custom_index, custom in enumerate(custom_rows):
        custom_row = custom_first_row + custom_index
        values = [
            custom["date"],
            custom["iacuc"],
            custom["range"],
            custom["quantity"],
            custom["unitPrice"],
            None,
            custom["note"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(custom_row, column_index)
            if column_index == 6:
                cell.value = f"=ROUND(D{custom_row}*E{custom_row},2)"
                cell.number_format = _MONEY_FORMAT
            else:
                cell.value = value
                if column_index == 4:
                    cell.number_format = _INTEGER_FORMAT
                elif column_index == 5:
                    cell.number_format = _MONEY_FORMAT
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(
                horizontal="center" if column_index != 7 else "left",
                vertical="center",
                wrap_text=column_index == 7,
            )
    custom_total_row = custom_last_row + 1
    worksheet.merge_cells(start_row=custom_total_row, start_column=1, end_row=custom_total_row, end_column=5)
    custom_total_label = worksheet.cell(custom_total_row, 1, "自定义收费合计（元）")
    custom_total_label.font = _HEADER_FONT
    custom_total_label.fill = _TOTAL_FILL
    custom_total_label.alignment = Alignment(horizontal="right", vertical="center")
    for column_index in range(1, 6):
        worksheet.cell(custom_total_row, column_index).border = _BORDER
        worksheet.cell(custom_total_row, column_index).fill = _TOTAL_FILL
    custom_total_cell = worksheet.cell(
        custom_total_row,
        6,
        f"=SUM(F{custom_first_row}:F{custom_last_row})",
    )
    custom_total_cell.font = _HEADER_FONT
    custom_total_cell.fill = _TOTAL_FILL
    custom_total_cell.border = _BORDER
    custom_total_cell.number_format = _MONEY_FORMAT
    custom_total_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(custom_total_row, 7).border = _BORDER
    worksheet.cell(custom_total_row, 7).fill = _TOTAL_FILL
    return f"F{custom_total_row}", custom_total_row + 1


def build_settlement_workbook(entries):
    """entries: [(statement, lines), ...]"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for statement, lines in entries:
        pi = str(statement.get("pi") or "-")
        month = str(statement.get("month") or "")
        sheet_name = _sanitize_sheet_name(f"{pi} {month} 结算单")
        _write_statement_sheet(workbook.create_sheet(sheet_name), statement, lines)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _count_label(unit):
    if unit == "animal_day":
        return "只数"
    if unit == "mixed":
        return "总量"
    return "笼数"


def _species_label(value):
    return _SPECIES_LABELS.get(str(value or "").strip().lower(), str(value or "动物"))


def _leading_count_label(unit):
    if unit == "animal_day":
        return "总数量"
    if unit == "mixed":
        return "总量"
    return "总笼数"


def _amount_formula(count_cell, free_cell, tier_cell, base_price, overage_price, data_row):
    def ref(cell):
        return f"{get_column_letter(cell)}{data_row}"

    count_ref = ref(count_cell)
    billable_expr = count_ref
    if free_cell:
        billable_expr = f"({count_ref}-{ref(free_cell)}"
        if tier_cell:
            billable_expr += f"-{ref(tier_cell)}"
        billable_expr += ")"
    elif tier_cell:
        billable_expr = f"({count_ref}-{ref(tier_cell)})"
    tier_expr = ref(tier_cell) if tier_cell else "0"
    return f"=ROUND({billable_expr}*{base_price:.2f}+{tier_expr}*{overage_price:.2f},2)"


def _write_statement_sheet(worksheet, statement, lines):
    columns = _statement_columns(statement, lines)
    rows = _statement_rows(lines, columns)
    custom_rows = [custom for day in rows for custom in day["custom"]]

    totals = {column["key"]: {"count": 0, "free": 0, "tier1": 0, "tier2": 0, "amount": 0.0} for column in columns}
    for day in rows:
        for key, cell in day["perColumn"].items():
            for field in ("count", "free", "tier1", "tier2"):
                totals[key][field] += cell[field]
            totals[key]["amount"] += cell["amount"]
    for column in columns:
        summary = totals[column["key"]]
        column["showFree"] = column["species"] == "mouse" and summary["free"] > 0
        column["showTier"] = column["species"] == "mouse" and summary["tier2"] > 0
    mouse_columns = [column for column in columns if column["species"] == "mouse"]
    leading_free = any(column["showFree"] for column in mouse_columns)
    leading_tier = any(column["showTier"] for column in mouse_columns)
    has_mouse_summary = bool(mouse_columns)
    unit = str(statement.get("billingUnit") or "cage_day")

    row = 1
    title = f"{statement.get('pi') or '-'}课题组实验动物饲养费核算汇总表"

    # 计算总列数：日期 + 小鼠汇总块（仅小鼠） + 各品系 IACUC 分列。
    leading_width = (2 + (1 if leading_free else 0) + (1 if leading_tier else 0)) if has_mouse_summary else 0
    table_columns = (
        1 + leading_width + sum(2 + (1 if c["showFree"] else 0) + (1 if c["showTier"] else 0) for c in columns)
    )
    total_columns = max(table_columns, 8)

    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
    title_cell = worksheet.cell(row, 1, title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[row].height = 28
    row += 1

    info_pairs = [
        ("结算月份", str(statement.get("month") or ""), "项目负责人", str(statement.get("pi") or "")),
        ("计费单位", _UNIT_LABELS.get(unit, unit), "实验负责人", str(statement.get("owner") or "-")),
    ]
    for label_left, value_left, label_right, value_right in info_pairs:
        left_label = worksheet.cell(row, 1, label_left)
        left_label.font = _HEADER_FONT
        left_label.fill = _HEADER_FILL
        left_label.border = _BORDER
        left_value = worksheet.cell(row, 2, value_left)
        left_value.font = _BODY_FONT
        left_value.border = _BORDER
        worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        right_label = worksheet.cell(row, 4, label_right)
        right_label.font = _HEADER_FONT
        right_label.fill = _HEADER_FILL
        right_label.border = _BORDER
        right_value = worksheet.cell(row, 5, value_right)
        right_value.font = _BODY_FONT
        right_value.border = _BORDER
        worksheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        row += 1

    iacucs = "、".join(statement.get("iacucs") or []) or "-"
    full_exemption = [column["iacuc"] for column in columns if column["fullExemption"]]
    if full_exemption:
        iacucs = f"{iacucs}（全额减免：{'、'.join(dict.fromkeys(full_exemption))}）"
    worksheet.cell(row, 1, "IACUC 编号").font = _HEADER_FONT
    worksheet.cell(row, 1).fill = _HEADER_FILL
    worksheet.cell(row, 1).border = _BORDER
    worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    iacuc_cell = worksheet.cell(row, 2, iacucs)
    iacuc_cell.font = _BODY_FONT
    iacuc_cell.border = _BORDER
    iacuc_cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[row].height = 30
    row += 1

    funding = str(statement.get("funding") or "-")
    worksheet.cell(row, 1, "支撑经费").font = _HEADER_FONT
    worksheet.cell(row, 1).fill = _HEADER_FILL
    worksheet.cell(row, 1).border = _BORDER
    worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    funding_cell = worksheet.cell(row, 2, funding)
    funding_cell.font = _BODY_FONT
    funding_cell.border = _BORDER
    row += 1

    notes = statement.get("notes") or ""
    note_text = (
        "合计与自定义收费金额为 Excel 公式；每日金额由系统按减免与梯度规则计算，"
        "含自定义收费的列金额与 PDF 一致，自定义部分可在附表按 数量 × 单价 核对。"
    )
    if notes:
        note_text = f"{note_text} {notes}"
    worksheet.cell(row, 1, "说明").font = _HEADER_FONT
    worksheet.cell(row, 1).fill = _HEADER_FILL
    worksheet.cell(row, 1).border = _BORDER
    worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    note_cell = worksheet.cell(row, 2, note_text)
    note_cell.font = Font(name="Microsoft YaHei", size=9)
    note_cell.border = _BORDER
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[row].height = 30
    row += 1

    if not columns:
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        all_custom_note = worksheet.cell(row, 1, "本月计费全部为自定义收费，金额详见下方自定义收费明细。")
        all_custom_note.font = _BODY_FONT
        all_custom_note.fill = _CUSTOM_FILL
        all_custom_note.border = _BORDER
        row += 1
        custom_ref, next_row = _write_custom_appendix(worksheet, custom_rows, row + 1)
        payable_row = next_row + 1
        worksheet.merge_cells(start_row=payable_row, start_column=1, end_row=payable_row, end_column=5)
        payable_label = worksheet.cell(payable_row, 1, "本月待缴纳饲养费总计（元）")
        payable_label.font = _HEADER_FONT
        payable_label.fill = _TOTAL_FILL
        payable_label.alignment = Alignment(horizontal="right", vertical="center")
        for column_index in range(1, 6):
            worksheet.cell(payable_row, column_index).border = _BORDER
            worksheet.cell(payable_row, column_index).fill = _TOTAL_FILL
        payable_value = worksheet.cell(payable_row, 6)
        payable_value.value = f"={custom_ref}" if custom_ref else "0"
        payable_value.font = _HEADER_FONT
        payable_value.fill = _TOTAL_FILL
        payable_value.border = _BORDER
        payable_value.number_format = _MONEY_FORMAT
        payable_value.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.cell(payable_row, 7).border = _BORDER
        worksheet.cell(payable_row, 7).fill = _TOTAL_FILL
        worksheet.freeze_panes = f"A{payable_row}"
        for column, width in (("A", 12), ("B", 20), ("C", 20), ("D", 10), ("E", 10), ("F", 13), ("G", 40)):
            worksheet.column_dimensions[column].width = width
        return

    # 计算各列实际位置
    current = 2
    leading = {}
    if has_mouse_summary:
        leading["count"] = current
        current += 1
        if leading_free:
            leading["free"] = current
            current += 1
        if leading_tier:
            leading["tier"] = current
            current += 1
        leading["amount"] = current
        current += 1
    positions = []
    for column in columns:
        position = {"count": current}
        current += 1
        if column["showFree"]:
            position["free"] = current
            current += 1
        if column["showTier"]:
            position["tier"] = current
            current += 1
        position["amount"] = current
        current += 1
        positions.append(position)

    header_species_row = row
    header_group_row = row + 1
    header_sub_row = row + 2
    first_data_row = row + 3
    last_data_row = first_data_row + max(len(rows) - 1, 0)

    worksheet.merge_cells(start_row=header_species_row, start_column=1, end_row=header_sub_row, end_column=1)
    date_header = worksheet.cell(header_species_row, 1, "日期")
    date_header.font = _HEADER_FONT
    date_header.fill = _HEADER_FILL
    date_header.alignment = Alignment(horizontal="center", vertical="center")
    date_header.border = _BORDER
    for header_row in (header_group_row, header_sub_row):
        worksheet.cell(header_row, 1).border = _BORDER
        worksheet.cell(header_row, 1).fill = _HEADER_FILL

    species_ranges = []
    for column, position in zip(columns, positions, strict=True):
        start = position["count"]
        end = position["amount"]
        if species_ranges and species_ranges[-1]["species"] == column["species"]:
            species_ranges[-1]["end"] = end
        else:
            species_ranges.append({"species": column["species"], "start": start, "end": end})
    if has_mouse_summary:
        species_ranges[0]["start"] = leading["count"]
    for item in species_ranges:
        worksheet.merge_cells(
            start_row=header_species_row,
            start_column=item["start"],
            end_row=header_species_row,
            end_column=item["end"],
        )
        cell = worksheet.cell(header_species_row, item["start"], _species_label(item["species"]))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index in range(item["start"], item["end"] + 1):
            worksheet.cell(header_species_row, column_index).border = _BORDER
            worksheet.cell(header_species_row, column_index).fill = _HEADER_FILL

    if has_mouse_summary:
        leading_start = leading["count"]
        leading_end = leading["amount"]
        worksheet.merge_cells(
            start_row=header_group_row,
            start_column=leading_start,
            end_row=header_group_row,
            end_column=leading_end,
        )
        leading_header = worksheet.cell(header_group_row, leading_start, "汇总")
        leading_header.font = _HEADER_FONT
        leading_header.fill = _HEADER_FILL
        leading_header.alignment = Alignment(horizontal="center", vertical="center")
        for column_index in range(leading_start, leading_end + 1):
            worksheet.cell(header_group_row, column_index).border = _BORDER
            worksheet.cell(header_group_row, column_index).fill = _HEADER_FILL
        leading_sub = [(leading["count"], _leading_count_label("cage_day"))]
        if leading_free:
            leading_sub.append((leading["free"], "减免总笼数"))
        if leading_tier:
            leading_sub.append((leading["tier"], "阶梯总笼数"))
        leading_sub.append((leading["amount"], "缴纳（元）"))
        for column_index, label in leading_sub:
            cell = worksheet.cell(header_sub_row, column_index, label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    base_counts = {}
    for column in columns:
        base = str(column["iacuc"] or "-")
        base_counts[base] = base_counts.get(base, 0) + 1

    for group_index, column in enumerate(columns):
        start = positions[group_index]["count"]
        end = positions[group_index]["amount"]
        label = str(column["iacuc"] or "-")
        if base_counts[label] > 1:
            label = f"{label} / ¥{column['basePrice']:.2f}"
        if column["showTier"]:
            label = f"{label}（梯度收费）"
        if column["fullExemption"]:
            label = f"{label}（全额减免）"
        if column["hasCustom"]:
            label = f"{label}（含自定义收费）"
        worksheet.merge_cells(start_row=header_group_row, start_column=start, end_row=header_group_row, end_column=end)
        group_cell = worksheet.cell(header_group_row, start, label)
        group_cell.font = _HEADER_FONT
        group_cell.fill = _HEADER_FILL
        group_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index in range(start, end + 1):
            worksheet.cell(header_group_row, column_index).border = _BORDER
            worksheet.cell(header_group_row, column_index).fill = _HEADER_FILL
        sub_headers = [_count_label(column["unit"])]
        if column["showFree"]:
            sub_headers.append("减免")
        if column["showTier"]:
            sub_headers.append("阶梯")
        sub_headers.append("缴纳（元）")
        for sub_index, sub_label in enumerate(sub_headers):
            cell = worksheet.cell(header_sub_row, start + sub_index, sub_label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    for data_index, day in enumerate(rows):
        data_row = first_data_row + data_index
        date_cell = worksheet.cell(data_row, 1, day["date"])
        date_cell.font = _BODY_FONT
        date_cell.border = _BORDER
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        mouse_count_refs = []
        mouse_free_refs = []
        mouse_tier_refs = []
        mouse_amount_refs = []
        for group_index, column in enumerate(columns):
            position = positions[group_index]
            cell = day["perColumn"].get(column["key"], {"count": 0, "free": 0, "tier1": 0, "tier2": 0, "amount": 0.0})
            count_cell = worksheet.cell(data_row, position["count"], cell["count"])
            count_cell.font = _BODY_FONT
            count_cell.border = _BORDER
            count_cell.number_format = _INTEGER_FORMAT
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            if column["species"] == "mouse":
                mouse_count_refs.append(position["count"])
            if column["showFree"]:
                free_cell = worksheet.cell(data_row, position["free"], cell["free"])
                free_cell.font = _BODY_FONT
                free_cell.border = _BORDER
                free_cell.number_format = _INTEGER_FORMAT
                free_cell.alignment = Alignment(horizontal="center", vertical="center")
                if column["species"] == "mouse":
                    mouse_free_refs.append(position["free"])
            if column["showTier"]:
                tier_cell = worksheet.cell(data_row, position["tier"], cell["tier2"])
                tier_cell.font = _BODY_FONT
                tier_cell.border = _BORDER
                tier_cell.number_format = _INTEGER_FORMAT
                tier_cell.alignment = Alignment(horizontal="center", vertical="center")
                if column["species"] == "mouse":
                    mouse_tier_refs.append(position["tier"])
            amount_cell = worksheet.cell(data_row, position["amount"])
            if column["hasCustom"]:
                amount_cell.value = cell["amount"]
            else:
                amount_cell.value = _amount_formula(
                    position["count"],
                    position.get("free"),
                    position.get("tier"),
                    column["basePrice"],
                    column["overagePrice"],
                    data_row,
                )
            amount_cell.font = _BODY_FONT
            amount_cell.border = _BORDER
            amount_cell.number_format = _MONEY_FORMAT
            if column["species"] == "mouse":
                mouse_amount_refs.append(position["amount"])

        def ref_list(cells, current_row=data_row):
            return "+".join(f"{get_column_letter(cell)}{current_row}" for cell in cells)

        if has_mouse_summary:
            leading_count_cell = worksheet.cell(data_row, leading["count"], f"={ref_list(mouse_count_refs)}")
            leading_count_cell.font = _BODY_FONT
            leading_count_cell.border = _BORDER
            leading_count_cell.number_format = _INTEGER_FORMAT
            leading_count_cell.alignment = Alignment(horizontal="center", vertical="center")
            if leading_free:
                cell = worksheet.cell(data_row, leading["free"], f"={ref_list(mouse_free_refs)}")
                cell.font = _BODY_FONT
                cell.border = _BORDER
                cell.number_format = _INTEGER_FORMAT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if leading_tier:
                cell = worksheet.cell(data_row, leading["tier"], f"={ref_list(mouse_tier_refs)}")
                cell.font = _BODY_FONT
                cell.border = _BORDER
                cell.number_format = _INTEGER_FORMAT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            leading_amount_cell = worksheet.cell(data_row, leading["amount"], f"={ref_list(mouse_amount_refs)}")
            leading_amount_cell.font = _BODY_FONT
            leading_amount_cell.border = _BORDER
            leading_amount_cell.number_format = _MONEY_FORMAT
            leading_amount_cell.alignment = Alignment(horizontal="center", vertical="center")

    totals_row = last_data_row + 1
    summary_label = worksheet.cell(totals_row, 1, "单项合计")
    summary_label.font = _HEADER_FONT
    summary_label.fill = _TOTAL_FILL
    summary_label.border = _BORDER

    def sum_cell(target_row, target_column, number_format):
        cell = worksheet.cell(target_row, target_column)
        cell.value = f"=SUM({get_column_letter(target_column)}{first_data_row}:{get_column_letter(target_column)}{last_data_row})"
        cell.font = _HEADER_FONT
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER
        cell.number_format = number_format
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if has_mouse_summary:
        sum_cell(totals_row, leading["count"], _INTEGER_FORMAT)
        if leading_free:
            sum_cell(totals_row, leading["free"], _INTEGER_FORMAT)
        if leading_tier:
            sum_cell(totals_row, leading["tier"], _INTEGER_FORMAT)
        sum_cell(totals_row, leading["amount"], _MONEY_FORMAT)
    for position in positions:
        for field in ("count", "free", "tier"):
            if field in position:
                sum_cell(totals_row, position[field], _INTEGER_FORMAT)
        sum_cell(totals_row, position["amount"], _MONEY_FORMAT)

    payable_row = totals_row + 1
    payable_amount_column = leading["amount"] if has_mouse_summary else positions[0]["amount"]
    worksheet.merge_cells(
        start_row=payable_row, start_column=1, end_row=payable_row, end_column=payable_amount_column - 1
    )
    payable_label = worksheet.cell(payable_row, 1, "本月待缴纳饲养费总计（元）")
    payable_label.font = _HEADER_FONT
    payable_label.fill = _TOTAL_FILL
    payable_label.alignment = Alignment(horizontal="right", vertical="center")
    for column_index in range(1, payable_amount_column):
        worksheet.cell(payable_row, column_index).border = _BORDER
        worksheet.cell(payable_row, column_index).fill = _TOTAL_FILL
    payable_cell = worksheet.cell(payable_row, payable_amount_column)
    amount_columns = [position["amount"] for position in positions]
    payable_cell.value = f"=SUM({','.join(f'{get_column_letter(column)}{totals_row}' for column in amount_columns)})"
    payable_cell.font = _HEADER_FONT
    payable_cell.fill = _TOTAL_FILL
    payable_cell.border = _BORDER
    payable_cell.number_format = _MONEY_FORMAT
    payable_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(payable_row, payable_amount_column + 1).border = _BORDER
    worksheet.cell(payable_row, payable_amount_column + 1).fill = _TOTAL_FILL

    _write_custom_appendix(worksheet, custom_rows, payable_row + 2)

    # 签字区
    sign_row = payable_row + (2 if not custom_rows else 6)
    sign_headers = ["项目负责人", "实验负责人/经办人", "日期"]
    worksheet.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=2)
    worksheet.merge_cells(start_row=sign_row, start_column=3, end_row=sign_row, end_column=4)
    worksheet.merge_cells(start_row=sign_row, start_column=5, end_row=sign_row, end_column=6)
    for index, label in enumerate(sign_headers):
        cell = worksheet.cell(sign_row, 1 + index * 2, label)
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        worksheet.cell(sign_row, 2 + index * 2).border = _BORDER

    worksheet.freeze_panes = f"B{first_data_row}"
    worksheet.column_dimensions["A"].width = 12
    for column_index in range(2, total_columns + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 9
    worksheet.column_dimensions[get_column_letter(payable_amount_column)].width = 12
    for position in positions:
        worksheet.column_dimensions[get_column_letter(position["amount"])].width = 12
