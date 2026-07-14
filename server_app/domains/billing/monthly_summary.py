import io
import re
from collections import defaultdict
from datetime import date, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - deployment dependency validation
    Workbook = None

from server_app.domains.iacuc import normalize_application_date, normalize_iacuc_number
from server_app.shared import clean_text

from .allowance import billing_free_cages_for_pi
from .profiles import billing_profile_for_room
from .statements import quantity_sheet_statement_lines

MONTH_PATTERN = re.compile(r"\d{4}-\d{2}\Z")
EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FACILITY_LABELS = {
    "zhujiang": "珠江新城",
    "bioisland": "生物岛",
}
SPECIES_LABELS = {
    "mouse": "小鼠",
    "mouse_standard": "小鼠",
    "mouse_diabetic": "小鼠",
    "rat": "大鼠",
    "rat_standard": "大鼠",
    "rat_diabetic": "大鼠",
    "guinea_pig": "豚鼠",
    "rabbit": "兔",
    "monkey": "猴",
    "pig": "猪",
    "dog": "犬",
}


def build_monthly_summary_rows(
    month,
    sheets,
    rooms,
    applications_by_iacuc,
    principal_type_by_pi,
    reimbursement_by_pi,
):
    """Build monthly export rows from the same daily breakdown used by PI settlement."""
    if not MONTH_PATTERN.fullmatch(str(month or "")):
        raise ValueError("结算月份格式应为 YYYY-MM")

    room_by_id = {room.get("id"): room for room in rooms or []}
    contexts = {}
    sheets_by_pi = defaultdict(list)
    errors = []
    for sheet in sheets or []:
        pi_name = clean_text(sheet.get("pi", ""))
        iacuc = normalize_iacuc_number(sheet.get("iacuc", ""))
        if not pi_name or not iacuc:
            errors.append(
                f"项目负责人 {pi_name or '未填写'} / IACUC {iacuc or '未填写'}："
                f"缺少{'项目负责人' if not pi_name else 'IACUC 编号'}"
            )
            continue
        room = room_by_id.get(sheet.get("roomId")) or next(
            (item for item in rooms or [] if clean_text(item.get("name")) == clean_text(sheet.get("roomName"))), {}
        )
        profile = billing_profile_for_room(room or {}, sheet.get("billingUnit"))
        key = (pi_name, iacuc, clean_text(profile.get("facility")) or "未设置设施")
        current = contexts.setdefault(
            key,
            {
                "pi": pi_name,
                "iacuc": iacuc,
                "facility": _facility_label(key[2]),
                "species": set(),
                "roomNames": set(),
                "amount": 0.0,
                "supportAmount": 0.0,
                "payableAmount": 0.0,
                "expiredAllowance": {},
            },
        )
        species = _species_label(clean_text(room.get("defaultSpecies")) or clean_text(profile.get("species")))
        if species:
            current["species"].add(species)
        if clean_text(sheet.get("roomName")):
            current["roomNames"].add(clean_text(sheet.get("roomName")))
        sheets_by_pi[pi_name].append(sheet)

    for pi_name, pi_sheets in sheets_by_pi.items():
        free_cages = billing_free_cages_for_pi(principal_type_by_pi, pi_name)
        lines = quantity_sheet_statement_lines(pi_sheets, free_cages, rooms, applications_by_iacuc)
        for line in lines:
            for item in line.get("iacucBreakdown", []):
                iacuc = normalize_iacuc_number(item.get("iacuc", ""))
                facility = clean_text(item.get("facility")) or "未设置设施"
                key = (pi_name, iacuc, facility)
                current = contexts.get(key)
                if current is None:
                    continue
                current["amount"] += float(item.get("amount") or 0)
                current["supportAmount"] += float(item.get("supportAmount") or 0)
                current["payableAmount"] += float(item.get("payableAmount") or 0)
                expiry = clean_text(item.get("freeAllowanceExpiryDate"))
                if expiry and (item.get("freeAllowance") or item.get("fullExemption")):
                    current["expiredAllowance"].setdefault(iacuc, (expiry, line.get("date", "")))

    if errors:
        raise ValueError("；".join(sorted(set(errors))))

    rows = []
    for current in contexts.values():
        application = applications_by_iacuc.get(current["iacuc"], {})
        reimbursement = reimbursement_by_pi.get(current["pi"], {})
        rows.append(
            {
                "pi": current["pi"],
                "facility": current["facility"],
                "iacuc": current["iacuc"],
                "funding": clean_text(application.get("funding", "")),
                "species": "、".join(sorted(current["species"])),
                "amount": round(current["amount"], 2),
                "supportAmount": round(current["supportAmount"], 2),
                "payableAmount": round(current["payableAmount"], 2),
                "fundBookNo": clean_text(reimbursement.get("fundBookNo", "")),
                "reimbursementFormNo": clean_text(reimbursement.get("reimbursementFormNo", "")),
                "projectStartDate": normalize_application_date(application.get("projectStartDate", "")),
                "projectEndDate": normalize_application_date(application.get("projectEndDate", "")),
                "notes": _expired_allowance_note(current["expiredAllowance"], month),
                "roomNames": "、".join(sorted(current["roomNames"])),
            }
        )
    return sorted(rows, key=lambda item: (item["pi"], item["facility"], item["iacuc"]))


def build_monthly_summary_xlsx(month, rows):
    if Workbook is None:
        raise ValueError("当前运行环境缺少 Excel 导出依赖 openpyxl")
    year, month_number = month.split("-")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:O1")
    worksheet["A1"] = f"{year}年{int(month_number)}月动物饲养费汇总"
    worksheet["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    headers = [
        "序号",
        "项目负责人",
        "设施",
        "IACUC编号",
        "伦理对应的经费",
        "动物品系",
        f"{int(month_number)}月1日至{_month_days(month)}日产生的饲养费（元）",
        "单位支持的饲养费（元）",
        "缴纳的饲养费（元）",
        "科研报销单经费本号",
        "科研报销单\n单号",
        "实验开始时间",
        "实验结束时间",
        "备注",
        "房间号",
    ]
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(2, column, value)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[2].height = 36

    widths = [9, 13, 9, 13, 33.5, 16.25, 17, 17, 15, 18.5, 19.25, 13.25, 16, 24, 14]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[_excel_column(index)].width = width

    for index, row in enumerate(rows, start=1):
        excel_row = index + 2
        values = [
            index,
            row["pi"],
            row["facility"],
            row["iacuc"],
            row["funding"],
            row["species"],
            row["amount"],
            row["supportAmount"],
            f"=G{excel_row}-H{excel_row}",
            row["fundBookNo"],
            row["reimbursementFormNo"],
            row["projectStartDate"],
            row["projectEndDate"],
            row["notes"],
            row["roomNames"],
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(excel_row, column, value)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=column in (5, 6, 14, 15))
            if column in (7, 8, 9):
                cell.number_format = "0.00"
        worksheet.row_dimensions[excel_row].height = 24

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:O{max(2, len(rows) + 2)}"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def monthly_summary_filename(month):
    year, month_number = month.split("-")
    return f"{year}年{int(month_number)}月动物饲养费汇总.xlsx"


def _month_days(month):
    year, month_number = (int(value) for value in month.split("-"))
    if month_number == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month_number + 1
    return (date(next_year, next_month, 1) - date(year, month_number, 1)).days


def _excel_column(index):
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def _facility_label(value):
    return FACILITY_LABELS.get(clean_text(value), clean_text(value))


def _species_label(value):
    return SPECIES_LABELS.get(clean_text(value), clean_text(value))


def _expired_allowance_note(items, month):
    month_start = date.fromisoformat(f"{month}-01")
    notes = []
    for iacuc, (expiry, first_ineligible_date) in sorted((items or {}).items()):
        expiry_date = _parse_date(expiry)
        if expiry_date and expiry_date < month_start:
            notes.append(f"{iacuc} 已于 {expiry} 到期，本月不参与减免")
            continue
        next_day = (expiry_date + timedelta(days=1)).isoformat() if expiry_date else first_ineligible_date
        notes.append(f"{iacuc} 于 {expiry} 到期，自 {next_day} 起不参与减免")
    return "；".join(notes)


def _parse_date(value):
    try:
        return date.fromisoformat(clean_text(value))
    except ValueError:
        return None
