"""自动扫描 inbox 中的最新 IACUC 汇总表并导入系统。

数据流：群晖上的汇总表由 scripts/sync_iacuc_summary_nas.sh 复制到
<DATA_ROOT>/inbox/iacuc/，本模块按固定间隔扫描该目录，取最新文件
（xlsx 或 csv），转换为系统 IACUC 索引并入库，成功后移入
<DATA_ROOT>/archive/iacuc/ 防止重复导入。
"""

import csv
import io
import os
import shutil
import threading
import time
from datetime import date, datetime
from pathlib import Path

from server_app.cache import invalidate_data_cache, invalidate_data_cache_prefixes
from server_app.composition import save_iacuc_index_file
from server_app.config import DATA_ROOT
from server_app.db import connect_db
from server_app.domains.administration import audit_event, write_audit_events
from server_app.domains.iacuc.importer import parse_iacuc_csv
from server_app.domains.iacuc.sync import (
    application_payload,
    invalidate_all_quantity_sheet_candidate_snapshots,
    read_current_applications,
    sync_project_derived_fields_after_iacuc_upload,
    write_experiment_applications,
)
from server_app.shared import now_iso

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - deployment dependency validation
    load_workbook = None

DEFAULT_INBOX_DIR = DATA_ROOT / "inbox" / "iacuc"
DEFAULT_ARCHIVE_DIR = DATA_ROOT / "archive" / "iacuc"
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
SYSTEM_ACTOR = {
    "id": "system",
    "username": "system",
    "displayName": "系统自动导入",
    "role": "admin",
    "roomIds": [],
}
# 表头行识别提示：与 importer 必填列保持一致。
HEADER_HINTS = ("动物伦理编号", "动物实验名称", "项目负责人", "实验负责人")


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def xlsx_to_csv_bytes(path):
    """读取 xlsx 第一个工作表并转换为 CSV bytes（utf-8-sig）。"""
    if load_workbook is None:
        raise ValueError("当前运行环境缺少 Excel 读取依赖 openpyxl")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            values = [_cell_text(value) for value in row]
            if any(value.strip() for value in values):
                rows.append(values)
        if not rows:
            raise ValueError(f"汇总表 {path.name} 没有可读取的数据行")
        header_index = 0
        for index, row in enumerate(rows):
            if any(any(hint in value for value in row) for hint in HEADER_HINTS):
                header_index = index
                break
        headers = rows[header_index]
        data = [row for row in rows[header_index + 1 :] if any(value.strip() for value in row)]
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data)
        return output.getvalue().encode("utf-8-sig")
    finally:
        workbook.close()


def import_summary_file(path, *, conn, now, actor=SYSTEM_ACTOR):
    """把单个汇总表文件解析并写入系统，返回导入摘要。"""
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        raw = Path(path).read_bytes()
    else:
        raw = xlsx_to_csv_bytes(path)
    parsed = parse_iacuc_csv(raw)
    if not parsed["items"]:
        raise ValueError(f"汇总表 {Path(path).name} 没有可导入的 IACUC 行")
    file_items = [application_payload(item, now) for item in parsed["items"]]
    old_items = read_current_applications(conn)
    write_experiment_applications(conn, parsed["items"], now)
    save_iacuc_index_file(file_items)
    sync_summary = sync_project_derived_fields_after_iacuc_upload(conn, old_items, file_items, actor, now)
    invalidate_all_quantity_sheet_candidate_snapshots(conn)
    event = audit_event(
        actor,
        "iacuc_index.auto_imported",
        "iacuc_index",
        "iacuc-inbox",
        f"{actor['displayName']} 自动导入 {Path(path).name}：{parsed['summary']['count']} 条 IACUC",
        [],
        now,
        None,
        {
            "source": Path(path).name,
            "rowCount": parsed["summary"]["rowCount"],
            "count": parsed["summary"]["count"],
            "emptyIacucCount": parsed["summary"]["emptyIacucCount"],
            "syncSummary": sync_summary,
            "importedAt": now,
        },
    )
    write_audit_events(conn, [event])
    conn.commit()
    return {"source": Path(path).name, **parsed["summary"], "syncSummary": sync_summary}


def _archive_imported(path, archive_dir, now):
    archive = Path(archive_dir)
    archive.mkdir(parents=True, exist_ok=True)
    timestamp = now.replace(":", "").replace("-", "")[:14]
    target = archive / f"{Path(path).stem}-{timestamp}{Path(path).suffix}"
    shutil.move(str(path), str(target))
    return target


def latest_summary_path(inbox_dir):
    """返回 inbox 中修改时间最新的汇总表文件；目录不存在或为空返回 None。"""
    inbox = Path(inbox_dir)
    if not inbox.is_dir():
        return None
    candidates = [item for item in inbox.iterdir() if item.is_file() and item.suffix.lower() in SUPPORTED_SUFFIXES]
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.stat().st_mtime)


def scan_and_import_once(inbox_dir=DEFAULT_INBOX_DIR, archive_dir=DEFAULT_ARCHIVE_DIR, now=None):
    """扫描 inbox 取最新汇总表并导入，成功返回摘要，无文件返回 None。"""
    latest = latest_summary_path(inbox_dir)
    if latest is None:
        return None
    current = now or now_iso()
    with connect_db() as conn:
        result = import_summary_file(latest, conn=conn, now=current)
        _archive_imported(latest, archive_dir, current)
    invalidate_data_cache("assembled_state", "iacuc_index", "principal_identities", "principal_types_by_pi")
    invalidate_data_cache_prefixes(
        "bootstrap_summary::",
        "billing_occupancies::",
        "quantity_sheets::",
        "billing_workflows::",
        "billing_statements::",
        "reimbursement_records::",
        "intake_batches::",
        "placement_tasks::",
    )
    return result


def auto_import_interval_seconds():
    minutes = int(os.environ.get("CAGELEDGER_IACUC_AUTO_IMPORT_INTERVAL_MINUTES", "5"))
    return max(30, minutes * 60)


def start_auto_import_watcher(*, enabled=None, interval_seconds=None):
    """启动后台扫描线程；重复调用返回 None。"""
    if enabled is None:
        enabled = os.environ.get("CAGELEDGER_IACUC_AUTO_IMPORT_ENABLED", "1").strip().lower() not in {
            "0",
            "false",
            "no",
            "off",
        }
    if not enabled:
        return None
    if interval_seconds is None:
        interval_seconds = auto_import_interval_seconds()

    def loop():
        while True:
            try:
                result = scan_and_import_once()
                if result:
                    print(f"[auto-import] 已导入 {result['source']}：{result['count']} 条 IACUC")
            except Exception as exc:  # 后台扫描不因单次失败退出
                print(f"[auto-import] 导入失败：{exc}")
            time.sleep(interval_seconds)

    thread = threading.Thread(target=loop, name="iacuc-auto-import", daemon=True)
    thread.start()
    return thread
