#!/usr/bin/env python3
import calendar
import hashlib
import io
import json
import re
import sqlite3
import time
from datetime import UTC, date, datetime
from email.utils import format_datetime
from http import HTTPStatus
from http.server import ThreadingHTTPServer
from urllib.parse import parse_qs, unquote, urlparse

try:
    import openpyxl
except ImportError:
    openpyxl = None

from server_app.cache import (
    cache_get,
    cache_key,
    cache_set,
    invalidate_data_cache,
    invalidate_data_cache_prefixes,
    log_perf,
)
from server_app.config import (
    DB_PATH,
    HOST,
    IACUC_INDEX_PATH,
    LEGACY_IACUC_INDEX_PATH,
    PORT,
    SESSION_COOKIE,
    frontend_root,
)
from server_app.db import configure_database, connect_db, ensure_database_ready
from server_app.domains.administration import (
    action_label,
    audit_event,
    authenticate,
    create_session,
    create_user,
    delete_session,
    delete_user,
    ensure_default_admin,
    list_users,
    merge_audit_logs,
    system_info,
    system_update_status,
    update_user,
    user_from_token,
    write_audit_events,
)
from server_app.domains.billing import (
    BILLING_PRINCIPAL_INDEPENDENT,
    BILLING_TIER_BASE_PRICE,
    BILLING_TIER_LIMIT,
    BILLING_TIER_OVER_PRICE,
    add_charge_group,
    allocate_daily_free_cages_by_iacuc,
    apply_free_cage_allocations,
    billing_free_cages_for_pi,
    billing_profile_for_occupancy,
    billing_profile_for_room,
    combined_daily_charge,
    dates_in_month,
    free_cages_for_principal_type,
    iacuc_free_allowance_eligible,
    normalize_principal_type,
    occupancy_active_on_date,
    occupancy_animal_count,
    pi_for_iacuc,
    principal_type_label,
    quantity_sheet_statement_lines,
    statement_application_snapshot,
    statement_billing_unit_from_lines,
    statement_pi_snapshot,
)
from server_app.domains.billing.candidates import (
    list_settlement_candidates,
    update_settlement_candidate_snapshot_from_statement,
)
from server_app.domains.iacuc import (
    normalize_application_amount,
    normalize_application_date,
    normalize_iacuc_number,
    parse_iacuc_csv,
)
from server_app.domains.intake import (
    animal_age_text,
    cage_card_qr_id_from_batch_card,
    cage_card_status_label,
    collect_cage_card_qr_ids,
    is_cage_card_qr_id,
    legacy_cage_card_qr_id,
    next_cage_card_qr_id,
    species_label,
)
from server_app.domains.reimbursement.importer import (
    arrears_summary_columns,
    build_monthly_import_groups,
    distinct_text_list,
    ensure_excel_import_supported,
    next_imported_record_id,
)
from server_app.persistence import SchemaRegistry, SchemaStep
from server_app.persistence.backfills import backfill_quantity_sheet_staff
from server_app.persistence.base_schema import initialize_base_schema
from server_app.persistence.indexes import create_performance_indexes
from server_app.repositories.billing import (
    billing_workflow_detail_item as billing_workflow_detail_item_repository,
)
from server_app.repositories.billing import (
    delete_billing_workflow_tree as delete_billing_workflow_tree_repository,
)
from server_app.repositories.billing import (
    delete_quantity_sheet_by_id as delete_quantity_sheet_by_id_repository,
)
from server_app.repositories.billing import (
    get_billing_version as get_billing_version_repository,
)
from server_app.repositories.billing import (
    get_billing_workflow as get_billing_workflow_repository,
)
from server_app.repositories.billing import (
    get_billing_workflow_by_key as get_billing_workflow_by_key_repository,
)
from server_app.repositories.billing import (
    get_billing_workflow_detail as get_billing_workflow_detail_repository,
)
from server_app.repositories.billing import (
    get_current_billing_statement as get_current_billing_statement_repository,
)
from server_app.repositories.billing import (
    get_quantity_sheet as get_quantity_sheet_repository,
)
from server_app.repositories.billing import (
    insert_billing_version as insert_billing_version_repository,
)
from server_app.repositories.billing import (
    insert_billing_workflow as insert_billing_workflow_repository,
)
from server_app.repositories.billing import (
    insert_billing_workflow_event as insert_billing_workflow_event_repository,
)
from server_app.repositories.billing import (
    insert_quantity_sheet as insert_quantity_sheet_repository,
)
from server_app.repositories.billing import (
    list_billing_statement_line_summaries_for_version as list_billing_statement_line_summaries_for_version_repository,
)
from server_app.repositories.billing import (
    list_billing_statement_lines_for_version as list_billing_statement_lines_for_version_repository,
)
from server_app.repositories.billing import (
    list_billing_workflow_events as list_billing_workflow_events_repository,
)
from server_app.repositories.billing import (
    list_billing_workflow_versions as list_billing_workflow_versions_repository,
)
from server_app.repositories.billing import (
    list_billing_workflows as list_billing_workflows_repository,
)
from server_app.repositories.billing import (
    list_billing_workflows_page as list_billing_workflows_page_repository,
)
from server_app.repositories.billing import (
    list_current_billing_statements as list_current_billing_statements_repository,
)
from server_app.repositories.billing import (
    list_quantity_sheet_filter_options as list_quantity_sheet_filter_options_repository,
)
from server_app.repositories.billing import (
    list_quantity_sheets as list_quantity_sheets_repository,
)
from server_app.repositories.billing import (
    list_quantity_sheets_by_month_iacuc as list_quantity_sheets_by_month_iacuc_repository,
)
from server_app.repositories.billing import (
    list_quantity_sheets_by_month_pi as list_quantity_sheets_by_month_pi_repository,
)
from server_app.repositories.billing import (
    list_quantity_sheets_page as list_quantity_sheets_page_repository,
)
from server_app.repositories.billing import (
    replace_billing_statement_version_lines as replace_billing_statement_version_lines_repository,
)
from server_app.repositories.billing import (
    select_quantity_sheets_for_transfer as select_quantity_sheets_for_transfer_repository,
)
from server_app.repositories.billing import (
    update_billing_version as update_billing_version_repository,
)
from server_app.repositories.billing import (
    update_billing_workflow as update_billing_workflow_repository,
)
from server_app.repositories.billing import (
    update_quantity_sheet as update_quantity_sheet_repository,
)
from server_app.repositories.billing_candidates import (
    mark_all_billing_candidate_snapshots_stale,
    mark_billing_candidate_snapshots_stale,
    mark_billing_candidate_snapshots_stale_by_pi,
)
from server_app.repositories.entities import (
    delete_intake_batch as delete_intake_batch_repository,
)
from server_app.repositories.entities import (
    delete_placement_task as delete_placement_task_repository,
)
from server_app.repositories.entities import (
    list_audit_events_page,
    list_distinct_principal_names,
    list_intake_batch_filter_options,
    list_intake_batches_page,
    list_placement_tasks_page,
    read_principal_identity_payloads,
    upsert_principal_identity,
)
from server_app.repositories.entities import (
    read_principal_type_by_pi as read_principal_type_by_pi_repository,
)
from server_app.repositories.entities import (
    upsert_intake_batch as upsert_intake_batch_repository,
)
from server_app.repositories.entities import (
    upsert_placement_task as upsert_placement_task_repository,
)
from server_app.repositories.iacuc import read_iacuc_index as read_iacuc_index_repository
from server_app.repositories.iacuc import (
    replace_experiment_applications,
)
from server_app.repositories.iacuc import (
    save_iacuc_index_file as save_iacuc_index_file_repository,
)
from server_app.repositories.infrastructure import (
    delete_rack_record,
    delete_room_record,
    delete_slot_record,
    insert_rack_record,
    insert_room_record,
    insert_slot_record,
    update_rack_record,
    update_room_record,
)
from server_app.repositories.payload import (
    dump_json,
    read_payloads,
    read_setting,
    read_updated_at,
    set_setting,
    table_has_rows,
)
from server_app.repositories.reimbursement import (
    delete_reimbursement_record as delete_reimbursement_record_repository,
)
from server_app.repositories.reimbursement import (
    get_reimbursement_record as get_reimbursement_record_repository,
)
from server_app.repositories.reimbursement import (
    get_reimbursement_record_by_key as get_reimbursement_record_by_key_repository,
)
from server_app.repositories.reimbursement import (
    get_reimbursement_record_by_workflow_id as get_reimbursement_record_by_workflow_id_repository,
)
from server_app.repositories.reimbursement import (
    list_reimbursement_record_summaries_for_pi as list_reimbursement_record_summaries_for_pi_repository,
)
from server_app.repositories.reimbursement import (
    list_reimbursement_records_for_pi as list_reimbursement_records_for_pi_repository,
)
from server_app.repositories.reimbursement import (
    list_reimbursement_records_page as list_reimbursement_records_page_repository,
)
from server_app.repositories.reimbursement import (
    reimbursement_record_list_item,
)
from server_app.repositories.reimbursement import (
    upsert_reimbursement_record as upsert_reimbursement_record_repository,
)
from server_app.repositories.state import (
    assemble_state as assemble_state_repository,
)
from server_app.repositories.state import (
    read_applications_by_iacuc as read_applications_by_iacuc_repository,
)
from server_app.repositories.state import (
    read_cached_state as read_cached_state_repository,
)
from server_app.services.billing import (
    save_billing_statement_workflow as save_billing_statement_workflow_service,
)
from server_app.services.billing import (
    update_workflow_status as update_workflow_status_service,
)
from server_app.services.intake import confirm_intake_receipt as confirm_intake_receipt_service
from server_app.services.placement import (
    move_in_placement_task as move_in_placement_task_service,
)
from server_app.services.placement import (
    reassign_placement_task_room as reassign_placement_task_room_service,
)
from server_app.services.placement import (
    reserve_placement_task as reserve_placement_task_service,
)
from server_app.services.placement import (
    sync_slot_statuses,
)
from server_app.services.quantity import sync_quantity_sheet_transfer_rows as sync_quantity_sheet_transfer_rows_service
from server_app.services.reimbursement import (
    REIMBURSEMENT_STATUS_COMPLETED,
    REIMBURSEMENT_STATUS_PENDING,
    REIMBURSEMENT_STATUS_REIMBURSING,
    infer_import_status,
    merge_reimbursement_edit,
    normalize_reimbursement_status,
    reimbursement_business_key,
    reimbursement_has_manual_entry,
    summarize_statement,
)
from server_app.services.reimbursement import (
    coerce_money as coerce_reimbursement_money,
)
from server_app.shared import as_float, as_int, clean_text, new_id, now_iso, today_iso
from server_app.static import send_frontend_asset
from server_app.web import CageLedgerHttpHandler, JsonResponse, Router

WORKFLOW_STATUS_IN_FEEDING = "in_feeding"
WORKFLOW_STATUS_GENERATED = "statement_generated"
WORKFLOW_STATUS_SENT = "statement_sent"
WORKFLOW_STATUS_SIGNED = "statement_signed_returned"
WORKFLOW_STATUS_FINANCE = "submitted_to_finance"
REIMBURSEMENT_MIGRATION_KEY = "reimbursementRecordMigrationDone"
WORKFLOW_STATUSES = (
    WORKFLOW_STATUS_IN_FEEDING,
    WORKFLOW_STATUS_GENERATED,
    WORKFLOW_STATUS_SENT,
    WORKFLOW_STATUS_SIGNED,
    WORKFLOW_STATUS_FINANCE,
)
VERSION_STATUS_ACTIVE = "active"
VERSION_STATUS_VOIDED = "voided"
BILLING_WORKFLOW_MIGRATION_KEY = "billingWorkflowMigrationDone"
TABLES = (
    "audit_logs",
    "billing_adjustments",
    "billing_rules",
    "occupancies",
    "placement_tasks",
    "cage_slots",
    "intake_batches",
    "racks",
    "rooms",
    "app_settings",
)
ENTITY_ENDPOINTS = {
    "/api/rooms": "rooms",
    "/api/racks": "racks",
    "/api/cage-slots": "cage_slots",
    "/api/occupancies": "occupancies",
    "/api/placement-tasks": "placement_tasks",
    "/api/billing-rules": "billing_rules",
    "/api/billing-adjustments": "billing_adjustments",
    "/api/intake-batches": "intake_batches",
    "/api/experiment-applications": "experiment_applications",
    "/api/billing-statements": "billing_statements",
    "/api/billing-statement-lines": "billing_statement_lines",
    "/api/audit-events": "audit_events",
}
ENTITY_ORDER_BY = {
    "rooms": "rowid",
    "racks": "room_id, index_no, rowid",
    "cage_slots": "rack_id, row_no, col_no, rowid",
    "occupancies": "start_date, rowid",
    "placement_tasks": "planned_move_in_date, rowid",
    "billing_rules": "rowid",
    "billing_adjustments": "rowid",
    "intake_batches": "updated_at DESC, rowid DESC",
    "experiment_applications": "rowid",
    "billing_statements": "month DESC, iacuc, rowid DESC",
    "billing_statement_lines": "statement_id, line_date, rowid",
}
WRITABLE_ENTITY_ENDPOINTS = {
    "/api/rooms": {"collection": "rooms", "id_prefix": "room"},
    "/api/racks": {"collection": "racks", "id_prefix": "rack"},
    "/api/cage-slots": {"collection": "slots", "id_prefix": "slot"},
    "/api/occupancies": {"collection": "occupancies", "id_prefix": "occ"},
    "/api/placement-tasks": {"collection": "placementTasks", "id_prefix": "ptask"},
    "/api/billing-rules": {"collection": "billingRules", "id_prefix": "rule"},
    "/api/billing-adjustments": {"collection": "adjustments", "id_prefix": "adj"},
    "/api/intake-batches": {"collection": "intakeBatches", "id_prefix": "batch"},
}


def initialize_legacy_schema(conn):
    initialize_base_schema(
        conn,
        migrate_schema=migrate_schema,
        repair_missing_cage_slots=repair_missing_cage_slots,
        create_performance_indexes=create_performance_indexes,
        ensure_default_admin=ensure_default_admin,
    )


def initialize_schema(conn):
    registry = SchemaRegistry([SchemaStep("legacy-schema", initialize_legacy_schema)])
    registry.apply(conn)


def migrate_schema(conn):
    ensure_experiment_applications_duplicate_schema(conn)
    ensure_occupancies_history_schema(conn)
    ensure_occupancies_structured_columns(conn)
    ensure_intake_batch_structured_columns(conn)
    migrate_billing_workflow_schema(conn)
    backfill_billing_workflow_scope(conn)
    migrate_reimbursement_record_schema(conn)
    backfill_quantity_sheet_staff(conn)


def ensure_experiment_applications_duplicate_schema(conn):
    columns = table_columns(conn, "experiment_applications")
    pk_column = next((name for name, info in columns.items() if info.get("pk")), "")
    if "id" in columns and pk_column == "id":
        return

    conn.execute("ALTER TABLE experiment_applications RENAME TO experiment_applications_legacy")
    conn.execute(
        """
        CREATE TABLE experiment_applications (
            id TEXT PRIMARY KEY,
            iacuc TEXT NOT NULL,
            raw_iacuc TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            imported_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    rows = conn.execute("SELECT rowid, * FROM experiment_applications_legacy ORDER BY rowid").fetchall()
    for index, row in enumerate(rows, start=1):
        payload = json.loads(row["payload"])
        raw_iacuc = clean_text(payload.get("rawIacuc", "") or row["raw_iacuc"] or row["iacuc"])
        item = {
            **payload,
            "id": payload.get("id") or f"app-{index:06d}",
            "iacuc": clean_text(payload.get("iacuc", "") or raw_iacuc or row["iacuc"]),
            "rawIacuc": raw_iacuc,
        }
        conn.execute(
            """
            INSERT INTO experiment_applications (
                id, iacuc, raw_iacuc, project, pi, owner, funding, imported_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item["id"],
                item["iacuc"],
                item.get("rawIacuc", ""),
                item.get("project", ""),
                item.get("pi", ""),
                item.get("owner", ""),
                item.get("funding", ""),
                row["imported_at"],
                dump_json(item),
            ),
        )
    conn.execute("DROP TABLE experiment_applications_legacy")


def ensure_occupancies_structured_columns(conn):
    columns = table_columns(conn, "occupancies")
    additions = {
        "room_id": "TEXT",
        "rack_id": "TEXT",
        "species": "TEXT",
        "billing_item": "TEXT",
        "customer_type": "TEXT",
        "animal_count": "INTEGER",
    }
    changed = False
    for column, column_type in additions.items():
        if column in columns:
            continue
        conn.execute(f"ALTER TABLE occupancies ADD COLUMN {column} {column_type}")
        changed = True
    if changed:
        backfill_occupancy_structured_columns(conn)


def backfill_occupancy_structured_columns(conn):
    state = assemble_state(conn) or empty_state()
    applications_by_iacuc = read_applications_by_iacuc(conn)
    for occupancy in state.get("occupancies", []):
        normalized = occupancy_with_snapshots(occupancy, state, applications_by_iacuc)
        values = occupancy_structured_values(normalized)
        conn.execute(
            """
            UPDATE occupancies
            SET room_id = ?, rack_id = ?, species = ?, billing_item = ?,
                customer_type = ?, animal_count = ?, payload = ?
            WHERE id = ?
            """,
            (
                values["room_id"],
                values["rack_id"],
                values["species"],
                values["billing_item"],
                values["customer_type"],
                values["animal_count"],
                dump_json(normalized),
                normalized.get("id"),
            ),
        )


def ensure_intake_batch_structured_columns(conn):
    columns = table_columns(conn, "intake_batches")
    additions = {
        "pi": "TEXT",
        "owner": "TEXT",
        "quantity": "INTEGER",
        "card_count": "INTEGER",
    }
    changed = False
    for column, column_type in additions.items():
        if column in columns:
            continue
        conn.execute(f"ALTER TABLE intake_batches ADD COLUMN {column} {column_type}")
        changed = True
    if changed:
        backfill_intake_batch_structured_columns(conn)


def backfill_intake_batch_structured_columns(conn):
    rows = conn.execute("SELECT id, payload FROM intake_batches").fetchall()
    for row in rows:
        payload = json.loads(row["payload"])
        conn.execute(
            """
            UPDATE intake_batches
            SET pi = ?, owner = ?, quantity = ?, card_count = ?
            WHERE id = ?
            """,
            (
                clean_text(payload.get("pi", "")),
                clean_text(payload.get("owner", "")),
                as_int(payload.get("quantity")) or 0,
                as_int(payload.get("finalCardCount")) or 0,
                row["id"],
            ),
        )


def ensure_occupancies_history_schema(conn):
    columns = table_columns(conn, "occupancies")
    foreign_keys = conn.execute("PRAGMA foreign_key_list(occupancies)").fetchall()
    slot_not_null = bool(columns.get("slot_id", {}).get("notnull"))
    required_columns = {"funding", "room_name", "rack_name", "slot_code"}
    if not foreign_keys and not slot_not_null and required_columns.issubset(columns):
        return

    conn.execute("ALTER TABLE occupancies RENAME TO occupancies_legacy")
    conn.execute(
        """
        CREATE TABLE occupancies (
            id TEXT PRIMARY KEY,
            slot_id TEXT,
            room_id TEXT,
            rack_id TEXT,
            cage_code TEXT,
            status TEXT NOT NULL,
            iacuc TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            species TEXT,
            billing_item TEXT,
            customer_type TEXT,
            animal_count INTEGER,
            room_name TEXT,
            rack_name TEXT,
            slot_code TEXT,
            start_date TEXT,
            end_date TEXT,
            end_reason TEXT,
            notes TEXT,
            updated_at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    rows = conn.execute("SELECT * FROM occupancies_legacy").fetchall()
    for row in rows:
        payload = json.loads(row["payload"])
        conn.execute(
            """
            INSERT INTO occupancies (
                id, slot_id, room_id, rack_id, cage_code, status, iacuc, project, pi, owner, funding,
                species, billing_item, customer_type, animal_count,
                room_name, rack_name, slot_code, start_date, end_date, end_reason,
                notes, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["id"],
                row["slot_id"],
                payload.get("roomId", ""),
                payload.get("rackId", ""),
                row["cage_code"],
                row["status"],
                row["iacuc"],
                row["project"],
                row["pi"],
                row["owner"],
                payload.get("funding", ""),
                payload.get("species", ""),
                payload.get("billingItem", ""),
                payload.get("customerType", ""),
                as_int(payload.get("animalCount")),
                payload.get("roomName", ""),
                payload.get("rackName", ""),
                payload.get("slotCode", ""),
                row["start_date"],
                row["end_date"],
                row["end_reason"],
                row["notes"],
                row["updated_at"],
                dump_json(payload),
            ),
        )
    conn.execute("DROP TABLE occupancies_legacy")


def migrate_billing_workflow_schema(conn):
    if read_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, False):
        return
    if table_has_rows(conn, "billing_workflows"):
        set_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, True, now_iso())
        return
    if not table_has_rows(conn, "billing_statements"):
        set_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, True, now_iso())
        return

    rows = conn.execute(
        "SELECT id, payload, generated_at, rowid FROM billing_statements ORDER BY month, iacuc, generated_at, rowid"
    ).fetchall()
    grouped = {}
    for row in rows:
        statement = json.loads(row["payload"])
        source_type = normalize_workflow_source(statement.get("sourceType", ""))
        scope_type, scope_key = workflow_scope_for_statement(statement)
        business_key = billing_workflow_business_key(scope_type, scope_key, statement.get("month", ""), source_type)
        grouped.setdefault(business_key, []).append((row, statement, source_type))

    for items in grouped.values():
        first_statement = items[0][1]
        source_type = items[0][2]
        workflow_id = new_id("bwf")
        conn.execute(
            """
            INSERT INTO billing_workflows (
                id, business_key, iacuc, month, source_type, workflow_status,
                current_version_id, current_version_no, latest_event_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                workflow_id,
                billing_workflow_business_key(
                    *workflow_scope_for_statement(first_statement), first_statement.get("month", ""), source_type
                ),
                first_statement.get("iacuc", ""),
                first_statement.get("month", ""),
                source_type,
                WORKFLOW_STATUS_GENERATED,
                "",
                0,
                "",
                dump_json({"id": workflow_id, "migrationPending": True}),
            ),
        )
        versions = []
        latest_at = ""
        for index, (row, statement, grouped_source) in enumerate(items, start=1):  # noqa: B007
            version_id = row["id"]
            generated_at = statement.get("generatedAt") or row["generated_at"] or now_iso()
            lines = [
                json.loads(line_row["payload"])
                for line_row in conn.execute(
                    "SELECT payload FROM billing_statement_lines WHERE statement_id = ? ORDER BY line_date, rowid",
                    (row["id"],),
                ).fetchall()
            ]
            version_status = VERSION_STATUS_ACTIVE if index == len(items) else VERSION_STATUS_VOIDED
            workflow_status = WORKFLOW_STATUS_GENERATED
            document_number = statement.get("documentNumber") or make_statement_document_number(statement, index)
            enriched = enrich_statement_for_workflow(
                statement,
                workflow_id=workflow_id,
                version_id=version_id,
                version_no=index,
                version_status=version_status,
                workflow_status=workflow_status,
                document_number=document_number,
            )
            version_payload = build_version_payload(
                enriched,
                workflow_id,
                index,
                version_status,
                workflow_status,
                generated_at,
                "",
                "",
                "",
            )
            conn.execute(
                """
                INSERT INTO billing_statement_versions (
                    id, workflow_id, version_no, version_status, workflow_status,
                    generated_at, voided_at, created_by, payload
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    version_id,
                    workflow_id,
                    index,
                    version_status,
                    workflow_status,
                    generated_at,
                    generated_at if version_status == VERSION_STATUS_VOIDED else "",
                    "",
                    dump_json(version_payload),
                ),
            )
            for line in lines:
                conn.execute(
                    """
                    INSERT INTO billing_statement_version_lines (id, version_id, line_date, payload)
                    VALUES (?, ?, ?, ?)
                    """,
                    (line.get("id") or new_id("line"), version_id, line.get("date", ""), dump_json(line)),
                )
            event_at = generated_at
            event_payload = build_workflow_event_payload(
                new_id("wevt"),
                workflow_id,
                version_id,
                "statement_generated",
                "",
                WORKFLOW_STATUS_GENERATED,
                {
                    "displayName": "系统迁移",
                    "username": "system",
                    "id": "system",
                },
                event_at,
                "system",
                "由旧版 billing_statements 迁移生成",
            )
            conn.execute(
                """
                INSERT INTO billing_workflow_events (
                    id, workflow_id, version_id, event_type, from_status, to_status, at, payload
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_payload["id"],
                    workflow_id,
                    version_id,
                    event_payload["eventType"],
                    event_payload["fromStatus"],
                    event_payload["toStatus"],
                    event_payload["at"],
                    dump_json(event_payload),
                ),
            )
            versions.append(version_payload)
            latest_at = max(latest_at, generated_at)

        current_version = versions[-1]
        workflow_payload = build_workflow_payload(
            workflow_id,
            first_statement.get("iacuc", ""),
            first_statement.get("month", ""),
            source_type,
            WORKFLOW_STATUS_GENERATED,
            current_version,
            latest_at,
        )
        conn.execute(
            """
            UPDATE billing_workflows
            SET business_key = ?, iacuc = ?, month = ?, source_type = ?,
                workflow_status = ?, current_version_id = ?, current_version_no = ?,
                latest_event_at = ?, payload = ?
            WHERE id = ?
            """,
            (
                workflow_payload.get("businessKey", ""),
                first_statement.get("iacuc", ""),
                first_statement.get("month", ""),
                grouped_source,
                WORKFLOW_STATUS_GENERATED,
                current_version["id"],
                current_version["versionNo"],
                latest_at,
                dump_json(workflow_payload),
                workflow_id,
            ),
        )
    set_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, True, now_iso())


def backfill_billing_workflow_scope(conn):
    rows = conn.execute("SELECT id, payload FROM billing_workflows").fetchall()
    for row in rows:
        workflow = json.loads(row["payload"])
        current_version = workflow.get("currentVersion") or {}
        statement = current_version.get("statement") or {}
        scope_type, scope_key = workflow_scope_for_statement(statement)
        desired_key = billing_workflow_business_key(
            scope_type, scope_key, workflow.get("month", ""), workflow.get("sourceType", "")
        )
        conflict = conn.execute(
            "SELECT id FROM billing_workflows WHERE business_key = ? AND id != ?",
            (desired_key, row["id"]),
        ).fetchone()
        if conflict:
            desired_key = f"{desired_key}|legacy|{row['id']}"
        if (
            workflow.get("scopeType") == scope_type
            and workflow.get("scopeKey") == scope_key
            and desired_key == workflow.get("businessKey")
        ):
            continue
        workflow["scopeType"] = scope_type
        workflow["scopeKey"] = scope_key
        workflow["businessKey"] = desired_key
        workflow["iacucs"] = statement.get("iacucs", workflow.get("iacucs", []))
        conn.execute(
            """
            UPDATE billing_workflows
            SET business_key = ?, payload = ?
            WHERE id = ?
            """,
            (desired_key, dump_json(workflow), row["id"]),
        )


def table_columns(conn, table):
    return {
        row["name"]: {"type": row["type"], "notnull": bool(row["notnull"]), "pk": int(row["pk"] or 0)}
        for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
    }


def repair_missing_cage_slots(conn):
    racks = conn.execute("SELECT id, rows, cols FROM racks").fetchall()
    if not racks:
        return

    existing_rows = conn.execute("SELECT rack_id, row_no, col_no FROM cage_slots").fetchall()
    existing_positions = {(row["rack_id"], int(row["row_no"] or 0), int(row["col_no"] or 0)) for row in existing_rows}
    for rack in racks:
        rows = max(as_int(rack["rows"]) or 0, 0)
        cols = max(as_int(rack["cols"]) or 0, 0)
        for row_no in range(1, rows + 1):
            for col_no in range(1, cols + 1):
                position = (rack["id"], row_no, col_no)
                if position in existing_positions:
                    continue
                slot = {
                    "id": slot_id_for_rack(rack["id"], row_no, col_no),
                    "rackId": rack["id"],
                    "row": row_no,
                    "col": col_no,
                    "code": f"{column_label(col_no)}{row_no}",
                    "status": "empty",
                }
                conn.execute(
                    """
                    INSERT INTO cage_slots (id, rack_id, row_no, col_no, code, status, payload)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        slot["id"],
                        slot["rackId"],
                        slot["row"],
                        slot["col"],
                        slot["code"],
                        slot["status"],
                        dump_json(slot),
                    ),
                )
                existing_positions.add(position)


def slot_id_for_rack(rack_id, row_no, col_no):
    suffix = str(rack_id).removeprefix("rack-")
    return f"slot-{suffix}-{row_no}-{col_no}"


def column_label(index):
    value = int(index)
    label = ""
    while value > 0:
        value -= 1
        label = chr(65 + (value % 26)) + label
        value //= 26
    return label or "A"


def read_state():
    with connect_db() as conn:
        migrate_legacy_state(conn)
        state = assemble_state(conn)
        updated_at = read_updated_at(conn)
    if not state:
        return {"state": None, "updatedAt": None}
    return {"state": state, "updatedAt": updated_at}


def current_occupancy_by_slot(state):
    current = {}
    for item in state.get("occupancies", []):
        if item.get("slotId") and item.get("status") in ("active", "reserved"):
            current[item.get("slotId")] = item
    return current


def public_cage_card_payload(conn, qr_id):
    target = clean_text(qr_id).upper()
    if not target:
        raise LookupError("二维码地址无效")
    state = assemble_state(conn) or empty_state()
    applications_by_iacuc = read_applications_by_iacuc(conn)
    rooms_by_id = {item.get("id"): item for item in state.get("rooms", [])}
    racks_by_id = {item.get("id"): item for item in state.get("racks", [])}
    slots_by_id = {item.get("id"): item for item in state.get("slots", [])}
    tasks_by_batch_and_sequence = {
        (task.get("sourceBatchId"), as_int(task.get("cardSequence")) or 0): task
        for task in state.get("placementTasks", [])
    }
    occupancies_by_id = {item.get("id"): item for item in state.get("occupancies", [])}
    occupancies_by_qr = {
        clean_text(item.get("qrId")).upper(): item
        for item in state.get("occupancies", [])
        if clean_text(item.get("qrId"))
    }

    for batch in state.get("intakeBatches", []):
        card_count = max(
            as_int(batch.get("finalCardCount")) or 0,
            as_int(batch.get("suggestedCardCount")) or 0,
            len(batch.get("cards") or []) if isinstance(batch.get("cards"), list) else 0,
        )
        for sequence in range(1, card_count + 1):
            task = tasks_by_batch_and_sequence.get((batch.get("id"), sequence))
            candidate_ids = {
                cage_card_qr_id_from_batch_card(batch, sequence).upper(),
                legacy_cage_card_qr_id(batch, sequence).upper(),
            }
            cards = batch.get("cards") if isinstance(batch.get("cards"), list) else []
            card_index = sequence - 1
            if 0 <= card_index < len(cards) and isinstance(cards[card_index], dict):
                stored_qr_id = clean_text(cards[card_index].get("qrId")).upper()
                if stored_qr_id:
                    candidate_ids.add(stored_qr_id)
            if task and clean_text(task.get("qrId")):
                candidate_ids.add(clean_text(task.get("qrId")).upper())
            if target not in candidate_ids:
                continue
            occupancy = occupancies_by_qr.get(target)
            if not occupancy and task and task.get("reservedOccupancyId"):
                occupancy = occupancies_by_id.get(task.get("reservedOccupancyId"))
            slot = slots_by_id.get((occupancy or {}).get("slotId"))
            rack = racks_by_id.get((slot or {}).get("rackId") or (occupancy or {}).get("rackId"))
            room = rooms_by_id.get(
                (rack or {}).get("roomId") or (task or {}).get("targetRoomId") or (occupancy or {}).get("roomId")
            )
            iacuc = normalize_iacuc_number(
                batch.get("iacuc") or (task or {}).get("iacuc") or (occupancy or {}).get("iacuc")
            )
            application = applications_by_iacuc.get(iacuc, {})
            animal_count = (task or {}).get("animalCount")
            if animal_count in (None, ""):
                animal_count = (occupancy or {}).get("animalCount")
            if animal_count in (None, ""):
                per_cage = max(as_int(batch.get("suggestedAnimalsPerCage")) or 1, 1)
                quantity = max(as_int(batch.get("quantity")) or 0, 0)
                remainder = quantity % per_cage if quantity and per_cage else 0
                animal_count = "" if remainder and sequence == card_count else per_cage
            birth_date = (occupancy or {}).get("birthDate", "")
            item = {
                "qrId": target,
                "batchNo": batch.get("batchNo", ""),
                "cageCode": (occupancy or {}).get("cageCode", ""),
                "roomName": (room or {}).get("name")
                or (task or {}).get("targetRoomName")
                or batch.get("roomName", "")
                or (occupancy or {}).get("roomName", ""),
                "rackName": (rack or {}).get("name") or (occupancy or {}).get("rackName", ""),
                "slotCode": (slot or {}).get("code") or (occupancy or {}).get("slotCode", ""),
                "iacuc": iacuc,
                "project": batch.get("project")
                or (task or {}).get("project")
                or (occupancy or {}).get("project")
                or application.get("project", ""),
                "pi": batch.get("pi")
                or (task or {}).get("pi")
                or (occupancy or {}).get("pi")
                or application.get("pi", ""),
                "owner": batch.get("owner")
                or (task or {}).get("owner")
                or (occupancy or {}).get("owner")
                or application.get("owner", ""),
                "species": batch.get("species") or (task or {}).get("species") or (occupancy or {}).get("species", ""),
                "speciesLabel": species_label(
                    batch.get("species") or (task or {}).get("species") or (occupancy or {}).get("species", "")
                ),
                "strainStandard": batch.get("strainStandard")
                or (task or {}).get("strainStandard")
                or (occupancy or {}).get("strainStandard", ""),
                "animalCount": animal_count,
                "sex": batch.get("sex") or (occupancy or {}).get("sex", ""),
                "birthDate": birth_date,
                "age": animal_age_text(birth_date),
                "startDate": (occupancy or {}).get("startDate") or (task or {}).get("actualMoveInDate", ""),
                "actualMoveInDate": (task or {}).get("actualMoveInDate", ""),
                "endDate": (occupancy or {}).get("endDate") or batch.get("endDate", ""),
                "statusLabel": cage_card_status_label(batch, task, occupancy),
            }
            return {"item": item}
    raise LookupError("该二维码没有匹配到笼卡记录")


def occupancy_period_tone(occupancy):
    if not occupancy or occupancy.get("status") != "active":
        return ""
    end_date = clean_text(occupancy.get("endDate", ""))
    if not end_date:
        return "open"
    return "overdue" if today_iso() > end_date else "normal"


def summarize_infrastructure(state):
    current_by_slot = current_occupancy_by_slot(state)
    rack_by_id = {rack.get("id"): rack for rack in state.get("racks", [])}
    room_by_id = {room.get("id"): room for room in state.get("rooms", [])}
    slot_to_rack_id = {slot.get("id"): slot.get("rackId") for slot in state.get("slots", [])}
    facility_summaries = {}
    current_month = date.today().strftime("%Y-%m")
    room_summaries = {
        room.get("id"): {
            "roomId": room.get("id"),
            "rackCount": 0,
            "slotCount": 0,
            "activeCount": 0,
            "reservedCount": 0,
            "emptyCount": 0,
            "periodOpenCount": 0,
            "periodNormalCount": 0,
            "periodOverdueCount": 0,
            "occupancyRecordCount": 0,
        }
        for room in state.get("rooms", [])
    }
    rack_summaries = {
        rack.get("id"): {
            "rackId": rack.get("id"),
            "roomId": rack.get("roomId"),
            "slotCount": 0,
            "activeCount": 0,
            "reservedCount": 0,
            "emptyCount": 0,
            "periodOpenCount": 0,
            "periodNormalCount": 0,
            "periodOverdueCount": 0,
            "occupancyRecordCount": 0,
        }
        for rack in state.get("racks", [])
    }
    dashboard = {
        "total": 0,
        "active": 0,
        "reserved": 0,
        "empty": 0,
        "periodOpen": 0,
        "periodNormal": 0,
        "periodOverdue": 0,
        "intakePendingCount": 0,
        "openPlacementTaskCount": 0,
        "currentMonthWorkflowTodoCount": 0,
        "currentMonthWorkflowDoneCount": 0,
        "unmatchedIntakeCount": 0,
        "overduePlacementCount": 0,
        "stalledWorkflowCount": 0,
        "exceptionCount": 0,
    }

    def ensure_facility_summary(facility):
        key = clean_text(facility or "zhujiang") or "zhujiang"
        if key not in facility_summaries:
            facility_summaries[key] = {
                "facility": key,
                "roomCount": 0,
                "activeCageCount": 0,
                "activeAnimalCount": 0,
                "openPlacementTaskCount": 0,
                "currentMonthWorkflowTodoCount": 0,
                "currentMonthWorkflowDoneCount": 0,
            }
        return facility_summaries[key]

    for room in state.get("rooms", []):
        ensure_facility_summary(room.get("facility"))["roomCount"] += 1

    for rack in state.get("racks", []):
        room_summary = room_summaries.get(rack.get("roomId"))
        if room_summary:
            room_summary["rackCount"] += 1

    for slot in state.get("slots", []):
        rack_summary = rack_summaries.get(slot.get("rackId"))
        rack = rack_by_id.get(slot.get("rackId"), {})
        room_summary = room_summaries.get(rack.get("roomId"))
        occupancy = current_by_slot.get(slot.get("id"))
        tone = occupancy_period_tone(occupancy)
        status = clean_text(slot.get("status", ""))
        dashboard["total"] += 1
        if rack_summary:
            rack_summary["slotCount"] += 1
        if room_summary:
            room_summary["slotCount"] += 1
        if status == "active":
            dashboard["active"] += 1
            ensure_facility_summary(
                (room_summary and room_by_id.get(room_summary["roomId"], {}).get("facility")) or rack.get("facility")
            )["activeCageCount"] += 1
            if rack_summary:
                rack_summary["activeCount"] += 1
            if room_summary:
                room_summary["activeCount"] += 1
        elif status == "reserved":
            dashboard["reserved"] += 1
            if rack_summary:
                rack_summary["reservedCount"] += 1
            if room_summary:
                room_summary["reservedCount"] += 1
        else:
            dashboard["empty"] += 1
            if rack_summary:
                rack_summary["emptyCount"] += 1
            if room_summary:
                room_summary["emptyCount"] += 1
        if tone == "open":
            dashboard["periodOpen"] += 1
            if rack_summary:
                rack_summary["periodOpenCount"] += 1
            if room_summary:
                room_summary["periodOpenCount"] += 1
        elif tone == "normal":
            dashboard["periodNormal"] += 1
            if rack_summary:
                rack_summary["periodNormalCount"] += 1
            if room_summary:
                room_summary["periodNormalCount"] += 1
        elif tone == "overdue":
            dashboard["periodOverdue"] += 1
            if rack_summary:
                rack_summary["periodOverdueCount"] += 1
            if room_summary:
                room_summary["periodOverdueCount"] += 1

    for item in state.get("occupancies", []):
        slot_id = item.get("slotId")
        rack_id = item.get("rackId") or slot_to_rack_id.get(slot_id)
        room_id = item.get("roomId") or rack_by_id.get(rack_id, {}).get("roomId")
        room = room_by_id.get(room_id, {})
        if item.get("status") == "active":
            facility_profile = billing_profile_for_room(room)
            ensure_facility_summary(facility_profile.get("facility"))["activeAnimalCount"] += occupancy_animal_count(
                item, facility_profile
            )
        if rack_id and rack_id in rack_summaries:
            rack_summaries[rack_id]["occupancyRecordCount"] += 1
        if room_id and room_id in room_summaries:
            room_summaries[room_id]["occupancyRecordCount"] += 1

    for batch in state.get("intakeBatches", []):
        if clean_text(batch.get("status")) != "received":
            dashboard["intakePendingCount"] += 1
        if clean_text(batch.get("roomName")) and not batch.get("roomMatched"):
            dashboard["unmatchedIntakeCount"] += 1

    for task in state.get("placementTasks", []):
        status = clean_text(task.get("status"))
        if status in ("active", "cancelled"):
            continue
        room = room_by_id.get(task.get("targetRoomId"), {})
        ensure_facility_summary(billing_profile_for_room(room).get("facility"))["openPlacementTaskCount"] += 1
        dashboard["openPlacementTaskCount"] += 1
        planned_move_in = clean_text(task.get("plannedMoveInDate"))
        if planned_move_in and planned_move_in < today_iso():
            dashboard["overduePlacementCount"] += 1

    for workflow in state.get("billingWorkflows", []):
        if clean_text(workflow.get("month")) != current_month:
            continue
        workflow_status = clean_text(workflow.get("workflowStatus"))
        facility_keys = set()
        statement = (workflow.get("currentVersion") or {}).get("statement") or {}
        room_name = clean_text(statement.get("roomName"))
        if room_name:
            room = next((item for item in state.get("rooms", []) if clean_text(item.get("name")) == room_name), None)
            if room:
                facility_keys.add(billing_profile_for_room(room).get("facility"))
        if not facility_keys and clean_text(workflow.get("sourceType")) == "cage_map":
            pi = clean_text(workflow.get("pi"))
            if pi:
                for item in state.get("occupancies", []):
                    if clean_text(item.get("pi")) != pi or not occupancy_overlaps_month(item, current_month):
                        continue
                    rack_id = item.get("rackId") or slot_to_rack_id.get(item.get("slotId"))
                    room_id = item.get("roomId") or rack_by_id.get(rack_id, {}).get("roomId")
                    facility_keys.add(billing_profile_for_room(room_by_id.get(room_id, {})).get("facility"))
        if not facility_keys:
            facility_keys.add("zhujiang")
        if workflow_status == WORKFLOW_STATUS_FINANCE:
            dashboard["currentMonthWorkflowDoneCount"] += 1
            for key in facility_keys:
                ensure_facility_summary(key)["currentMonthWorkflowDoneCount"] += 1
            continue
        dashboard["currentMonthWorkflowTodoCount"] += 1
        if workflow_status in (WORKFLOW_STATUS_GENERATED, WORKFLOW_STATUS_SENT, WORKFLOW_STATUS_SIGNED):
            dashboard["stalledWorkflowCount"] += 1
        for key in facility_keys:
            ensure_facility_summary(key)["currentMonthWorkflowTodoCount"] += 1

    dashboard["exceptionCount"] = (
        dashboard["unmatchedIntakeCount"] + dashboard["overduePlacementCount"] + dashboard["stalledWorkflowCount"]
    )

    return {
        "dashboardSummary": dashboard,
        "roomSummaries": list(room_summaries.values()),
        "rackSummaries": list(rack_summaries.values()),
        "facilitySummaries": list(facility_summaries.values()),
    }


def read_bootstrap_state(conn, actor, scope="summary", room_id=""):
    started_at = time.perf_counter()
    state = filter_state_for_actor(read_cached_state(conn), actor)
    actor_scope = actor_cache_scope(actor)
    if scope == "summary":
        key = cache_key("bootstrap_summary", actor=actor_scope)
        cached = cache_get(key)
        if cached is not None:
            log_perf("bootstrap", started_at, scope=scope, cached=1, rooms=len(cached.get("rooms", [])))
            return cached
    summary = summarize_infrastructure(state)
    payload = {
        "rooms": state.get("rooms", []),
        "racks": state.get("racks", []),
        **summary,
    }
    if scope == "full":
        payload["billingRules"] = state.get("billingRules", [])
        payload["adjustments"] = state.get("adjustments", [])
        payload["intakeBatches"] = state.get("intakeBatches", [])
        payload["placementTasks"] = state.get("placementTasks", [])
        payload["slots"] = state.get("slots", [])
        payload["occupancies"] = state.get("occupancies", [])
        log_perf("bootstrap", started_at, scope=scope, rooms=len(payload["rooms"]), slots=len(payload["slots"]))
        return payload
    if scope == "room":
        room_id = clean_text(room_id)
        payload["placementTasks"] = [
            item for item in state.get("placementTasks", []) if item.get("targetRoomId") == room_id
        ]
        rack_ids = {rack.get("id") for rack in state.get("racks", []) if rack.get("roomId") == room_id}
        payload["slots"] = [slot for slot in state.get("slots", []) if slot.get("rackId") in rack_ids]
        slot_ids = {slot.get("id") for slot in payload["slots"]}
        payload["occupancies"] = [item for item in state.get("occupancies", []) if item.get("slotId") in slot_ids]
        log_perf(
            "bootstrap",
            started_at,
            scope=scope,
            room_id=room_id,
            rooms=len(payload["rooms"]),
            slots=len(payload["slots"]),
        )
        return payload
    payload["slots"] = []
    payload["occupancies"] = []
    cached_payload = cache_set(key, payload) if scope == "summary" else payload
    log_perf("bootstrap", started_at, scope=scope, rooms=len(payload["rooms"]))
    return cached_payload


def actor_cache_scope(actor):
    if not actor:
        return "anonymous"
    role = actor.get("role", "")
    rooms = ",".join(sorted(clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)))
    return f"{role}:{rooms}"


def month_range(month):
    normalized = clean_text(month)
    if not re.fullmatch(r"\d{4}-\d{2}", normalized):
        raise ValueError("结算月份格式应为 YYYY-MM")
    year, month_no = normalized.split("-")
    start = f"{year}-{month_no}-01"
    last_day = calendar.monthrange(int(year), int(month_no))[1]
    end = f"{year}-{month_no}-{last_day:02d}"
    return start, end


def occupancy_overlaps_month(item, month):
    start, end = month_range(month)
    start_date = clean_text(item.get("startDate") or item.get("start_date"))
    end_date = clean_text(item.get("endDate") or item.get("end_date"))
    if not start_date:
        return False
    if start_date > end:
        return False
    if end_date and end_date < start:
        return False
    return True


def read_billing_occupancies(conn, actor, filters):
    started_at = time.perf_counter()
    month = clean_text(filters.get("month", ""))
    iacuc = clean_text(filters.get("iacuc", ""))
    pi = clean_text(filters.get("pi", ""))
    if not month:
        raise ValueError("请提供结算月份")
    key = cache_key(
        "billing_occupancies",
        actor=actor_cache_scope(actor),
        month=month,
        iacuc=iacuc,
        pi=pi,
    )
    cached = cache_get(key)
    if cached is not None:
        log_perf(
            "billing_occupancies", started_at, cached=1, month=month, occupancies=len(cached.get("occupancies", []))
        )
        return cached

    occupancies = read_occupancies_for_billing(conn, month, iacuc=iacuc, pi=pi)
    if actor and actor.get("role") != "admin":
        allowed_room_ids = {clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)}
        occupancies = [item for item in occupancies if clean_text(item.get("roomId", "")) in allowed_room_ids]
    state = read_billing_state_for_occupancies(conn, occupancies)
    matched_slot_ids = {item.get("slotId") for item in occupancies if item.get("slotId")}
    payload = {
        "month": month,
        "pi": pi,
        "iacuc": iacuc,
        "slots": [slot for slot in state["slots"] if slot.get("id") in matched_slot_ids],
        "occupancies": occupancies,
    }
    cache_set(key, payload)
    log_perf("billing_occupancies", started_at, month=month, occupancies=len(occupancies), slots=len(payload["slots"]))
    return payload


def read_occupancies_for_billing(conn, month, iacuc="", pi=""):
    start, end = month_range(month)
    clauses = [
        "status IN ('active', 'ended')",
        "start_date <> ''",
        "start_date <= ?",
        "(end_date IS NULL OR end_date = '' OR end_date >= ?)",
    ]
    params = [end, start]
    normalized_iacuc = normalize_iacuc_number(iacuc)
    normalized_pi = clean_text(pi)
    if normalized_iacuc:
        clauses.append("iacuc = ?")
        params.append(normalized_iacuc)
    if normalized_pi:
        clauses.append("pi = ?")
        params.append(normalized_pi)

    rows = conn.execute(
        f"SELECT payload FROM occupancies WHERE {' AND '.join(clauses)} ORDER BY start_date, rowid",
        tuple(params),
    ).fetchall()
    items = [json.loads(row["payload"]) for row in rows]
    slot_ids = {item.get("slotId") for item in items if item.get("slotId")}
    slots = read_payloads_by_ids(conn, "cage_slots", slot_ids, "rack_id, row_no, col_no, rowid")
    rack_ids = {slot.get("rackId") for slot in slots if slot.get("rackId")}
    racks = read_payloads_by_ids(conn, "racks", rack_ids, "room_id, index_no, rowid")
    room_ids = {rack.get("roomId") for rack in racks if rack.get("roomId")}
    rooms = read_payloads_by_ids(conn, "rooms", room_ids, "rowid")
    applications_by_iacuc = read_applications_by_iacuc(conn)
    state = {"rooms": rooms, "racks": racks, "slots": slots}
    return [occupancy_with_snapshots(item, state, applications_by_iacuc) for item in items]


def read_payloads_by_ids(conn, table, ids, order_by="rowid"):
    id_list = sorted({clean_text(item) for item in ids if clean_text(item)})
    if not id_list:
        return []
    rows = conn.execute(
        f"SELECT payload FROM {table} WHERE id IN ({placeholders(id_list)}) ORDER BY {order_by}",
        tuple(id_list),
    ).fetchall()
    return [json.loads(row["payload"]) for row in rows]


def read_billing_state_for_occupancies(conn, occupancies):
    slot_ids = {item.get("slotId") for item in occupancies if item.get("slotId")}
    slots = read_payloads_by_ids(conn, "cage_slots", slot_ids, "rack_id, row_no, col_no, rowid")
    rack_ids = {slot.get("rackId") for slot in slots if slot.get("rackId")}
    rack_ids.update(item.get("rackId") for item in occupancies if item.get("rackId"))
    racks = read_payloads_by_ids(conn, "racks", rack_ids, "room_id, index_no, rowid")
    room_ids = {rack.get("roomId") for rack in racks if rack.get("roomId")}
    room_ids.update(item.get("roomId") for item in occupancies if item.get("roomId"))
    rooms = read_payloads_by_ids(conn, "rooms", room_ids, "rowid")
    return {"rooms": rooms, "racks": racks, "slots": slots}


def filter_state_for_actor(state, actor):
    if not state or not actor or actor.get("role") == "admin":
        return state

    allowed_rooms = set(actor.get("roomIds", []))
    rooms = [room for room in state.get("rooms", []) if room.get("id") in allowed_rooms]
    room_ids = {room.get("id") for room in rooms}
    racks = [rack for rack in state.get("racks", []) if rack.get("roomId") in room_ids]
    rack_ids = {rack.get("id") for rack in racks}
    slots = [slot for slot in state.get("slots", []) if slot.get("rackId") in rack_ids]
    slot_ids = {slot.get("id") for slot in slots}
    occupancies = [item for item in state.get("occupancies", []) if item.get("slotId") in slot_ids]
    placement_tasks = [item for item in state.get("placementTasks", []) if item.get("targetRoomId") in room_ids]

    return {
        **state,
        "rooms": rooms,
        "racks": racks,
        "slots": slots,
        "occupancies": occupancies,
        "placementTasks": placement_tasks,
    }


def filter_entity_payloads_for_actor(collection, items, actor):
    if not actor or actor.get("role") == "admin":
        return items

    allowed_rooms = {clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)}
    if collection == "rooms":
        return [item for item in items if item.get("id") in allowed_rooms]
    if collection == "racks":
        return [item for item in items if item.get("roomId") in allowed_rooms]
    if collection == "placementTasks":
        return [item for item in items if item.get("targetRoomId") in allowed_rooms]
    if collection == "slots":
        rack_ids = {clean_text(item.get("rackId", "")) for item in items if clean_text(item.get("rackId", ""))}
        rack_rooms = read_rack_room_map(rack_ids)
        return [item for item in items if rack_rooms.get(item.get("rackId")) in allowed_rooms]
    if collection == "occupancies":
        direct = [item for item in items if item.get("roomId") in allowed_rooms]
        unresolved = [item for item in items if not item.get("roomId")]
        if not unresolved:
            return direct
        slot_ids = {clean_text(item.get("slotId", "")) for item in unresolved if clean_text(item.get("slotId", ""))}
        slot_rooms = read_slot_room_map(slot_ids)
        return direct + [item for item in unresolved if slot_rooms.get(item.get("slotId")) in allowed_rooms]
    return items


def placeholders(values):
    return ", ".join("?" for _ in values)


def read_rack_room_map(rack_ids):
    ids = sorted({clean_text(item) for item in rack_ids if clean_text(item)})
    if not ids:
        return {}
    with connect_db() as conn:
        rows = conn.execute(f"SELECT id, room_id FROM racks WHERE id IN ({placeholders(ids)})", ids).fetchall()
    return {row["id"]: row["room_id"] for row in rows}


def read_slot_room_map(slot_ids):
    ids = sorted({clean_text(item) for item in slot_ids if clean_text(item)})
    if not ids:
        return {}
    with connect_db() as conn:
        rows = conn.execute(
            f"""
            SELECT slots.id AS slot_id, racks.room_id AS room_id
            FROM cage_slots AS slots
            JOIN racks ON racks.id = slots.rack_id
            WHERE slots.id IN ({placeholders(ids)})
            """,
            ids,
        ).fetchall()
    return {row["slot_id"]: row["room_id"] for row in rows}


def write_state(state, actor, skip_permission=False):
    updated_at = datetime.now(UTC).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or {}
        if not skip_permission:
            validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        state["auditLogs"] = merge_audit_logs(state.get("auditLogs", []), events)
        write_normalized_state(conn, state, updated_at)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state", "principal_identities")
    invalidate_data_cache_prefixes(
        "bootstrap_summary::", "billing_occupancies::", "quantity_sheets::", "billing_workflows::"
    )
    return {"ok": True, "updatedAt": updated_at, "auditLogs": merge_audit_logs([], events)}


def write_entity_state(endpoint, method, item_id, payload, actor):
    spec = WRITABLE_ENTITY_ENDPOINTS[endpoint]
    collection = spec["collection"]
    if collection == "occupancies":
        return write_occupancy_entity_state(method, item_id, payload, actor, spec)
    if collection == "intakeBatches":
        return write_intake_batch_entity_state(method, item_id, payload, actor, spec)
    if collection == "placementTasks":
        return write_placement_task_entity_state(method, item_id, payload, actor, spec)
    if collection in ("rooms", "racks", "slots"):
        return write_infrastructure_entity_state(collection, method, item_id, payload, actor, spec)

    current = read_state()
    state = current.get("state") or empty_state()

    item = normalize_entity_payload(collection, payload, item_id, method, spec["id_prefix"])
    status = HTTPStatus.OK
    if method == "POST":
        insert_entity(state, collection, item)
        status = HTTPStatus.CREATED
    elif method == "PUT":
        replace_entity(state, collection, item_id, item)
    elif method == "DELETE":
        item = delete_entity(state, collection, item_id)
    else:
        raise ValueError("Unsupported entity write method")

    if collection == "occupancies":
        sync_slot_statuses(state)

    result = write_state(state, actor)
    response = {"item": item, "updatedAt": result["updatedAt"], "auditLogs": result["auditLogs"]}
    if collection == "intakeBatches" and method != "DELETE":
        response["placementTasks"] = [
            task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == item.get("id")
        ]
    return response, status


def write_infrastructure_entity_state(collection, method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(UTC).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        item = normalize_entity_payload(collection, payload, item_id, method, spec["id_prefix"])
        if method == "POST":
            insert_entity(state, collection, item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, collection, item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, collection, item_id)
        else:
            raise ValueError("Unsupported entity write method")

        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        if collection == "rooms":
            if method == "POST":
                insert_room_record(conn, item)
            elif method == "PUT":
                update_room_record(conn, item)
            else:
                delete_room_record(conn, item_id)
        elif collection == "racks":
            if method == "POST":
                insert_rack_record(conn, item)
            elif method == "PUT":
                update_rack_record(conn, item)
            else:
                delete_rack_record(conn, item_id)
        elif collection == "slots":
            if method == "POST":
                insert_slot_record(conn, item)
            elif method == "PUT":
                update_slot_record(conn, item)
            else:
                delete_slot_record(conn, item_id)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes(
        "bootstrap_summary::", "billing_occupancies::", "intake_batches::", "placement_tasks::"
    )
    response = {
        "item": item,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1, collection=collection, method=method),
    }
    log_perf("infrastructure_entity.save", started_at, collection=collection, method=method)
    return response, status


def write_intake_batch_entity_state(method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(UTC).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        item = normalize_entity_payload("intakeBatches", payload, item_id, method, spec["id_prefix"])
        if method == "POST":
            insert_entity(state, "intakeBatches", item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, "intakeBatches", item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, "intakeBatches", item_id)
        else:
            raise ValueError("Unsupported entity write method")

        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        if method == "DELETE":
            delete_intake_batch_repository(conn, item_id)
            conn.execute("DELETE FROM placement_tasks WHERE source_batch_id = ?", (item_id,))
        else:
            saved_item = next(
                (entry for entry in state.get("intakeBatches", []) if entry.get("id") == item.get("id")), item
            )
            upsert_intake_batch_repository(conn, saved_item)
            item = saved_item
            next_task_ids = {
                task.get("id")
                for task in state.get("placementTasks", [])
                if task.get("sourceBatchId") == item.get("id")
            }
            old_task_ids = {
                task.get("id")
                for task in old_state.get("placementTasks", [])
                if task.get("sourceBatchId") == item.get("id")
            }
            for task_id in sorted(old_task_ids - next_task_ids):
                delete_placement_task_repository(conn, task_id)
            for task in state.get("placementTasks", []):
                if task.get("sourceBatchId") == item.get("id"):
                    upsert_placement_task_repository(conn, task)
        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes(
        "bootstrap_summary::", "billing_occupancies::", "intake_batches::", "placement_tasks::"
    )
    response = {
        "item": item,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1, method=method),
    }
    if method != "DELETE":
        response["placementTasks"] = [
            task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == item.get("id")
        ]
        response["perf"] = write_perf_summary(
            started_at, rows_changed=1 + len(response["placementTasks"]), method=method
        )
    log_perf("intake_batch.save", started_at, method=method, tasks=len(response.get("placementTasks", [])))
    return response, status


def write_placement_task_entity_state(method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(UTC).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        item = normalize_entity_payload("placementTasks", payload, item_id, method, spec["id_prefix"])
        old_item = next((entry for entry in old_state.get("placementTasks", []) if entry.get("id") == item_id), None)
        removed_occupancy_id = ""
        if method == "POST":
            insert_entity(state, "placementTasks", item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, "placementTasks", item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, "placementTasks", item_id)
            removed_occupancy_id = clean_text((old_item or item).get("reservedOccupancyId", ""))
        else:
            raise ValueError("Unsupported entity write method")

        affected_slot_ids = {
            clean_text(occ.get("slotId", ""))
            for occ in old_state.get("occupancies", [])
            if removed_occupancy_id and occ.get("id") == removed_occupancy_id
        }
        sync_slot_statuses(state)
        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        if method == "DELETE":
            delete_placement_task_repository(conn, item_id)
            if removed_occupancy_id:
                conn.execute("DELETE FROM occupancies WHERE id = ?", (removed_occupancy_id,))
        else:
            saved_item = next(
                (entry for entry in state.get("placementTasks", []) if entry.get("id") == item.get("id")), item
            )
            upsert_placement_task_repository(conn, saved_item)
            item = saved_item
        affected_slots = [slot for slot in state.get("slots", []) if slot.get("id") in affected_slot_ids]
        for slot in affected_slots:
            update_slot_record(conn, slot)
        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "placement_tasks::")
    response = {
        "item": item,
        "affectedSlots": affected_slots,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + len(affected_slots), method=method),
    }
    log_perf("placement_task.save", started_at, method=method, slot_count=len(affected_slots))
    return response, status


def persist_intake_receipt_confirmation(batch_id, body, actor):
    started_at = time.perf_counter()
    updated_at = datetime.now(UTC).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        batch, receipt, tasks = confirm_intake_receipt(state, batch_id, body, actor)
        events = build_audit_events(actor, old_state, state, updated_at)
        upsert_intake_batch_repository(conn, batch)
        for task in tasks:
            upsert_placement_task_repository(conn, task)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes(
        "bootstrap_summary::", "billing_occupancies::", "intake_batches::", "placement_tasks::"
    )
    log_perf("intake_batch.confirm", started_at, tasks=len(tasks))
    return {
        "batch": batch,
        "receipt": receipt,
        "tasks": tasks,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + len(tasks), tasks=len(tasks)),
    }


def persist_placement_action(task_id, actor, mutator):
    started_at = time.perf_counter()
    updated_at = datetime.now(UTC).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        result = mutator(state)
        task = result[0] if isinstance(result, tuple) else result
        occupancy = result[1] if isinstance(result, tuple) and len(result) > 1 else None
        affected_slot_ids = changed_placement_slot_ids(old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        upsert_placement_task_repository(conn, task)
        if occupancy:
            applications_by_iacuc = read_applications_by_iacuc(conn)
            occupancy = occupancy_with_snapshots(occupancy, state, applications_by_iacuc)
            upsert_occupancy_record(conn, occupancy)
        affected_slots = [slot for slot in state.get("slots", []) if slot.get("id") in affected_slot_ids]
        for slot in affected_slots:
            update_slot_record(conn, slot)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::")
    log_perf("placement_task.action", started_at, task_id=task_id, slot_count=len(affected_slots))
    payload = {
        "task": task,
        "affectedSlots": affected_slots,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(
            started_at, rows_changed=1 + (1 if occupancy else 0) + len(affected_slots), slot_count=len(affected_slots)
        ),
    }
    if occupancy:
        payload["occupancy"] = occupancy
    return payload


def changed_placement_slot_ids(old_state, state):
    old_occupancies = {item.get("id"): item for item in old_state.get("occupancies", [])}
    new_occupancies = {item.get("id"): item for item in state.get("occupancies", [])}
    slot_ids = set()
    for occupancy_id in changed_keys(old_occupancies, new_occupancies):
        old_item = old_occupancies.get(occupancy_id) or {}
        new_item = new_occupancies.get(occupancy_id) or {}
        for item in (old_item, new_item):
            slot_id = clean_text(item.get("slotId", ""))
            if slot_id:
                slot_ids.add(slot_id)
    return slot_ids


def write_occupancy_entity_state(method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(UTC).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        old_item = next((item for item in old_state.get("occupancies", []) if item.get("id") == item_id), None)
        item = normalize_entity_payload("occupancies", payload, item_id, method, spec["id_prefix"])
        if method == "POST":
            insert_entity(state, "occupancies", item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, "occupancies", item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, "occupancies", item_id)
        else:
            raise ValueError("Unsupported entity write method")

        affected_slot_ids = {clean_text((old_item or {}).get("slotId", "")), clean_text((item or {}).get("slotId", ""))}
        affected_slot_ids.discard("")
        sync_slot_statuses(state)
        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        applications_by_iacuc = read_applications_by_iacuc(conn)
        saved_item = None
        if method == "DELETE":
            conn.execute("DELETE FROM occupancies WHERE id = ?", (item_id,))
        else:
            saved_item = next(
                (entry for entry in state.get("occupancies", []) if entry.get("id") == item.get("id")), item
            )
            saved_item = occupancy_with_snapshots(saved_item, state, applications_by_iacuc)
            upsert_occupancy_record(conn, saved_item)

        affected_slots = [slot for slot in state.get("slots", []) if slot.get("id") in affected_slot_ids]
        for slot in affected_slots:
            update_slot_record(conn, slot)
        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::")
    response_item = item if method == "DELETE" else saved_item
    response = {
        "item": response_item,
        "affectedSlots": affected_slots,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + len(affected_slots), method=method),
    }
    log_perf("occupancy.save", started_at, method=method, slot_count=len(affected_slots))
    return response, status


def upsert_occupancy_record(conn, occupancy):
    structured = occupancy_structured_values(occupancy)
    conn.execute(
        """
        INSERT INTO occupancies (
            id, slot_id, room_id, rack_id, cage_code, status, iacuc, project, pi, owner, funding,
            species, billing_item, customer_type, animal_count, room_name, rack_name, slot_code,
            start_date, end_date, end_reason, notes, updated_at, payload
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            slot_id = excluded.slot_id,
            room_id = excluded.room_id,
            rack_id = excluded.rack_id,
            cage_code = excluded.cage_code,
            status = excluded.status,
            iacuc = excluded.iacuc,
            project = excluded.project,
            pi = excluded.pi,
            owner = excluded.owner,
            funding = excluded.funding,
            species = excluded.species,
            billing_item = excluded.billing_item,
            customer_type = excluded.customer_type,
            animal_count = excluded.animal_count,
            room_name = excluded.room_name,
            rack_name = excluded.rack_name,
            slot_code = excluded.slot_code,
            start_date = excluded.start_date,
            end_date = excluded.end_date,
            end_reason = excluded.end_reason,
            notes = excluded.notes,
            updated_at = excluded.updated_at,
            payload = excluded.payload
        """,
        (
            occupancy.get("id"),
            occupancy.get("slotId"),
            structured["room_id"],
            structured["rack_id"],
            occupancy.get("cageCode", ""),
            occupancy.get("status", ""),
            occupancy.get("iacuc", ""),
            occupancy.get("project", ""),
            occupancy.get("pi", ""),
            occupancy.get("owner", ""),
            occupancy.get("funding", ""),
            structured["species"],
            structured["billing_item"],
            structured["customer_type"],
            structured["animal_count"],
            occupancy.get("roomName", ""),
            occupancy.get("rackName", ""),
            occupancy.get("slotCode", ""),
            occupancy.get("startDate", ""),
            occupancy.get("endDate", ""),
            occupancy.get("endReason", ""),
            occupancy.get("notes", ""),
            occupancy.get("updatedAt", ""),
            dump_json(occupancy),
        ),
    )


def update_slot_record(conn, slot):
    conn.execute(
        """
        UPDATE cage_slots
        SET rack_id = ?, row_no = ?, col_no = ?, code = ?, status = ?, payload = ?
        WHERE id = ?
        """,
        (
            slot.get("rackId"),
            as_int(slot.get("row")),
            as_int(slot.get("col")),
            slot.get("code", ""),
            slot.get("status", "empty"),
            dump_json(slot),
            slot.get("id"),
        ),
    )


def write_infrastructure_state(payload, actor):
    if not isinstance(payload, dict):
        raise ValueError("Request body must be a JSON object")

    created_rooms = normalize_entity_batch("rooms", payload.get("rooms", []), "POST")
    updated_rooms = normalize_entity_batch("rooms", payload.get("roomUpdates", []), "PUT")
    created_racks = normalize_entity_batch("racks", payload.get("racks", []), "POST")
    updated_racks = normalize_entity_batch("racks", payload.get("rackUpdates", []), "PUT")
    deleted_rack_ids = normalize_id_batch(payload.get("rackDeletes", []), "rackDeletes")
    created_slots = normalize_entity_batch("slots", payload.get("slots", []), "POST")
    deleted_slot_ids = normalize_id_batch(payload.get("slotDeletes", []), "slotDeletes")

    updated_at = datetime.now(UTC).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))

        for room in created_rooms:
            insert_entity(state, "rooms", room)
        for room in updated_rooms:
            replace_entity(state, "rooms", room["id"], room)
        for rack in created_racks:
            insert_entity(state, "racks", rack)
        for rack in updated_racks:
            replace_entity(state, "racks", rack["id"], rack)
        for slot in created_slots:
            insert_entity(state, "slots", slot)
        validate_infrastructure_slot_deletes(state, deleted_slot_ids)
        for slot_id in deleted_slot_ids:
            delete_entity(state, "slots", slot_id)
        for rack_id in deleted_rack_ids:
            delete_entity(state, "racks", rack_id)

        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)

        for room in created_rooms:
            insert_room_record(conn, room)
        for room in updated_rooms:
            update_room_record(conn, room)
        backfill_quantity_sheet_staff(conn, [room["id"] for room in created_rooms + updated_rooms])
        for rack in created_racks:
            insert_rack_record(conn, rack)
        for rack in updated_racks:
            update_rack_record(conn, rack)
        for slot in created_slots:
            insert_slot_record(conn, slot)
        for slot_id in deleted_slot_ids:
            conn.execute("DELETE FROM cage_slots WHERE id = ?", (slot_id,))
        for rack_id in deleted_rack_ids:
            conn.execute("DELETE FROM racks WHERE id = ?", (rack_id,))

        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "quantity_sheets::")
    return {
        "rooms": created_rooms,
        "roomUpdates": updated_rooms,
        "racks": created_racks,
        "rackUpdates": updated_racks,
        "rackDeletes": deleted_rack_ids,
        "slots": created_slots,
        "slotDeletes": deleted_slot_ids,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
    }


def normalize_entity_batch(collection, items, method):
    if items is None:
        return []
    if not isinstance(items, list):
        raise ValueError("批量保存内容必须是数组")

    spec = {
        "rooms": {"id_prefix": "room"},
        "racks": {"id_prefix": "rack"},
        "slots": {"id_prefix": "slot"},
    }[collection]
    normalized = []
    for item in items:
        item_id = item.get("id") if isinstance(item, dict) else None
        normalized.append(normalize_entity_payload(collection, item, item_id, method, spec["id_prefix"]))
    return normalized


def normalize_id_batch(items, label):
    if items is None:
        return []
    if not isinstance(items, list):
        raise ValueError(f"{label} 必须是数组")
    return [str(item) for item in items if str(item).strip()]


def validate_infrastructure_slot_deletes(state, slot_ids):
    if not slot_ids:
        return
    deleting = set(slot_ids)
    active = [
        item
        for item in state.get("occupancies", [])
        if item.get("slotId") in deleting and item.get("status") in ("active", "reserved")
    ]
    if active:
        raise ValueError("不能移除仍在用或已预约的笼位")


def empty_state():
    return {
        "baseRate": 4.5,
        "billingMonth": "",
        "billingIacuc": "",
        "rooms": [],
        "racks": [],
        "slots": [],
        "occupancies": [],
        "placementTasks": [],
        "billingRules": [],
        "adjustments": [],
        "intakeBatches": [],
        "auditLogs": [],
    }


def normalize_entity_payload(collection, payload, item_id, method, id_prefix):
    if method == "DELETE":
        return {}
    if not isinstance(payload, dict):
        raise ValueError("Request body must be a JSON object")

    item = dict(payload.get("item") if isinstance(payload.get("item"), dict) else payload)
    if method == "POST":
        item["id"] = str(item.get("id") or new_id(id_prefix))
    else:
        if not item_id:
            raise ValueError("Entity id is required")
        item["id"] = item_id

    validate_entity_payload(collection, item)
    return item


def validate_entity_payload(collection, item):
    if not item.get("id"):
        raise ValueError("实体 id 不能为空")
    if collection == "rooms":
        require_text(item, "name", "饲养间名称不能为空")
    elif collection == "racks":
        require_text(item, "roomId", "笼架必须关联饲养间")
        require_text(item, "name", "笼架名称不能为空")
    elif collection == "slots":
        require_text(item, "rackId", "笼位必须关联笼架")
        status = item.get("status", "empty")
        if status not in ("empty", "reserved", "active"):
            raise ValueError("笼位状态只能是 empty、reserved 或 active")
    elif collection == "occupancies":
        require_text(item, "slotId", "占用记录必须关联笼位")
        status = item.get("status")
        if status not in ("reserved", "active", "ended"):
            raise ValueError("占用状态只能是 reserved、active 或 ended")
    elif collection == "placementTasks":
        require_text(item, "sourceBatchId", "来源批次不能为空")
        require_text(item, "sourceReceiptId", "来源接收记录不能为空")
        require_text(item, "plannedMoveInDate", "计划入驻日期不能为空")
        status = item.get("status")
        if status not in ("pending", "reserved", "active", "cancelled"):
            raise ValueError("待进驻状态只能是 pending、reserved、active 或 cancelled")
    elif collection == "billingRules":
        require_text(item, "unit", "计费规则单位不能为空")
    elif collection == "adjustments":
        require_text(item, "targetType", "减免规则目标类型不能为空")
        require_text(item, "targetId", "减免规则目标不能为空")
    elif collection == "intakeBatches":
        require_text(item, "supplier", "购买单位不能为空")
        require_text(item, "iacuc", "IACUC 编号不能为空")
        require_text(item, "pi", "项目负责人不能为空")
        require_text(item, "owner", "实验负责人不能为空")
        require_text(item, "roomName", "房间不能为空")
        require_text(item, "intakeDate", "接收日期不能为空")
        require_text(item, "status", "批次状态不能为空")
        quantity = as_int(item.get("quantity"))
        if quantity is not None and quantity <= 0:
            raise ValueError("动物数量必须大于 0")


def require_text(item, key, message):
    if not str(item.get(key, "")).strip():
        raise ValueError(message)


def insert_entity(state, collection, item):
    items = state.setdefault(collection, [])
    if any(existing.get("id") == item["id"] for existing in items):
        raise sqlite3.IntegrityError(f"Duplicate id: {item['id']}")
    if collection == "intakeBatches":
        item = ensure_intake_batch_card_qr_ids(state, item)
    validate_entity_references(state, collection, item)
    items.append(item)


def intake_card_suggested_quantity(batch, index, card_count):
    per_cage = max(as_int(batch.get("suggestedAnimalsPerCage")) or 1, 1)
    quantity = max(as_int(batch.get("quantity")) or 0, 0)
    remainder = quantity % per_cage if quantity and per_cage else 0
    is_last = index == card_count - 1
    if not quantity or (remainder and is_last):
        return ""
    return str(per_cage)


def ensure_intake_batch_card_qr_ids(state, batch):
    next_batch = dict(batch)
    card_count = max(as_int(next_batch.get("finalCardCount")) or as_int(next_batch.get("suggestedCardCount")) or 0, 0)
    existing_cards = next_batch.get("cards") if isinstance(next_batch.get("cards"), list) else []
    used_qr_ids = collect_cage_card_qr_ids(state, next_batch.get("id", ""))
    cards = []
    for index in range(card_count):
        existing = (
            existing_cards[index] if index < len(existing_cards) and isinstance(existing_cards[index], dict) else {}
        )
        existing_qr_id = clean_text(existing.get("qrId")).upper()
        qr_id = (
            existing_qr_id
            if is_cage_card_qr_id(existing_qr_id) and existing_qr_id not in used_qr_ids
            else next_cage_card_qr_id(used_qr_ids)
        )
        used_qr_ids.add(qr_id)
        cards.append(
            {
                **existing,
                "id": clean_text(existing.get("id")) or f"{next_batch.get('id')}-card-{index + 1}",
                "index": index + 1,
                "label": f"{index + 1}/{card_count}",
                "suggestedQuantity": clean_text(existing.get("suggestedQuantity"))
                or intake_card_suggested_quantity(next_batch, index, card_count),
                "qrId": qr_id,
            }
        )
    next_batch["cards"] = cards
    return next_batch


def reconcile_intake_batch_update(state, old_item, item):
    next_item = {**old_item, **item}
    next_item["receipts"] = [
        dict(receipt)
        for receipt in (
            item.get("receipts") if isinstance(item.get("receipts"), list) else old_item.get("receipts", [])
        )
    ]
    final_count = max(as_int(next_item.get("finalCardCount")) or 0, 0)
    confirmed_count = max(
        as_int(next_item.get("confirmedCardCount"))
        or sum(max(as_int(receipt.get("cardCount")) or 0, 0) for receipt in next_item.get("receipts", [])),
        0,
    )
    next_item["confirmedCardCount"] = confirmed_count
    next_item["remainingCardCount"] = max(
        as_int(next_item.get("remainingCardCount"))
        if next_item.get("remainingCardCount") not in (None, "")
        else final_count - confirmed_count,
        0,
    )

    old_status = clean_text(old_item.get("status", ""))
    new_status = clean_text(next_item.get("status", ""))
    old_room_name = clean_text(old_item.get("roomName", ""))
    new_room_name = clean_text(next_item.get("roomName", ""))
    if old_status == "received" and new_status == "printed":
        related_tasks = [
            task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == old_item.get("id")
        ]
        blocking = [task for task in related_tasks if task.get("status") in ("reserved", "active")]
        if blocking:
            raise ValueError("该批次已有已预留或已入驻的待进驻任务，请先处理相关任务后再回退为已打印")
        state["placementTasks"] = [
            task for task in state.get("placementTasks", []) if task.get("sourceBatchId") != old_item.get("id")
        ]
        next_item["receipts"] = []
        next_item["confirmedCardCount"] = 0
        next_item["remainingCardCount"] = final_count
    if old_room_name != new_room_name:
        related_tasks = [
            task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == old_item.get("id")
        ]
        blocking = [task for task in related_tasks if task.get("status") in ("reserved", "active")]
        if blocking:
            raise ValueError("该批次已有已预留或已入驻的待进驻任务，请先处理相关任务后再调整房间")
        target_room = (
            next((room for room in state.get("rooms", []) if clean_text(room.get("name", "")) == new_room_name), None)
            if new_room_name
            else None
        )
        if related_tasks and new_room_name and not target_room:
            raise ValueError("房间尚未在系统中配置，请先选择已配置饲养间后再保存")
        for task in related_tasks:
            task["targetRoomId"] = target_room.get("id", "") if target_room else ""
            task["targetRoomName"] = target_room.get("name", "") if target_room else new_room_name
            task["updatedAt"] = next_item.get("updatedAt") or now_iso()
    return ensure_intake_batch_card_qr_ids(state, next_item)


def replace_entity(state, collection, item_id, item):
    items = state.setdefault(collection, [])
    for index, existing in enumerate(items):
        if existing.get("id") == item_id:
            if collection == "intakeBatches":
                item = reconcile_intake_batch_update(state, existing, item)
            validate_entity_references(state, collection, item)
            items[index] = item
            return
    raise LookupError("实体不存在")


def delete_entity(state, collection, item_id):
    items = state.setdefault(collection, [])
    deleted = None
    kept = []
    for item in items:
        if item.get("id") == item_id:
            deleted = item
        else:
            kept.append(item)
    if deleted is None:
        raise LookupError("实体不存在")
    state[collection] = kept

    if collection == "rooms":
        rack_ids = {rack.get("id") for rack in state.get("racks", []) if rack.get("roomId") == item_id}
        state["racks"] = [rack for rack in state.get("racks", []) if rack.get("roomId") != item_id]
        state["slots"] = [slot for slot in state.get("slots", []) if slot.get("rackId") not in rack_ids]
    elif collection == "racks":
        state["slots"] = [slot for slot in state.get("slots", []) if slot.get("rackId") != item_id]
    elif collection == "placementTasks":
        reserved_occupancy_id = clean_text(deleted.get("reservedOccupancyId", ""))
        task_status = clean_text(deleted.get("status", ""))
        if task_status == "active":
            raise ValueError("已正式入驻的待进驻任务不能直接删除")
        if reserved_occupancy_id:
            state["occupancies"] = [
                item for item in state.get("occupancies", []) if item.get("id") != reserved_occupancy_id
            ]
            sync_slot_statuses(state)

    return deleted


def validate_entity_references(state, collection, item):
    if collection == "racks" and not entity_exists(state, "rooms", item.get("roomId")):
        raise ValueError("关联的饲养间不存在")
    if collection == "slots" and not entity_exists(state, "racks", item.get("rackId")):
        raise ValueError("关联的笼架不存在")
    if collection == "occupancies" and not entity_exists(state, "slots", item.get("slotId")):
        raise ValueError("关联的笼位不存在")
    if (
        collection == "placementTasks"
        and item.get("targetRoomId")
        and not entity_exists(state, "rooms", item.get("targetRoomId"))
    ):
        raise ValueError("关联的目标饲养间不存在")


def entity_exists(state, collection, item_id):
    return any(item.get("id") == item_id for item in state.get(collection, []))


def intake_service_deps():
    return {
        "as_int": as_int,
        "cage_card_qr_id": cage_card_qr_id_from_batch_card,
        "clean_text": clean_text,
        "new_id": new_id,
        "now_iso": now_iso,
    }


def placement_service_deps():
    return {
        "clean_text": clean_text,
        "new_id": new_id,
        "now_iso": now_iso,
    }


def quantity_service_deps():
    return {
        "as_int": as_int,
        "audit_event": audit_event,
        "clean_text": clean_text,
        "insert_quantity_sheet": insert_quantity_sheet_repository,
        "new_id": new_id,
        "normalize_iacuc_number": normalize_iacuc_number,
        "quantity_sheet_db_values": quantity_sheet_db_values,
        "read_applications_by_iacuc": read_applications_by_iacuc,
        "select_quantity_sheets_for_transfer": select_quantity_sheets_for_transfer_repository,
        "update_quantity_sheet": update_quantity_sheet_repository,
    }


def billing_workflow_service_deps():
    return {
        "VERSION_STATUS_ACTIVE": VERSION_STATUS_ACTIVE,
        "VERSION_STATUS_VOIDED": VERSION_STATUS_VOIDED,
        "WORKFLOW_STATUS_FINANCE": WORKFLOW_STATUS_FINANCE,
        "WORKFLOW_STATUS_GENERATED": WORKFLOW_STATUS_GENERATED,
        "WORKFLOW_STATUS_SENT": WORKFLOW_STATUS_SENT,
        "WORKFLOW_STATUS_SIGNED": WORKFLOW_STATUS_SIGNED,
        "billing_workflow_business_key": billing_workflow_business_key,
        "build_version_payload": build_version_payload,
        "build_workflow_event_payload": build_workflow_event_payload,
        "build_workflow_payload": build_workflow_payload,
        "enrich_statement_for_workflow": enrich_statement_for_workflow,
        "get_billing_version": get_billing_version,
        "get_billing_workflow": get_billing_workflow,
        "get_billing_workflow_by_key": get_billing_workflow_by_key,
        "insert_billing_version": insert_billing_version,
        "insert_billing_workflow": insert_billing_workflow,
        "insert_billing_workflow_event": insert_billing_workflow_event,
        "make_statement_document_number": make_statement_document_number,
        "new_id": new_id,
        "normalize_workflow_source": normalize_workflow_source,
        "now_iso": now_iso,
        "replace_version_lines": replace_version_lines,
        "update_billing_version": update_billing_version,
        "update_billing_workflow": update_billing_workflow,
        "workflow_scope_for_statement": workflow_scope_for_statement,
    }


def confirm_intake_receipt(state, batch_id, payload, actor):
    return confirm_intake_receipt_service(state, batch_id, payload, actor, intake_service_deps())


def reserve_placement_task(state, task_id, slot_id, actor):
    return reserve_placement_task_service(state, task_id, slot_id, actor, placement_service_deps())


def move_in_placement_task(state, task_id, actual_move_in_date, actor):
    return move_in_placement_task_service(state, task_id, actual_move_in_date, actor, placement_service_deps())


def reassign_placement_task_room(state, task_id, room_id, actor):
    return reassign_placement_task_room_service(state, task_id, room_id, actor, placement_service_deps())


def migrate_legacy_state(conn):
    if table_has_rows(conn, "rooms"):
        return
    row = conn.execute("SELECT payload, updated_at FROM app_state WHERE id = ?", ("default",)).fetchone()
    if not row:
        return
    write_normalized_state(conn, json.loads(row["payload"]), row["updated_at"])


def write_normalized_state(conn, state, updated_at):
    for table in TABLES:
        conn.execute(f"DELETE FROM {table}")

    applications_by_iacuc = read_applications_by_iacuc(conn)
    set_setting(conn, "baseRate", state.get("baseRate", 4.5), updated_at)
    set_setting(conn, "billingMonth", state.get("billingMonth", ""), updated_at)
    set_setting(conn, "billingIacuc", state.get("billingIacuc", ""), updated_at)

    for room in state.get("rooms", []):
        conn.execute(
            """
            INSERT INTO rooms (id, name, area, rack_count, rows, cols, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                room.get("id"),
                room.get("name", ""),
                room.get("area", ""),
                as_int(room.get("rackCount")),
                as_int(room.get("rows")),
                as_int(room.get("cols")),
                dump_json(room),
            ),
        )

    for rack in state.get("racks", []):
        conn.execute(
            """
            INSERT INTO racks (id, room_id, name, rows, cols, index_no, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rack.get("id"),
                rack.get("roomId"),
                rack.get("name", ""),
                as_int(rack.get("rows")),
                as_int(rack.get("cols")),
                as_int(rack.get("index")),
                dump_json(rack),
            ),
        )

    for slot in state.get("slots", []):
        conn.execute(
            """
            INSERT INTO cage_slots (id, rack_id, row_no, col_no, code, status, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                slot.get("id"),
                slot.get("rackId"),
                as_int(slot.get("row")),
                as_int(slot.get("col")),
                slot.get("code", ""),
                slot.get("status", "empty"),
                dump_json(slot),
            ),
        )

    for occupancy_item in state.get("occupancies", []):
        occupancy = occupancy_with_snapshots(occupancy_item, state, applications_by_iacuc)
        structured = occupancy_structured_values(occupancy)
        conn.execute(
            """
            INSERT INTO occupancies (
                id, slot_id, room_id, rack_id, cage_code, status, iacuc, project, pi, owner, funding,
                species, billing_item, customer_type, animal_count, room_name, rack_name, slot_code,
                start_date, end_date, end_reason, notes, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                occupancy.get("id"),
                occupancy.get("slotId"),
                structured["room_id"],
                structured["rack_id"],
                occupancy.get("cageCode", ""),
                occupancy.get("status", ""),
                occupancy.get("iacuc", ""),
                occupancy.get("project", ""),
                occupancy.get("pi", ""),
                occupancy.get("owner", ""),
                occupancy.get("funding", ""),
                structured["species"],
                structured["billing_item"],
                structured["customer_type"],
                structured["animal_count"],
                occupancy.get("roomName", ""),
                occupancy.get("rackName", ""),
                occupancy.get("slotCode", ""),
                occupancy.get("startDate", ""),
                occupancy.get("endDate", ""),
                occupancy.get("endReason", ""),
                occupancy.get("notes", ""),
                occupancy.get("updatedAt", ""),
                dump_json(occupancy),
            ),
        )

    for task in state.get("placementTasks", []):
        conn.execute(
            """
            INSERT INTO placement_tasks (
                id, source_batch_id, source_receipt_id, target_room_id, planned_move_in_date,
                status, reserved_occupancy_id, actual_move_in_date, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                task.get("id"),
                task.get("sourceBatchId", ""),
                task.get("sourceReceiptId", ""),
                task.get("targetRoomId", ""),
                task.get("plannedMoveInDate", ""),
                task.get("status", "pending"),
                task.get("reservedOccupancyId", ""),
                task.get("actualMoveInDate", ""),
                task.get("updatedAt", updated_at),
                dump_json(task),
            ),
        )

    for rule in state.get("billingRules", []):
        conn.execute(
            """
            INSERT INTO billing_rules (id, name, unit, price, effective_start, effective_end, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rule.get("id"),
                rule.get("name", ""),
                rule.get("unit", ""),
                as_float(rule.get("price")),
                rule.get("effectiveStart", ""),
                rule.get("effectiveEnd", ""),
                dump_json(rule),
            ),
        )

    for adjustment in state.get("adjustments", []):
        conn.execute(
            """
            INSERT INTO billing_adjustments (
                id, target_type, target_id, adjustment_type, value, reason,
                effective_start, effective_end, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                adjustment.get("id"),
                adjustment.get("targetType", ""),
                adjustment.get("targetId", ""),
                adjustment.get("type", ""),
                as_float(adjustment.get("value")),
                adjustment.get("reason", ""),
                adjustment.get("effectiveStart", ""),
                adjustment.get("effectiveEnd", ""),
                dump_json(adjustment),
            ),
        )

    for batch in state.get("intakeBatches", []):
        conn.execute(
            """
            INSERT INTO intake_batches (
                id, batch_no, iacuc, supplier, pi, owner, quantity, card_count,
                room_name, intake_date, status, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch.get("id"),
                batch.get("batchNo", ""),
                batch.get("iacuc", ""),
                batch.get("supplier", ""),
                batch.get("pi", ""),
                batch.get("owner", ""),
                as_int(batch.get("quantity")) or 0,
                as_int(batch.get("finalCardCount")) or 0,
                batch.get("roomName", ""),
                batch.get("intakeDate", ""),
                batch.get("status", "draft"),
                batch.get("updatedAt", updated_at),
                dump_json(batch),
            ),
        )

    for log in state.get("auditLogs", []):
        conn.execute(
            """
            INSERT INTO audit_logs (id, message, at, payload)
            VALUES (?, ?, ?, ?)
            """,
            (log.get("id"), log.get("message", ""), log.get("at", ""), dump_json(log)),
        )


def migrate_reimbursement_record_schema(conn):
    if read_setting(conn, REIMBURSEMENT_MIGRATION_KEY, False):
        return
    rows = conn.execute("SELECT id, payload FROM billing_workflows ORDER BY month, rowid").fetchall()
    for row in rows:
        workflow = json.loads(row["payload"])
        current_version = workflow.get("currentVersion") or {}
        statement = current_version.get("statement") or {}
        if not statement:
            continue
        lines = list_billing_statement_lines_for_version(conn, current_version.get("id", ""))
        detail_context = reimbursement_detail_context_from_workflow(conn, workflow, statement)
        upsert_reimbursement_record_from_statement(
            conn,
            workflow,
            statement,
            lines,
            detail_context,
            source="workflow" if workflow.get("sourceType") else "imported",
        )
    recalculate_all_reimbursement_accumulations(conn)
    set_setting(conn, REIMBURSEMENT_MIGRATION_KEY, True, now_iso())


def assemble_state(conn):
    return assemble_state_repository(conn)


def read_cached_state(conn):
    return read_cached_state_repository(conn, empty_state)


def read_applications_by_iacuc(conn):
    return read_applications_by_iacuc_repository(conn, normalize_iacuc_number)


def occupancy_with_snapshots(occupancy, state, applications_by_iacuc):
    item = dict(occupancy)
    iacuc = normalize_iacuc_number(item.get("iacuc", ""))
    if iacuc:
        item["iacuc"] = iacuc
    application = applications_by_iacuc.get(iacuc, {})
    for key in ("project", "pi", "owner", "funding", "projectStartDate", "projectEndDate", "applicationApprovalDate"):
        if not item.get(key) and application.get(key):
            item[key] = application.get(key, "")

    slot_context = slot_snapshot_context(state, item.get("slotId"))
    for key, value in slot_context.items():
        if value and not item.get(key):
            item[key] = value
    profile = billing_profile_for_occupancy(item, state)
    item["roomId"] = item.get("roomId") or slot_context.get("roomId", "")
    item["rackId"] = item.get("rackId") or slot_context.get("rackId", "")
    item["species"] = item.get("species") or profile.get("species", "")
    item["billingItem"] = item.get("billingItem") or profile.get("billingItem", "")
    item["customerType"] = item.get("customerType") or profile.get("customerType", "")
    if item.get("animalCount") in (None, ""):
        item["animalCount"] = occupancy_animal_count(item, profile) if profile.get("unit") == "animal_day" else None
    return item


def slot_snapshot_context(state, slot_id):
    slot = next((item for item in state.get("slots", []) if item.get("id") == slot_id), None)
    if not slot:
        return {}
    rack = next((item for item in state.get("racks", []) if item.get("id") == slot.get("rackId")), None)
    room = next((item for item in state.get("rooms", []) if item.get("id") == (rack or {}).get("roomId")), None)
    return {
        "roomId": (room or {}).get("id", ""),
        "rackId": (rack or {}).get("id", ""),
        "roomName": (room or {}).get("name", ""),
        "rackName": (rack or {}).get("name", ""),
        "slotCode": slot.get("code", ""),
    }


def occupancy_structured_values(occupancy):
    return {
        "room_id": occupancy.get("roomId", ""),
        "rack_id": occupancy.get("rackId", ""),
        "species": occupancy.get("species", ""),
        "billing_item": occupancy.get("billingItem", ""),
        "customer_type": occupancy.get("customerType", ""),
        "animal_count": as_int(occupancy.get("animalCount")),
    }


def validate_state_write_permission(conn, actor, old_state, new_state):
    if actor["role"] == "admin":
        return

    allowed_rooms = set(actor.get("roomIds", []))
    if not allowed_rooms:
        raise PermissionError("当前账号没有可编辑的饲养间")

    old_rooms = {item.get("id"): item for item in old_state.get("rooms", [])}
    new_rooms = {item.get("id"): item for item in new_state.get("rooms", [])}
    if set(old_rooms) != set(new_rooms):
        raise PermissionError("房间管理员不能新增或删除饲养间")
    for room_id in changed_keys(old_rooms, new_rooms):
        if room_id not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间配置")
        old_room = old_rooms.get(room_id, {})
        new_room = new_rooms.get(room_id, {})
        allowed_room_update_keys = {"rackCount"}
        changed_fields = {key for key in set(old_room) | set(new_room) if old_room.get(key) != new_room.get(key)}
        if not changed_fields.issubset(allowed_room_update_keys):
            raise PermissionError("房间管理员不能修改饲养间基础信息")

    old_racks = {item.get("id"): item for item in old_state.get("racks", [])}
    new_racks = {item.get("id"): item for item in new_state.get("racks", [])}
    rack_rooms = rack_room_map(old_state, new_state)
    for rack_id in changed_keys(old_racks, new_racks):
        if rack_rooms.get(rack_id) not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼架配置")
    if comparable_items(old_state.get("billingRules", [])) != comparable_items(new_state.get("billingRules", [])):
        raise PermissionError("房间管理员不能修改计费规则")
    if old_state.get("baseRate") != new_state.get("baseRate"):
        raise PermissionError("房间管理员不能修改计费规则")
    if comparable_items(old_state.get("adjustments", [])) != comparable_items(new_state.get("adjustments", [])):
        raise PermissionError("房间管理员不能修改减免规则")

    old_slots = {item.get("id"): item for item in old_state.get("slots", [])}
    new_slots = {item.get("id"): item for item in new_state.get("slots", [])}
    slot_rooms = slot_room_map(new_state)
    old_slot_rooms = slot_room_map(old_state)
    for slot_id in changed_keys(old_slots, new_slots):
        room_id = slot_rooms.get(slot_id) or old_slot_rooms.get(slot_id)
        if room_id not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼位结构")
    for slot_id, new_slot in new_slots.items():
        old_slot = dict(old_slots.get(slot_id, {}))
        if not old_slot:
            continue
        comparable_old = {k: v for k, v in old_slot.items() if k != "status"}
        comparable_new = {k: v for k, v in new_slot.items() if k != "status"}
        if comparable_old != comparable_new and slot_rooms.get(slot_id) not in allowed_rooms:
            raise PermissionError("房间管理员不能修改笼位结构")
        if old_slot.get("status") != new_slot.get("status") and slot_rooms.get(slot_id) not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼位状态")

    old_occupancies = {item.get("id"): item for item in old_state.get("occupancies", [])}
    new_occupancies = {item.get("id"): item for item in new_state.get("occupancies", [])}
    changed_ids = changed_keys(old_occupancies, new_occupancies)
    for occupancy_id in changed_ids:
        item = new_occupancies.get(occupancy_id) or old_occupancies.get(occupancy_id) or {}
        if slot_rooms.get(item.get("slotId")) not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼位信息")

    old_tasks = {item.get("id"): item for item in old_state.get("placementTasks", [])}
    new_tasks = {item.get("id"): item for item in new_state.get("placementTasks", [])}
    for task_id in changed_keys(old_tasks, new_tasks):
        item = new_tasks.get(task_id) or old_tasks.get(task_id) or {}
        if item.get("targetRoomId") not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的待进驻任务")


def comparable_items(items):
    return {item.get("id"): item for item in items}


def changed_keys(old_items, new_items):
    keys = set(old_items) | set(new_items)
    return {key for key in keys if old_items.get(key) != new_items.get(key)}


def slot_room_map(state):
    rack_rooms = {rack.get("id"): rack.get("roomId") for rack in state.get("racks", [])}
    return {slot.get("id"): rack_rooms.get(slot.get("rackId")) for slot in state.get("slots", [])}


def rack_room_map(*states):
    rooms = {}
    for state in states:
        for rack in state.get("racks", []):
            rooms[rack.get("id")] = rack.get("roomId")
    return rooms


def build_audit_events(actor, old_state, new_state, at):
    events = []
    slot_labels = slot_label_map(new_state)
    old_occupancies = {item.get("id"): item for item in old_state.get("occupancies", [])}
    new_occupancies = {item.get("id"): item for item in new_state.get("occupancies", [])}
    for occupancy_id in sorted(changed_keys(old_occupancies, new_occupancies)):
        old_item = old_occupancies.get(occupancy_id)
        new_item = new_occupancies.get(occupancy_id)
        item = new_item or old_item or {}
        slot_id = item.get("slotId", "")
        label = slot_labels.get(slot_id, slot_id)
        if old_item is None:
            action = "occupancy.created"
            message = f"{actor['displayName']} 新增笼位 {label} 的占用记录"
        elif new_item is None:
            action = "occupancy.deleted"
            message = f"{actor['displayName']} 删除笼位 {label} 的占用记录"
        elif (
            new_item.get("status") == "ended"
            and new_item.get("endReason") == "sampled"
            and old_item.get("status") != "ended"
        ):
            action = "occupancy.sampled"
            message = f"{actor['displayName']} 将笼位 {label} 标记为已取材，最后计费日期 {new_item.get('endDate', '')}"
        elif (
            new_item.get("status") == "ended"
            and new_item.get("endReason") == "cleared"
            and old_item.get("status") != "ended"
        ):
            action = "occupancy.cleared"
            message = f"{actor['displayName']} 将笼位 {label} 设为空"
        else:
            action = "occupancy.updated"
            message = f"{actor['displayName']} 更新笼位 {label} 的占用信息"
        events.append(audit_event(actor, action, "occupancy", occupancy_id, message, [slot_id], at, old_item, new_item))

    old_rooms = {item.get("id"): item for item in old_state.get("rooms", [])}
    new_rooms = {item.get("id"): item for item in new_state.get("rooms", [])}
    for room_id in sorted(changed_keys(old_rooms, new_rooms)):
        old_item = old_rooms.get(room_id)
        new_item = new_rooms.get(room_id)
        action = "room.updated"
        if old_item is None:
            action = "room.created"
        elif new_item is None:
            action = "room.deleted"
        name = (new_item or old_item or {}).get("name", room_id)
        message = f"{actor['displayName']} {action_label(action)}饲养间 {name}"
        events.append(audit_event(actor, action, "room", room_id, message, [], at, old_item, new_item))

    old_batches = {item.get("id"): item for item in old_state.get("intakeBatches", [])}
    new_batches = {item.get("id"): item for item in new_state.get("intakeBatches", [])}
    for batch_id in sorted(changed_keys(old_batches, new_batches)):
        old_item = old_batches.get(batch_id)
        new_item = new_batches.get(batch_id)
        action = "intake_batch.updated"
        if old_item is None:
            action = "intake_batch.created"
        elif new_item is None:
            action = "intake_batch.deleted"
        label = (new_item or old_item or {}).get("batchNo", batch_id)
        message = f"{actor['displayName']} {action_label(action)}待接收批次 {label}"
        events.append(audit_event(actor, action, "intake_batch", batch_id, message, [], at, old_item, new_item))

    old_tasks = {item.get("id"): item for item in old_state.get("placementTasks", [])}
    new_tasks = {item.get("id"): item for item in new_state.get("placementTasks", [])}
    for task_id in sorted(changed_keys(old_tasks, new_tasks)):
        old_item = old_tasks.get(task_id)
        new_item = new_tasks.get(task_id)
        action = "placement_task.updated"
        if old_item is None:
            action = "placement_task.created"
        elif new_item is None:
            action = "placement_task.deleted"
        label = (new_item or old_item or {}).get("batchNo", task_id)
        message = f"{actor['displayName']} {action_label(action)}待进驻任务 {label}"
        events.append(audit_event(actor, action, "placement_task", task_id, message, [], at, old_item, new_item))

    return events[:100]


def write_perf_summary(started_at, rows_changed=0, **fields):
    return {
        "total_ms": round((time.perf_counter() - started_at) * 1000, 1),
        "rows_changed": rows_changed,
        **{key: value for key, value in fields.items() if value not in (None, "")},
    }


def slot_label_map(state):
    rack_by_id = {rack.get("id"): rack for rack in state.get("racks", [])}
    room_by_id = {room.get("id"): room for room in state.get("rooms", [])}
    labels = {}
    for slot in state.get("slots", []):
        rack = rack_by_id.get(slot.get("rackId"), {})
        room = room_by_id.get(rack.get("roomId"), {})
        rack_index = rack.get("index", "")
        rack_code = str(rack_index).zfill(2) if str(rack_index).isdigit() else str(rack_index)
        labels[slot.get("id")] = f"{room.get('name', '')}-{rack_code}-{slot.get('code', '')}".strip("-")
    return labels


def format_http_date(value):
    return format_datetime(value, usegmt=True)


def read_iacuc_index():
    with connect_db() as conn:
        return read_iacuc_index_repository(conn, IACUC_INDEX_PATH, LEGACY_IACUC_INDEX_PATH)


def write_experiment_applications(conn, items, imported_at):
    replace_experiment_applications(conn, items, imported_at, application_payload)


PROJECT_DERIVED_FIELDS = (
    "project",
    "pi",
    "owner",
    "funding",
    "projectStartDate",
    "projectEndDate",
    "applicationApprovalDate",
    "iacucApprovalDate",
    "fundCode",
    "supportProjectPeriod",
    "experimentNo",
    "species",
    "facility",
    "maxFeedingPeriod",
    "approvedFeedingFee",
    "approvalLeader",
    "actualFeedingFee",
    "pendingReimbursementFee",
    "assistant",
    "applicationDate",
)


def application_by_iacuc(items):
    applications = {}
    for item in items:
        iacuc = normalize_iacuc_number((item or {}).get("iacuc", ""))
        if iacuc:
            applications[iacuc] = item
    return applications


def project_field_snapshot(item):
    return {field: clean_text((item or {}).get(field, "")) for field in PROJECT_DERIVED_FIELDS}


def changed_project_fields(before, after):
    return {
        field
        for field in PROJECT_DERIVED_FIELDS
        if clean_text(before.get(field, "")) != clean_text(after.get(field, ""))
    }


def read_current_applications(conn):
    rows = conn.execute("SELECT payload FROM experiment_applications ORDER BY rowid").fetchall()
    return [json.loads(row["payload"]) for row in rows]


def source_iacuc_for_placement_tasks(conn):
    rows = conn.execute("SELECT id, iacuc, payload FROM intake_batches").fetchall()
    by_batch = {}
    for row in rows:
        payload = json.loads(row["payload"])
        by_batch[row["id"]] = normalize_iacuc_number(payload.get("iacuc") or row["iacuc"] or "")
    return by_batch


def sync_project_fields_for_table(conn, table, applications, changed_iacucs, imported_at, source_iacuc_by_batch=None):
    rows = conn.execute(f"SELECT * FROM {table}").fetchall()
    changes = []
    for row in rows:
        payload = json.loads(row["payload"])
        if table == "placement_tasks":
            source_iacuc = (source_iacuc_by_batch or {}).get(clean_text(payload.get("sourceBatchId", "")), "")
            iacuc = normalize_iacuc_number(payload.get("iacuc", "") or source_iacuc)
        else:
            iacuc = normalize_iacuc_number(payload.get("iacuc", "") or (row["iacuc"] if "iacuc" in row.keys() else ""))
        if not iacuc or iacuc not in changed_iacucs or iacuc not in applications:
            continue

        before = project_field_snapshot(payload)
        after = project_field_snapshot(applications[iacuc])
        changed_fields = sorted(changed_project_fields(before, after))
        if not changed_fields:
            continue

        for field in PROJECT_DERIVED_FIELDS:
            payload[field] = after[field]
        payload["projectSyncedAt"] = imported_at
        payload["projectSyncSource"] = "iacuc_upload"
        payload["updatedAt"] = imported_at
        payload_json = dump_json(payload)

        if table == "quantity_sheets":
            conn.execute(
                """
                UPDATE quantity_sheets
                SET project = ?, pi = ?, owner = ?, funding = ?, updated_at = ?, payload = ?
                WHERE id = ?
                """,
                (after["project"], after["pi"], after["owner"], after["funding"], imported_at, payload_json, row["id"]),
            )
        elif table == "occupancies":
            conn.execute(
                """
                UPDATE occupancies
                SET project = ?, pi = ?, owner = ?, funding = ?, species = ?, updated_at = ?, payload = ?
                WHERE id = ?
                """,
                (
                    after["project"],
                    after["pi"],
                    after["owner"],
                    after["funding"],
                    after["species"],
                    imported_at,
                    payload_json,
                    row["id"],
                ),
            )
        elif table == "intake_batches":
            conn.execute(
                "UPDATE intake_batches SET pi = ?, owner = ?, updated_at = ?, payload = ? WHERE id = ?",
                (after["pi"], after["owner"], imported_at, payload_json, row["id"]),
            )
        elif table == "placement_tasks":
            conn.execute(
                "UPDATE placement_tasks SET updated_at = ?, payload = ? WHERE id = ?",
                (imported_at, payload_json, row["id"]),
            )

        changes.append(
            {
                "table": table,
                "id": row["id"],
                "iacuc": iacuc,
                "changedFields": changed_fields,
                "before": {field: before[field] for field in changed_fields},
                "after": {field: after[field] for field in changed_fields},
            }
        )
    return changes


def sync_project_derived_fields_after_iacuc_upload(conn, old_items, new_items, actor, imported_at):
    old_by_iacuc = application_by_iacuc(old_items)
    new_by_iacuc = application_by_iacuc(new_items)
    changed_iacucs = {
        iacuc
        for iacuc, new_item in new_by_iacuc.items()
        if iacuc not in old_by_iacuc
        or changed_project_fields(project_field_snapshot(old_by_iacuc[iacuc]), project_field_snapshot(new_item))
    }
    if not changed_iacucs:
        return {"changedIacucCount": 0, "updatedRecordCount": 0, "tableCounts": {}, "snapshotId": ""}

    source_iacuc_by_batch = source_iacuc_for_placement_tasks(conn)
    changes = []
    for table in ("quantity_sheets", "occupancies", "intake_batches", "placement_tasks"):
        changes.extend(
            sync_project_fields_for_table(conn, table, new_by_iacuc, changed_iacucs, imported_at, source_iacuc_by_batch)
        )

    table_counts = {}
    for item in changes:
        table_counts[item["table"]] = table_counts.get(item["table"], 0) + 1

    snapshot_id = ""
    if changes:
        snapshot_id = new_id("project-sync")
        summary = {
            "changedIacucCount": len(changed_iacucs),
            "updatedRecordCount": len(changes),
            "tableCounts": table_counts,
        }
        payload = {
            **summary,
            "changedIacucs": sorted(changed_iacucs),
            "changes": changes,
        }
        conn.execute(
            """
            INSERT INTO project_sync_snapshots (
                id, imported_at, actor_user_id, actor_display_name, summary, payload
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                snapshot_id,
                imported_at,
                actor.get("id", ""),
                actor.get("displayName", ""),
                dump_json(summary),
                dump_json(payload),
            ),
        )

    return {
        "changedIacucCount": len(changed_iacucs),
        "updatedRecordCount": len(changes),
        "tableCounts": table_counts,
        "snapshotId": snapshot_id,
    }


def list_principal_identities(conn):
    identity_by_pi = {
        clean_text(item.get("pi", "")): item
        for item in read_principal_identity_payloads(conn)
        if clean_text(item.get("pi", ""))
    }
    principal_names = {clean_text(name) for name in list_distinct_principal_names(conn)}
    principal_names.update(identity_by_pi.keys())
    items = []
    for pi_name in sorted((name for name in principal_names if name), key=lambda value: value.lower()):
        saved = identity_by_pi.get(pi_name, {})
        principal_type = normalize_principal_type(saved.get("principalType", BILLING_PRINCIPAL_INDEPENDENT))
        items.append(
            {
                "pi": pi_name,
                "principalType": principal_type,
                "freeCageAllowance": free_cages_for_principal_type(principal_type),
                "updatedAt": saved.get("updatedAt", ""),
            }
        )
    return items


def read_principal_identities():
    cached = cache_get("principal_identities")
    if cached is not None:
        return cached
    with connect_db() as conn:
        return cache_set("principal_identities", list_principal_identities(conn))


def save_principal_identity(conn, payload, actor, pi_name):
    pi_name = clean_text(pi_name or payload.get("pi", ""))
    if not pi_name:
        raise ValueError("项目负责人不能为空")
    principal_type = normalize_principal_type(payload.get("principalType", ""))
    now = now_iso()
    item = {
        "pi": pi_name,
        "principalType": principal_type,
        "freeCageAllowance": free_cages_for_principal_type(principal_type),
        "updatedAt": now,
    }
    upsert_principal_identity(conn, pi_name, principal_type, now, dump_json(item))
    event = audit_event(
        actor,
        "principal_identity.updated",
        "principal_identity",
        pi_name,
        f"{actor['displayName']} 更新项目负责人 {pi_name} 身份为 {principal_type_label(principal_type)}",
        [],
        now,
        None,
        item,
    )
    write_audit_events(conn, [event])
    invalidate_data_cache("principal_identities")
    invalidate_data_cache_prefixes("quantity_sheets::", "billing_workflows::")
    return item, merge_audit_logs([], [event])


def invalidate_quantity_sheet_candidate_snapshots(conn, sheets_or_keys):
    keys = []
    for item in sheets_or_keys:
        if isinstance(item, dict):
            month = clean_text(item.get("month", ""))
            pi_name = clean_text(item.get("pi", ""))
        else:
            month, pi_name = item
            month = clean_text(month)
            pi_name = clean_text(pi_name)
        if month and pi_name:
            keys.append((month, pi_name))
    mark_billing_candidate_snapshots_stale(conn, "quantity_sheet", keys, now_iso())


def invalidate_quantity_sheet_candidate_snapshots_by_pi(conn, pi_name):
    mark_billing_candidate_snapshots_stale_by_pi(conn, "quantity_sheet", clean_text(pi_name), now_iso())


def invalidate_all_quantity_sheet_candidate_snapshots(conn):
    mark_all_billing_candidate_snapshots_stale(conn, "quantity_sheet", now_iso())


def application_payload(item, imported_at):
    raw_iacuc = clean_text(item.get("rawIacuc", "") or item.get("iacuc", ""))
    normalized = {
        "id": clean_text(item.get("id", "")),
        "iacuc": clean_text(item.get("iacuc", "") or raw_iacuc),
        "rawIacuc": raw_iacuc,
        "project": clean_text(item.get("project", "")),
        "pi": clean_text(item.get("pi", "")),
        "owner": clean_text(item.get("owner", "")),
        "funding": clean_text(item.get("funding", "")),
        "fundCode": clean_text(item.get("fundCode", "")),
        "supportProjectPeriod": clean_text(item.get("supportProjectPeriod", "")),
        "experimentNo": clean_text(item.get("experimentNo", "")),
        "species": clean_text(item.get("species", "")),
        "facility": clean_text(item.get("facility", "")),
        "maxFeedingPeriod": clean_text(item.get("maxFeedingPeriod", "")),
        "iacucApprovalDate": normalize_application_date(item.get("iacucApprovalDate", "")),
        "applicationApprovalDate": normalize_application_date(item.get("applicationApprovalDate", "")),
        "projectStartDate": normalize_application_date(item.get("projectStartDate", "")),
        "projectEndDate": normalize_application_date(item.get("projectEndDate", "")),
        "approvedFeedingFee": normalize_application_amount(item.get("approvedFeedingFee", "")),
        "approvalLeader": clean_text(item.get("approvalLeader", "")),
        "actualFeedingFee": normalize_application_amount(item.get("actualFeedingFee", "")),
        "pendingReimbursementFee": normalize_application_amount(item.get("pendingReimbursementFee", "")),
        "assistant": clean_text(item.get("assistant", "")),
        "notes": clean_text(item.get("notes", "")),
        "applicationDate": normalize_application_date(item.get("applicationDate", "")),
        "rawFields": item.get("rawFields") if isinstance(item.get("rawFields"), dict) else {},
        "importedAt": imported_at,
    }
    return normalized


def save_iacuc_index_file(items):
    save_iacuc_index_file_repository(IACUC_INDEX_PATH, items)


def bounded_int(value, default, min_value, max_value):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return max(min_value, min(number, max_value))


def filtered_where(filter_specs, filters):
    where = []
    params = []
    for key, expression in filter_specs:
        value = clean_text(filters.get(key, ""))
        if not value:
            continue
        where.append(expression)
        params.append(value)
    return " AND ".join(where), tuple(params)


def list_quantity_sheets(conn):
    return list_quantity_sheets_repository(conn)


def list_quantity_sheets_by_month_iacuc(conn, month, iacuc):
    return list_quantity_sheets_by_month_iacuc_repository(conn, month, iacuc)


def list_quantity_sheets_by_month_pi(conn, month, pi):
    return list_quantity_sheets_by_month_pi_repository(conn, month, pi)


def list_quantity_sheets_page(conn, filters):
    return list_quantity_sheets_page_repository(conn, filters, filtered_where)


def list_quantity_sheet_filter_options(conn, filters, column):
    return list_quantity_sheet_filter_options_repository(conn, filters, filtered_where, column)


def get_quantity_sheet(conn, sheet_id):
    row = get_quantity_sheet_repository(conn, sheet_id)
    if not row:
        raise LookupError("数量统计表不存在")
    return row


def get_current_billing_statement(conn, statement_id):
    return get_current_billing_statement_repository(conn, clean_text(statement_id))


def save_quantity_sheet(conn, payload, actor, sheet_id=None):
    started_at = time.perf_counter()
    now = now_iso()
    sheet = normalize_quantity_sheet(payload, sheet_id, now)
    room = read_room_payload(conn, sheet.get("roomId", "")) or {}
    sheet["manager"] = clean_text(actor.get("displayName", ""))
    sheet["roomManager"] = clean_text(room.get("roomManager", ""))
    validate_quantity_sheet_permission(actor, sheet)
    validate_quantity_sheet_animal_requirements(conn, sheet)
    validate_quantity_sheet_free_cage_settings(conn, sheet)
    validate_quantity_sheet_tier_priority(conn, sheet)
    validate_quantity_sheet_custom_billing(sheet)
    previous_sheet = get_quantity_sheet(conn, sheet["id"])
    exists = previous_sheet is not None
    db_values = quantity_sheet_db_values(sheet)
    if exists:
        update_quantity_sheet_repository(conn, sheet, db_values)
        action = "quantity_sheet.updated"
        message = f"{actor['displayName']} 更新 {sheet['iacuc']} {sheet['month']} 数量统计表"
        status = HTTPStatus.OK
    else:
        insert_quantity_sheet_repository(conn, sheet, db_values)
        action = "quantity_sheet.created"
        message = f"{actor['displayName']} 创建 {sheet['iacuc']} {sheet['month']} 数量统计表"
        status = HTTPStatus.CREATED

    transfer_events, affected_sheets = sync_quantity_sheet_transfer_rows(conn, sheet, actor, now)
    event = audit_event(actor, action, "quantity_sheet", sheet["id"], message, [], now, None, sheet)
    events = [event, *transfer_events]
    write_audit_events(conn, events)
    log_perf(
        "quantity_sheet.save",
        started_at,
        sheet_id=sheet["id"],
        affected=len(affected_sheets),
        rows=len(sheet.get("rows", [])),
    )
    perf = write_perf_summary(
        started_at,
        rows_changed=1 + len(affected_sheets),
        affected=len(affected_sheets),
        rows=len(sheet.get("rows", [])),
    )
    return sheet, previous_sheet, affected_sheets, merge_audit_logs([], events), status, perf


def delete_quantity_sheet(conn, actor, sheet_id):
    sheet = get_quantity_sheet(conn, sheet_id)
    validate_quantity_sheet_permission(actor, sheet)
    now = now_iso()
    delete_quantity_sheet_by_id_repository(conn, sheet_id)
    event = audit_event(
        actor,
        "quantity_sheet.deleted",
        "quantity_sheet",
        sheet_id,
        f"{actor['displayName']} 删除 {sheet.get('iacuc', '')} {sheet.get('month', '')} 数量统计表",
        [],
        now,
        sheet,
        None,
    )
    write_audit_events(conn, [event])
    return sheet, merge_audit_logs([], [event])


def quantity_sheet_db_values(sheet):
    return (
        sheet["month"],
        sheet["iacuc"],
        sheet.get("roomId", ""),
        sheet.get("roomName", ""),
        sheet.get("manager", ""),
        sheet.get("project", ""),
        sheet.get("pi", ""),
        sheet.get("owner", ""),
        sheet.get("funding", ""),
        sheet["updatedAt"],
        dump_json(sheet),
    )


def sync_quantity_sheet_transfer_rows(conn, source_sheet, actor, now):
    return sync_quantity_sheet_transfer_rows_service(conn, source_sheet, actor, now, quantity_service_deps())


def normalize_quantity_sheet(payload, sheet_id, updated_at):
    source = payload.get("sheet") if isinstance(payload, dict) and isinstance(payload.get("sheet"), dict) else payload
    if not isinstance(source, dict):
        raise ValueError("数量统计表必须是 JSON 对象")

    month = clean_text(source.get("month", ""))
    iacuc = clean_text(source.get("iacuc", ""))
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("结算月份格式应为 YYYY-MM")
    if not normalize_iacuc_number(iacuc):
        raise ValueError("IACUC 编号不能为空")

    rows = source.get("rows", [])
    if not isinstance(rows, list):
        raise ValueError("统计表明细必须是数组")

    sheet = {
        "id": clean_text(sheet_id or source.get("id") or new_id("qsheet")),
        "month": month,
        "roomId": clean_text(source.get("roomId", "")),
        "roomName": clean_text(source.get("roomName", "")),
        "manager": clean_text(source.get("manager", "")),
        "roomManager": clean_text(source.get("roomManager", "")),
        "iacuc": iacuc,
        "project": clean_text(source.get("project", "")),
        "pi": clean_text(source.get("pi", "")),
        "owner": clean_text(source.get("owner", "")),
        "contact": clean_text(source.get("contact", "")),
        "funding": clean_text(source.get("funding", "")),
        "preferredFreeCages": max(as_int(source.get("preferredFreeCages")) or 0, 0)
        if source.get("preferredFreeCages") not in (None, "")
        else None,
        "freeCagePriority": max(as_int(source.get("freeCagePriority")) or 0, 0)
        if source.get("freeCagePriority") not in (None, "")
        else None,
        "tierCagePriority": max(as_int(source.get("tierCagePriority")) or 0, 0)
        if source.get("tierCagePriority") not in (None, "")
        else None,
        "fullExemption": parse_bool(source.get("fullExemption")),
        "customBillingEnabled": parse_bool(source.get("customBillingEnabled")),
        "customUnitPrice": max(as_float(source.get("customUnitPrice")) or 0, 0)
        if source.get("customUnitPrice") not in (None, "")
        else None,
        "billingUnit": "animal_day" if clean_text(source.get("billingUnit", "")) == "animal_day" else "cage_day",
        "animalDetailEnabled": parse_bool(source.get("animalDetailEnabled")),
        "initialAnimalCount": as_int(source.get("initialAnimalCount")),
        "initialCageCount": as_int(source.get("initialCageCount")),
        "pageCount": max(as_int(source.get("pageCount")) or 1, 1),
        "rows": [normalize_quantity_sheet_row(row, month) for row in rows],
        "updatedAt": updated_at,
    }
    if sheet["fullExemption"]:
        sheet["preferredFreeCages"] = None
        sheet["freeCagePriority"] = None
        sheet["tierCagePriority"] = None
    sheet["rows"] = sorted(sheet["rows"], key=lambda item: (item["date"], item["id"]))
    return sheet


def parse_bool(value):
    if isinstance(value, bool):
        return value
    return clean_text(value).lower() in ("1", "true", "yes", "on")


def read_room_payload(conn, room_id):
    room_key = clean_text(room_id)
    if not room_key:
        return None
    row = conn.execute("SELECT payload FROM rooms WHERE id = ?", (room_key,)).fetchone()
    return json.loads(row["payload"]) if row else None


def read_room_payloads_for_context(conn, room_ids=None, room_names=None):
    ids = [clean_text(item) for item in (room_ids or []) if clean_text(item)]
    names = [clean_text(item) for item in (room_names or []) if clean_text(item)]
    clauses = []
    params = []
    if ids:
        clauses.append(f"id IN ({placeholders(ids)})")
        params.extend(ids)
    if names:
        clauses.append(f"name IN ({placeholders(names)})")
        params.extend(names)
    if not clauses:
        return []
    rows = conn.execute(
        f"SELECT payload FROM rooms WHERE {' OR '.join(clauses)} ORDER BY rowid", tuple(params)
    ).fetchall()
    return [json.loads(row["payload"]) for row in rows]


def validate_quantity_sheet_animal_requirements(conn, sheet):
    room = read_room_payload(conn, sheet.get("roomId"))
    profile = billing_profile_for_room(room, sheet.get("billingUnit"))
    if profile["unit"] != "animal_day":
        return
    has_animal_balance = any((row.get("animalCount") or 0) > 0 for row in sheet.get("rows", []))
    if not has_animal_balance:
        raise ValueError("该房间按只/天计费，请打开动物数量并补充结余总数")


def validate_quantity_sheet_free_cage_settings(conn, sheet):
    preferred = max(as_int(sheet.get("preferredFreeCages")) or 0, 0)
    if preferred <= 0:
        return
    pi_name = clean_text(sheet.get("pi", ""))
    if not pi_name:
        raise ValueError("设置优先减免笼数前，请先填写项目负责人")
    principal_type_by_pi = read_principal_type_by_pi(conn)
    allowance = billing_free_cages_for_pi(principal_type_by_pi, pi_name)
    if preferred > allowance:
        raise ValueError(f"优先减免笼数不能超过 {pi_name} 的每日总减免额度 {allowance} 笼")
    total = preferred
    for item in list_quantity_sheets_by_month_pi(conn, sheet.get("month"), pi_name):
        if item.get("id") == sheet.get("id"):
            continue
        total += max(as_int(item.get("preferredFreeCages")) or 0, 0)
    if total > allowance:
        raise ValueError(f"{pi_name} 本月已指定优先减免 {total} 笼/天，超过总额度 {allowance} 笼/天")


def validate_quantity_sheet_tier_priority(conn, sheet):
    priority = as_int(sheet.get("tierCagePriority"))
    if priority is None:
        return
    room = read_room_payload(conn, sheet.get("roomId"))
    profile = billing_profile_for_room(room, sheet.get("billingUnit"))
    if profile["unit"] != "cage_day" or not profile.get("tiered"):
        raise ValueError("当前房间计费口径不支持优先梯度")
    pi_name = clean_text(sheet.get("pi", ""))
    if not pi_name:
        raise ValueError("设置优先梯度前，请先填写项目负责人")
    enabled_count = 1
    for item in list_quantity_sheets_by_month_pi(conn, sheet.get("month"), pi_name):
        if item.get("id") == sheet.get("id"):
            continue
        if as_int(item.get("tierCagePriority")) is not None:
            enabled_count += 1
    if enabled_count > 1:
        raise ValueError(f"{pi_name} 在 {sheet.get('month')} 仅能指定一个优先梯度伦理")


def validate_quantity_sheet_custom_billing(sheet):
    if not sheet.get("customBillingEnabled"):
        return
    unit_price = as_float(sheet.get("customUnitPrice"))
    if unit_price is None or unit_price <= 0:
        raise ValueError("启用自定义饲养费后，请填写大于 0 的收费标准")


def read_rooms_for_quantity_sheets(conn, sheets):
    return read_room_payloads_for_context(
        conn,
        room_ids=[sheet.get("roomId", "") for sheet in sheets],
        room_names=[sheet.get("roomName", "") for sheet in sheets],
    )


def normalize_quantity_sheet_row(row, month):
    if not isinstance(row, dict):
        raise ValueError("统计表明细行必须是 JSON 对象")
    date = normalize_sheet_date(row.get("date", ""), month)
    return {
        "id": clean_text(row.get("id", "")) or new_id("qrow"),
        "date": date,
        "addedCount": as_int(row.get("addedCount")),
        "addedType": clean_text(row.get("addedType", "")),
        "removedCount": as_int(row.get("removedCount")),
        "removedType": clean_text(row.get("removedType", "")),
        "transferInFromIacuc": normalize_iacuc_number(row.get("transferInFromIacuc", "")),
        "transferOutToIacuc": normalize_iacuc_number(row.get("transferOutToIacuc", "")),
        "animalCount": as_int(row.get("animalCount")),
        "cageCount": as_int(row.get("cageCount")),
        "handler": clean_text(row.get("handler", "")),
        "balanceSource": "manual" if clean_text(row.get("balanceSource", "")) == "manual" else "auto",
        "notes": clean_text(row.get("notes", "")),
        "transferSourceSheetId": clean_text(row.get("transferSourceSheetId", "")),
        "transferSourceIacuc": normalize_iacuc_number(row.get("transferSourceIacuc", "")),
        "transferMirrorContrib": {
            clean_text(key): max(as_int(value) or 0, 0)
            for key, value in (row.get("transferMirrorContrib") or {}).items()
            if clean_text(key)
        }
        if isinstance(row.get("transferMirrorContrib"), dict)
        else {},
    }


def normalize_sheet_date(value, month):
    text = clean_text(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        date = text
    elif re.fullmatch(r"\d{1,2}[./-]\d{1,2}", text):
        year = month.split("-", 1)[0]
        month_no, day = [int(part) for part in re.split(r"[./-]", text)]
        date = f"{int(year):04d}-{month_no:02d}-{day:02d}"
    elif re.fullmatch(r"\d{1,2}", text):
        date = f"{month}-{int(text):02d}"
    else:
        raise ValueError("统计表日期格式应为 YYYY-MM-DD、M.D 或当月日期")
    if not date.startswith(month + "-"):
        raise ValueError("统计表明细日期必须属于结算月份")
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("统计表明细日期无效") from exc
    return date


def validate_quantity_sheet_permission(actor, sheet):
    if actor:
        return
    raise PermissionError("请先登录")


def generate_quantity_sheet_statement(conn, sheet_id, payload, actor):
    sheet = get_quantity_sheet(conn, sheet_id)
    validate_quantity_sheet_permission(actor, sheet)
    status = clean_text(payload.get("status", "draft")) or "draft"
    persist = bool(payload.get("persist"))
    if status not in ("draft", "locked"):
        raise ValueError("结算单状态只能是 draft 或 locked")

    sheet_iacuc = normalize_iacuc_number(sheet.get("iacuc", ""))
    if not sheet_iacuc:
        raise ValueError("数量统计表缺少伦理号，无法生成按伦理号拆分的结算单")
    sheets = list_quantity_sheets_by_month_iacuc(conn, sheet["month"], sheet_iacuc)
    for item in sheets:
        validate_quantity_sheet_permission(actor, item)
    pi_name = clean_text(sheet.get("pi", ""))
    principal_type_by_pi = read_principal_type_by_pi(conn)
    principal_type = principal_type_by_pi.get(pi_name, BILLING_PRINCIPAL_INDEPENDENT)
    # IACUC 分表阶段不应用 PI 免费笼位，避免跨伦理号结算失真。
    free_cages = 0
    rooms = read_rooms_for_quantity_sheets(conn, sheets)
    applications_by_iacuc = read_applications_by_iacuc(conn)
    lines = quantity_sheet_statement_lines(sheets, free_cages, rooms, applications_by_iacuc)
    generated_at = now_iso()
    iacucs = sorted({normalize_iacuc_number(item.get("iacuc", "")) for item in sheets if item.get("iacuc")})
    statement_iacuc = iacucs[0] if iacucs else sheet_iacuc
    statement = {
        "id": new_id("stmt"),
        "iacuc": statement_iacuc,
        "iacucs": iacucs,
        "month": sheet["month"],
        "project": "、".join(sorted({item.get("project", "") for item in sheets if item.get("project")})),
        "pi": pi_name,
        "owner": sheet.get("owner", ""),
        "funding": "、".join(sorted({item.get("funding", "") for item in sheets if item.get("funding")})),
        "sourceType": "quantity_sheet",
        "sourceId": sheet["id"],
        "sourceIds": [item["id"] for item in sheets],
        "sourceLabel": "数量统计表",
        "roomName": "、".join(sorted({item.get("roomName", "") for item in sheets if item.get("roomName")})),
        "manager": "、".join(sorted({item.get("manager", "") for item in sheets if item.get("manager")})),
        "billingUnit": "cage_day",
        "principalType": principal_type,
        "freeCageAllowance": free_cages,
        "tierLimit": BILLING_TIER_LIMIT,
        "baseUnitPrice": BILLING_TIER_BASE_PRICE,
        "overageUnitPrice": BILLING_TIER_OVER_PRICE,
        "totalCageDays": sum(line["cageCount"] for line in lines),
        "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
        "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
        "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
        "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
        "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
        "totalAmount": lines[-1]["cumulative"] if lines else 0,
        "status": status,
        "generatedAt": generated_at,
        "lockedAt": generated_at if status == "locked" else "",
    }
    for line in lines:
        line["statementId"] = statement["id"]

    if not persist:
        return statement, lines, []

    workflow, version, statement, lines, workflow_events = save_billing_statement_workflow(
        conn,
        statement,
        lines,
        actor,
        f"根据数量统计表生成 {statement_iacuc} {sheet['month']} 饲养费结算单",
    )
    event = audit_event(
        actor,
        "billing_statement.generated_from_quantity_sheet",
        "billing_workflow",
        workflow["id"],
        f"{actor['displayName']} 根据数量统计表生成 {statement_iacuc} {sheet['month']} 饲养费结算单",
        [],
        generated_at,
        sheet,
        {"workflow": workflow, "version": version},
    )
    write_audit_events(conn, [event])
    return statement, lines, merge_audit_logs([], [event])


def generate_billing_statement(conn, payload, actor):
    iacuc = normalize_iacuc_number(payload.get("iacuc", ""))
    requested_pi = clean_text(payload.get("pi", ""))
    month = clean_text(payload.get("month", ""))
    status = clean_text(payload.get("status", "draft")) or "draft"
    persist = bool(payload.get("persist"))
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("结算月份格式应为 YYYY-MM")
    if status not in ("draft", "locked"):
        raise ValueError("结算单状态只能是 draft 或 locked")
    if not iacuc:
        raise ValueError("请先选择伦理号后再生成结算单")

    occupancies = read_occupancies_for_billing(conn, month, iacuc=iacuc)
    applications_by_iacuc = read_applications_by_iacuc(conn)
    state = read_billing_state_for_occupancies(conn, occupancies)
    rooms = state["rooms"]
    dates = dates_in_month(month)
    generated_at = now_iso()
    cumulative = 0
    lines = []
    pi_name = requested_pi or pi_for_iacuc(iacuc, applications_by_iacuc, occupancies)
    if not pi_name:
        raise ValueError("项目负责人不能为空")
    principal_type_by_pi = read_principal_type_by_pi(conn)
    principal_type = principal_type_by_pi.get(pi_name, BILLING_PRINCIPAL_INDEPENDENT)
    # IACUC 分表阶段不应用 PI 免费笼位，避免跨伦理号结算失真。
    free_cages = 0
    iacucs = [iacuc]

    for line_date in dates:
        active_items = [
            item
            for item in occupancies
            if normalize_iacuc_number(item.get("iacuc", "")) == iacuc and occupancy_active_on_date(item, line_date)
        ]
        charge_groups = {}
        cage_count = 0
        animal_count = 0
        for item in active_items:
            profile = billing_profile_for_occupancy(item, state)
            if profile["unit"] == "animal_day":
                count = occupancy_animal_count(item, profile)
                animal_count += count
            else:
                count = 1
                cage_count += 1
            add_charge_group(charge_groups, profile, count)
        charges = combined_daily_charge(charge_groups, free_cages)
        amount = charges["amount"]
        cumulative += amount
        breakdown = []
        if active_items:
            for item in active_items:
                profile = billing_profile_for_occupancy(item, state)
                found = next(
                    (
                        entry
                        for entry in breakdown
                        if entry["iacuc"] == iacuc
                        and entry.get("billingItem") == profile["billingItem"]
                        and entry.get("customerType") == profile["customerType"]
                    ),
                    None,
                )
                if not found:
                    found = {
                        "iacuc": iacuc,
                        "project": statement_application_snapshot(iacuc, applications_by_iacuc, occupancies).get(
                            "project", ""
                        ),
                        "animalCount": 0,
                        "cageCount": 0,
                        "billingItem": profile["billingItem"],
                        "billingUnit": profile["unit"],
                        "customerType": profile["customerType"],
                        "unitPrice": profile["unitPrice"],
                        "overageUnitPrice": BILLING_TIER_OVER_PRICE if profile["tiered"] else 0,
                        "tiered": bool(profile["tiered"]),
                        "freeAllowance": bool(profile["freeAllowance"]),
                    }
                    breakdown.append(found)
                if profile["unit"] == "animal_day":
                    found["animalCount"] += occupancy_animal_count(item, profile)
                else:
                    found["cageCount"] += 1
        line = {
            "id": new_id("line"),
            "date": line_date,
            "animalCount": animal_count,
            "cageCount": cage_count,
            **charges,
            "amount": amount,
            "cumulative": cumulative,
            "iacucBreakdown": breakdown,
            "occupancyIds": [item.get("id") for item in active_items if item.get("id")],
        }
        lines.append(line)

    application = statement_application_snapshot(iacuc, applications_by_iacuc, occupancies)
    statement = {
        "id": new_id("stmt"),
        "iacuc": iacuc,
        "iacucs": iacucs,
        "month": month,
        "project": application.get("project", ""),
        "pi": pi_name,
        "owner": application.get("owner", ""),
        "funding": application.get("funding", ""),
        "sourceType": "cage_map",
        "sourceLabel": "动态笼位图",
        "billingUnit": statement_billing_unit_from_lines(lines),
        "principalType": principal_type,
        "freeCageAllowance": free_cages,
        "tierLimit": BILLING_TIER_LIMIT,
        "baseUnitPrice": BILLING_TIER_BASE_PRICE,
        "overageUnitPrice": BILLING_TIER_OVER_PRICE,
        "totalCageDays": sum(line["cageCount"] for line in lines),
        "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
        "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
        "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
        "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
        "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
        "totalAmount": cumulative,
        "status": status,
        "generatedAt": generated_at,
        "lockedAt": generated_at if status == "locked" else "",
    }
    for line in lines:
        line["statementId"] = statement["id"]

    if not persist:
        return statement, lines, []

    workflow, version, statement, lines, workflow_events = save_billing_statement_workflow(
        conn,
        statement,
        lines,
        actor,
        f"生成 {pi_name} {month} 饲养费结算单",
    )
    detail_context = occupancy_detail_context(occupancies, rooms)
    upsert_reimbursement_record_from_statement(conn, workflow, statement, lines, detail_context, "workflow")
    recalculate_reimbursement_accumulations(conn, pi_name)
    event = audit_event(
        actor,
        "billing_statement.generated",
        "billing_workflow",
        workflow["id"],
        f"{actor['displayName']} 生成 {pi_name} {month} 饲养费结算单",
        [],
        generated_at,
        None,
        {"workflow": workflow, "version": version},
    )
    write_audit_events(conn, [event])
    return statement, lines, merge_audit_logs([], [event])


def generate_billing_statement_by_pi(conn, payload, actor):
    month = clean_text(payload.get("month", ""))
    pi_name = clean_text(payload.get("pi", ""))
    status = clean_text(payload.get("status", "draft")) or "draft"
    source_type = clean_text(payload.get("sourceType", "cage_map")) or "cage_map"
    persist = bool(payload.get("persist"))
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("结算月份格式应为 YYYY-MM")
    if status not in ("draft", "locked"):
        raise ValueError("结算单状态只能是 draft 或 locked")
    if source_type not in ("cage_map", "quantity_sheet"):
        raise ValueError("sourceType 只能是 cage_map 或 quantity_sheet")
    if not pi_name:
        raise ValueError("按 PI 合表需要提供项目负责人")

    principal_type_by_pi = read_principal_type_by_pi(conn)
    principal_type = principal_type_by_pi.get(pi_name, BILLING_PRINCIPAL_INDEPENDENT)
    free_cages = billing_free_cages_for_pi(principal_type_by_pi, pi_name)
    generated_at = now_iso()
    iacucs = []

    if source_type == "cage_map":
        applications_by_iacuc = read_applications_by_iacuc(conn)
        occupancies = read_occupancies_for_billing(conn, month, pi=pi_name)
        state = read_billing_state_for_occupancies(conn, occupancies)
        rooms = state["rooms"]
        iacucs = sorted(
            {
                normalize_iacuc_number(item.get("iacuc", ""))
                for item in occupancies
                if clean_text(item.get("pi", "")) == pi_name and normalize_iacuc_number(item.get("iacuc", ""))
            }
        )
        cumulative = 0
        lines = []
        for date in dates_in_month(month):
            active_items = [
                item
                for item in occupancies
                if clean_text(item.get("pi", "")) == pi_name and occupancy_active_on_date(item, date)
            ]
            charge_groups = {}
            cage_count = 0
            animal_count = 0
            for item in active_items:
                profile = billing_profile_for_occupancy(item, state)
                if profile["unit"] == "animal_day":
                    count = occupancy_animal_count(item, profile)
                    animal_count += count
                else:
                    count = 1
                    cage_count += 1
                add_charge_group(charge_groups, profile, count)
            breakdown = []
            for item in active_items:
                item_iacuc = normalize_iacuc_number(item.get("iacuc", ""))
                profile = billing_profile_for_occupancy(item, state)
                found = next(
                    (
                        entry
                        for entry in breakdown
                        if entry["iacuc"] == item_iacuc
                        and entry.get("billingItem") == profile["billingItem"]
                        and entry.get("customerType") == profile["customerType"]
                    ),
                    None,
                )
                if not found:
                    found = {
                        "iacuc": item_iacuc,
                        "project": item.get("project", ""),
                        "animalCount": 0,
                        "cageCount": 0,
                        "billingItem": profile["billingItem"],
                        "billingUnit": profile["unit"],
                        "customerType": profile["customerType"],
                        "unitPrice": profile["unitPrice"],
                        "overageUnitPrice": BILLING_TIER_OVER_PRICE if profile["tiered"] else 0,
                        "tiered": bool(profile["tiered"]),
                        "freeAllowance": bool(profile["freeAllowance"]),
                        "freeEligible": iacuc_free_allowance_eligible(
                            applications_by_iacuc.get(item_iacuc, item), date
                        ),
                        "freeCages": 0,
                    }
                    breakdown.append(found)
                if profile["unit"] == "animal_day":
                    found["animalCount"] += occupancy_animal_count(item, profile)
                else:
                    found["cageCount"] += 1
            free_allocations = allocate_daily_free_cages_by_iacuc(breakdown, free_cages)
            apply_free_cage_allocations(breakdown, free_allocations)
            charges = combined_daily_charge(charge_groups, sum(free_allocations.values()))
            cumulative += charges["amount"]
            lines.append(
                {
                    "id": new_id("line"),
                    "date": date,
                    "animalCount": animal_count,
                    "cageCount": cage_count,
                    **charges,
                    "amount": charges["amount"],
                    "cumulative": cumulative,
                    "iacucBreakdown": breakdown,
                    "occupancyIds": [item.get("id") for item in active_items if item.get("id")],
                }
            )
        application = statement_pi_snapshot(pi_name, applications_by_iacuc, occupancies)
        statement = {
            "id": new_id("stmt"),
            "iacuc": f"pi::{pi_name}",
            "iacucs": iacucs,
            "month": month,
            "project": application.get("project", ""),
            "pi": pi_name,
            "owner": application.get("owner", ""),
            "funding": application.get("funding", ""),
            "sourceType": "pi_merged_cage_map",
            "sourceLabel": "动态笼位图（按 PI 合表）",
            "billingUnit": statement_billing_unit_from_lines(lines),
            "principalType": principal_type,
            "freeCageAllowance": free_cages,
            "tierLimit": BILLING_TIER_LIMIT,
            "baseUnitPrice": BILLING_TIER_BASE_PRICE,
            "overageUnitPrice": BILLING_TIER_OVER_PRICE,
            "totalCageDays": sum(line["cageCount"] for line in lines),
            "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
            "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
            "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
            "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
            "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
            "totalAmount": cumulative,
            "status": status,
            "generatedAt": generated_at,
            "lockedAt": generated_at if status == "locked" else "",
        }
        detail_context = occupancy_detail_context(occupancies, rooms)
    else:
        sheets = list_quantity_sheets_by_month_pi(conn, month, pi_name)
        if not sheets:
            raise ValueError("未找到该 PI 在结算月份内的数量统计表")
        for item in sheets:
            validate_quantity_sheet_permission(actor, item)
        iacucs = sorted({normalize_iacuc_number(item.get("iacuc", "")) for item in sheets if item.get("iacuc")})
        rooms = read_rooms_for_quantity_sheets(conn, sheets)
        applications_by_iacuc = read_applications_by_iacuc(conn)
        lines = quantity_sheet_statement_lines(sheets, free_cages, rooms, applications_by_iacuc)
        statement = {
            "id": new_id("stmt"),
            "iacuc": f"pi::{pi_name}",
            "iacucs": iacucs,
            "month": month,
            "project": "、".join(sorted({item.get("project", "") for item in sheets if item.get("project")})),
            "pi": pi_name,
            "owner": "、".join(sorted({item.get("owner", "") for item in sheets if item.get("owner")})),
            "funding": "、".join(sorted({item.get("funding", "") for item in sheets if item.get("funding")})),
            "sourceType": "pi_merged_quantity_sheet",
            "sourceIds": [item["id"] for item in sheets],
            "sourceLabel": "数量统计表（按 PI 合表）",
            "roomName": "、".join(sorted({item.get("roomName", "") for item in sheets if item.get("roomName")})),
            "manager": "、".join(sorted({item.get("manager", "") for item in sheets if item.get("manager")})),
            "billingUnit": statement_billing_unit_from_lines(lines),
            "principalType": principal_type,
            "freeCageAllowance": free_cages,
            "tierLimit": BILLING_TIER_LIMIT,
            "baseUnitPrice": BILLING_TIER_BASE_PRICE,
            "overageUnitPrice": BILLING_TIER_OVER_PRICE,
            "totalCageDays": sum(line["cageCount"] for line in lines),
            "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
            "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
            "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
            "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
            "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
            "totalAmount": lines[-1]["cumulative"] if lines else 0,
            "status": status,
            "generatedAt": generated_at,
            "lockedAt": generated_at if status == "locked" else "",
        }
        detail_context = quantity_sheet_detail_context(sheets, rooms)

    for line in lines:
        line["statementId"] = statement["id"]
    if not persist:
        return statement, lines, []
    workflow, version, statement, lines, workflow_events = save_billing_statement_workflow(
        conn,
        statement,
        lines,
        actor,
        f"按 PI 合表生成 {pi_name} {month} 饲养费结算单",
    )
    upsert_reimbursement_record_from_statement(conn, workflow, statement, lines, detail_context, "workflow")
    recalculate_reimbursement_accumulations(conn, pi_name)
    event = audit_event(
        actor,
        "billing_statement.generated_by_pi",
        "billing_workflow",
        workflow["id"],
        f"{actor['displayName']} 按 PI 合表生成 {pi_name} {month} 饲养费结算单",
        [],
        generated_at,
        None,
        {"workflow": workflow, "version": version},
    )
    write_audit_events(conn, [event])
    return statement, lines, merge_audit_logs([], [event])


def read_principal_type_by_pi(conn):
    return {
        clean_text(pi_name): normalize_principal_type(principal_type)
        for pi_name, principal_type in read_principal_type_by_pi_repository(conn).items()
        if clean_text(pi_name)
    }


def normalize_workflow_source(source_type):
    text = clean_text(source_type or "")
    return text if text else "cage_map"


def workflow_scope_for_statement(statement):
    source_type = normalize_workflow_source(statement.get("sourceType", ""))
    pi_name = clean_text(statement.get("pi", ""))
    if source_type.startswith("pi_merged_") and pi_name:
        return "pi", f"pi::{pi_name}"
    if source_type in ("cage_map", "quantity_sheet") and pi_name:
        return "pi", f"pi::{pi_name}"
    iacuc = clean_text(statement.get("iacuc", ""))
    return "iacuc", iacuc


def billing_workflow_business_key(scope_type, scope_key, month, source_type):
    return "|".join(
        [clean_text(scope_type), clean_text(scope_key), clean_text(month), normalize_workflow_source(source_type)]
    )


def make_statement_document_number(statement, version_no):
    source = normalize_workflow_source(statement.get("sourceType", ""))
    source_code = {
        "quantity_sheet": "QS",
        "pi_merged_quantity_sheet": "PQS",
        "pi_merged_cage_map": "PCM",
    }.get(source, "CM")
    month = re.sub(r"\D", "", clean_text(statement.get("month", "")) or "000000")
    iacuc = re.sub(r"[^A-Za-z0-9]", "", clean_text(statement.get("iacuc", "")) or "UNKNOWN").upper()
    return f"CL-{source_code}-{month}-{iacuc}-V{int(version_no):02d}"


def enrich_statement_for_workflow(
    statement,
    *,
    workflow_id,
    version_id,
    version_no,
    version_status,
    workflow_status,
    document_number,
):
    return {
        **statement,
        "id": version_id,
        "workflowId": workflow_id,
        "versionId": version_id,
        "versionNo": version_no,
        "versionStatus": version_status,
        "workflowStatus": workflow_status,
        "documentNumber": document_number,
    }


def build_version_payload(
    statement,
    workflow_id,
    version_no,
    version_status,
    workflow_status,
    generated_at,
    voided_at,
    voided_by,
    void_reason,
):
    return {
        "id": statement["id"],
        "workflowId": workflow_id,
        "versionNo": version_no,
        "versionStatus": version_status,
        "workflowStatus": workflow_status,
        "generatedAt": generated_at,
        "voidedAt": voided_at,
        "voidedBy": voided_by,
        "voidReason": void_reason,
        "documentNumber": statement.get("documentNumber", ""),
        "statement": statement,
        "summary": {
            "iacuc": statement.get("iacuc", ""),
            "iacucs": statement.get("iacucs", []),
            "month": statement.get("month", ""),
            "sourceType": statement.get("sourceType", ""),
            "pi": statement.get("pi", ""),
            "project": statement.get("project", ""),
            "owner": statement.get("owner", ""),
            "funding": statement.get("funding", ""),
            "totalAmount": statement.get("totalAmount", 0),
            "totalCageDays": statement.get("totalCageDays", 0),
            "status": statement.get("status", "draft"),
        },
    }


def build_workflow_payload(workflow_id, iacuc, month, source_type, workflow_status, current_version, latest_event_at):
    statement = current_version.get("statement", {})
    scope_type, scope_key = workflow_scope_for_statement(statement)
    timestamps = {
        "generatedAt": current_version.get("generatedAt", ""),
        "sentAt": statement.get("sentAt", ""),
        "signedReturnedAt": statement.get("signedReturnedAt", ""),
        "submittedToFinanceAt": statement.get("submittedToFinanceAt", ""),
    }
    return {
        "id": workflow_id,
        "businessKey": billing_workflow_business_key(scope_type, scope_key, month, source_type),
        "scopeType": scope_type,
        "scopeKey": scope_key,
        "iacuc": iacuc,
        "iacucs": statement.get("iacucs", []),
        "month": month,
        "sourceType": source_type,
        "workflowStatus": workflow_status,
        "currentVersionId": current_version.get("id", ""),
        "currentVersionNo": current_version.get("versionNo", 0),
        "currentVersion": current_version,
        "latestEventAt": latest_event_at,
        "pi": statement.get("pi", ""),
        "project": statement.get("project", ""),
        "owner": statement.get("owner", ""),
        "funding": statement.get("funding", ""),
        "totalAmount": statement.get("totalAmount", 0),
        "totalCageDays": statement.get("totalCageDays", 0),
        **timestamps,
    }


def build_workflow_event_payload(
    event_id, workflow_id, version_id, event_type, from_status, to_status, actor, at, channel, note
):
    return {
        "id": event_id,
        "workflowId": workflow_id,
        "versionId": version_id,
        "eventType": event_type,
        "fromStatus": from_status,
        "toStatus": to_status,
        "actor": {
            "id": actor.get("id", ""),
            "username": actor.get("username", ""),
            "displayName": actor.get("displayName", ""),
        },
        "channel": channel,
        "note": note,
        "at": at,
    }


def get_billing_workflow_by_key(conn, business_key):
    return get_billing_workflow_by_key_repository(conn, business_key)


def get_billing_workflow(conn, workflow_id):
    return get_billing_workflow_repository(conn, workflow_id)


def get_billing_workflow_detail(conn, workflow_id):
    return get_billing_workflow_detail_repository(conn, workflow_id)


def get_billing_version(conn, version_id):
    return get_billing_version_repository(conn, version_id)


def list_billing_workflows(conn):
    return list_billing_workflows_repository(conn)


def list_billing_workflows_page(conn, filters):
    return list_billing_workflows_page_repository(conn, filters, clean_text, WORKFLOW_STATUS_FINANCE)


def list_billing_workflow_lines(conn, workflow_id, version_id=""):
    workflow = get_billing_workflow(conn, workflow_id)
    if not workflow:
        raise LookupError("结算流程不存在")
    selected_version_id = clean_text(version_id) or clean_text(workflow.get("currentVersionId", ""))
    if not selected_version_id:
        return {"workflowId": workflow_id, "versionId": "", "lines": []}
    return {
        "workflowId": workflow_id,
        "versionId": selected_version_id,
        "lines": list_billing_statement_line_summaries_for_version_repository(conn, selected_version_id),
    }


def delete_billing_workflow(conn, workflow_id):
    workflow = get_billing_workflow(conn, workflow_id)
    if not workflow:
        raise LookupError("结算流程不存在")
    delete_billing_workflow_tree_repository(conn, workflow_id)
    return workflow


def list_billing_workflow_versions(conn, workflow_id):
    return list_billing_workflow_versions_repository(conn, workflow_id)


def billing_workflow_detail_item(workflow):
    return billing_workflow_detail_item_repository(workflow)


def list_billing_workflow_events(conn, workflow_id):
    return list_billing_workflow_events_repository(conn, workflow_id)


def list_billing_statement_lines_for_version(conn, version_id):
    return list_billing_statement_lines_for_version_repository(conn, version_id)


def list_current_billing_statements(conn):
    return list_current_billing_statements_repository(conn)


def list_reimbursement_records_page(conn, filters):
    return list_reimbursement_records_page_repository(conn, filters, clean_text)


def get_reimbursement_record(conn, record_id):
    return get_reimbursement_record_repository(conn, record_id)


def get_reimbursement_record_by_key(conn, business_key):
    return get_reimbursement_record_by_key_repository(conn, business_key)


def get_reimbursement_record_by_workflow_id(conn, workflow_id):
    return get_reimbursement_record_by_workflow_id_repository(conn, workflow_id)


def list_reimbursement_records_for_pi(conn, pi_name):
    return list_reimbursement_records_for_pi_repository(conn, pi_name)


def list_reimbursement_record_summaries_for_pi(conn, pi_name):
    return list_reimbursement_record_summaries_for_pi_repository(conn, pi_name)


def upsert_reimbursement_record(conn, payload):
    existing_by_key = get_reimbursement_record_by_key(conn, payload.get("businessKey", ""))
    if existing_by_key and existing_by_key.get("id"):
        payload["id"] = existing_by_key["id"]
    else:
        row = conn.execute(
            "SELECT payload FROM reimbursement_records WHERE id = ?", (payload.get("id", ""),)
        ).fetchone()
        if row:
            existing_by_id = json.loads(row["payload"])
            payload["id"] = existing_by_id.get("id", payload.get("id", ""))
            if not payload.get("businessKey") and existing_by_id.get("businessKey"):
                payload["businessKey"] = existing_by_id["businessKey"]
    upsert_reimbursement_record_repository(conn, payload)


def delete_reimbursement_record(conn, record_id):
    delete_reimbursement_record_repository(conn, record_id)


def reimbursement_status_label(value):
    return {
        REIMBURSEMENT_STATUS_PENDING: "待提交",
        REIMBURSEMENT_STATUS_REIMBURSING: "报销中",
        REIMBURSEMENT_STATUS_COMPLETED: "已完成",
    }.get(normalize_reimbursement_status(value), "待提交")


def reimbursement_record_id(month, pi_name):
    raw = f"{clean_text(month)}|{clean_text(pi_name)}"
    slug = re.sub(r"[^A-Za-z0-9]+", "-", clean_text(month)).strip("-") or "record"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]
    return f"rrc-{slug}-{digest}"


def quantity_sheet_detail_context(sheets, rooms):
    room_by_id = {room.get("id"): room for room in rooms}
    context = {}
    for sheet in sheets:
        iacuc = normalize_iacuc_number(sheet.get("iacuc", ""))
        if not iacuc:
            continue
        room = room_by_id.get(sheet.get("roomId")) or next(
            (item for item in rooms if clean_text(item.get("name")) == clean_text(sheet.get("roomName"))), {}
        )
        profile = billing_profile_for_room(room or {})
        current = context.get(iacuc) or {
            "facility": profile.get("facility", ""),
            "funding": "",
            "species": room.get("defaultSpecies", ""),
            "project": "",
            "owner": "",
            "roomNames": set(),
        }
        if sheet.get("funding"):
            current["funding"] = join_distinct_text(current.get("funding", ""), sheet.get("funding", ""))
        if sheet.get("project"):
            current["project"] = join_distinct_text(current.get("project", ""), sheet.get("project", ""))
        if sheet.get("owner"):
            current["owner"] = join_distinct_text(current.get("owner", ""), sheet.get("owner", ""))
        if room.get("defaultSpecies"):
            current["species"] = join_distinct_text(current.get("species", ""), room.get("defaultSpecies", ""))
        if sheet.get("roomName"):
            current["roomNames"].add(sheet.get("roomName"))
        context[iacuc] = current
    return finalize_reimbursement_detail_context(context)


def occupancy_detail_context(occupancies, rooms):
    room_by_id = {room.get("id"): room for room in rooms}
    context = {}
    for item in occupancies:
        iacuc = normalize_iacuc_number(item.get("iacuc", ""))
        if not iacuc:
            continue
        room = room_by_id.get(item.get("roomId"), {})
        profile = billing_profile_for_room(room or {})
        current = context.get(iacuc) or {
            "facility": profile.get("facility", ""),
            "funding": "",
            "species": "",
            "project": "",
            "owner": "",
            "roomNames": set(),
        }
        current["funding"] = join_distinct_text(current.get("funding", ""), item.get("funding", ""))
        current["species"] = join_distinct_text(current.get("species", ""), item.get("species", ""))
        current["project"] = join_distinct_text(current.get("project", ""), item.get("project", ""))
        current["owner"] = join_distinct_text(current.get("owner", ""), item.get("owner", ""))
        if item.get("roomName"):
            current["roomNames"].add(item.get("roomName"))
        context[iacuc] = current
    return finalize_reimbursement_detail_context(context)


def finalize_reimbursement_detail_context(context):
    finalized = {}
    for iacuc, item in context.items():
        finalized[iacuc] = {
            **item,
            "roomNames": sorted(item.get("roomNames", set())),
        }
    return finalized


def join_distinct_text(current, value):
    values = [part for part in [clean_text(current), clean_text(value)] if part]
    if not values:
        return ""
    return "、".join(sorted(set(values)))


def reimbursement_detail_context_from_workflow(conn, workflow, statement):
    detail_context = {}
    iacucs = [normalize_iacuc_number(value) for value in statement.get("iacucs", []) if normalize_iacuc_number(value)]
    applications_by_iacuc = read_applications_by_iacuc(conn)
    if statement.get("sourceType") == "pi_merged_quantity_sheet":
        sheets = list_quantity_sheets_by_month_pi(conn, statement.get("month"), clean_text(statement.get("pi")))
        rooms = read_room_payloads_for_context(
            conn,
            room_ids=[item.get("roomId", "") for item in sheets],
            room_names=[item.get("roomName", "") for item in sheets],
        )
        detail_context = quantity_sheet_detail_context(sheets, rooms)
    elif statement.get("sourceType") == "quantity_sheet":
        sheets = list_quantity_sheets_by_month_iacuc(
            conn, statement.get("month"), normalize_iacuc_number(statement.get("iacuc"))
        )
        rooms = read_room_payloads_for_context(
            conn,
            room_ids=[item.get("roomId", "") for item in sheets],
            room_names=[item.get("roomName", "") for item in sheets],
        )
        detail_context = quantity_sheet_detail_context(sheets, rooms)
    else:
        occupancies = read_occupancies_for_billing(
            conn,
            statement.get("month", ""),
            iacuc=""
            if clean_text(statement.get("sourceType", "")).startswith("pi_merged_")
            else statement.get("iacuc", ""),
            pi=statement.get("pi", "") if clean_text(statement.get("pi", "")) else "",
        )
        rooms = read_room_payloads_for_context(conn, room_ids=[item.get("roomId", "") for item in occupancies])
        detail_context = occupancy_detail_context(occupancies, rooms)
    for iacuc in iacucs:
        snapshot = statement_application_snapshot(iacuc, applications_by_iacuc, [])
        current = detail_context.get(iacuc, {"roomNames": []})
        detail_context[iacuc] = {
            **current,
            "funding": current.get("funding") or snapshot.get("funding", ""),
            "project": current.get("project") or snapshot.get("project", ""),
            "owner": current.get("owner") or snapshot.get("owner", ""),
        }
    return detail_context


def build_reimbursement_record_payload(existing, workflow, statement, lines, detail_context_by_iacuc, source):
    summary = summarize_statement(statement, lines, detail_context_by_iacuc, BILLING_TIER_LIMIT)
    month = clean_text(statement.get("month", ""))
    pi_name = clean_text(statement.get("pi", ""))
    business_key = reimbursement_business_key(month, pi_name)
    current_month_amount = coerce_reimbursement_money(statement.get("totalAmount", 0))
    support_amount = coerce_reimbursement_money(summary.get("supportAmount", 0))
    payable_amount = coerce_reimbursement_money(summary.get("payableAmount", current_month_amount - support_amount))
    paid_amount = coerce_reimbursement_money(existing.get("paidAmount", 0) if existing else 0)
    reimbursement_status = normalize_reimbursement_status(existing.get("reimbursementStatus") if existing else "")
    payload = {
        "id": existing.get("id") if existing else reimbursement_record_id(month, pi_name),
        "businessKey": business_key,
        "month": month,
        "pi": pi_name,
        "workflowId": workflow.get("id", "") if workflow else existing.get("workflowId", "") if existing else "",
        "workflowStatus": workflow.get("workflowStatus", "")
        if workflow
        else existing.get("workflowStatus", "")
        if existing
        else "",
        "reimbursementStatus": reimbursement_status,
        "currentMonthAmount": current_month_amount,
        "supportAmount": support_amount,
        "payableAmount": payable_amount,
        "paidAmount": paid_amount,
        "unpaidAmount": coerce_reimbursement_money(max(payable_amount - paid_amount, 0)),
        "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0) if existing else 0),
        "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0) if existing else 0),
        "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0) if existing else 0),
        "fundBookNo": clean_text(existing.get("fundBookNo", "") if existing else ""),
        "reimbursementFormNo": clean_text(existing.get("reimbursementFormNo", "") if existing else ""),
        "approvedBudget": existing.get("approvedBudget", "") if existing else "",
        "notes": clean_text(existing.get("notes", "") if existing else ""),
        "completedAt": clean_text(existing.get("completedAt", "") if existing else ""),
        "source": source,
        "latestEventAt": workflow.get("latestEventAt", "") if workflow else statement.get("generatedAt", ""),
        "updatedAt": now_iso(),
        "details": summary.get("details", []),
        "iacucs": [detail.get("iacuc", "") for detail in summary.get("details", []) if detail.get("iacuc")],
        "statementVersionId": statement.get("versionId", "") or statement.get("id", ""),
        "documentNumber": statement.get("documentNumber", ""),
        "billingUnit": statement.get("billingUnit", ""),
        "project": statement.get("project", ""),
        "owner": statement.get("owner", ""),
        "funding": statement.get("funding", ""),
    }
    if (
        payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED
        and payload["paidAmount"] + 1e-9 < payload["payableAmount"]
    ):
        payload["reimbursementStatus"] = (
            REIMBURSEMENT_STATUS_REIMBURSING
            if (payload["fundBookNo"] or payload["reimbursementFormNo"])
            else REIMBURSEMENT_STATUS_PENDING
        )
        payload["completedAt"] = ""
    return payload


def upsert_reimbursement_record_from_statement(
    conn, workflow, statement, lines, detail_context_by_iacuc, source="workflow"
):
    business_key = reimbursement_business_key(statement.get("month", ""), statement.get("pi", ""))
    existing = get_reimbursement_record_by_key(conn, business_key) or {}
    payload = build_reimbursement_record_payload(existing, workflow, statement, lines, detail_context_by_iacuc, source)
    upsert_reimbursement_record(conn, payload)
    return payload


def recalculate_reimbursement_accumulations(conn, pi_name):
    records = list_reimbursement_records_for_pi(conn, clean_text(pi_name))
    records = sorted(
        records,
        key=lambda item: (
            clean_text(item.get("month", "")),
            clean_text(item.get("latestEventAt", "")),
            clean_text(item.get("id", "")),
        ),
    )
    accumulated_payable = 0.0
    accumulated_paid = 0.0
    for record in records:
        payable_amount = coerce_reimbursement_money(record.get("payableAmount", 0))
        paid_amount = coerce_reimbursement_money(record.get("paidAmount", 0))
        record["unpaidAmount"] = coerce_reimbursement_money(max(payable_amount - paid_amount, 0))
        accumulated_payable += payable_amount
        accumulated_paid += paid_amount
        record["accumulatedPayable"] = coerce_reimbursement_money(accumulated_payable)
        record["accumulatedPaid"] = coerce_reimbursement_money(accumulated_paid)
        record["accumulatedUnpaid"] = coerce_reimbursement_money(accumulated_payable - accumulated_paid)
        record["updatedAt"] = now_iso()
        upsert_reimbursement_record(conn, record)


def recalculate_all_reimbursement_accumulations(conn):
    rows = conn.execute("SELECT DISTINCT pi FROM reimbursement_records WHERE TRIM(COALESCE(pi, '')) != ''").fetchall()
    for row in rows:
        recalculate_reimbursement_accumulations(conn, row["pi"])


def import_monthly_reimbursement_workbook(conn, file_body, actor):
    ensure_excel_import_supported()
    workbook = openpyxl.load_workbook(io.BytesIO(file_body), data_only=True)
    groups = build_monthly_import_groups(workbook)
    if not groups:
        raise ValueError("未识别到可导入的月度汇总数据")
    saved = []
    audits = []
    imported_months = set()
    imported_pis = set()
    for entry in groups.values():
        existing = get_reimbursement_record_by_key(conn, reimbursement_business_key(entry["month"], entry["pi"])) or {}
        fund_book_no = "；".join(distinct_text_list(entry["fundBookNos"]))
        reimbursement_form_no = "；".join(distinct_text_list(entry["reimbursementFormNos"]))
        notes = "；".join(distinct_text_list(entry["notes"]))
        payload = {
            "id": existing.get("id") or next_imported_record_id(),
            "businessKey": reimbursement_business_key(entry["month"], entry["pi"]),
            "month": entry["month"],
            "pi": entry["pi"],
            "workflowId": existing.get("workflowId", ""),
            "workflowStatus": existing.get("workflowStatus", ""),
            "reimbursementStatus": normalize_reimbursement_status(
                existing.get("reimbursementStatus") or infer_import_status(fund_book_no, reimbursement_form_no, notes)
            ),
            "currentMonthAmount": coerce_reimbursement_money(entry["currentMonthAmount"]),
            "supportAmount": coerce_reimbursement_money(entry["supportAmount"]),
            "payableAmount": coerce_reimbursement_money(entry["payableAmount"]),
            "paidAmount": coerce_reimbursement_money(existing.get("paidAmount", 0)),
            "unpaidAmount": 0,
            "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0)),
            "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0)),
            "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0)),
            "fundBookNo": fund_book_no or clean_text(existing.get("fundBookNo", "")),
            "reimbursementFormNo": reimbursement_form_no or clean_text(existing.get("reimbursementFormNo", "")),
            "approvedBudget": existing.get("approvedBudget", ""),
            "notes": notes or clean_text(existing.get("notes", "")),
            "completedAt": clean_text(existing.get("completedAt", "")),
            "source": "imported",
            "latestEventAt": now_iso(),
            "updatedAt": now_iso(),
            "details": entry["details"],
            "iacucs": distinct_text_list(detail.get("iacuc", "") for detail in entry["details"]),
            "statementVersionId": existing.get("statementVersionId", ""),
            "documentNumber": existing.get("documentNumber", ""),
            "billingUnit": existing.get("billingUnit", ""),
            "project": clean_text(existing.get("project", "")),
            "owner": clean_text(existing.get("owner", "")),
            "funding": "；".join(distinct_text_list(entry["funding"])) or clean_text(existing.get("funding", "")),
        }
        payload["unpaidAmount"] = coerce_reimbursement_money(max(payload["payableAmount"] - payload["paidAmount"], 0))
        if (
            payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED
            and payload["paidAmount"] + 1e-9 < payload["payableAmount"]
        ):
            payload["reimbursementStatus"] = infer_import_status(
                payload["fundBookNo"], payload["reimbursementFormNo"], payload["notes"]
            )
            payload["completedAt"] = ""
        if payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED and not payload["completedAt"]:
            payload["completedAt"] = now_iso()
        upsert_reimbursement_record(conn, payload)
        saved.append(payload)
        imported_months.add(payload["month"])
        imported_pis.add(payload["pi"])
    for pi_name in imported_pis:
        recalculate_reimbursement_accumulations(conn, pi_name)
    event_time = now_iso()
    audit = audit_event(
        actor,
        "reimbursement.import_monthly",
        "reimbursement_record",
        f"monthly:{len(saved)}",
        f"{actor['displayName']} 导入月度报销台账 {len(saved)} 条",
        [],
        event_time,
        None,
        {"months": sorted(imported_months), "count": len(saved)},
    )
    write_audit_events(conn, [audit])
    audits.append(audit)
    invalidate_data_cache_prefixes("reimbursement_records::", "billing_workflows::")
    return {
        "items": saved,
        "auditLogs": merge_audit_logs([], audits),
        "count": len(saved),
        "months": sorted(imported_months),
    }


def import_arrears_reimbursement_workbook(conn, file_body, actor):
    ensure_excel_import_supported()
    workbook = openpyxl.load_workbook(io.BytesIO(file_body), data_only=True)
    if not workbook.worksheets:
        raise ValueError("欠缴工作簿为空")
    sheet = workbook.worksheets[0]
    month_columns = arrears_summary_columns(sheet)
    if not month_columns:
        raise ValueError("未识别到欠缴月份列")
    saved = []
    imported_pis = set()
    empty_rows = 0
    for row_index in range(2, 1200):
        pi_name = clean_text(sheet.cell(row_index, 2).value)
        total_amount = coerce_reimbursement_money(sheet.cell(row_index, 5).value)
        if not pi_name and total_amount <= 0:
            empty_rows += 1
            if empty_rows >= 20:
                break
            continue
        empty_rows = 0
        if not pi_name:
            continue
        species = clean_text(sheet.cell(row_index, 3).value)
        funding = clean_text(sheet.cell(row_index, 4).value)
        for column_index, month in month_columns:
            monthly_unpaid = coerce_reimbursement_money(sheet.cell(row_index, column_index).value)
            if monthly_unpaid <= 0:
                continue
            business_key = reimbursement_business_key(month, pi_name)
            existing = get_reimbursement_record_by_key(conn, business_key) or {}
            if existing and existing.get("source") != "imported":
                continue
            details = existing.get("details") or [
                {
                    "iacuc": "",
                    "facility": "",
                    "funding": funding,
                    "species": species,
                    "project": "",
                    "owner": "",
                    "amount": monthly_unpaid,
                    "supportAmount": 0,
                    "payableAmount": monthly_unpaid,
                    "roomNames": [],
                    "statementVersionId": "",
                }
            ]
            payload = {
                "id": existing.get("id") or next_imported_record_id("reim-arrears"),
                "businessKey": business_key,
                "month": month,
                "pi": pi_name,
                "workflowId": existing.get("workflowId", ""),
                "workflowStatus": existing.get("workflowStatus", ""),
                "reimbursementStatus": normalize_reimbursement_status(
                    existing.get("reimbursementStatus") or REIMBURSEMENT_STATUS_PENDING
                ),
                "currentMonthAmount": coerce_reimbursement_money(
                    existing.get("currentMonthAmount", monthly_unpaid) or monthly_unpaid
                ),
                "supportAmount": coerce_reimbursement_money(existing.get("supportAmount", 0)),
                "payableAmount": coerce_reimbursement_money(
                    existing.get("payableAmount", monthly_unpaid) or monthly_unpaid
                ),
                "paidAmount": coerce_reimbursement_money(existing.get("paidAmount", 0)),
                "unpaidAmount": 0,
                "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0)),
                "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0)),
                "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0)),
                "fundBookNo": clean_text(existing.get("fundBookNo", "")),
                "reimbursementFormNo": clean_text(existing.get("reimbursementFormNo", "")),
                "approvedBudget": existing.get("approvedBudget", ""),
                "notes": join_distinct_text(existing.get("notes", ""), f"欠缴汇算导入：累计欠缴 {total_amount:.2f} 元"),
                "completedAt": clean_text(existing.get("completedAt", "")),
                "source": "imported",
                "latestEventAt": now_iso(),
                "updatedAt": now_iso(),
                "details": details,
                "iacucs": distinct_text_list(detail.get("iacuc", "") for detail in details),
                "statementVersionId": existing.get("statementVersionId", ""),
                "documentNumber": existing.get("documentNumber", ""),
                "billingUnit": existing.get("billingUnit", ""),
                "project": clean_text(existing.get("project", "")),
                "owner": clean_text(existing.get("owner", "")),
                "funding": funding or clean_text(existing.get("funding", "")),
            }
            payload["unpaidAmount"] = coerce_reimbursement_money(
                max(payload["payableAmount"] - payload["paidAmount"], 0)
            )
            upsert_reimbursement_record(conn, payload)
            saved.append(payload)
            imported_pis.add(pi_name)
    if not saved:
        raise ValueError("未识别到可导入的欠缴记录")
    for pi_name in imported_pis:
        recalculate_reimbursement_accumulations(conn, pi_name)
    event_time = now_iso()
    audit = audit_event(
        actor,
        "reimbursement.import_arrears",
        "reimbursement_record",
        f"arrears:{len(saved)}",
        f"{actor['displayName']} 导入欠缴汇算 {len(saved)} 条",
        [],
        event_time,
        None,
        {"count": len(saved)},
    )
    write_audit_events(conn, [audit])
    invalidate_data_cache_prefixes("reimbursement_records::", "billing_workflows::")
    return {"items": saved, "auditLogs": merge_audit_logs([], [audit]), "count": len(saved)}


def reimbursement_detail_payload(conn, record):
    record_id = clean_text(record.get("id", ""))
    workflow_id = clean_text(record.get("workflowId", ""))
    cache_key_value = cache_key(
        "reimbursement_records::detail",
        record_id=record_id,
        workflow_id=workflow_id,
        updated_at=clean_text(record.get("updatedAt", "")),
        latest_event_at=clean_text(record.get("latestEventAt", "")),
    )
    cached = cache_get(cache_key_value)
    if cached is not None:
        return cached
    workflow = get_billing_workflow_detail(conn, record.get("workflowId", "")) if record.get("workflowId") else None
    workflow_versions = list_billing_workflow_versions(conn, workflow["id"]) if workflow else []
    workflow_events = list_billing_workflow_events(conn, workflow["id"]) if workflow else []
    history = list_reimbursement_record_summaries_for_pi(conn, record.get("pi", ""))
    payload = {
        "item": record,
        "workflow": workflow if workflow else None,
        "workflowVersions": workflow_versions,
        "workflowEvents": workflow_events,
        "history": history,
    }
    return cache_set(cache_key_value, payload)


def save_billing_statement_workflow(conn, statement, lines, actor, note=""):
    result = save_billing_statement_workflow_service(
        conn, statement, lines, actor, note, billing_workflow_service_deps()
    )
    invalidate_data_cache_prefixes("billing_workflows::", "billing_statements::", "reimbursement_records::")
    return result


def insert_billing_workflow(conn, payload):
    insert_billing_workflow_repository(
        conn,
        payload,
        payload.get(
            "businessKey",
            billing_workflow_business_key(
                payload.get("scopeType", ""),
                payload.get("scopeKey", ""),
                payload.get("month", ""),
                payload.get("sourceType", ""),
            ),
        ),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
        as_int(payload.get("currentVersionNo")) or 0,
    )


def update_billing_workflow(conn, payload):
    update_billing_workflow_repository(
        conn,
        payload,
        payload.get(
            "businessKey",
            billing_workflow_business_key(
                payload.get("scopeType", ""),
                payload.get("scopeKey", ""),
                payload.get("month", ""),
                payload.get("sourceType", ""),
            ),
        ),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
        as_int(payload.get("currentVersionNo")) or 0,
    )


def insert_billing_version(conn, payload):
    insert_billing_version_repository(
        conn,
        payload,
        as_int(payload.get("versionNo")) or 1,
        payload.get("versionStatus", VERSION_STATUS_ACTIVE),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
    )


def update_billing_version(conn, payload):
    update_billing_version_repository(
        conn,
        payload,
        as_int(payload.get("versionNo")) or 1,
        payload.get("versionStatus", VERSION_STATUS_ACTIVE),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
    )


def replace_version_lines(conn, version_id, lines):
    replace_billing_statement_version_lines_repository(conn, version_id, lines)


def insert_billing_workflow_event(conn, payload):
    insert_billing_workflow_event_repository(conn, payload)


def update_workflow_status(conn, workflow_id, next_status, actor, note=""):
    return update_workflow_status_service(conn, workflow_id, next_status, actor, note, billing_workflow_service_deps())


def parse_multipart_upload(content_type, raw):
    if "multipart/form-data" not in content_type:
        raise ValueError("请使用 multipart/form-data 上传文件")
    boundary = multipart_boundary(content_type)
    delimiter = b"--" + boundary
    for part in raw.split(delimiter):
        part = part.strip(b"\r\n")
        if not part or part == b"--" or b"\r\n\r\n" not in part:
            continue
        header_blob, body = part.split(b"\r\n\r\n", 1)
        headers = header_blob.decode("utf-8", errors="replace")
        disposition = next(
            (line for line in headers.split("\r\n") if line.lower().startswith("content-disposition:")), ""
        )
        if 'name="file"' not in disposition:
            continue
        filename = multipart_filename(disposition)
        return filename, body.rstrip(b"\r\n")
    raise ValueError("没有找到上传字段 file")


def multipart_boundary(content_type):
    for segment in content_type.split(";"):
        segment = segment.strip()
        if segment.startswith("boundary="):
            value = segment.split("=", 1)[1].strip()
            if value.startswith('"') and value.endswith('"'):
                value = value[1:-1]
            return value.encode("utf-8")
    raise ValueError("上传请求缺少 multipart boundary")


def multipart_filename(disposition):
    match = re.search(r'filename="([^"]*)"', disposition)
    return match.group(1) if match else ""


API_ROUTER = Router()
API_ROUTER.add(
    "GET",
    r"/api/health",
    lambda handler, params: JsonResponse({"ok": True, "database": str(DB_PATH), "system": system_info()}),
)
API_ROUTER.add("GET", r"/api/system/info", lambda handler, params: JsonResponse(system_info()))


def current_session_response(handler, params):
    user = handler.current_user()
    if not user:
        return JsonResponse({"user": None}, HTTPStatus.UNAUTHORIZED)
    return JsonResponse({"user": user})


API_ROUTER.add("GET", r"/api/auth/me", current_session_response)


class CageLedgerHandler(CageLedgerHttpHandler):
    server_version = "CageLedger/0.2"

    def do_GET(self):
        path = urlparse(self.path).path
        routed = API_ROUTER.dispatch("GET", path, self)
        if routed:
            self.send_json(routed.payload, routed.status)
            return
        if path.startswith("/api/public/cage-card/"):
            qr_id = unquote(path.rsplit("/", 1)[-1])
            try:
                with connect_db() as conn:
                    self.send_json(public_cage_card_payload(conn, qr_id))
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            return
        if path == "/api/state":
            user = self.require_user()
            if not user:
                return
            payload = read_state()
            self.send_json({**payload, "state": filter_state_for_actor(payload.get("state"), user)})
            return
        if path == "/api/bootstrap":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                query = parse_qs(urlparse(self.path).query)
                scope = clean_text(query.get("scope", ["summary"])[0]).lower() or "summary"
                room_id = clean_text(query.get("roomId", [""])[0])
                self.send_json(read_bootstrap_state(conn, user, scope, room_id))
            return
        if path == "/api/iacuc-index":
            if not self.require_user():
                return
            self.send_json(read_iacuc_index())
            return
        if path == "/api/iacuc-index/status":
            if not self.require_user():
                return
            payload = read_iacuc_index()
            self.send_json({key: payload[key] for key in ("count", "updatedAt", "source")})
            return
        if path == "/api/system/update-check":
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                self.send_json(system_update_status())
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_GATEWAY)
            return
        if path == "/api/users":
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            with connect_db() as conn:
                self.send_json({"users": list_users(conn)})
            return
        if path == "/api/quantity-sheets":
            if not self.require_user():
                return
            with connect_db() as conn:
                self.send_json(list_quantity_sheets_page(conn, self.list_filters()))
            return
        if path == "/api/quantity-sheets/filter-options":
            if not self.require_user():
                return
            query = parse_qs(urlparse(self.path).query)
            column = clean_text(query.get("column", [""])[0])
            with connect_db() as conn:
                self.send_json(list_quantity_sheet_filter_options(conn, self.list_filters(), column))
            return
        if path == "/api/quantity-sheet-rooms":
            if not self.require_user():
                return
            with connect_db() as conn:
                rooms = read_payloads(conn, "rooms", "rowid")
            self.send_json({"items": rooms})
            return
        sheet_id = self.quantity_sheet_route(path)
        if sheet_id:
            user = self.require_user()
            if not user:
                return
            try:
                with connect_db() as conn:
                    sheet = get_quantity_sheet(conn, sheet_id)
                    validate_quantity_sheet_permission(user, sheet)
                    self.send_json({"item": sheet})
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            return
        if path == "/api/principal-identities":
            user = self.require_user()
            if not user:
                return
            self.send_json({"items": read_principal_identities()})
            return
        if path == "/api/infrastructure/occupancies":
            user = self.require_user()
            if not user:
                return
            query = parse_qs(urlparse(self.path).query)
            filters = {
                "month": clean_text(query.get("month", [""])[0]),
                "iacuc": clean_text(query.get("iacuc", [""])[0]),
                "pi": clean_text(query.get("pi", [""])[0]),
            }
            try:
                with connect_db() as conn:
                    self.send_json(read_billing_occupancies(conn, user, filters))
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        if path == "/api/billing-workflows":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                self.send_json(list_billing_workflows_page(conn, self.list_filters()))
            return
        if path == "/api/reimbursement-records":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                self.send_json(list_reimbursement_records_page(conn, self.list_filters()))
            return
        workflow_id = self.billing_workflow_lines_route(path)
        if workflow_id:
            user = self.require_user()
            if not user:
                return
            query = parse_qs(urlparse(self.path).query)
            version_id = clean_text(query.get("versionId", [""])[0])
            try:
                with connect_db() as conn:
                    self.send_json(list_billing_workflow_lines(conn, workflow_id, version_id))
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            return
        workflow_id = self.billing_workflow_route(path)
        if workflow_id:
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                workflow = get_billing_workflow_detail(conn, workflow_id)
                if not workflow:
                    self.send_json({"error": "结算流程不存在"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json(
                    {
                        "workflow": workflow,
                        "versions": list_billing_workflow_versions(conn, workflow_id),
                        "events": list_billing_workflow_events(conn, workflow_id),
                    }
                )
            return
        reimbursement_id = self.reimbursement_record_route(path)
        if reimbursement_id:
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                record = get_reimbursement_record(conn, reimbursement_id)
                if not record:
                    self.send_json({"error": "报销台账不存在"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json(reimbursement_detail_payload(conn, record))
            return
        if path == "/api/billing-statements":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                self.send_json({"items": list_current_billing_statements(conn)})
            return
        if path == "/api/billing-settlement-candidates":
            user = self.require_user()
            if not user:
                return
            filters = self.list_filters(default_limit=10, max_limit=100)
            with connect_db() as conn:

                def calculate(month, pi):
                    payload = {"month": month, "pi": pi, "sourceType": "quantity_sheet"}
                    return generate_billing_statement_by_pi(conn, payload, user)[0]

                self.send_json(list_settlement_candidates(conn, filters, calculate, "quantity_sheet", now_iso()))
            return
        if path == "/api/intake-batches/filter-options":
            if not self.require_user():
                return
            query = parse_qs(urlparse(self.path).query)
            column = clean_text(query.get("column", [""])[0])
            with connect_db() as conn:
                self.send_json(list_intake_batch_filter_options(conn, self.list_filters(), filtered_where, column))
            return
        statement_id = self.billing_statement_route(path)
        if statement_id:
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                statement = get_current_billing_statement(conn, statement_id)
                if not statement:
                    self.send_json({"error": "结算单不存在"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json({"item": statement})
            return
        if path in ENTITY_ENDPOINTS:
            user = self.require_user()
            if not user:
                return
            self.send_entity_list(ENTITY_ENDPOINTS[path], user)
            return
        endpoint, item_id = self.entity_route(path)
        if endpoint and item_id:
            user = self.require_user()
            if not user:
                return
            collection = WRITABLE_ENTITY_ENDPOINTS[endpoint]["collection"]
            with connect_db() as conn:
                state = filter_state_for_actor(read_cached_state(conn), user)
            item = next((entry for entry in state.get(collection, []) if entry.get("id") == item_id), None)
            if not item:
                self.send_json({"error": "记录不存在"}, HTTPStatus.NOT_FOUND)
                return
            self.send_json({"item": item})
            return
        if path.startswith("/scan/cage-card/") or path.startswith("/c/"):
            self.send_spa_index()
            return
        if send_frontend_asset(self, frontend_root()):
            return
        if not path.startswith("/api/") and "." not in path.rsplit("/", 1)[-1]:
            self.send_spa_index()
            return
        super().do_GET()

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/auth/login":
            self.handle_login()
            return
        if path == "/api/auth/logout":
            self.handle_logout()
            return
        if path == "/api/iacuc-index/upload":
            self.handle_iacuc_upload()
            return
        if path == "/api/billing-statements/generate":
            self.handle_billing_statement_generate()
            return
        if path == "/api/billing-statements/generate-by-pi":
            self.handle_billing_statement_generate_by_pi()
            return
        if path == "/api/billing-workflows/advance":
            self.handle_billing_workflow_advance()
            return
        if path == "/api/reimbursement-records/import-monthly":
            self.handle_reimbursement_monthly_import()
            return
        if path == "/api/reimbursement-records/import-arrears":
            self.handle_reimbursement_arrears_import()
            return
        if path == "/api/infrastructure":
            self.handle_infrastructure_write()
            return
        sheet_id = self.quantity_sheet_generate_route(path)
        if sheet_id:
            self.handle_quantity_sheet_statement_generate(sheet_id)
            return
        if path == "/api/quantity-sheets":
            self.handle_quantity_sheet_save(None)
            return
        batch_id = self.intake_batch_confirm_route(path)
        if batch_id:
            self.handle_intake_batch_confirm(batch_id)
            return
        task_id = self.placement_task_action_route(path, "reserve")
        if task_id:
            self.handle_placement_task_reserve(task_id)
            return
        task_id = self.placement_task_action_route(path, "move-in")
        if task_id:
            self.handle_placement_task_move_in(task_id)
            return
        task_id = self.placement_task_action_route(path, "reassign-room")
        if task_id:
            self.handle_placement_task_reassign_room(task_id)
            return
        if path == "/api/users":
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                body = self.read_json_body()
                with connect_db() as conn:
                    created = create_user(conn, body)
                self.send_json({"user": created}, HTTPStatus.CREATED)
            except sqlite3.IntegrityError:
                self.send_json({"error": "用户名已存在"}, HTTPStatus.CONFLICT)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        if path in WRITABLE_ENTITY_ENDPOINTS:
            self.handle_entity_write("POST", path, None)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_PUT(self):
        path = urlparse(self.path).path
        if path == "/api/state":
            user = self.require_user()
            if not user:
                return

            try:
                body = self.read_json_body()
                state = body.get("state")
                if not isinstance(state, dict):
                    self.send_error(HTTPStatus.BAD_REQUEST, "Request body must contain a state object")
                    return
                self.send_json(write_state(state, user))
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            except ValueError as exc:
                self.send_error(HTTPStatus.BAD_REQUEST, str(exc))
            return

        user_id = self.user_route(path)
        if user_id:
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                body = self.read_json_body()
                with connect_db() as conn:
                    updated = update_user(conn, user, user_id, body)
                self.send_json({"user": updated})
            except sqlite3.IntegrityError:
                self.send_json({"error": "用户名已存在"}, HTTPStatus.CONFLICT)
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return

        endpoint, item_id = self.entity_route(path)
        if endpoint and item_id:
            self.handle_entity_write("PUT", endpoint, item_id)
            return
        principal_name = self.principal_identity_route(path)
        if principal_name:
            self.handle_principal_identity_save(principal_name)
            return
        sheet_id = self.quantity_sheet_route(path)
        if sheet_id:
            self.handle_quantity_sheet_save(sheet_id)
            return
        reimbursement_id = self.reimbursement_record_route(path)
        if reimbursement_id:
            self.handle_reimbursement_record_update(reimbursement_id)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_DELETE(self):
        path = urlparse(self.path).path
        user_id = self.user_route(path)
        if user_id:
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                with connect_db() as conn:
                    delete_user(conn, user, user_id)
                self.send_json({"ok": True})
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            return

        endpoint, item_id = self.entity_route(path)
        if endpoint and item_id:
            self.handle_entity_write("DELETE", endpoint, item_id)
            return
        sheet_id = self.quantity_sheet_route(path)
        if sheet_id:
            self.handle_quantity_sheet_delete(sheet_id)
            return
        workflow_id = self.billing_workflow_route(path)
        if workflow_id:
            self.handle_billing_workflow_delete(workflow_id)
            return
        reimbursement_id = self.reimbursement_record_route(path)
        if reimbursement_id:
            self.handle_reimbursement_record_delete(reimbursement_id)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def handle_login(self):
        try:
            body = self.read_json_body()
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        username = str(body.get("username", "")).strip()
        password = str(body.get("password", ""))
        with connect_db() as conn:
            user = authenticate(conn, username, password)
            if not user:
                self.send_json({"error": "用户名或密码错误"}, HTTPStatus.UNAUTHORIZED)
                return
            token, expires_at = create_session(conn, user["id"])
            now = now_iso()
            event = audit_event(
                user,
                "auth.login",
                "session",
                user["id"],
                f"{user['displayName']} 登录系统",
                [],
                now,
                None,
                {
                    "username": user["username"],
                    "role": user["role"],
                    "clientIp": self.client_address[0] if self.client_address else "",
                    "userAgent": self.headers.get("User-Agent", ""),
                },
            )
            write_audit_events(conn, [event])
            conn.commit()
        self.send_response(HTTPStatus.OK)
        body_bytes = json.dumps({"user": user}, ensure_ascii=False).encode("utf-8")
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body_bytes)))
        self.send_header("Cache-Control", "no-store")
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE}={token}; Path=/; HttpOnly; SameSite=Lax; Expires={format_http_date(expires_at)}",
        )
        self.end_headers()
        self.wfile.write(body_bytes)

    def handle_logout(self):
        token = self.session_token()
        with connect_db() as conn:
            delete_session(conn, token)
        self.send_response(HTTPStatus.OK)
        body = b'{"ok": true}'
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")
        self.end_headers()
        self.wfile.write(body)

    def handle_iacuc_upload(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return

        try:
            raw = self.read_raw_body()
            filename, file_body = parse_multipart_upload(self.headers.get("Content-Type", ""), raw)
            if filename and not filename.lower().endswith(".csv"):
                raise ValueError("目前只支持上传 CSV 文件")
            parsed = parse_iacuc_csv(file_body)
            now = now_iso()
            file_items = [application_payload(item, now) for item in parsed["items"]]
            save_iacuc_index_file(file_items)
            event = audit_event(
                user,
                "iacuc_index.uploaded",
                "iacuc_index",
                "iacuc-index",
                f"{user['displayName']} 上传 IACUC 索引 {len(parsed['items'])} 条",
                [],
                now,
                None,
                {"filename": filename, **parsed["summary"]},
            )
            with connect_db() as conn:
                old_items = read_current_applications(conn)
                write_experiment_applications(conn, parsed["items"], now)
                sync_summary = sync_project_derived_fields_after_iacuc_upload(conn, old_items, file_items, user, now)
                invalidate_all_quantity_sheet_candidate_snapshots(conn)
                write_audit_events(conn, [event])
                conn.commit()
            invalidate_data_cache("assembled_state", "iacuc_index", "principal_identities")
            invalidate_data_cache_prefixes(
                "bootstrap_summary::",
                "billing_occupancies::",
                "quantity_sheets::",
                "billing_workflows::",
                "billing_statements::",
                "reimbursement_records::",
                "intake_batches::",
                "placement_tasks::",
            )
            self.send_json(
                {
                    "ok": True,
                    "filename": filename,
                    "updatedAt": now,
                    **parsed["summary"],
                    "syncSummary": sync_summary,
                    "items": file_items,
                    "auditLogs": merge_audit_logs([], [event]),
                }
            )
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_statement_generate(self):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        try:
            with connect_db() as conn:
                statement, lines, audit_logs = generate_billing_statement(conn, body, user)
                workflow = (
                    get_billing_workflow(conn, statement.get("workflowId", "")) if statement.get("workflowId") else None
                )
                conn.commit()
            self.send_json(
                {"statement": statement, "lines": lines, "workflow": workflow, "auditLogs": audit_logs},
                HTTPStatus.CREATED,
            )
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_statement_generate_by_pi(self):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        try:
            refresh_candidate_cache = clean_text(body.get("sourceType", "cage_map")) == "quantity_sheet"
            with connect_db() as conn:
                statement, lines, audit_logs = generate_billing_statement_by_pi(conn, body, user)
                if refresh_candidate_cache:
                    update_settlement_candidate_snapshot_from_statement(
                        conn,
                        clean_text(body.get("month", "")),
                        clean_text(body.get("pi", "")),
                        statement,
                        "quantity_sheet",
                        statement.get("generatedAt", "") or now_iso(),
                    )
                workflow = (
                    get_billing_workflow(conn, statement.get("workflowId", "")) if statement.get("workflowId") else None
                )
                reimbursement = (
                    get_reimbursement_record_by_workflow_id(conn, statement.get("workflowId", ""))
                    if statement.get("workflowId")
                    else None
                )
                conn.commit()
            if refresh_candidate_cache:
                invalidate_data_cache_prefixes("quantity_sheets::settlement_candidates::")
            self.send_json(
                {
                    "statement": statement,
                    "lines": lines,
                    "workflow": workflow,
                    "reimbursementItem": reimbursement_record_list_item(reimbursement) if reimbursement else None,
                    "auditLogs": audit_logs,
                },
                HTTPStatus.CREATED,
            )
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_workflow_advance(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            body = self.read_json_body()
            workflow_id = clean_text(body.get("workflowId", ""))
            to_status = clean_text(body.get("toStatus", ""))
            note = clean_text(body.get("note", ""))
            with connect_db() as conn:
                workflow, version, event = update_workflow_status(conn, workflow_id, to_status, user, note)
                reimbursement = get_reimbursement_record_by_workflow_id(conn, workflow_id)
                if reimbursement:
                    reimbursement["workflowStatus"] = workflow.get("workflowStatus", "")
                    reimbursement["latestEventAt"] = workflow.get("latestEventAt", "") or event.get("at", "")
                    reimbursement["updatedAt"] = now_iso()
                    upsert_reimbursement_record(conn, reimbursement)
                audit = audit_event(
                    user,
                    f"billing_workflow.{to_status}",
                    "billing_workflow",
                    workflow_id,
                    f"{user['displayName']} 更新 {workflow.get('iacuc', '')} {workflow.get('month', '')} 结算流程状态",
                    [],
                    event["at"],
                    None,
                    {"workflow": workflow, "version": version, "event": event},
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("billing_workflows::", "billing_statements::", "reimbursement_records::")
            self.send_json(
                {
                    "workflow": workflow,
                    "event": event,
                    "reimbursementItem": reimbursement_record_list_item(reimbursement) if reimbursement else None,
                    "auditLogs": merge_audit_logs([], [audit]),
                }
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_workflow_delete(self, workflow_id):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            with connect_db() as conn:
                workflow = delete_billing_workflow(conn, workflow_id)
                reimbursement = get_reimbursement_record_by_workflow_id(conn, workflow_id)
                deleted_reimbursement_id = ""
                if reimbursement:
                    if reimbursement_has_manual_entry(reimbursement) or reimbursement.get("source") == "imported":
                        reimbursement["workflowId"] = ""
                        reimbursement["workflowStatus"] = "workflow_deleted"
                        reimbursement["notes"] = join_distinct_text(reimbursement.get("notes", ""), "原流程已删除")
                        reimbursement["latestEventAt"] = now_iso()
                        reimbursement["updatedAt"] = now_iso()
                        upsert_reimbursement_record(conn, reimbursement)
                    else:
                        deleted_reimbursement_id = reimbursement.get("id", "")
                        delete_reimbursement_record(conn, reimbursement["id"])
                at = now_iso()
                audit = audit_event(
                    user,
                    "billing_workflow.deleted",
                    "billing_workflow",
                    workflow_id,
                    f"{user['displayName']} 删除 {workflow.get('pi') or workflow.get('iacuc', '')} {workflow.get('month', '')} 结算流程",
                    [],
                    at,
                    workflow,
                    None,
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("billing_workflows::", "billing_statements::", "reimbursement_records::")
            self.send_json(
                {
                    "ok": True,
                    "workflow": workflow,
                    "reimbursementItem": reimbursement_record_list_item(reimbursement)
                    if reimbursement and not deleted_reimbursement_id
                    else None,
                    "deletedReimbursementId": deleted_reimbursement_id,
                    "auditLogs": merge_audit_logs([], [audit]),
                }
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)

    def handle_reimbursement_monthly_import(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            raw = self.read_raw_body()
            filename, file_body = parse_multipart_upload(self.headers.get("Content-Type", ""), raw)
            if filename and not filename.lower().endswith(".xlsx"):
                raise ValueError("请上传月汇总 Excel 文件")
            with connect_db() as conn:
                payload = import_monthly_reimbursement_workbook(conn, file_body, user)
                conn.commit()
            self.send_json({**payload, "filename": filename}, HTTPStatus.CREATED)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_reimbursement_arrears_import(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            raw = self.read_raw_body()
            filename, file_body = parse_multipart_upload(self.headers.get("Content-Type", ""), raw)
            if filename and not filename.lower().endswith(".xlsx"):
                raise ValueError("请上传欠缴汇算 Excel 文件")
            with connect_db() as conn:
                payload = import_arrears_reimbursement_workbook(conn, file_body, user)
                conn.commit()
            self.send_json({**payload, "filename": filename}, HTTPStatus.CREATED)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_reimbursement_record_update(self, record_id):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            patch = self.read_json_body()
            with connect_db() as conn:
                existing = get_reimbursement_record(conn, record_id)
                if not existing:
                    raise LookupError("报销台账不存在")
                updated = merge_reimbursement_edit(existing, patch)
                updated["latestEventAt"] = now_iso()
                updated["updatedAt"] = now_iso()
                upsert_reimbursement_record(conn, updated)
                recalculate_reimbursement_accumulations(conn, updated.get("pi", ""))
                refreshed = get_reimbursement_record(conn, record_id) or updated
                audit = audit_event(
                    user,
                    "reimbursement.updated",
                    "reimbursement_record",
                    record_id,
                    f"{user['displayName']} 更新 {refreshed.get('pi', '')} {refreshed.get('month', '')} 报销台账",
                    [],
                    refreshed.get("updatedAt", now_iso()),
                    existing,
                    refreshed,
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("reimbursement_records::")
            with connect_db() as conn:
                detail = reimbursement_detail_payload(conn, get_reimbursement_record(conn, record_id) or refreshed)
            self.send_json({**detail, "auditLogs": merge_audit_logs([], [audit])})
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_reimbursement_record_delete(self, record_id):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            with connect_db() as conn:
                record = get_reimbursement_record(conn, record_id)
                if not record:
                    raise LookupError("报销台账不存在")
                pi_name = record.get("pi", "")
                delete_reimbursement_record(conn, record_id)
                recalculate_reimbursement_accumulations(conn, pi_name)
                audit = audit_event(
                    user,
                    "reimbursement.deleted",
                    "reimbursement_record",
                    record_id,
                    f"{user['displayName']} 删除 {record.get('pi', '')} {record.get('month', '')} 报销台账",
                    [],
                    now_iso(),
                    record,
                    None,
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("reimbursement_records::")
            self.send_json({"ok": True, "item": record, "auditLogs": merge_audit_logs([], [audit])})
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)

    def current_user(self):
        with connect_db() as conn:
            return user_from_token(conn, self.session_token())

    def require_user(self):
        user = self.current_user()
        if not user:
            self.send_json({"error": "请先登录"}, HTTPStatus.UNAUTHORIZED)
            return None
        return user

    def session_token(self):
        cookie = self.headers.get("Cookie", "")
        for part in cookie.split(";"):
            if "=" not in part:
                continue
            key, value = part.strip().split("=", 1)
            if key == SESSION_COOKIE:
                return value
        return ""

    def list_filters(self, default_limit=10000, max_limit=10000):
        query = parse_qs(urlparse(self.path).query)

        def value(key):
            return query.get(key, [""])[0]

        column_filters = {}
        raw_column_filters = value("columnFilters") or value("filters")
        if raw_column_filters:
            try:
                parsed = json.loads(raw_column_filters)
            except json.JSONDecodeError:
                parsed = {}
            if isinstance(parsed, dict):
                for key, values in parsed.items():
                    if isinstance(values, list):
                        cleaned = [clean_text(item) for item in values if clean_text(item)]
                    else:
                        cleaned = [clean_text(values)] if clean_text(values) else []
                    if cleaned:
                        column_filters[clean_text(key)] = cleaned
        return {
            "limit": bounded_int(value("limit"), default_limit, 1, max_limit),
            "offset": bounded_int(value("offset"), 0, 0, 1_000_000),
            "sortKey": value("sortKey"),
            "sortDir": value("sortDir"),
            "columnFilters": column_filters,
            "status": value("status"),
            "month": value("month"),
            "iacuc": value("iacuc"),
            "pi": value("pi"),
            "roomId": value("roomId"),
            "roomName": value("roomName"),
            "sourceType": value("sourceType"),
            "entityType": value("entityType"),
            "action": value("action"),
            "onlyUnpaid": value("onlyUnpaid"),
        }

    def send_entity_list(self, table, actor):
        with connect_db() as conn:
            if table == "audit_events":
                filters = self.list_filters(default_limit=500, max_limit=1000)
                payload = list_audit_events_page(conn, filters, filtered_where)
                items = payload["items"]
                page = payload["page"]
            elif table == "intake_batches":
                filters = self.list_filters()
                payload = list_intake_batches_page(conn, filters, filtered_where, ENTITY_ORDER_BY)
                items = payload["items"]
                page = payload["page"]
            elif table == "placement_tasks":
                filters = self.list_filters()
                if actor.get("role") != "admin":
                    allowed_rooms = [clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)]
                    requested_room_id = clean_text(filters.get("roomId", ""))
                    if requested_room_id and requested_room_id not in allowed_rooms:
                        filters["roomIds"] = []
                        filters["roomId"] = ""
                    elif not requested_room_id:
                        filters["roomIds"] = allowed_rooms
                payload = list_placement_tasks_page(conn, filters, ENTITY_ORDER_BY, clean_text)
                items = payload["items"]
                page = payload["page"]
            else:
                rows = conn.execute(
                    f"SELECT payload FROM {table} ORDER BY {ENTITY_ORDER_BY.get(table, 'rowid')}"
                ).fetchall()
                items = [json.loads(row["payload"]) for row in rows]
                page = None
        collection = {
            "rooms": "rooms",
            "racks": "racks",
            "cage_slots": "slots",
            "occupancies": "occupancies",
            "placement_tasks": "placementTasks",
        }.get(table)
        if collection:
            items = filter_entity_payloads_for_actor(collection, items, actor)
        response = {"items": items}
        if page:
            response["page"] = page
        self.send_json(response)

    def entity_route(self, path):
        for endpoint in WRITABLE_ENTITY_ENDPOINTS:
            prefix = endpoint + "/"
            if path.startswith(prefix):
                item_id = unquote(path[len(prefix) :])
                if "/" not in item_id and item_id:
                    return endpoint, item_id
        return None, None

    def user_route(self, path):
        prefix = "/api/users/"
        if not path.startswith(prefix):
            return None
        user_id = unquote(path[len(prefix) :])
        if "/" in user_id or not user_id:
            return None
        return user_id

    def quantity_sheet_route(self, path):
        prefix = "/api/quantity-sheets/"
        if not path.startswith(prefix):
            return None
        sheet_id = unquote(path[len(prefix) :])
        if "/" in sheet_id or not sheet_id:
            return None
        return sheet_id

    def billing_statement_route(self, path):
        prefix = "/api/billing-statements/"
        if not path.startswith(prefix):
            return None
        statement_id = unquote(path[len(prefix) :])
        if "/" in statement_id or not statement_id:
            return None
        return statement_id

    def principal_identity_route(self, path):
        prefix = "/api/principal-identities/"
        if not path.startswith(prefix):
            return None
        pi_name = unquote(path[len(prefix) :])
        if "/" in pi_name or not pi_name:
            return None
        return pi_name

    def billing_workflow_route(self, path):
        prefix = "/api/billing-workflows/"
        if not path.startswith(prefix):
            return None
        workflow_id = unquote(path[len(prefix) :])
        if "/" in workflow_id or not workflow_id:
            return None
        return workflow_id

    def billing_workflow_lines_route(self, path):
        prefix = "/api/billing-workflows/"
        suffix = "/lines"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        workflow_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in workflow_id or not workflow_id:
            return None
        return workflow_id

    def reimbursement_record_route(self, path):
        prefix = "/api/reimbursement-records/"
        if not path.startswith(prefix):
            return None
        record_id = unquote(path[len(prefix) :])
        if "/" in record_id or not record_id:
            return None
        return record_id

    def intake_batch_confirm_route(self, path):
        prefix = "/api/intake-batches/"
        suffix = "/confirm-receipt"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        batch_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in batch_id or not batch_id:
            return None
        return batch_id

    def placement_task_action_route(self, path, action):
        prefix = "/api/placement-tasks/"
        suffix = f"/{action}"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        task_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in task_id or not task_id:
            return None
        return task_id

    def quantity_sheet_generate_route(self, path):
        prefix = "/api/quantity-sheets/"
        suffix = "/generate-statement"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        sheet_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in sheet_id or not sheet_id:
            return None
        return sheet_id

    def handle_entity_write(self, method, endpoint, item_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_optional_json_body() if method == "DELETE" else self.read_json_body()
            payload, status = write_entity_state(endpoint, method, item_id, body, user)
            self.send_json(payload, status)
        except sqlite3.IntegrityError:
            self.send_json({"error": "实体 id 已存在"}, HTTPStatus.CONFLICT)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_infrastructure_write(self):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            payload = write_infrastructure_state(body, user)
            self.send_json(payload, HTTPStatus.CREATED)
        except sqlite3.IntegrityError:
            self.send_json({"error": "实体 id 已存在"}, HTTPStatus.CONFLICT)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_quantity_sheet_save(self, sheet_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            with connect_db() as conn:
                sheet, previous_sheet, affected_sheets, audit_logs, status, perf = save_quantity_sheet(
                    conn, body, user, sheet_id
                )
                invalidate_quantity_sheet_candidate_snapshots(
                    conn, [item for item in (previous_sheet, sheet, *affected_sheets) if item]
                )
                conn.commit()
            invalidate_data_cache("principal_identities")
            invalidate_data_cache_prefixes("quantity_sheets::", "billing_workflows::")
            self.send_json(
                {"item": sheet, "affectedItems": affected_sheets, "auditLogs": audit_logs, "perf": perf}, status
            )
        except sqlite3.IntegrityError:
            self.send_json({"error": "数量统计表 id 已存在"}, HTTPStatus.CONFLICT)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_principal_identity_save(self, pi_name):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            body = self.read_json_body()
            with connect_db() as conn:
                item, audit_logs = save_principal_identity(conn, body, user, pi_name)
                invalidate_quantity_sheet_candidate_snapshots_by_pi(conn, pi_name)
                conn.commit()
            self.send_json({"item": item, "auditLogs": audit_logs})
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_quantity_sheet_delete(self, sheet_id):
        user = self.require_user()
        if not user:
            return
        try:
            with connect_db() as conn:
                sheet, audit_logs = delete_quantity_sheet(conn, user, sheet_id)
                invalidate_quantity_sheet_candidate_snapshots(conn, [sheet])
                conn.commit()
            invalidate_data_cache("principal_identities")
            invalidate_data_cache_prefixes("quantity_sheets::", "billing_workflows::")
            self.send_json({"ok": True, "auditLogs": audit_logs})
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)

    def handle_quantity_sheet_statement_generate(self, sheet_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_optional_json_body()
            with connect_db() as conn:
                statement, lines, audit_logs = generate_quantity_sheet_statement(conn, sheet_id, body, user)
                conn.commit()
            self.send_json({"statement": statement, "lines": lines, "auditLogs": audit_logs}, HTTPStatus.CREATED)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_intake_batch_confirm(self, batch_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(persist_intake_receipt_confirmation(batch_id, body, user), HTTPStatus.CREATED)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_placement_task_reserve(self, task_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(
                persist_placement_action(
                    task_id,
                    user,
                    lambda state: reserve_placement_task(state, task_id, clean_text(body.get("slotId")), user),
                )
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_placement_task_move_in(self, task_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(
                persist_placement_action(
                    task_id,
                    user,
                    lambda state: move_in_placement_task(state, task_id, body.get("actualMoveInDate"), user),
                )
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_placement_task_reassign_room(self, task_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(
                persist_placement_action(
                    task_id,
                    user,
                    lambda state: reassign_placement_task_room(state, task_id, clean_text(body.get("roomId")), user),
                )
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)


configure_database(initialize_schema)


def main():
    ensure_database_ready()
    server = ThreadingHTTPServer((HOST, PORT), CageLedgerHandler)
    print(f"CageLedger server listening on http://{HOST}:{PORT}")
    print(f"SQLite database: {DB_PATH}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
