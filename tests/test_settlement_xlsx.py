import io
import unittest

from openpyxl import load_workbook

from server_app.domains.billing.settlement_xlsx import build_settlement_workbook


def standard_item(iacuc, count, free, tier1, tier2, **overrides):
    item = {
        "iacuc": iacuc,
        "species": "mouse",
        "billingItem": "mouse_standard",
        "billingUnit": "cage_day",
        "cageCount": count,
        "animalCount": 0,
        "statementUnitPrice": 1.0,
        "statementOverageUnitPrice": 1.5,
        "statementTiered": True,
        "statementFreeAllowance": True,
        "statementFullExemption": False,
        "tiered": True,
        "freeAllowance": True,
        "fullExemption": False,
        "freeCages": free,
        "tier1BillableCages": tier1,
        "tier2BillableCages": tier2,
        "supportAmount": 0.0,
        "payableAmount": 0.0,
        "unitPrice": 1.0,
        "overageUnitPrice": 1.5,
        "customBilling": False,
    }
    item.update(overrides)
    return item


def sample_entry(include_custom=False):
    statement = {
        "id": "stmt-1",
        "iacuc": "pi::李教授",
        "iacucs": ["Z1", "Z2"],
        "month": "2026-07",
        "project": "样本课题",
        "pi": "李教授",
        "owner": "张三",
        "funding": "F-1",
        "sourceType": "pi_merged_quantity_sheet",
        "billingUnit": "cage_day",
        "freeCageAllowance": 20,
        "tierLimit": 160,
        "baseUnitPrice": 1.0,
        "overageUnitPrice": 1.5,
        "totalCageDays": 18,
        "totalAmount": 16.5,
        "notes": "",
    }
    day_one_breakdown = [
        standard_item("Z1", 6, 2, 4, 0),
        standard_item("Z2", 4, 0, 4, 0),
    ]
    if include_custom:
        day_one_breakdown.append(
            standard_item(
                "Z1",
                0,
                0,
                0,
                0,
                customBilling=True,
                cageCount=2,
                unitPrice=2.0,
                statementUnitPrice=1.0,
                customBillingStartDate="2026-07-01",
                customBillingEndDate="2026-07-01",
                customBillingNote="临时加收",
            )
        )
    lines = [
        {
            "date": "2026-07-01",
            "cageCount": 12,
            "amount": 12.0,
            "cumulative": 12.0,
            "iacucBreakdown": day_one_breakdown,
        },
        {
            "date": "2026-07-02",
            "cageCount": 8,
            "amount": 8.5,
            "cumulative": 20.5,
            "iacucBreakdown": [
                standard_item("Z1", 5, 0, 5, 0),
                standard_item("Z2", 3, 0, 2, 1),
            ],
        },
        {
            "date": "2026-07-03",
            "cageCount": 0,
            "amount": 0.0,
            "cumulative": 20.5,
            "iacucBreakdown": [],
        },
    ]
    return statement, lines


class SettlementXlsxTests(unittest.TestCase):
    def test_workbook_contains_formula_amount_cells_and_totals(self):
        workbook = load_workbook(
            io.BytesIO(build_settlement_workbook([sample_entry()])),
        )
        worksheet = workbook["李教授 2026-07 结算单"]

        self.assertEqual(worksheet["A1"].value, "李教授课题组实验动物饲养费核算汇总表")
        amount_formula_cells = [
            cell
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
            and cell.value.startswith("=ROUND(")
            and not cell.value.startswith("=ROUND(D")
        ]
        # 2 天 × 2 个 IACUC 列的每日金额公式
        self.assertEqual(len(amount_formula_cells), 4)
        for cell in amount_formula_cells:
            self.assertRegex(str(cell.value), r"^=ROUND\(\(?[A-Z]+\d+")

        payable_label = [
            cell for row in worksheet.iter_rows() for cell in row if cell.value == "本月待缴纳饲养费总计（元）"
        ][0]
        payable_cell = [
            cell
            for cell in worksheet[payable_label.row]
            if isinstance(cell.value, str) and cell.value.startswith("=SUM(")
        ][0]
        self.assertTrue(str(payable_cell.value).startswith("=SUM("))
        self.assertNotIn("+", str(payable_cell.value))

        totals_row = [cell for row in worksheet.iter_rows() for cell in row if cell.value == "单项合计"][0].row
        self.assertTrue(str(worksheet.cell(totals_row, 2).value).startswith("=SUM("))
        self.assertNotIn(
            "—",
            [worksheet.cell(totals_row, column).value for column in range(2, worksheet.max_column + 1)],
        )

    def test_custom_billing_uses_quantity_times_price_formula(self):
        workbook = load_workbook(
            io.BytesIO(build_settlement_workbook([sample_entry(include_custom=True)])),
        )
        worksheet = workbook["李教授 2026-07 结算单"]

        custom_amounts = [
            cell
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value == f"=ROUND(D{cell.row}*E{cell.row},2)"
        ]
        self.assertEqual(len(custom_amounts), 1)

        payable_label = [
            cell for row in worksheet.iter_rows() for cell in row if cell.value == "本月待缴纳饲养费总计（元）"
        ][0]
        payable_cell = [
            cell
            for cell in worksheet[payable_label.row]
            if isinstance(cell.value, str) and cell.value.startswith("=SUM(")
        ][0]
        # 自定义金额已包含在每日列金额中，应付总计不得再加自定义合计（避免重复计入）
        self.assertNotIn("+", str(payable_cell.value))
        self.assertIn("自定义收费明细", str(worksheet.cell(payable_label.row + 2, 1).value))

    def test_groups_iacuc_columns_by_species_and_limits_summary_to_mice(self):
        statement, lines = sample_entry()
        lines[0]["iacucBreakdown"].extend(
            [
                standard_item(
                    "R1",
                    3,
                    0,
                    0,
                    0,
                    species="rat",
                    billingItem="rat_standard",
                    billingUnit="cage_day",
                    statementTiered=False,
                    statementFreeAllowance=False,
                    tiered=False,
                    freeAllowance=False,
                ),
                standard_item(
                    "M1",
                    2,
                    0,
                    0,
                    0,
                    species="monkey",
                    billingItem="monkey",
                    billingUnit="animal_day",
                    animalCount=2,
                    cageCount=0,
                    statementTiered=False,
                    statementFreeAllowance=False,
                    tiered=False,
                    freeAllowance=False,
                ),
            ]
        )
        statement["iacucs"] = ["R1", "Z2", "M1", "Z1"]
        workbook = load_workbook(io.BytesIO(build_settlement_workbook([(statement, lines)])))
        worksheet = workbook["李教授 2026-07 结算单"]

        self.assertEqual(worksheet["B7"].value, "小鼠")
        self.assertEqual(worksheet["L7"].value, "大鼠")
        self.assertEqual(worksheet["N7"].value, "猴")
        self.assertEqual(worksheet["B8"].value, "汇总")
        self.assertEqual(worksheet["L8"].value, "R1")
        self.assertEqual(worksheet["N8"].value, "M1")
        self.assertEqual(worksheet["L9"].value, "笼数")
        self.assertEqual(worksheet["M9"].value, "缴纳（元）")
        self.assertEqual(worksheet["N9"].value, "数量")
        self.assertEqual(worksheet["O9"].value, "缴纳（元）")
        self.assertNotIn(
            "减免", [worksheet["L9"].value, worksheet["M9"].value, worksheet["N9"].value, worksheet["O9"].value]
        )


if __name__ == "__main__":
    unittest.main()
