import csv
import json
import os
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

OUTPUT = Path("src/iacuc-data.local.json")


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    if text.endswith(" 00:00:00"):
        text = text[: -len(" 00:00:00")]
    return text


def compact_field_name(value):
    return re.sub(r"\s+", "", clean(value))


def field_name_for_labels(field_by_name, field_by_compact, labels):
    for label in labels:
        clean_label = clean(label)
        if clean_label in field_by_name:
            return field_by_name[clean_label]
        compact_label = compact_field_name(label)
        if compact_label in field_by_compact:
            return field_by_compact[compact_label]
    return None


def raw_fields_from_dict(row):
    return {clean(key): clean(value) for key, value in row.items() if clean(key)}


def raw_fields_from_row(headers, row):
    fields = {}
    for index, header in enumerate(headers):
        name = clean(header)
        if not name:
            continue
        fields[name] = clean(row[index] if index < len(row) else "")
    return fields


def value_from_dict(row, field_name):
    return clean(row.get(field_name)) if field_name else ""


def value_from_row(row, column_index):
    return clean(row[column_index]) if column_index is not None and column_index < len(row) else ""


def normalize_application_date(value):
    text = clean(value)
    if not text:
        return ""
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = re.sub(r"[./]", "-", text)
    text = re.sub(r"\s*00:00:00$", "", text)
    match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", text)
    if not match:
        return text
    year, month, day = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return clean(value)


def normalize_application_amount(value):
    text = clean(value)
    if not text:
        return None
    normalized = re.sub(r"[,\s￥¥元]", "", text)
    try:
        return round(float(normalized), 2)
    except ValueError:
        return text


def normalize_application_record(record):
    for key in (
        "iacucApprovalDate",
        "applicationApprovalDate",
        "projectStartDate",
        "projectEndDate",
        "applicationDate",
    ):
        record[key] = normalize_application_date(record.get(key, ""))
    for key in ("approvedFeedingFee", "actualFeedingFee", "pendingReimbursementFee"):
        record[key] = normalize_application_amount(record.get(key, ""))
    return record


def normalize_iacuc(value):
    text = clean(value)
    text = re.sub(r"（.*?）", "", text)
    text = re.sub(r"\(.*?\)", "", text)
    return text.strip()


def main():
    source_arg = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("IACUC_SOURCE_XLSX")
    if not source_arg:
        raise SystemExit("Usage: python3 scripts/generate_iacuc_index.py /path/to/iacuc-summary.xlsx|csv")

    source = Path(source_arg)
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else OUTPUT

    required = {
        "iacuc": "动物伦理编号",
        "project": "动物实验名称",
        "pi": "项目负责人",
        "owner": "实验负责人",
    }
    optional_aliases = {
        "funding": ("项目来源", "支撑经费", "经费来源", "课题来源", "经费项目"),
        "fundCode": ("经费号", "经费本编号", "经费编号"),
        "supportProjectPeriod": ("支撑科研项目起止时间", "科研项目起止时间", "项目起止时间"),
        "experimentNo": ("动物实验编号", "实验编号"),
        "species": ("动物品系", "品系", "动物种类"),
        "facility": ("动物实验场所", "实验场所", "动物房"),
        "maxFeedingPeriod": ("最长饲养周期", "饲养周期"),
        "iacucApprovalDate": ("动物伦理通过日期", "伦理通过日期", "伦理批准日期"),
        "applicationApprovalDate": ("实验申请通过日期", "申请通过日期", "实验批准日期"),
        "projectStartDate": ("实验开始日期", "项目开始日期", "开始日期"),
        "projectEndDate": ("实验结束日期", "项目结束日期", "结束日期", "有效期至"),
        "approvedFeedingFee": ("实验审核通过饲养费（元）", "审核通过饲养费", "批准饲养费"),
        "approvalLeader": ("中心审批领导", "审批领导"),
        "actualFeedingFee": ("实际产生饲养费（元）", "实际产生饲养费", "实际饲养费"),
        "pendingReimbursementFee": ("待报销饲养费（元）", "待报销饲养费", "欠缴饲养费"),
        "assistant": ("实验助手", "助手"),
        "notes": ("备注", "说明"),
        "applicationDate": ("申请日期", "提交日期"),
    }
    records = []

    if source.suffix.lower() == ".csv":
        text = source.read_text(encoding="utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        field_by_name = {clean(name): name for name in reader.fieldnames or [] if clean(name)}
        field_by_compact = {
            compact_field_name(name): name for name in reader.fieldnames or [] if compact_field_name(name)
        }
        required_fields = {
            key: field_name_for_labels(field_by_name, field_by_compact, [label]) for key, label in required.items()
        }
        missing = [required[key] for key, field_name in required_fields.items() if not field_name]
        if missing:
            raise SystemExit(f"Missing required columns: {', '.join(missing)}")
        optional_fields = {
            key: field_name_for_labels(field_by_name, field_by_compact, labels)
            for key, labels in optional_aliases.items()
        }
        for index, row in enumerate(reader, start=1):
            raw_iacuc = value_from_dict(row, required_fields["iacuc"])
            iacuc = normalize_iacuc(raw_iacuc)
            if not iacuc:
                continue
            record = {
                "id": f"app-{index:06d}",
                "iacuc": raw_iacuc,
                "rawIacuc": raw_iacuc,
                "project": value_from_dict(row, required_fields["project"]),
                "pi": value_from_dict(row, required_fields["pi"]),
                "owner": value_from_dict(row, required_fields["owner"]),
                "rawFields": raw_fields_from_dict(row),
            }
            record.update({key: value_from_dict(row, field_name) for key, field_name in optional_fields.items()})
            records.append(normalize_application_record(record))
    else:
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        column_by_name = {name: index for index, name in enumerate(headers) if name}
        column_by_compact = {
            compact_field_name(name): index for index, name in enumerate(headers) if compact_field_name(name)
        }
        required_columns = {
            key: column_by_name.get(label)
            if label in column_by_name
            else column_by_compact.get(compact_field_name(label))
            for key, label in required.items()
        }
        missing = [required[key] for key, column_index in required_columns.items() if column_index is None]
        if missing:
            raise SystemExit(f"Missing required columns: {', '.join(missing)}")
        optional_columns = {
            key: next(
                (
                    column_by_name.get(clean(label))
                    if clean(label) in column_by_name
                    else column_by_compact.get(compact_field_name(label))
                    for label in labels
                    if clean(label) in column_by_name or compact_field_name(label) in column_by_compact
                ),
                None,
            )
            for key, labels in optional_aliases.items()
        }

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            raw_iacuc = value_from_row(row, required_columns["iacuc"])
            iacuc = normalize_iacuc(raw_iacuc)
            if not iacuc:
                continue

            record = {
                "id": f"app-{index:06d}",
                "iacuc": raw_iacuc,
                "rawIacuc": raw_iacuc,
                "project": value_from_row(row, required_columns["project"]),
                "pi": value_from_row(row, required_columns["pi"]),
                "owner": value_from_row(row, required_columns["owner"]),
                "rawFields": raw_fields_from_row(headers, row),
            }
            record.update({key: value_from_row(row, column_index) for key, column_index in optional_columns.items()})
            records.append(normalize_application_record(record))

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(records)} IACUC records to {output}")


if __name__ == "__main__":
    main()
