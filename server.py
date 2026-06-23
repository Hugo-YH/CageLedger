#!/usr/bin/env python3
import base64
import calendar
import csv
import hashlib
import hmac
import io
import json
import re
import secrets
import sqlite3
import time
from datetime import date, datetime, timedelta, timezone
from email.utils import format_datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import Request, urlopen

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel as openpyxl_from_excel
except ImportError:
    openpyxl = None
    openpyxl_from_excel = None

from server_app.cache import cache_get, cache_key, cache_set, invalidate_data_cache, invalidate_data_cache_prefixes, log_perf
from server_app.config import (
    CAGELEDGER_APP_VERSION,
    CAGELEDGER_BRANCH,
    CAGELEDGER_CONTACT_EMAIL,
    CAGELEDGER_COPYRIGHT,
    CAGELEDGER_DEPARTMENT,
    CAGELEDGER_DEVELOPER,
    CAGELEDGER_GITEA_TOKEN,
    CAGELEDGER_LICENSE,
    CAGELEDGER_ORGANIZATION,
    CAGELEDGER_REPOSITORY_URL,
    CAGELEDGER_UPDATE_CHECK_ENABLED,
    CAGELEDGER_VERSION,
    DB_PATH,
    DEFAULT_ADMIN_PASSWORD,
    DEFAULT_ADMIN_USERNAME,
    HOST,
    IACUC_INDEX_PATH,
    LEGACY_IACUC_INDEX_PATH,
    MAX_BODY_BYTES,
    PORT,
    ROOT,
    SESSION_COOKIE,
    SESSION_TTL_DAYS,
)
from server_app.db import configure_database, connect_db, ensure_database_ready
from server_app.http import add_default_headers, send_json as send_json_response
from server_app.repositories.audit import insert_audit_events
from server_app.repositories.billing import (
    delete_billing_workflow_tree as delete_billing_workflow_tree_repository,
    delete_quantity_sheet_by_id as delete_quantity_sheet_by_id_repository,
    get_current_billing_statement as get_current_billing_statement_repository,
    billing_workflow_detail_item as billing_workflow_detail_item_repository,
    get_billing_workflow_detail as get_billing_workflow_detail_repository,
    get_billing_version as get_billing_version_repository,
    get_billing_workflow as get_billing_workflow_repository,
    get_billing_workflow_by_key as get_billing_workflow_by_key_repository,
    get_quantity_sheet as get_quantity_sheet_repository,
    insert_billing_version as insert_billing_version_repository,
    insert_billing_workflow as insert_billing_workflow_repository,
    insert_billing_workflow_event as insert_billing_workflow_event_repository,
    insert_quantity_sheet as insert_quantity_sheet_repository,
    list_quantity_sheets_by_month_iacuc as list_quantity_sheets_by_month_iacuc_repository,
    list_quantity_sheets_by_month_pi as list_quantity_sheets_by_month_pi_repository,
    list_billing_workflows as list_billing_workflows_repository,
    list_billing_statement_lines_for_version as list_billing_statement_lines_for_version_repository,
    list_billing_statement_line_summaries_for_version as list_billing_statement_line_summaries_for_version_repository,
    list_billing_workflow_events as list_billing_workflow_events_repository,
    list_billing_workflow_versions as list_billing_workflow_versions_repository,
    list_billing_workflows_page as list_billing_workflows_page_repository,
    list_current_billing_statements as list_current_billing_statements_repository,
    list_quantity_sheets as list_quantity_sheets_repository,
    list_quantity_sheets_page as list_quantity_sheets_page_repository,
    replace_billing_statement_version_lines as replace_billing_statement_version_lines_repository,
    select_quantity_sheets_for_transfer as select_quantity_sheets_for_transfer_repository,
    update_billing_version as update_billing_version_repository,
    update_billing_workflow as update_billing_workflow_repository,
    update_quantity_sheet as update_quantity_sheet_repository,
)
from server_app.repositories.entities import (
    delete_intake_batch as delete_intake_batch_repository,
    delete_placement_task as delete_placement_task_repository,
    list_audit_events_page,
    list_intake_batches_page,
    list_placement_tasks_page,
    list_distinct_principal_names,
    read_principal_identity_payloads,
    read_principal_type_by_pi as read_principal_type_by_pi_repository,
    upsert_intake_batch as upsert_intake_batch_repository,
    upsert_placement_task as upsert_placement_task_repository,
    upsert_principal_identity,
)
from server_app.repositories.infrastructure import (
    delete_rack_record,
    delete_room_record,
    delete_slot_record,
    insert_rack_record,
    insert_room_record,
    insert_slot_record,
    update_rack_record,
    update_room_record,
)
from server_app.repositories.iacuc import read_iacuc_index as read_iacuc_index_repository
from server_app.repositories.iacuc import replace_experiment_applications, save_iacuc_index_file as save_iacuc_index_file_repository
from server_app.repositories.payload import (
    cached_paginated_payloads,
    dump_json,
    paginated_payloads,
    read_payloads,
    read_setting,
    read_updated_at,
    set_setting,
    table_has_rows,
)
from server_app.repositories.reimbursement import (
    delete_reimbursement_record as delete_reimbursement_record_repository,
    get_reimbursement_record as get_reimbursement_record_repository,
    get_reimbursement_record_by_key as get_reimbursement_record_by_key_repository,
    get_reimbursement_record_by_workflow_id as get_reimbursement_record_by_workflow_id_repository,
    list_reimbursement_record_summaries_for_pi as list_reimbursement_record_summaries_for_pi_repository,
    list_reimbursement_records_for_pi as list_reimbursement_records_for_pi_repository,
    list_reimbursement_records_page as list_reimbursement_records_page_repository,
    reimbursement_record_list_item,
    upsert_reimbursement_record as upsert_reimbursement_record_repository,
)
from server_app.repositories.state import assemble_state as assemble_state_repository, read_applications_by_iacuc as read_applications_by_iacuc_repository, read_cached_state as read_cached_state_repository
from server_app.repositories.users import (
    delete_session_by_token_hash,
    delete_sessions_by_user_id,
    delete_user_by_id,
    get_active_user_by_username,
    has_any_user,
    get_user_by_id,
    get_user_by_session_token_hash,
    insert_user,
    insert_session,
    list_users as list_users_repository,
    update_user_with_password,
    update_user_without_password,
)
from server_app.services.billing import save_billing_statement_workflow as save_billing_statement_workflow_service, update_workflow_status as update_workflow_status_service
from server_app.services.intake import confirm_intake_receipt as confirm_intake_receipt_service
from server_app.services.placement import (
    move_in_placement_task as move_in_placement_task_service,
    reassign_placement_task_room as reassign_placement_task_room_service,
    reserve_placement_task as reserve_placement_task_service,
    sync_slot_statuses,
)
from server_app.services.quantity import sync_quantity_sheet_transfer_rows as sync_quantity_sheet_transfer_rows_service
from server_app.services.reimbursement import (
    REIMBURSEMENT_STATUS_COMPLETED,
    REIMBURSEMENT_STATUS_PENDING,
    REIMBURSEMENT_STATUS_REIMBURSING,
    coerce_money as coerce_reimbursement_money,
    infer_import_status,
    merge_reimbursement_edit,
    normalize_reimbursement_status,
    reimbursement_business_key,
    reimbursement_has_manual_entry,
    summarize_statement,
)

BILLING_PRINCIPAL_PI = "pi"
BILLING_PRINCIPAL_INDEPENDENT = "independent"
FREE_CAGES_PI = 20
FREE_CAGES_INDEPENDENT = 10
FREE_CAGES_DEFAULT = FREE_CAGES_PI
BILLING_TIER_LIMIT = 160
BILLING_TIER_BASE_PRICE = 4.5
BILLING_TIER_OVER_PRICE = 6.5
BILLING_RULES = {
    "mouse_standard": {"species": "mouse", "unit": "cage_day", "internalPrice": 4.5, "externalPrice": 13.5, "tiered": True, "freeAllowance": True},
    "mouse_diabetic": {"species": "mouse", "unit": "cage_day", "internalPrice": 7.2, "externalPrice": 21.6, "tiered": False, "freeAllowance": False},
    "rat_standard": {"species": "rat", "unit": "cage_day", "internalPrice": 8.5, "externalPrice": 25.5, "tiered": False, "freeAllowance": False},
    "rat_diabetic": {"species": "rat", "unit": "cage_day", "internalPrice": 14, "externalPrice": 42, "tiered": False, "freeAllowance": False},
    "guinea_pig": {"species": "guinea_pig", "unit": "animal_day", "internalPrice": 3, "externalPrice": 9, "tiered": False, "freeAllowance": False},
    "rabbit": {"species": "rabbit", "unit": "animal_day", "internalPrice": 5, "externalPrice": 15, "tiered": False, "freeAllowance": False},
    "monkey": {"species": "monkey", "unit": "animal_day", "internalPrice": 35, "externalPrice": 65, "tiered": False, "freeAllowance": False},
    "pig": {"species": "pig", "unit": "animal_day", "internalPrice": 15, "externalPrice": 45, "tiered": False, "freeAllowance": False},
    "dog": {"species": "dog", "unit": "animal_day", "internalPrice": 15, "externalPrice": 45, "tiered": False, "freeAllowance": False},
}
WORKFLOW_STATUS_IN_FEEDING = "in_feeding"
WORKFLOW_STATUS_GENERATED = "statement_generated"
WORKFLOW_STATUS_SENT = "statement_sent"
WORKFLOW_STATUS_SIGNED = "statement_signed_returned"
WORKFLOW_STATUS_FINANCE = "submitted_to_finance"
REIMBURSEMENT_MIGRATION_KEY = "reimbursementRecordMigrationDone"
WORKFLOW_STATUSES = (
    WORKFLOW_STATUS_IN_FEEDING,
    WORKFLOW_STATUS_GENERATED,
    WORKFLOW_STATUS_SENT,
    WORKFLOW_STATUS_SIGNED,
    WORKFLOW_STATUS_FINANCE,
)
VERSION_STATUS_ACTIVE = "active"
VERSION_STATUS_VOIDED = "voided"
BILLING_WORKFLOW_MIGRATION_KEY = "billingWorkflowMigrationDone"
TABLES = (
    "audit_logs",
    "billing_adjustments",
    "billing_rules",
    "occupancies",
    "placement_tasks",
    "cage_slots",
    "intake_batches",
    "racks",
    "rooms",
    "app_settings",
)
ENTITY_ENDPOINTS = {
    "/api/rooms": "rooms",
    "/api/racks": "racks",
    "/api/cage-slots": "cage_slots",
    "/api/occupancies": "occupancies",
    "/api/placement-tasks": "placement_tasks",
    "/api/billing-rules": "billing_rules",
    "/api/billing-adjustments": "billing_adjustments",
    "/api/intake-batches": "intake_batches",
    "/api/experiment-applications": "experiment_applications",
    "/api/billing-statements": "billing_statements",
    "/api/billing-statement-lines": "billing_statement_lines",
    "/api/audit-events": "audit_events",
}
ENTITY_ORDER_BY = {
    "rooms": "rowid",
    "racks": "room_id, index_no, rowid",
    "cage_slots": "rack_id, row_no, col_no, rowid",
    "occupancies": "start_date, rowid",
    "placement_tasks": "planned_move_in_date, rowid",
    "billing_rules": "rowid",
    "billing_adjustments": "rowid",
    "intake_batches": "updated_at DESC, rowid DESC",
    "experiment_applications": "rowid",
    "billing_statements": "month DESC, iacuc, rowid DESC",
    "billing_statement_lines": "statement_id, line_date, rowid",
}
WRITABLE_ENTITY_ENDPOINTS = {
    "/api/rooms": {"collection": "rooms", "id_prefix": "room"},
    "/api/racks": {"collection": "racks", "id_prefix": "rack"},
    "/api/cage-slots": {"collection": "slots", "id_prefix": "slot"},
    "/api/occupancies": {"collection": "occupancies", "id_prefix": "occ"},
    "/api/placement-tasks": {"collection": "placementTasks", "id_prefix": "ptask"},
    "/api/billing-rules": {"collection": "billingRules", "id_prefix": "rule"},
    "/api/billing-adjustments": {"collection": "adjustments", "id_prefix": "adj"},
    "/api/intake-batches": {"collection": "intakeBatches", "id_prefix": "batch"},
}
def initialize_schema(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_state (
            id TEXT PRIMARY KEY,
            payload TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS rooms (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            area TEXT,
            rack_count INTEGER,
            rows INTEGER,
            cols INTEGER,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS racks (
            id TEXT PRIMARY KEY,
            room_id TEXT NOT NULL,
            name TEXT NOT NULL,
            rows INTEGER,
            cols INTEGER,
            index_no INTEGER,
            payload TEXT NOT NULL,
            FOREIGN KEY(room_id) REFERENCES rooms(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cage_slots (
            id TEXT PRIMARY KEY,
            rack_id TEXT NOT NULL,
            row_no INTEGER,
            col_no INTEGER,
            code TEXT,
            status TEXT,
            payload TEXT NOT NULL,
            FOREIGN KEY(rack_id) REFERENCES racks(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS occupancies (
            id TEXT PRIMARY KEY,
            slot_id TEXT,
            room_id TEXT,
            rack_id TEXT,
            cage_code TEXT,
            status TEXT NOT NULL,
            iacuc TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            species TEXT,
            billing_item TEXT,
            customer_type TEXT,
            animal_count INTEGER,
            room_name TEXT,
            rack_name TEXT,
            slot_code TEXT,
            start_date TEXT,
            end_date TEXT,
            end_reason TEXT,
            notes TEXT,
            updated_at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS placement_tasks (
            id TEXT PRIMARY KEY,
            source_batch_id TEXT NOT NULL,
            source_receipt_id TEXT NOT NULL,
            target_room_id TEXT,
            planned_move_in_date TEXT NOT NULL,
            status TEXT NOT NULL,
            reserved_occupancy_id TEXT,
            actual_move_in_date TEXT,
            updated_at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS experiment_applications (
            id TEXT PRIMARY KEY,
            iacuc TEXT NOT NULL,
            raw_iacuc TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            imported_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS principal_identities (
            pi TEXT PRIMARY KEY,
            principal_type TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_rules (
            id TEXT PRIMARY KEY,
            name TEXT,
            unit TEXT,
            price REAL,
            effective_start TEXT,
            effective_end TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_adjustments (
            id TEXT PRIMARY KEY,
            target_type TEXT,
            target_id TEXT,
            adjustment_type TEXT,
            value REAL,
            reason TEXT,
            effective_start TEXT,
            effective_end TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS quantity_sheets (
            id TEXT PRIMARY KEY,
            month TEXT NOT NULL,
            iacuc TEXT NOT NULL,
            room_id TEXT,
            room_name TEXT,
            manager TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            updated_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS intake_batches (
            id TEXT PRIMARY KEY,
            batch_no TEXT NOT NULL,
            iacuc TEXT,
            supplier TEXT,
            room_name TEXT,
            intake_date TEXT,
            status TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_statements (
            id TEXT PRIMARY KEY,
            iacuc TEXT NOT NULL,
            month TEXT NOT NULL,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            total_cage_days INTEGER,
            total_amount REAL,
            status TEXT NOT NULL,
            generated_at TEXT NOT NULL,
            locked_at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_statement_lines (
            id TEXT PRIMARY KEY,
            statement_id TEXT NOT NULL,
            line_date TEXT NOT NULL,
            cage_count INTEGER,
            unit_price REAL,
            discount_percent REAL,
            amount REAL,
            cumulative REAL,
            occupancy_ids TEXT,
            payload TEXT NOT NULL,
            FOREIGN KEY(statement_id) REFERENCES billing_statements(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_workflows (
            id TEXT PRIMARY KEY,
            business_key TEXT NOT NULL UNIQUE,
            iacuc TEXT NOT NULL,
            month TEXT NOT NULL,
            source_type TEXT NOT NULL,
            workflow_status TEXT NOT NULL,
            current_version_id TEXT,
            current_version_no INTEGER NOT NULL DEFAULT 0,
            latest_event_at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_statement_versions (
            id TEXT PRIMARY KEY,
            workflow_id TEXT NOT NULL,
            version_no INTEGER NOT NULL,
            version_status TEXT NOT NULL,
            workflow_status TEXT NOT NULL,
            generated_at TEXT NOT NULL,
            voided_at TEXT,
            created_by TEXT,
            payload TEXT NOT NULL,
            FOREIGN KEY(workflow_id) REFERENCES billing_workflows(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_statement_version_lines (
            id TEXT PRIMARY KEY,
            version_id TEXT NOT NULL,
            line_date TEXT NOT NULL,
            payload TEXT NOT NULL,
            FOREIGN KEY(version_id) REFERENCES billing_statement_versions(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS billing_workflow_events (
            id TEXT PRIMARY KEY,
            workflow_id TEXT NOT NULL,
            version_id TEXT,
            event_type TEXT NOT NULL,
            from_status TEXT,
            to_status TEXT,
            at TEXT NOT NULL,
            payload TEXT NOT NULL,
            FOREIGN KEY(workflow_id) REFERENCES billing_workflows(id) ON DELETE CASCADE,
            FOREIGN KEY(version_id) REFERENCES billing_statement_versions(id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS reimbursement_records (
            id TEXT PRIMARY KEY,
            business_key TEXT NOT NULL UNIQUE,
            month TEXT NOT NULL,
            pi TEXT NOT NULL,
            workflow_id TEXT,
            workflow_status TEXT,
            reimbursement_status TEXT NOT NULL,
            current_month_amount REAL NOT NULL DEFAULT 0,
            support_amount REAL NOT NULL DEFAULT 0,
            payable_amount REAL NOT NULL DEFAULT 0,
            paid_amount REAL NOT NULL DEFAULT 0,
            unpaid_amount REAL NOT NULL DEFAULT 0,
            accumulated_payable REAL NOT NULL DEFAULT 0,
            accumulated_paid REAL NOT NULL DEFAULT 0,
            accumulated_unpaid REAL NOT NULL DEFAULT 0,
            source TEXT NOT NULL,
            latest_event_at TEXT,
            updated_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_logs (
            id TEXT PRIMARY KEY,
            message TEXT,
            at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            display_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            room_ids TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sessions (
            token_hash TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_events (
            id TEXT PRIMARY KEY,
            actor_user_id TEXT,
            actor_username TEXT,
            actor_display_name TEXT,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id TEXT,
            message TEXT NOT NULL,
            slot_ids TEXT NOT NULL,
            at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    migrate_schema(conn)
    repair_missing_cage_slots(conn)
    create_performance_indexes(conn)
    conn.commit()
    ensure_default_admin(conn)


def create_performance_indexes(conn):
    index_statements = [
        "CREATE INDEX IF NOT EXISTS idx_racks_room_id ON racks(room_id)",
        "CREATE INDEX IF NOT EXISTS idx_cage_slots_rack_id ON cage_slots(rack_id)",
        "CREATE INDEX IF NOT EXISTS idx_cage_slots_status ON cage_slots(status)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_slot_id ON occupancies(slot_id)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_room_id ON occupancies(room_id)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_rack_id ON occupancies(rack_id)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_iacuc ON occupancies(iacuc)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_pi ON occupancies(pi)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_status ON occupancies(status)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_species ON occupancies(species)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_billing_item ON occupancies(billing_item)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_customer_type ON occupancies(customer_type)",
        "CREATE INDEX IF NOT EXISTS idx_occupancies_start_end ON occupancies(start_date, end_date)",
        "CREATE INDEX IF NOT EXISTS idx_placement_tasks_room_id ON placement_tasks(target_room_id)",
        "CREATE INDEX IF NOT EXISTS idx_placement_tasks_status ON placement_tasks(status)",
        "CREATE INDEX IF NOT EXISTS idx_placement_tasks_move_in_date ON placement_tasks(planned_move_in_date)",
        "CREATE INDEX IF NOT EXISTS idx_experiment_applications_iacuc ON experiment_applications(iacuc)",
        "CREATE INDEX IF NOT EXISTS idx_experiment_applications_raw_iacuc ON experiment_applications(raw_iacuc)",
        "CREATE INDEX IF NOT EXISTS idx_experiment_applications_pi ON experiment_applications(pi)",
        "CREATE INDEX IF NOT EXISTS idx_quantity_sheets_month ON quantity_sheets(month)",
        "CREATE INDEX IF NOT EXISTS idx_quantity_sheets_iacuc ON quantity_sheets(iacuc)",
        "CREATE INDEX IF NOT EXISTS idx_quantity_sheets_pi ON quantity_sheets(pi)",
        "CREATE INDEX IF NOT EXISTS idx_intake_batches_status ON intake_batches(status)",
        "CREATE INDEX IF NOT EXISTS idx_intake_batches_intake_date ON intake_batches(intake_date)",
        "CREATE INDEX IF NOT EXISTS idx_intake_batches_iacuc ON intake_batches(iacuc)",
        "CREATE INDEX IF NOT EXISTS idx_billing_statements_iacuc_month ON billing_statements(iacuc, month)",
        "CREATE INDEX IF NOT EXISTS idx_billing_statement_lines_statement_id ON billing_statement_lines(statement_id)",
        "CREATE INDEX IF NOT EXISTS idx_billing_workflows_month ON billing_workflows(month)",
        "CREATE INDEX IF NOT EXISTS idx_billing_workflows_status ON billing_workflows(workflow_status)",
        "CREATE INDEX IF NOT EXISTS idx_billing_versions_workflow_id ON billing_statement_versions(workflow_id)",
        "CREATE INDEX IF NOT EXISTS idx_billing_version_lines_version_id ON billing_statement_version_lines(version_id)",
        "CREATE INDEX IF NOT EXISTS idx_billing_events_workflow_id ON billing_workflow_events(workflow_id)",
        "CREATE INDEX IF NOT EXISTS idx_reimbursement_records_month ON reimbursement_records(month)",
        "CREATE INDEX IF NOT EXISTS idx_reimbursement_records_pi ON reimbursement_records(pi)",
        "CREATE INDEX IF NOT EXISTS idx_reimbursement_records_status ON reimbursement_records(reimbursement_status)",
        "CREATE INDEX IF NOT EXISTS idx_reimbursement_records_workflow_id ON reimbursement_records(workflow_id)",
        "CREATE INDEX IF NOT EXISTS idx_audit_events_at ON audit_events(at)",
        "CREATE INDEX IF NOT EXISTS idx_sessions_user_id ON sessions(user_id)",
        "CREATE INDEX IF NOT EXISTS idx_sessions_expires_at ON sessions(expires_at)",
    ]
    for statement in index_statements:
        conn.execute(statement)


def migrate_schema(conn):
    ensure_experiment_applications_duplicate_schema(conn)
    ensure_occupancies_history_schema(conn)
    ensure_occupancies_structured_columns(conn)
    migrate_billing_workflow_schema(conn)
    backfill_billing_workflow_scope(conn)
    migrate_reimbursement_record_schema(conn)


def ensure_experiment_applications_duplicate_schema(conn):
    columns = table_columns(conn, "experiment_applications")
    pk_column = next((name for name, info in columns.items() if info.get("pk")), "")
    if "id" in columns and pk_column == "id":
        return

    conn.execute("ALTER TABLE experiment_applications RENAME TO experiment_applications_legacy")
    conn.execute(
        """
        CREATE TABLE experiment_applications (
            id TEXT PRIMARY KEY,
            iacuc TEXT NOT NULL,
            raw_iacuc TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            imported_at TEXT NOT NULL,
            payload TEXT NOT NULL
        )
        """
    )
    rows = conn.execute("SELECT rowid, * FROM experiment_applications_legacy ORDER BY rowid").fetchall()
    for index, row in enumerate(rows, start=1):
        payload = json.loads(row["payload"])
        raw_iacuc = clean_text(payload.get("rawIacuc", "") or row["raw_iacuc"] or row["iacuc"])
        item = {
            **payload,
            "id": payload.get("id") or f"app-{index:06d}",
            "iacuc": clean_text(payload.get("iacuc", "") or raw_iacuc or row["iacuc"]),
            "rawIacuc": raw_iacuc,
        }
        conn.execute(
            """
            INSERT INTO experiment_applications (
                id, iacuc, raw_iacuc, project, pi, owner, funding, imported_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item["id"],
                item["iacuc"],
                item.get("rawIacuc", ""),
                item.get("project", ""),
                item.get("pi", ""),
                item.get("owner", ""),
                item.get("funding", ""),
                row["imported_at"],
                dump_json(item),
            ),
        )
    conn.execute("DROP TABLE experiment_applications_legacy")


def ensure_occupancies_structured_columns(conn):
    columns = table_columns(conn, "occupancies")
    additions = {
        "room_id": "TEXT",
        "rack_id": "TEXT",
        "species": "TEXT",
        "billing_item": "TEXT",
        "customer_type": "TEXT",
        "animal_count": "INTEGER",
    }
    changed = False
    for column, column_type in additions.items():
        if column in columns:
            continue
        conn.execute(f"ALTER TABLE occupancies ADD COLUMN {column} {column_type}")
        changed = True
    if changed:
        backfill_occupancy_structured_columns(conn)


def backfill_occupancy_structured_columns(conn):
    state = assemble_state(conn) or empty_state()
    applications_by_iacuc = read_applications_by_iacuc(conn)
    for occupancy in state.get("occupancies", []):
        normalized = occupancy_with_snapshots(occupancy, state, applications_by_iacuc)
        values = occupancy_structured_values(normalized)
        conn.execute(
            """
            UPDATE occupancies
            SET room_id = ?, rack_id = ?, species = ?, billing_item = ?,
                customer_type = ?, animal_count = ?, payload = ?
            WHERE id = ?
            """,
            (
                values["room_id"],
                values["rack_id"],
                values["species"],
                values["billing_item"],
                values["customer_type"],
                values["animal_count"],
                dump_json(normalized),
                normalized.get("id"),
            ),
        )


def ensure_occupancies_history_schema(conn):
    columns = table_columns(conn, "occupancies")
    foreign_keys = conn.execute("PRAGMA foreign_key_list(occupancies)").fetchall()
    slot_not_null = bool(columns.get("slot_id", {}).get("notnull"))
    required_columns = {"funding", "room_name", "rack_name", "slot_code"}
    if not foreign_keys and not slot_not_null and required_columns.issubset(columns):
        return

    conn.execute("ALTER TABLE occupancies RENAME TO occupancies_legacy")
    conn.execute(
        """
        CREATE TABLE occupancies (
            id TEXT PRIMARY KEY,
            slot_id TEXT,
            room_id TEXT,
            rack_id TEXT,
            cage_code TEXT,
            status TEXT NOT NULL,
            iacuc TEXT,
            project TEXT,
            pi TEXT,
            owner TEXT,
            funding TEXT,
            species TEXT,
            billing_item TEXT,
            customer_type TEXT,
            animal_count INTEGER,
            room_name TEXT,
            rack_name TEXT,
            slot_code TEXT,
            start_date TEXT,
            end_date TEXT,
            end_reason TEXT,
            notes TEXT,
            updated_at TEXT,
            payload TEXT NOT NULL
        )
        """
    )
    rows = conn.execute("SELECT * FROM occupancies_legacy").fetchall()
    for row in rows:
        payload = json.loads(row["payload"])
        conn.execute(
            """
            INSERT INTO occupancies (
                id, slot_id, room_id, rack_id, cage_code, status, iacuc, project, pi, owner, funding,
                species, billing_item, customer_type, animal_count,
                room_name, rack_name, slot_code, start_date, end_date, end_reason,
                notes, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["id"],
                row["slot_id"],
                payload.get("roomId", ""),
                payload.get("rackId", ""),
                row["cage_code"],
                row["status"],
                row["iacuc"],
                row["project"],
                row["pi"],
                row["owner"],
                payload.get("funding", ""),
                payload.get("species", ""),
                payload.get("billingItem", ""),
                payload.get("customerType", ""),
                as_int(payload.get("animalCount")),
                payload.get("roomName", ""),
                payload.get("rackName", ""),
                payload.get("slotCode", ""),
                row["start_date"],
                row["end_date"],
                row["end_reason"],
                row["notes"],
                row["updated_at"],
                dump_json(payload),
            ),
        )
    conn.execute("DROP TABLE occupancies_legacy")


def migrate_billing_workflow_schema(conn):
    if read_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, False):
        return
    if table_has_rows(conn, "billing_workflows"):
        set_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, True, now_iso())
        return
    if not table_has_rows(conn, "billing_statements"):
        set_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, True, now_iso())
        return

    rows = conn.execute(
        "SELECT id, payload, generated_at, rowid FROM billing_statements ORDER BY month, iacuc, generated_at, rowid"
    ).fetchall()
    grouped = {}
    for row in rows:
        statement = json.loads(row["payload"])
        source_type = normalize_workflow_source(statement.get("sourceType", ""))
        scope_type, scope_key = workflow_scope_for_statement(statement)
        business_key = billing_workflow_business_key(scope_type, scope_key, statement.get("month", ""), source_type)
        grouped.setdefault(business_key, []).append((row, statement, source_type))

    for items in grouped.values():
        first_statement = items[0][1]
        source_type = items[0][2]
        workflow_id = new_id("bwf")
        conn.execute(
            """
            INSERT INTO billing_workflows (
                id, business_key, iacuc, month, source_type, workflow_status,
                current_version_id, current_version_no, latest_event_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                workflow_id,
                billing_workflow_business_key(*workflow_scope_for_statement(first_statement), first_statement.get("month", ""), source_type),
                first_statement.get("iacuc", ""),
                first_statement.get("month", ""),
                source_type,
                WORKFLOW_STATUS_GENERATED,
                "",
                0,
                "",
                dump_json({"id": workflow_id, "migrationPending": True}),
            ),
        )
        versions = []
        latest_at = ""
        for index, (row, statement, grouped_source) in enumerate(items, start=1):
            version_id = row["id"]
            generated_at = statement.get("generatedAt") or row["generated_at"] or now_iso()
            lines = [
                json.loads(line_row["payload"])
                for line_row in conn.execute(
                    "SELECT payload FROM billing_statement_lines WHERE statement_id = ? ORDER BY line_date, rowid",
                    (row["id"],),
                ).fetchall()
            ]
            version_status = VERSION_STATUS_ACTIVE if index == len(items) else VERSION_STATUS_VOIDED
            workflow_status = WORKFLOW_STATUS_GENERATED
            document_number = statement.get("documentNumber") or make_statement_document_number(statement, index)
            enriched = enrich_statement_for_workflow(
                statement,
                workflow_id=workflow_id,
                version_id=version_id,
                version_no=index,
                version_status=version_status,
                workflow_status=workflow_status,
                document_number=document_number,
            )
            version_payload = build_version_payload(
                enriched,
                workflow_id,
                index,
                version_status,
                workflow_status,
                generated_at,
                "",
                "",
                "",
            )
            conn.execute(
                """
                INSERT INTO billing_statement_versions (
                    id, workflow_id, version_no, version_status, workflow_status,
                    generated_at, voided_at, created_by, payload
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    version_id,
                    workflow_id,
                    index,
                    version_status,
                    workflow_status,
                    generated_at,
                    generated_at if version_status == VERSION_STATUS_VOIDED else "",
                    "",
                    dump_json(version_payload),
                ),
            )
            for line in lines:
                conn.execute(
                    """
                    INSERT INTO billing_statement_version_lines (id, version_id, line_date, payload)
                    VALUES (?, ?, ?, ?)
                    """,
                    (line.get("id") or new_id("line"), version_id, line.get("date", ""), dump_json(line)),
                )
            event_at = generated_at
            event_payload = build_workflow_event_payload(
                new_id("wevt"),
                workflow_id,
                version_id,
                "statement_generated",
                "",
                WORKFLOW_STATUS_GENERATED,
                {
                    "displayName": "系统迁移",
                    "username": "system",
                    "id": "system",
                },
                event_at,
                "system",
                "由旧版 billing_statements 迁移生成",
            )
            conn.execute(
                """
                INSERT INTO billing_workflow_events (
                    id, workflow_id, version_id, event_type, from_status, to_status, at, payload
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_payload["id"],
                    workflow_id,
                    version_id,
                    event_payload["eventType"],
                    event_payload["fromStatus"],
                    event_payload["toStatus"],
                    event_payload["at"],
                    dump_json(event_payload),
                ),
            )
            versions.append(version_payload)
            latest_at = max(latest_at, generated_at)

        current_version = versions[-1]
        workflow_payload = build_workflow_payload(
            workflow_id,
            first_statement.get("iacuc", ""),
            first_statement.get("month", ""),
            source_type,
            WORKFLOW_STATUS_GENERATED,
            current_version,
            latest_at,
        )
        conn.execute(
            """
            UPDATE billing_workflows
            SET business_key = ?, iacuc = ?, month = ?, source_type = ?,
                workflow_status = ?, current_version_id = ?, current_version_no = ?,
                latest_event_at = ?, payload = ?
            WHERE id = ?
            """,
            (
                workflow_payload.get("businessKey", ""),
                first_statement.get("iacuc", ""),
                first_statement.get("month", ""),
                grouped_source,
                WORKFLOW_STATUS_GENERATED,
                current_version["id"],
                current_version["versionNo"],
                latest_at,
                dump_json(workflow_payload),
                workflow_id,
            ),
        )
    set_setting(conn, BILLING_WORKFLOW_MIGRATION_KEY, True, now_iso())


def backfill_billing_workflow_scope(conn):
    rows = conn.execute("SELECT id, payload FROM billing_workflows").fetchall()
    for row in rows:
        workflow = json.loads(row["payload"])
        current_version = workflow.get("currentVersion") or {}
        statement = current_version.get("statement") or {}
        scope_type, scope_key = workflow_scope_for_statement(statement)
        desired_key = billing_workflow_business_key(scope_type, scope_key, workflow.get("month", ""), workflow.get("sourceType", ""))
        conflict = conn.execute(
            "SELECT id FROM billing_workflows WHERE business_key = ? AND id != ?",
            (desired_key, row["id"]),
        ).fetchone()
        if conflict:
            desired_key = f"{desired_key}|legacy|{row['id']}"
        if workflow.get("scopeType") == scope_type and workflow.get("scopeKey") == scope_key and desired_key == workflow.get("businessKey"):
            continue
        workflow["scopeType"] = scope_type
        workflow["scopeKey"] = scope_key
        workflow["businessKey"] = desired_key
        workflow["iacucs"] = statement.get("iacucs", workflow.get("iacucs", []))
        conn.execute(
            """
            UPDATE billing_workflows
            SET business_key = ?, payload = ?
            WHERE id = ?
            """,
            (desired_key, dump_json(workflow), row["id"]),
        )


def table_columns(conn, table):
    return {
        row["name"]: {"type": row["type"], "notnull": bool(row["notnull"]), "pk": int(row["pk"] or 0)}
        for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
    }


def repair_missing_cage_slots(conn):
    racks = conn.execute("SELECT id, rows, cols FROM racks").fetchall()
    if not racks:
        return

    existing_rows = conn.execute("SELECT rack_id, row_no, col_no FROM cage_slots").fetchall()
    existing_positions = {
        (row["rack_id"], int(row["row_no"] or 0), int(row["col_no"] or 0))
        for row in existing_rows
    }
    for rack in racks:
        rows = max(as_int(rack["rows"]) or 0, 0)
        cols = max(as_int(rack["cols"]) or 0, 0)
        for row_no in range(1, rows + 1):
            for col_no in range(1, cols + 1):
                position = (rack["id"], row_no, col_no)
                if position in existing_positions:
                    continue
                slot = {
                    "id": slot_id_for_rack(rack["id"], row_no, col_no),
                    "rackId": rack["id"],
                    "row": row_no,
                    "col": col_no,
                    "code": f"{column_label(col_no)}{row_no}",
                    "status": "empty",
                }
                conn.execute(
                    """
                    INSERT INTO cage_slots (id, rack_id, row_no, col_no, code, status, payload)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (slot["id"], slot["rackId"], slot["row"], slot["col"], slot["code"], slot["status"], dump_json(slot)),
                )
                existing_positions.add(position)


def slot_id_for_rack(rack_id, row_no, col_no):
    suffix = str(rack_id).removeprefix("rack-")
    return f"slot-{suffix}-{row_no}-{col_no}"


def column_label(index):
    value = int(index)
    label = ""
    while value > 0:
        value -= 1
        label = chr(65 + (value % 26)) + label
        value //= 26
    return label or "A"


def read_state():
    with connect_db() as conn:
        migrate_legacy_state(conn)
        state = assemble_state(conn)
        updated_at = read_updated_at(conn)
    if not state:
        return {"state": None, "updatedAt": None}
    return {"state": state, "updatedAt": updated_at}


def current_occupancy_by_slot(state):
    current = {}
    for item in state.get("occupancies", []):
        if item.get("slotId") and item.get("status") in ("active", "reserved"):
            current[item.get("slotId")] = item
    return current


def base36_encode(value):
    digits = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    number = int(value or 0)
    if number <= 0:
        return "0"
    output = ""
    while number:
        number, remainder = divmod(number, 36)
        output = digits[remainder] + output
    return output


def hash_base36(value, length=8):
    hash_value = 1469598103934665603
    mask = (1 << 64) - 1
    for char in str(value or ""):
        hash_value ^= ord(char)
        hash_value = (hash_value * 1099511628211) & mask
    return base36_encode(hash_value).rjust(length, "0")[-length:]


CAGE_CARD_QR_ALPHABET = "23456789ABCDEFGHJKLMNPQRSTUVWXYZ"


def is_cage_card_qr_id(value):
    text = clean_text(value).upper()
    return bool(re.fullmatch(rf"[{CAGE_CARD_QR_ALPHABET}]{{4}}", text) or re.fullmatch(r"[A-Z0-9]{8}", text))


def cage_card_qr_id(batch, sequence):
    batch_id = clean_text(batch.get("id") or batch.get("sourceBatchId") or "batch") or "batch"
    card_no = str(max(as_int(sequence) or 0, 0)).zfill(2)
    return hash_base36(f"{batch_id}:{card_no}", 8)


def cage_card_qr_id_from_batch_card(batch, sequence):
    cards = batch.get("cards") if isinstance(batch.get("cards"), list) else []
    index = max(as_int(sequence) or 0, 0) - 1
    if 0 <= index < len(cards) and isinstance(cards[index], dict):
        qr_id = clean_text(cards[index].get("qrId")).upper()
        if is_cage_card_qr_id(qr_id):
            return qr_id
    return cage_card_qr_id(batch, sequence)


def legacy_cage_card_qr_id(batch, sequence):
    iacuc = clean_text(batch.get("iacuc") or "NOIACUC") or "NOIACUC"
    intake_date = clean_text(batch.get("intakeDate") or "nodate").replace("-", "")
    card_no = str(max(as_int(sequence) or 0, 0)).zfill(2)
    return "-".join([iacuc, intake_date, card_no])


def species_label(value):
    return {
        "mouse": "小鼠",
        "rat": "大鼠",
        "guinea_pig": "豚鼠",
        "rabbit": "兔",
        "monkey": "猴",
        "dog": "犬",
        "pig": "猪",
    }.get(clean_text(value), clean_text(value))


def cage_card_status_label(batch, task, occupancy):
    if occupancy:
        return {"reserved": "已预留", "active": "已入驻", "ended": "已结束"}.get(occupancy.get("status"), occupancy.get("status", ""))
    if task:
        return {"pending": "待进驻", "reserved": "已预留", "active": "已入驻", "cancelled": "已取消"}.get(task.get("status"), task.get("status", ""))
    return {"draft": "草稿", "pending_print": "未打印", "printed": "已打印", "received": "已接收"}.get(batch.get("status"), batch.get("status", ""))


def animal_age_text(birth_date, reference_date=None):
    raw = clean_text(birth_date)
    if not raw:
        return ""
    try:
        birth = date.fromisoformat(raw)
        reference = date.fromisoformat(clean_text(reference_date)) if reference_date else date.today()
    except ValueError:
        return ""
    if reference < birth:
        return ""
    months = (reference.year - birth.year) * 12 + reference.month - birth.month
    if reference.day < birth.day:
        months -= 1
    years = months // 12
    remaining_months = months % 12
    if years and remaining_months:
        return f"{years}岁{remaining_months}个月"
    if years:
        return f"{years}岁"
    return f"{max(months, 0)}个月"


def public_cage_card_payload(conn, qr_id):
    target = clean_text(qr_id).upper()
    if not target:
        raise LookupError("二维码地址无效")
    state = assemble_state(conn) or empty_state()
    applications_by_iacuc = read_applications_by_iacuc(conn)
    rooms_by_id = {item.get("id"): item for item in state.get("rooms", [])}
    racks_by_id = {item.get("id"): item for item in state.get("racks", [])}
    slots_by_id = {item.get("id"): item for item in state.get("slots", [])}
    tasks_by_batch_and_sequence = {
        (task.get("sourceBatchId"), as_int(task.get("cardSequence")) or 0): task
        for task in state.get("placementTasks", [])
    }
    occupancies_by_id = {item.get("id"): item for item in state.get("occupancies", [])}
    occupancies_by_qr = {clean_text(item.get("qrId")).upper(): item for item in state.get("occupancies", []) if clean_text(item.get("qrId"))}

    for batch in state.get("intakeBatches", []):
        card_count = max(
            as_int(batch.get("finalCardCount")) or 0,
            as_int(batch.get("suggestedCardCount")) or 0,
            len(batch.get("cards") or []) if isinstance(batch.get("cards"), list) else 0,
        )
        for sequence in range(1, card_count + 1):
            task = tasks_by_batch_and_sequence.get((batch.get("id"), sequence))
            candidate_ids = {
                cage_card_qr_id_from_batch_card(batch, sequence).upper(),
                legacy_cage_card_qr_id(batch, sequence).upper(),
            }
            if task and clean_text(task.get("qrId")):
                candidate_ids.add(clean_text(task.get("qrId")).upper())
            if target not in candidate_ids:
                continue
            occupancy = occupancies_by_qr.get(target)
            if not occupancy and task and task.get("reservedOccupancyId"):
                occupancy = occupancies_by_id.get(task.get("reservedOccupancyId"))
            slot = slots_by_id.get((occupancy or {}).get("slotId"))
            rack = racks_by_id.get((slot or {}).get("rackId") or (occupancy or {}).get("rackId"))
            room = rooms_by_id.get((rack or {}).get("roomId") or (task or {}).get("targetRoomId") or (occupancy or {}).get("roomId"))
            iacuc = normalize_iacuc_number(batch.get("iacuc") or (task or {}).get("iacuc") or (occupancy or {}).get("iacuc"))
            application = applications_by_iacuc.get(iacuc, {})
            animal_count = (task or {}).get("animalCount")
            if animal_count in (None, ""):
                animal_count = (occupancy or {}).get("animalCount")
            if animal_count in (None, ""):
                per_cage = max(as_int(batch.get("suggestedAnimalsPerCage")) or 1, 1)
                quantity = max(as_int(batch.get("quantity")) or 0, 0)
                remainder = quantity % per_cage if quantity and per_cage else 0
                animal_count = "" if remainder and sequence == card_count else per_cage
            birth_date = (occupancy or {}).get("birthDate", "")
            item = {
                "qrId": target,
                "batchNo": batch.get("batchNo", ""),
                "cageCode": (occupancy or {}).get("cageCode", ""),
                "roomName": (room or {}).get("name") or (task or {}).get("targetRoomName") or batch.get("roomName", "") or (occupancy or {}).get("roomName", ""),
                "rackName": (rack or {}).get("name") or (occupancy or {}).get("rackName", ""),
                "slotCode": (slot or {}).get("code") or (occupancy or {}).get("slotCode", ""),
                "iacuc": iacuc,
                "project": batch.get("project") or (task or {}).get("project") or (occupancy or {}).get("project") or application.get("project", ""),
                "pi": batch.get("pi") or (task or {}).get("pi") or (occupancy or {}).get("pi") or application.get("pi", ""),
                "owner": batch.get("owner") or (task or {}).get("owner") or (occupancy or {}).get("owner") or application.get("owner", ""),
                "species": batch.get("species") or (task or {}).get("species") or (occupancy or {}).get("species", ""),
                "speciesLabel": species_label(batch.get("species") or (task or {}).get("species") or (occupancy or {}).get("species", "")),
                "strainStandard": batch.get("strainStandard") or (task or {}).get("strainStandard") or (occupancy or {}).get("strainStandard", ""),
                "animalCount": animal_count,
                "sex": batch.get("sex") or (occupancy or {}).get("sex", ""),
                "birthDate": birth_date,
                "age": animal_age_text(birth_date),
                "startDate": (occupancy or {}).get("startDate") or (task or {}).get("actualMoveInDate", ""),
                "actualMoveInDate": (task or {}).get("actualMoveInDate", ""),
                "endDate": (occupancy or {}).get("endDate") or batch.get("endDate", ""),
                "statusLabel": cage_card_status_label(batch, task, occupancy),
            }
            return {"item": item}
    raise LookupError("该二维码没有匹配到笼卡记录")


def occupancy_period_tone(occupancy):
    if not occupancy or occupancy.get("status") != "active":
        return ""
    end_date = clean_text(occupancy.get("endDate", ""))
    if not end_date:
        return "open"
    return "overdue" if today_iso() > end_date else "normal"


def summarize_infrastructure(state):
    current_by_slot = current_occupancy_by_slot(state)
    rack_by_id = {rack.get("id"): rack for rack in state.get("racks", [])}
    room_by_id = {room.get("id"): room for room in state.get("rooms", [])}
    slot_to_rack_id = {slot.get("id"): slot.get("rackId") for slot in state.get("slots", [])}
    facility_summaries = {}
    current_month = date.today().strftime("%Y-%m")
    room_summaries = {
        room.get("id"): {
            "roomId": room.get("id"),
            "rackCount": 0,
            "slotCount": 0,
            "activeCount": 0,
            "reservedCount": 0,
            "emptyCount": 0,
            "periodOpenCount": 0,
            "periodNormalCount": 0,
            "periodOverdueCount": 0,
            "occupancyRecordCount": 0,
        }
        for room in state.get("rooms", [])
    }
    rack_summaries = {
        rack.get("id"): {
            "rackId": rack.get("id"),
            "roomId": rack.get("roomId"),
            "slotCount": 0,
            "activeCount": 0,
            "reservedCount": 0,
            "emptyCount": 0,
            "periodOpenCount": 0,
            "periodNormalCount": 0,
            "periodOverdueCount": 0,
            "occupancyRecordCount": 0,
        }
        for rack in state.get("racks", [])
    }
    dashboard = {
        "total": 0,
        "active": 0,
        "reserved": 0,
        "empty": 0,
        "periodOpen": 0,
        "periodNormal": 0,
        "periodOverdue": 0,
        "intakePendingCount": 0,
        "openPlacementTaskCount": 0,
        "currentMonthWorkflowTodoCount": 0,
        "currentMonthWorkflowDoneCount": 0,
        "unmatchedIntakeCount": 0,
        "overduePlacementCount": 0,
        "stalledWorkflowCount": 0,
        "exceptionCount": 0,
    }

    def ensure_facility_summary(facility):
        key = clean_text(facility or "zhujiang") or "zhujiang"
        if key not in facility_summaries:
            facility_summaries[key] = {
                "facility": key,
                "roomCount": 0,
                "activeCageCount": 0,
                "activeAnimalCount": 0,
                "openPlacementTaskCount": 0,
                "currentMonthWorkflowTodoCount": 0,
                "currentMonthWorkflowDoneCount": 0,
            }
        return facility_summaries[key]

    for room in state.get("rooms", []):
        ensure_facility_summary(room.get("facility"))["roomCount"] += 1

    for rack in state.get("racks", []):
        room_summary = room_summaries.get(rack.get("roomId"))
        if room_summary:
            room_summary["rackCount"] += 1

    for slot in state.get("slots", []):
        rack_summary = rack_summaries.get(slot.get("rackId"))
        rack = rack_by_id.get(slot.get("rackId"), {})
        room_summary = room_summaries.get(rack.get("roomId"))
        occupancy = current_by_slot.get(slot.get("id"))
        tone = occupancy_period_tone(occupancy)
        status = clean_text(slot.get("status", ""))
        dashboard["total"] += 1
        if rack_summary:
            rack_summary["slotCount"] += 1
        if room_summary:
            room_summary["slotCount"] += 1
        if status == "active":
            dashboard["active"] += 1
            ensure_facility_summary((room_summary and room_by_id.get(room_summary["roomId"], {}).get("facility")) or rack.get("facility"))["activeCageCount"] += 1
            if rack_summary:
                rack_summary["activeCount"] += 1
            if room_summary:
                room_summary["activeCount"] += 1
        elif status == "reserved":
            dashboard["reserved"] += 1
            if rack_summary:
                rack_summary["reservedCount"] += 1
            if room_summary:
                room_summary["reservedCount"] += 1
        else:
            dashboard["empty"] += 1
            if rack_summary:
                rack_summary["emptyCount"] += 1
            if room_summary:
                room_summary["emptyCount"] += 1
        if tone == "open":
            dashboard["periodOpen"] += 1
            if rack_summary:
                rack_summary["periodOpenCount"] += 1
            if room_summary:
                room_summary["periodOpenCount"] += 1
        elif tone == "normal":
            dashboard["periodNormal"] += 1
            if rack_summary:
                rack_summary["periodNormalCount"] += 1
            if room_summary:
                room_summary["periodNormalCount"] += 1
        elif tone == "overdue":
            dashboard["periodOverdue"] += 1
            if rack_summary:
                rack_summary["periodOverdueCount"] += 1
            if room_summary:
                room_summary["periodOverdueCount"] += 1

    for item in state.get("occupancies", []):
        slot_id = item.get("slotId")
        rack_id = item.get("rackId") or slot_to_rack_id.get(slot_id)
        room_id = item.get("roomId") or rack_by_id.get(rack_id, {}).get("roomId")
        room = room_by_id.get(room_id, {})
        if item.get("status") == "active":
            facility_profile = billing_profile_for_room(room)
            ensure_facility_summary(facility_profile.get("facility"))["activeAnimalCount"] += occupancy_animal_count(item, facility_profile)
        if rack_id and rack_id in rack_summaries:
            rack_summaries[rack_id]["occupancyRecordCount"] += 1
        if room_id and room_id in room_summaries:
            room_summaries[room_id]["occupancyRecordCount"] += 1

    for batch in state.get("intakeBatches", []):
        if clean_text(batch.get("status")) != "received":
            dashboard["intakePendingCount"] += 1
        if clean_text(batch.get("roomName")) and not batch.get("roomMatched"):
            dashboard["unmatchedIntakeCount"] += 1

    for task in state.get("placementTasks", []):
        status = clean_text(task.get("status"))
        if status in ("active", "cancelled"):
            continue
        room = room_by_id.get(task.get("targetRoomId"), {})
        ensure_facility_summary(billing_profile_for_room(room).get("facility"))["openPlacementTaskCount"] += 1
        dashboard["openPlacementTaskCount"] += 1
        planned_move_in = clean_text(task.get("plannedMoveInDate"))
        if planned_move_in and planned_move_in < today_iso():
            dashboard["overduePlacementCount"] += 1

    for workflow in state.get("billingWorkflows", []):
        if clean_text(workflow.get("month")) != current_month:
            continue
        workflow_status = clean_text(workflow.get("workflowStatus"))
        facility_keys = set()
        statement = (workflow.get("currentVersion") or {}).get("statement") or {}
        room_name = clean_text(statement.get("roomName"))
        if room_name:
            room = next((item for item in state.get("rooms", []) if clean_text(item.get("name")) == room_name), None)
            if room:
                facility_keys.add(billing_profile_for_room(room).get("facility"))
        if not facility_keys and clean_text(workflow.get("sourceType")) == "cage_map":
            pi = clean_text(workflow.get("pi"))
            if pi:
                for item in state.get("occupancies", []):
                    if clean_text(item.get("pi")) != pi or not occupancy_overlaps_month(item, current_month):
                        continue
                    rack_id = item.get("rackId") or slot_to_rack_id.get(item.get("slotId"))
                    room_id = item.get("roomId") or rack_by_id.get(rack_id, {}).get("roomId")
                    facility_keys.add(billing_profile_for_room(room_by_id.get(room_id, {})).get("facility"))
        if not facility_keys:
            facility_keys.add("zhujiang")
        if workflow_status == WORKFLOW_STATUS_FINANCE:
            dashboard["currentMonthWorkflowDoneCount"] += 1
            for key in facility_keys:
                ensure_facility_summary(key)["currentMonthWorkflowDoneCount"] += 1
            continue
        dashboard["currentMonthWorkflowTodoCount"] += 1
        if workflow_status in (WORKFLOW_STATUS_GENERATED, WORKFLOW_STATUS_SENT, WORKFLOW_STATUS_SIGNED):
            dashboard["stalledWorkflowCount"] += 1
        for key in facility_keys:
            ensure_facility_summary(key)["currentMonthWorkflowTodoCount"] += 1

    dashboard["exceptionCount"] = (
        dashboard["unmatchedIntakeCount"]
        + dashboard["overduePlacementCount"]
        + dashboard["stalledWorkflowCount"]
    )

    return {
        "dashboardSummary": dashboard,
        "roomSummaries": list(room_summaries.values()),
        "rackSummaries": list(rack_summaries.values()),
        "facilitySummaries": list(facility_summaries.values()),
    }


def read_bootstrap_state(conn, actor, scope="summary", room_id=""):
    started_at = time.perf_counter()
    state = filter_state_for_actor(read_cached_state(conn), actor)
    actor_scope = actor_cache_scope(actor)
    if scope == "summary":
        key = cache_key("bootstrap_summary", actor=actor_scope)
        cached = cache_get(key)
        if cached is not None:
            log_perf("bootstrap", started_at, scope=scope, cached=1, rooms=len(cached.get("rooms", [])))
            return cached
    summary = summarize_infrastructure(state)
    payload = {
        "rooms": state.get("rooms", []),
        "racks": state.get("racks", []),
        **summary,
    }
    if scope == "full":
        payload["billingRules"] = state.get("billingRules", [])
        payload["adjustments"] = state.get("adjustments", [])
        payload["intakeBatches"] = state.get("intakeBatches", [])
        payload["placementTasks"] = state.get("placementTasks", [])
        payload["slots"] = state.get("slots", [])
        payload["occupancies"] = state.get("occupancies", [])
        log_perf("bootstrap", started_at, scope=scope, rooms=len(payload["rooms"]), slots=len(payload["slots"]))
        return payload
    if scope == "room":
        room_id = clean_text(room_id)
        payload["placementTasks"] = [item for item in state.get("placementTasks", []) if item.get("targetRoomId") == room_id]
        rack_ids = {rack.get("id") for rack in state.get("racks", []) if rack.get("roomId") == room_id}
        payload["slots"] = [slot for slot in state.get("slots", []) if slot.get("rackId") in rack_ids]
        slot_ids = {slot.get("id") for slot in payload["slots"]}
        payload["occupancies"] = [item for item in state.get("occupancies", []) if item.get("slotId") in slot_ids]
        log_perf("bootstrap", started_at, scope=scope, room_id=room_id, rooms=len(payload["rooms"]), slots=len(payload["slots"]))
        return payload
    payload["slots"] = []
    payload["occupancies"] = []
    cached_payload = cache_set(key, payload) if scope == "summary" else payload
    log_perf("bootstrap", started_at, scope=scope, rooms=len(payload["rooms"]))
    return cached_payload


def actor_cache_scope(actor):
    if not actor:
        return "anonymous"
    role = actor.get("role", "")
    rooms = ",".join(sorted(clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)))
    return f"{role}:{rooms}"


def month_range(month):
    normalized = clean_text(month)
    if not re.fullmatch(r"\d{4}-\d{2}", normalized):
        raise ValueError("结算月份格式应为 YYYY-MM")
    year, month_no = normalized.split("-")
    start = f"{year}-{month_no}-01"
    last_day = calendar.monthrange(int(year), int(month_no))[1]
    end = f"{year}-{month_no}-{last_day:02d}"
    return start, end


def occupancy_overlaps_month(item, month):
    start, end = month_range(month)
    start_date = clean_text(item.get("startDate") or item.get("start_date"))
    end_date = clean_text(item.get("endDate") or item.get("end_date"))
    if not start_date:
        return False
    if start_date > end:
        return False
    if end_date and end_date < start:
        return False
    return True


def read_billing_occupancies(conn, actor, filters):
    started_at = time.perf_counter()
    month = clean_text(filters.get("month", ""))
    iacuc = clean_text(filters.get("iacuc", ""))
    pi = clean_text(filters.get("pi", ""))
    if not month:
        raise ValueError("请提供结算月份")
    key = cache_key(
        "billing_occupancies",
        actor=actor_cache_scope(actor),
        month=month,
        iacuc=iacuc,
        pi=pi,
    )
    cached = cache_get(key)
    if cached is not None:
        log_perf("billing_occupancies", started_at, cached=1, month=month, occupancies=len(cached.get("occupancies", [])))
        return cached

    occupancies = read_occupancies_for_billing(conn, month, iacuc=iacuc, pi=pi)
    if actor and actor.get("role") != "admin":
        allowed_room_ids = {clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)}
        occupancies = [item for item in occupancies if clean_text(item.get("roomId", "")) in allowed_room_ids]
    state = read_billing_state_for_occupancies(conn, occupancies)
    matched_slot_ids = {item.get("slotId") for item in occupancies if item.get("slotId")}
    payload = {
        "month": month,
        "pi": pi,
        "iacuc": iacuc,
        "slots": [slot for slot in state["slots"] if slot.get("id") in matched_slot_ids],
        "occupancies": occupancies,
    }
    cache_set(key, payload)
    log_perf("billing_occupancies", started_at, month=month, occupancies=len(occupancies), slots=len(payload["slots"]))
    return payload


def read_occupancies_for_billing(conn, month, iacuc="", pi=""):
    start, end = month_range(month)
    clauses = [
        "status IN ('active', 'ended')",
        "start_date <> ''",
        "start_date <= ?",
        "(end_date IS NULL OR end_date = '' OR end_date >= ?)",
    ]
    params = [end, start]
    normalized_iacuc = normalize_iacuc_number(iacuc)
    normalized_pi = clean_text(pi)
    if normalized_iacuc:
        clauses.append("iacuc = ?")
        params.append(normalized_iacuc)
    if normalized_pi:
        clauses.append("pi = ?")
        params.append(normalized_pi)

    rows = conn.execute(
        f"SELECT payload FROM occupancies WHERE {' AND '.join(clauses)} ORDER BY start_date, rowid",
        tuple(params),
    ).fetchall()
    items = [json.loads(row["payload"]) for row in rows]
    slot_ids = {item.get("slotId") for item in items if item.get("slotId")}
    slots = read_payloads_by_ids(conn, "cage_slots", slot_ids, "rack_id, row_no, col_no, rowid")
    rack_ids = {slot.get("rackId") for slot in slots if slot.get("rackId")}
    racks = read_payloads_by_ids(conn, "racks", rack_ids, "room_id, index_no, rowid")
    room_ids = {rack.get("roomId") for rack in racks if rack.get("roomId")}
    rooms = read_payloads_by_ids(conn, "rooms", room_ids, "rowid")
    applications_by_iacuc = read_applications_by_iacuc(conn)
    state = {"rooms": rooms, "racks": racks, "slots": slots}
    return [occupancy_with_snapshots(item, state, applications_by_iacuc) for item in items]


def read_payloads_by_ids(conn, table, ids, order_by="rowid"):
    id_list = sorted({clean_text(item) for item in ids if clean_text(item)})
    if not id_list:
        return []
    rows = conn.execute(
        f"SELECT payload FROM {table} WHERE id IN ({placeholders(id_list)}) ORDER BY {order_by}",
        tuple(id_list),
    ).fetchall()
    return [json.loads(row["payload"]) for row in rows]


def read_billing_state_for_occupancies(conn, occupancies):
    slot_ids = {item.get("slotId") for item in occupancies if item.get("slotId")}
    slots = read_payloads_by_ids(conn, "cage_slots", slot_ids, "rack_id, row_no, col_no, rowid")
    rack_ids = {slot.get("rackId") for slot in slots if slot.get("rackId")}
    rack_ids.update(item.get("rackId") for item in occupancies if item.get("rackId"))
    racks = read_payloads_by_ids(conn, "racks", rack_ids, "room_id, index_no, rowid")
    room_ids = {rack.get("roomId") for rack in racks if rack.get("roomId")}
    room_ids.update(item.get("roomId") for item in occupancies if item.get("roomId"))
    rooms = read_payloads_by_ids(conn, "rooms", room_ids, "rowid")
    return {"rooms": rooms, "racks": racks, "slots": slots}


def filter_state_for_actor(state, actor):
    if not state or not actor or actor.get("role") == "admin":
        return state

    allowed_rooms = set(actor.get("roomIds", []))
    rooms = [room for room in state.get("rooms", []) if room.get("id") in allowed_rooms]
    room_ids = {room.get("id") for room in rooms}
    racks = [rack for rack in state.get("racks", []) if rack.get("roomId") in room_ids]
    rack_ids = {rack.get("id") for rack in racks}
    slots = [slot for slot in state.get("slots", []) if slot.get("rackId") in rack_ids]
    slot_ids = {slot.get("id") for slot in slots}
    occupancies = [item for item in state.get("occupancies", []) if item.get("slotId") in slot_ids]
    placement_tasks = [item for item in state.get("placementTasks", []) if item.get("targetRoomId") in room_ids]

    return {
        **state,
        "rooms": rooms,
        "racks": racks,
        "slots": slots,
        "occupancies": occupancies,
        "placementTasks": placement_tasks,
    }


def filter_entity_payloads_for_actor(collection, items, actor):
    if not actor or actor.get("role") == "admin":
        return items

    allowed_rooms = {clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)}
    if collection == "rooms":
        return [item for item in items if item.get("id") in allowed_rooms]
    if collection == "racks":
        return [item for item in items if item.get("roomId") in allowed_rooms]
    if collection == "placementTasks":
        return [item for item in items if item.get("targetRoomId") in allowed_rooms]
    if collection == "slots":
        rack_ids = {clean_text(item.get("rackId", "")) for item in items if clean_text(item.get("rackId", ""))}
        rack_rooms = read_rack_room_map(rack_ids)
        return [item for item in items if rack_rooms.get(item.get("rackId")) in allowed_rooms]
    if collection == "occupancies":
        direct = [item for item in items if item.get("roomId") in allowed_rooms]
        unresolved = [item for item in items if not item.get("roomId")]
        if not unresolved:
            return direct
        slot_ids = {clean_text(item.get("slotId", "")) for item in unresolved if clean_text(item.get("slotId", ""))}
        slot_rooms = read_slot_room_map(slot_ids)
        return direct + [item for item in unresolved if slot_rooms.get(item.get("slotId")) in allowed_rooms]
    return items


def placeholders(values):
    return ", ".join("?" for _ in values)


def read_rack_room_map(rack_ids):
    ids = sorted({clean_text(item) for item in rack_ids if clean_text(item)})
    if not ids:
        return {}
    with connect_db() as conn:
        rows = conn.execute(f"SELECT id, room_id FROM racks WHERE id IN ({placeholders(ids)})", ids).fetchall()
    return {row["id"]: row["room_id"] for row in rows}


def read_slot_room_map(slot_ids):
    ids = sorted({clean_text(item) for item in slot_ids if clean_text(item)})
    if not ids:
        return {}
    with connect_db() as conn:
        rows = conn.execute(
            f"""
            SELECT slots.id AS slot_id, racks.room_id AS room_id
            FROM cage_slots AS slots
            JOIN racks ON racks.id = slots.rack_id
            WHERE slots.id IN ({placeholders(ids)})
            """,
            ids,
        ).fetchall()
    return {row["slot_id"]: row["room_id"] for row in rows}


def write_state(state, actor, skip_permission=False):
    updated_at = datetime.now(timezone.utc).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or {}
        if not skip_permission:
            validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        state["auditLogs"] = merge_audit_logs(state.get("auditLogs", []), events)
        write_normalized_state(conn, state, updated_at)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state", "principal_identities")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "quantity_sheets::", "billing_workflows::")
    return {"ok": True, "updatedAt": updated_at, "auditLogs": merge_audit_logs([], events)}


def write_entity_state(endpoint, method, item_id, payload, actor):
    spec = WRITABLE_ENTITY_ENDPOINTS[endpoint]
    collection = spec["collection"]
    if collection == "occupancies":
        return write_occupancy_entity_state(method, item_id, payload, actor, spec)
    if collection == "intakeBatches":
        return write_intake_batch_entity_state(method, item_id, payload, actor, spec)
    if collection == "placementTasks":
        return write_placement_task_entity_state(method, item_id, payload, actor, spec)
    if collection in ("rooms", "racks", "slots"):
        return write_infrastructure_entity_state(collection, method, item_id, payload, actor, spec)

    current = read_state()
    state = current.get("state") or empty_state()

    item = normalize_entity_payload(collection, payload, item_id, method, spec["id_prefix"])
    status = HTTPStatus.OK
    if method == "POST":
        insert_entity(state, collection, item)
        status = HTTPStatus.CREATED
    elif method == "PUT":
        replace_entity(state, collection, item_id, item)
    elif method == "DELETE":
        item = delete_entity(state, collection, item_id)
    else:
        raise ValueError("Unsupported entity write method")

    if collection == "occupancies":
        sync_slot_statuses(state)

    result = write_state(state, actor)
    response = {"item": item, "updatedAt": result["updatedAt"], "auditLogs": result["auditLogs"]}
    if collection == "intakeBatches" and method != "DELETE":
        response["placementTasks"] = [task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == item.get("id")]
    return response, status


def write_infrastructure_entity_state(collection, method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(timezone.utc).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        item = normalize_entity_payload(collection, payload, item_id, method, spec["id_prefix"])
        if method == "POST":
            insert_entity(state, collection, item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, collection, item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, collection, item_id)
        else:
            raise ValueError("Unsupported entity write method")

        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        if collection == "rooms":
            if method == "POST":
                insert_room_record(conn, item)
            elif method == "PUT":
                update_room_record(conn, item)
            else:
                delete_room_record(conn, item_id)
        elif collection == "racks":
            if method == "POST":
                insert_rack_record(conn, item)
            elif method == "PUT":
                update_rack_record(conn, item)
            else:
                delete_rack_record(conn, item_id)
        elif collection == "slots":
            if method == "POST":
                insert_slot_record(conn, item)
            elif method == "PUT":
                update_slot_record(conn, item)
            else:
                delete_slot_record(conn, item_id)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "intake_batches::", "placement_tasks::")
    response = {
        "item": item,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1, collection=collection, method=method),
    }
    log_perf("infrastructure_entity.save", started_at, collection=collection, method=method)
    return response, status


def write_intake_batch_entity_state(method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(timezone.utc).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        item = normalize_entity_payload("intakeBatches", payload, item_id, method, spec["id_prefix"])
        if method == "POST":
            insert_entity(state, "intakeBatches", item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, "intakeBatches", item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, "intakeBatches", item_id)
        else:
            raise ValueError("Unsupported entity write method")

        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        if method == "DELETE":
            delete_intake_batch_repository(conn, item_id)
            conn.execute("DELETE FROM placement_tasks WHERE source_batch_id = ?", (item_id,))
        else:
            saved_item = next((entry for entry in state.get("intakeBatches", []) if entry.get("id") == item.get("id")), item)
            upsert_intake_batch_repository(conn, saved_item)
            item = saved_item
            next_task_ids = {
                task.get("id")
                for task in state.get("placementTasks", [])
                if task.get("sourceBatchId") == item.get("id")
            }
            old_task_ids = {
                task.get("id")
                for task in old_state.get("placementTasks", [])
                if task.get("sourceBatchId") == item.get("id")
            }
            for task_id in sorted(old_task_ids - next_task_ids):
                delete_placement_task_repository(conn, task_id)
            for task in state.get("placementTasks", []):
                if task.get("sourceBatchId") == item.get("id"):
                    upsert_placement_task_repository(conn, task)
        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "intake_batches::", "placement_tasks::")
    response = {
        "item": item,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1, method=method),
    }
    if method != "DELETE":
        response["placementTasks"] = [task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == item.get("id")]
        response["perf"] = write_perf_summary(started_at, rows_changed=1 + len(response["placementTasks"]), method=method)
    log_perf("intake_batch.save", started_at, method=method, tasks=len(response.get("placementTasks", [])))
    return response, status


def write_placement_task_entity_state(method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(timezone.utc).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        item = normalize_entity_payload("placementTasks", payload, item_id, method, spec["id_prefix"])
        old_item = next((entry for entry in old_state.get("placementTasks", []) if entry.get("id") == item_id), None)
        removed_occupancy_id = ""
        if method == "POST":
            insert_entity(state, "placementTasks", item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, "placementTasks", item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, "placementTasks", item_id)
            removed_occupancy_id = clean_text((old_item or item).get("reservedOccupancyId", ""))
        else:
            raise ValueError("Unsupported entity write method")

        affected_slot_ids = {
            clean_text(occ.get("slotId", ""))
            for occ in old_state.get("occupancies", [])
            if removed_occupancy_id and occ.get("id") == removed_occupancy_id
        }
        sync_slot_statuses(state)
        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        if method == "DELETE":
            delete_placement_task_repository(conn, item_id)
            if removed_occupancy_id:
                conn.execute("DELETE FROM occupancies WHERE id = ?", (removed_occupancy_id,))
        else:
            saved_item = next((entry for entry in state.get("placementTasks", []) if entry.get("id") == item.get("id")), item)
            upsert_placement_task_repository(conn, saved_item)
            item = saved_item
        affected_slots = [slot for slot in state.get("slots", []) if slot.get("id") in affected_slot_ids]
        for slot in affected_slots:
            update_slot_record(conn, slot)
        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "placement_tasks::")
    response = {
        "item": item,
        "affectedSlots": affected_slots,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + len(affected_slots), method=method),
    }
    log_perf("placement_task.save", started_at, method=method, slot_count=len(affected_slots))
    return response, status


def persist_intake_receipt_confirmation(batch_id, body, actor):
    started_at = time.perf_counter()
    updated_at = datetime.now(timezone.utc).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        batch, receipt, tasks = confirm_intake_receipt(state, batch_id, body, actor)
        events = build_audit_events(actor, old_state, state, updated_at)
        upsert_intake_batch_repository(conn, batch)
        for task in tasks:
            upsert_placement_task_repository(conn, task)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::", "intake_batches::", "placement_tasks::")
    log_perf("intake_batch.confirm", started_at, tasks=len(tasks))
    return {
        "batch": batch,
        "receipt": receipt,
        "tasks": tasks,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + len(tasks), tasks=len(tasks)),
    }


def persist_placement_action(task_id, actor, mutator):
    started_at = time.perf_counter()
    updated_at = datetime.now(timezone.utc).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        result = mutator(state)
        task = result[0] if isinstance(result, tuple) else result
        occupancy = result[1] if isinstance(result, tuple) and len(result) > 1 else None
        affected_slot_ids = changed_placement_slot_ids(old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        upsert_placement_task_repository(conn, task)
        if occupancy:
            applications_by_iacuc = read_applications_by_iacuc(conn)
            occupancy = occupancy_with_snapshots(occupancy, state, applications_by_iacuc)
            upsert_occupancy_record(conn, occupancy)
        affected_slots = [slot for slot in state.get("slots", []) if slot.get("id") in affected_slot_ids]
        for slot in affected_slots:
            update_slot_record(conn, slot)
        write_audit_events(conn, events)
        conn.commit()
    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::")
    log_perf("placement_task.action", started_at, task_id=task_id, slot_count=len(affected_slots))
    payload = {
        "task": task,
        "affectedSlots": affected_slots,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + (1 if occupancy else 0) + len(affected_slots), slot_count=len(affected_slots)),
    }
    if occupancy:
        payload["occupancy"] = occupancy
    return payload


def changed_placement_slot_ids(old_state, state):
    old_occupancies = {item.get("id"): item for item in old_state.get("occupancies", [])}
    new_occupancies = {item.get("id"): item for item in state.get("occupancies", [])}
    slot_ids = set()
    for occupancy_id in changed_keys(old_occupancies, new_occupancies):
        old_item = old_occupancies.get(occupancy_id) or {}
        new_item = new_occupancies.get(occupancy_id) or {}
        for item in (old_item, new_item):
            slot_id = clean_text(item.get("slotId", ""))
            if slot_id:
                slot_ids.add(slot_id)
    return slot_ids


def write_occupancy_entity_state(method, item_id, payload, actor, spec):
    started_at = time.perf_counter()
    updated_at = datetime.now(timezone.utc).isoformat()
    status = HTTPStatus.OK
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))
        old_item = next((item for item in old_state.get("occupancies", []) if item.get("id") == item_id), None)
        item = normalize_entity_payload("occupancies", payload, item_id, method, spec["id_prefix"])
        if method == "POST":
            insert_entity(state, "occupancies", item)
            status = HTTPStatus.CREATED
        elif method == "PUT":
            replace_entity(state, "occupancies", item_id, item)
        elif method == "DELETE":
            item = delete_entity(state, "occupancies", item_id)
        else:
            raise ValueError("Unsupported entity write method")

        affected_slot_ids = {clean_text((old_item or {}).get("slotId", "")), clean_text((item or {}).get("slotId", ""))}
        affected_slot_ids.discard("")
        sync_slot_statuses(state)
        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)
        applications_by_iacuc = read_applications_by_iacuc(conn)
        saved_item = None
        if method == "DELETE":
            conn.execute("DELETE FROM occupancies WHERE id = ?", (item_id,))
        else:
            saved_item = next((entry for entry in state.get("occupancies", []) if entry.get("id") == item.get("id")), item)
            saved_item = occupancy_with_snapshots(saved_item, state, applications_by_iacuc)
            upsert_occupancy_record(conn, saved_item)

        affected_slots = [slot for slot in state.get("slots", []) if slot.get("id") in affected_slot_ids]
        for slot in affected_slots:
            update_slot_record(conn, slot)
        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::")
    response_item = item if method == "DELETE" else saved_item
    response = {
        "item": response_item,
        "affectedSlots": affected_slots,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
        "perf": write_perf_summary(started_at, rows_changed=1 + len(affected_slots), method=method),
    }
    log_perf("occupancy.save", started_at, method=method, slot_count=len(affected_slots))
    return response, status


def upsert_occupancy_record(conn, occupancy):
    structured = occupancy_structured_values(occupancy)
    conn.execute(
        """
        INSERT INTO occupancies (
            id, slot_id, room_id, rack_id, cage_code, status, iacuc, project, pi, owner, funding,
            species, billing_item, customer_type, animal_count, room_name, rack_name, slot_code,
            start_date, end_date, end_reason, notes, updated_at, payload
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            slot_id = excluded.slot_id,
            room_id = excluded.room_id,
            rack_id = excluded.rack_id,
            cage_code = excluded.cage_code,
            status = excluded.status,
            iacuc = excluded.iacuc,
            project = excluded.project,
            pi = excluded.pi,
            owner = excluded.owner,
            funding = excluded.funding,
            species = excluded.species,
            billing_item = excluded.billing_item,
            customer_type = excluded.customer_type,
            animal_count = excluded.animal_count,
            room_name = excluded.room_name,
            rack_name = excluded.rack_name,
            slot_code = excluded.slot_code,
            start_date = excluded.start_date,
            end_date = excluded.end_date,
            end_reason = excluded.end_reason,
            notes = excluded.notes,
            updated_at = excluded.updated_at,
            payload = excluded.payload
        """,
        (
            occupancy.get("id"),
            occupancy.get("slotId"),
            structured["room_id"],
            structured["rack_id"],
            occupancy.get("cageCode", ""),
            occupancy.get("status", ""),
            occupancy.get("iacuc", ""),
            occupancy.get("project", ""),
            occupancy.get("pi", ""),
            occupancy.get("owner", ""),
            occupancy.get("funding", ""),
            structured["species"],
            structured["billing_item"],
            structured["customer_type"],
            structured["animal_count"],
            occupancy.get("roomName", ""),
            occupancy.get("rackName", ""),
            occupancy.get("slotCode", ""),
            occupancy.get("startDate", ""),
            occupancy.get("endDate", ""),
            occupancy.get("endReason", ""),
            occupancy.get("notes", ""),
            occupancy.get("updatedAt", ""),
            dump_json(occupancy),
        ),
    )


def update_slot_record(conn, slot):
    conn.execute(
        """
        UPDATE cage_slots
        SET rack_id = ?, row_no = ?, col_no = ?, code = ?, status = ?, payload = ?
        WHERE id = ?
        """,
        (
            slot.get("rackId"),
            as_int(slot.get("row")),
            as_int(slot.get("col")),
            slot.get("code", ""),
            slot.get("status", "empty"),
            dump_json(slot),
            slot.get("id"),
        ),
    )


def write_infrastructure_state(payload, actor):
    if not isinstance(payload, dict):
        raise ValueError("Request body must be a JSON object")

    created_rooms = normalize_entity_batch("rooms", payload.get("rooms", []), "POST")
    updated_rooms = normalize_entity_batch("rooms", payload.get("roomUpdates", []), "PUT")
    created_racks = normalize_entity_batch("racks", payload.get("racks", []), "POST")
    updated_racks = normalize_entity_batch("racks", payload.get("rackUpdates", []), "PUT")
    deleted_rack_ids = normalize_id_batch(payload.get("rackDeletes", []), "rackDeletes")
    created_slots = normalize_entity_batch("slots", payload.get("slots", []), "POST")
    deleted_slot_ids = normalize_id_batch(payload.get("slotDeletes", []), "slotDeletes")

    updated_at = datetime.now(timezone.utc).isoformat()
    with connect_db() as conn:
        old_state = assemble_state(conn) or empty_state()
        state = json.loads(json.dumps(old_state))

        for room in created_rooms:
            insert_entity(state, "rooms", room)
        for room in updated_rooms:
            replace_entity(state, "rooms", room["id"], room)
        for rack in created_racks:
            insert_entity(state, "racks", rack)
        for rack in updated_racks:
            replace_entity(state, "racks", rack["id"], rack)
        for slot in created_slots:
            insert_entity(state, "slots", slot)
        validate_infrastructure_slot_deletes(state, deleted_slot_ids)
        for slot_id in deleted_slot_ids:
            delete_entity(state, "slots", slot_id)
        for rack_id in deleted_rack_ids:
            delete_entity(state, "racks", rack_id)

        validate_state_write_permission(conn, actor, old_state, state)
        events = build_audit_events(actor, old_state, state, updated_at)

        for room in created_rooms:
            insert_room_record(conn, room)
        for room in updated_rooms:
            update_room_record(conn, room)
        for rack in created_racks:
            insert_rack_record(conn, rack)
        for rack in updated_racks:
            update_rack_record(conn, rack)
        for slot in created_slots:
            insert_slot_record(conn, slot)
        for slot_id in deleted_slot_ids:
            conn.execute("DELETE FROM cage_slots WHERE id = ?", (slot_id,))
        for rack_id in deleted_rack_ids:
            conn.execute("DELETE FROM racks WHERE id = ?", (rack_id,))

        write_audit_events(conn, events)
        conn.commit()

    invalidate_data_cache("assembled_state")
    invalidate_data_cache_prefixes("bootstrap_summary::", "billing_occupancies::")
    return {
        "rooms": created_rooms,
        "roomUpdates": updated_rooms,
        "racks": created_racks,
        "rackUpdates": updated_racks,
        "rackDeletes": deleted_rack_ids,
        "slots": created_slots,
        "slotDeletes": deleted_slot_ids,
        "updatedAt": updated_at,
        "auditLogs": merge_audit_logs([], events),
    }


def normalize_entity_batch(collection, items, method):
    if items is None:
        return []
    if not isinstance(items, list):
        raise ValueError("批量保存内容必须是数组")

    spec = {
        "rooms": {"id_prefix": "room"},
        "racks": {"id_prefix": "rack"},
        "slots": {"id_prefix": "slot"},
    }[collection]
    normalized = []
    for item in items:
        item_id = item.get("id") if isinstance(item, dict) else None
        normalized.append(normalize_entity_payload(collection, item, item_id, method, spec["id_prefix"]))
    return normalized


def normalize_id_batch(items, label):
    if items is None:
        return []
    if not isinstance(items, list):
        raise ValueError(f"{label} 必须是数组")
    return [str(item) for item in items if str(item).strip()]


def validate_infrastructure_slot_deletes(state, slot_ids):
    if not slot_ids:
        return
    deleting = set(slot_ids)
    active = [
        item
        for item in state.get("occupancies", [])
        if item.get("slotId") in deleting and item.get("status") in ("active", "reserved")
    ]
    if active:
        raise ValueError("不能移除仍在用或已预约的笼位")


def empty_state():
    return {
        "baseRate": 4.5,
        "billingMonth": "",
        "billingIacuc": "",
        "rooms": [],
        "racks": [],
        "slots": [],
        "occupancies": [],
        "placementTasks": [],
        "billingRules": [],
        "adjustments": [],
        "intakeBatches": [],
        "auditLogs": [],
    }


def normalize_entity_payload(collection, payload, item_id, method, id_prefix):
    if method == "DELETE":
        return {}
    if not isinstance(payload, dict):
        raise ValueError("Request body must be a JSON object")

    item = dict(payload.get("item") if isinstance(payload.get("item"), dict) else payload)
    if method == "POST":
        item["id"] = str(item.get("id") or new_id(id_prefix))
    else:
        if not item_id:
            raise ValueError("Entity id is required")
        item["id"] = item_id

    validate_entity_payload(collection, item)
    return item


def validate_entity_payload(collection, item):
    if not item.get("id"):
        raise ValueError("实体 id 不能为空")
    if collection == "rooms":
        require_text(item, "name", "饲养间名称不能为空")
    elif collection == "racks":
        require_text(item, "roomId", "笼架必须关联饲养间")
        require_text(item, "name", "笼架名称不能为空")
    elif collection == "slots":
        require_text(item, "rackId", "笼位必须关联笼架")
        status = item.get("status", "empty")
        if status not in ("empty", "reserved", "active"):
            raise ValueError("笼位状态只能是 empty、reserved 或 active")
    elif collection == "occupancies":
        require_text(item, "slotId", "占用记录必须关联笼位")
        status = item.get("status")
        if status not in ("reserved", "active", "ended"):
            raise ValueError("占用状态只能是 reserved、active 或 ended")
    elif collection == "placementTasks":
        require_text(item, "sourceBatchId", "来源批次不能为空")
        require_text(item, "sourceReceiptId", "来源接收记录不能为空")
        require_text(item, "plannedMoveInDate", "计划入驻日期不能为空")
        status = item.get("status")
        if status not in ("pending", "reserved", "active", "cancelled"):
            raise ValueError("待进驻状态只能是 pending、reserved、active 或 cancelled")
    elif collection == "billingRules":
        require_text(item, "unit", "计费规则单位不能为空")
    elif collection == "adjustments":
        require_text(item, "targetType", "减免规则目标类型不能为空")
        require_text(item, "targetId", "减免规则目标不能为空")
    elif collection == "intakeBatches":
        require_text(item, "batchNo", "批次号不能为空")
        require_text(item, "supplier", "购买单位不能为空")
        require_text(item, "status", "批次状态不能为空")
        quantity = as_int(item.get("quantity"))
        if quantity is None or quantity <= 0:
            raise ValueError("动物数量必须大于 0")


def require_text(item, key, message):
    if not str(item.get(key, "")).strip():
        raise ValueError(message)


def insert_entity(state, collection, item):
    items = state.setdefault(collection, [])
    if any(existing.get("id") == item["id"] for existing in items):
        raise sqlite3.IntegrityError(f"Duplicate id: {item['id']}")
    validate_entity_references(state, collection, item)
    items.append(item)


def reconcile_intake_batch_update(state, old_item, item):
    next_item = {**old_item, **item}
    next_item["receipts"] = [dict(receipt) for receipt in (item.get("receipts") if isinstance(item.get("receipts"), list) else old_item.get("receipts", []))]
    final_count = max(as_int(next_item.get("finalCardCount")) or 0, 0)
    confirmed_count = max(as_int(next_item.get("confirmedCardCount")) or sum(max(as_int(receipt.get("cardCount")) or 0, 0) for receipt in next_item.get("receipts", [])), 0)
    next_item["confirmedCardCount"] = confirmed_count
    next_item["remainingCardCount"] = max(as_int(next_item.get("remainingCardCount")) if next_item.get("remainingCardCount") not in (None, "") else final_count - confirmed_count, 0)

    old_status = clean_text(old_item.get("status", ""))
    new_status = clean_text(next_item.get("status", ""))
    old_room_name = clean_text(old_item.get("roomName", ""))
    new_room_name = clean_text(next_item.get("roomName", ""))
    if old_status == "received" and new_status == "printed":
        related_tasks = [task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == old_item.get("id")]
        blocking = [task for task in related_tasks if task.get("status") in ("reserved", "active")]
        if blocking:
            raise ValueError("该批次已有已预留或已入驻的待进驻任务，请先处理相关任务后再回退为已打印")
        state["placementTasks"] = [task for task in state.get("placementTasks", []) if task.get("sourceBatchId") != old_item.get("id")]
        next_item["receipts"] = []
        next_item["confirmedCardCount"] = 0
        next_item["remainingCardCount"] = final_count
    if old_room_name != new_room_name:
        related_tasks = [task for task in state.get("placementTasks", []) if task.get("sourceBatchId") == old_item.get("id")]
        blocking = [task for task in related_tasks if task.get("status") in ("reserved", "active")]
        if blocking:
            raise ValueError("该批次已有已预留或已入驻的待进驻任务，请先处理相关任务后再调整房间")
        target_room = next((room for room in state.get("rooms", []) if clean_text(room.get("name", "")) == new_room_name), None) if new_room_name else None
        if related_tasks and new_room_name and not target_room:
            raise ValueError("房间尚未在系统中配置，请先选择已配置饲养间后再保存")
        for task in related_tasks:
            task["targetRoomId"] = target_room.get("id", "") if target_room else ""
            task["targetRoomName"] = target_room.get("name", "") if target_room else new_room_name
            task["updatedAt"] = next_item.get("updatedAt") or now_iso()
    return next_item


def replace_entity(state, collection, item_id, item):
    items = state.setdefault(collection, [])
    for index, existing in enumerate(items):
        if existing.get("id") == item_id:
            if collection == "intakeBatches":
                item = reconcile_intake_batch_update(state, existing, item)
            validate_entity_references(state, collection, item)
            items[index] = item
            return
    raise LookupError("实体不存在")


def delete_entity(state, collection, item_id):
    items = state.setdefault(collection, [])
    deleted = None
    kept = []
    for item in items:
        if item.get("id") == item_id:
            deleted = item
        else:
            kept.append(item)
    if deleted is None:
        raise LookupError("实体不存在")
    state[collection] = kept

    if collection == "rooms":
        rack_ids = {rack.get("id") for rack in state.get("racks", []) if rack.get("roomId") == item_id}
        state["racks"] = [rack for rack in state.get("racks", []) if rack.get("roomId") != item_id]
        state["slots"] = [slot for slot in state.get("slots", []) if slot.get("rackId") not in rack_ids]
    elif collection == "racks":
        state["slots"] = [slot for slot in state.get("slots", []) if slot.get("rackId") != item_id]
    elif collection == "placementTasks":
        reserved_occupancy_id = clean_text(deleted.get("reservedOccupancyId", ""))
        task_status = clean_text(deleted.get("status", ""))
        if task_status == "active":
            raise ValueError("已正式入驻的待进驻任务不能直接删除")
        if reserved_occupancy_id:
            state["occupancies"] = [item for item in state.get("occupancies", []) if item.get("id") != reserved_occupancy_id]
            sync_slot_statuses(state)

    return deleted


def validate_entity_references(state, collection, item):
    if collection == "racks" and not entity_exists(state, "rooms", item.get("roomId")):
        raise ValueError("关联的饲养间不存在")
    if collection == "slots" and not entity_exists(state, "racks", item.get("rackId")):
        raise ValueError("关联的笼架不存在")
    if collection == "occupancies" and not entity_exists(state, "slots", item.get("slotId")):
        raise ValueError("关联的笼位不存在")
    if collection == "placementTasks" and item.get("targetRoomId") and not entity_exists(state, "rooms", item.get("targetRoomId")):
        raise ValueError("关联的目标饲养间不存在")


def entity_exists(state, collection, item_id):
    return any(item.get("id") == item_id for item in state.get(collection, []))


def intake_service_deps():
    return {
        "as_int": as_int,
        "cage_card_qr_id": cage_card_qr_id_from_batch_card,
        "clean_text": clean_text,
        "new_id": new_id,
        "now_iso": now_iso,
    }


def placement_service_deps():
    return {
        "clean_text": clean_text,
        "new_id": new_id,
        "now_iso": now_iso,
    }


def quantity_service_deps():
    return {
        "as_int": as_int,
        "audit_event": audit_event,
        "clean_text": clean_text,
        "insert_quantity_sheet": insert_quantity_sheet_repository,
        "new_id": new_id,
        "normalize_iacuc_number": normalize_iacuc_number,
        "quantity_sheet_db_values": quantity_sheet_db_values,
        "read_applications_by_iacuc": read_applications_by_iacuc,
        "select_quantity_sheets_for_transfer": select_quantity_sheets_for_transfer_repository,
        "update_quantity_sheet": update_quantity_sheet_repository,
    }


def billing_workflow_service_deps():
    return {
        "VERSION_STATUS_ACTIVE": VERSION_STATUS_ACTIVE,
        "VERSION_STATUS_VOIDED": VERSION_STATUS_VOIDED,
        "WORKFLOW_STATUS_FINANCE": WORKFLOW_STATUS_FINANCE,
        "WORKFLOW_STATUS_GENERATED": WORKFLOW_STATUS_GENERATED,
        "WORKFLOW_STATUS_SENT": WORKFLOW_STATUS_SENT,
        "WORKFLOW_STATUS_SIGNED": WORKFLOW_STATUS_SIGNED,
        "billing_workflow_business_key": billing_workflow_business_key,
        "build_version_payload": build_version_payload,
        "build_workflow_event_payload": build_workflow_event_payload,
        "build_workflow_payload": build_workflow_payload,
        "enrich_statement_for_workflow": enrich_statement_for_workflow,
        "get_billing_version": get_billing_version,
        "get_billing_workflow": get_billing_workflow,
        "get_billing_workflow_by_key": get_billing_workflow_by_key,
        "insert_billing_version": insert_billing_version,
        "insert_billing_workflow": insert_billing_workflow,
        "insert_billing_workflow_event": insert_billing_workflow_event,
        "make_statement_document_number": make_statement_document_number,
        "new_id": new_id,
        "normalize_workflow_source": normalize_workflow_source,
        "now_iso": now_iso,
        "replace_version_lines": replace_version_lines,
        "update_billing_version": update_billing_version,
        "update_billing_workflow": update_billing_workflow,
        "workflow_scope_for_statement": workflow_scope_for_statement,
    }


def confirm_intake_receipt(state, batch_id, payload, actor):
    return confirm_intake_receipt_service(state, batch_id, payload, actor, intake_service_deps())


def reserve_placement_task(state, task_id, slot_id, actor):
    return reserve_placement_task_service(state, task_id, slot_id, actor, placement_service_deps())


def move_in_placement_task(state, task_id, actual_move_in_date, actor):
    return move_in_placement_task_service(state, task_id, actual_move_in_date, actor, placement_service_deps())


def reassign_placement_task_room(state, task_id, room_id, actor):
    return reassign_placement_task_room_service(state, task_id, room_id, actor, placement_service_deps())


def migrate_legacy_state(conn):
    if table_has_rows(conn, "rooms"):
        return
    row = conn.execute("SELECT payload, updated_at FROM app_state WHERE id = ?", ("default",)).fetchone()
    if not row:
        return
    write_normalized_state(conn, json.loads(row["payload"]), row["updated_at"])


def write_normalized_state(conn, state, updated_at):
    for table in TABLES:
        conn.execute(f"DELETE FROM {table}")

    applications_by_iacuc = read_applications_by_iacuc(conn)
    set_setting(conn, "baseRate", state.get("baseRate", 4.5), updated_at)
    set_setting(conn, "billingMonth", state.get("billingMonth", ""), updated_at)
    set_setting(conn, "billingIacuc", state.get("billingIacuc", ""), updated_at)

    for room in state.get("rooms", []):
        conn.execute(
            """
            INSERT INTO rooms (id, name, area, rack_count, rows, cols, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                room.get("id"),
                room.get("name", ""),
                room.get("area", ""),
                as_int(room.get("rackCount")),
                as_int(room.get("rows")),
                as_int(room.get("cols")),
                dump_json(room),
            ),
        )

    for rack in state.get("racks", []):
        conn.execute(
            """
            INSERT INTO racks (id, room_id, name, rows, cols, index_no, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rack.get("id"),
                rack.get("roomId"),
                rack.get("name", ""),
                as_int(rack.get("rows")),
                as_int(rack.get("cols")),
                as_int(rack.get("index")),
                dump_json(rack),
            ),
        )

    for slot in state.get("slots", []):
        conn.execute(
            """
            INSERT INTO cage_slots (id, rack_id, row_no, col_no, code, status, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                slot.get("id"),
                slot.get("rackId"),
                as_int(slot.get("row")),
                as_int(slot.get("col")),
                slot.get("code", ""),
                slot.get("status", "empty"),
                dump_json(slot),
            ),
        )

    for occupancy_item in state.get("occupancies", []):
        occupancy = occupancy_with_snapshots(occupancy_item, state, applications_by_iacuc)
        structured = occupancy_structured_values(occupancy)
        conn.execute(
            """
            INSERT INTO occupancies (
                id, slot_id, room_id, rack_id, cage_code, status, iacuc, project, pi, owner, funding,
                species, billing_item, customer_type, animal_count, room_name, rack_name, slot_code,
                start_date, end_date, end_reason, notes, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                occupancy.get("id"),
                occupancy.get("slotId"),
                structured["room_id"],
                structured["rack_id"],
                occupancy.get("cageCode", ""),
                occupancy.get("status", ""),
                occupancy.get("iacuc", ""),
                occupancy.get("project", ""),
                occupancy.get("pi", ""),
                occupancy.get("owner", ""),
                occupancy.get("funding", ""),
                structured["species"],
                structured["billing_item"],
                structured["customer_type"],
                structured["animal_count"],
                occupancy.get("roomName", ""),
                occupancy.get("rackName", ""),
                occupancy.get("slotCode", ""),
                occupancy.get("startDate", ""),
                occupancy.get("endDate", ""),
                occupancy.get("endReason", ""),
                occupancy.get("notes", ""),
                occupancy.get("updatedAt", ""),
                dump_json(occupancy),
            ),
        )

    for task in state.get("placementTasks", []):
        conn.execute(
            """
            INSERT INTO placement_tasks (
                id, source_batch_id, source_receipt_id, target_room_id, planned_move_in_date,
                status, reserved_occupancy_id, actual_move_in_date, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                task.get("id"),
                task.get("sourceBatchId", ""),
                task.get("sourceReceiptId", ""),
                task.get("targetRoomId", ""),
                task.get("plannedMoveInDate", ""),
                task.get("status", "pending"),
                task.get("reservedOccupancyId", ""),
                task.get("actualMoveInDate", ""),
                task.get("updatedAt", updated_at),
                dump_json(task),
            ),
        )

    for rule in state.get("billingRules", []):
        conn.execute(
            """
            INSERT INTO billing_rules (id, name, unit, price, effective_start, effective_end, payload)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rule.get("id"),
                rule.get("name", ""),
                rule.get("unit", ""),
                as_float(rule.get("price")),
                rule.get("effectiveStart", ""),
                rule.get("effectiveEnd", ""),
                dump_json(rule),
            ),
        )

    for adjustment in state.get("adjustments", []):
        conn.execute(
            """
            INSERT INTO billing_adjustments (
                id, target_type, target_id, adjustment_type, value, reason,
                effective_start, effective_end, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                adjustment.get("id"),
                adjustment.get("targetType", ""),
                adjustment.get("targetId", ""),
                adjustment.get("type", ""),
                as_float(adjustment.get("value")),
                adjustment.get("reason", ""),
                adjustment.get("effectiveStart", ""),
                adjustment.get("effectiveEnd", ""),
                dump_json(adjustment),
            ),
        )

    for batch in state.get("intakeBatches", []):
        conn.execute(
            """
            INSERT INTO intake_batches (
                id, batch_no, iacuc, supplier, room_name, intake_date, status, updated_at, payload
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch.get("id"),
                batch.get("batchNo", ""),
                batch.get("iacuc", ""),
                batch.get("supplier", ""),
                batch.get("roomName", ""),
                batch.get("intakeDate", ""),
                batch.get("status", "draft"),
                batch.get("updatedAt", updated_at),
                dump_json(batch),
            ),
        )

    for log in state.get("auditLogs", []):
        conn.execute(
            """
            INSERT INTO audit_logs (id, message, at, payload)
            VALUES (?, ?, ?, ?)
            """,
            (log.get("id"), log.get("message", ""), log.get("at", ""), dump_json(log)),
        )


def migrate_reimbursement_record_schema(conn):
    if read_setting(conn, REIMBURSEMENT_MIGRATION_KEY, False):
        return
    rows = conn.execute("SELECT id, payload FROM billing_workflows ORDER BY month, rowid").fetchall()
    for row in rows:
        workflow = json.loads(row["payload"])
        current_version = workflow.get("currentVersion") or {}
        statement = current_version.get("statement") or {}
        if not statement:
            continue
        lines = list_billing_statement_lines_for_version(conn, current_version.get("id", ""))
        detail_context = reimbursement_detail_context_from_workflow(conn, workflow, statement)
        upsert_reimbursement_record_from_statement(
            conn,
            workflow,
            statement,
            lines,
            detail_context,
            source="workflow" if workflow.get("sourceType") else "imported",
        )
    recalculate_all_reimbursement_accumulations(conn)
    set_setting(conn, REIMBURSEMENT_MIGRATION_KEY, True, now_iso())


def assemble_state(conn):
    return assemble_state_repository(conn)


def read_cached_state(conn):
    return read_cached_state_repository(conn, empty_state)


def read_applications_by_iacuc(conn):
    return read_applications_by_iacuc_repository(conn, normalize_iacuc_number)


def occupancy_with_snapshots(occupancy, state, applications_by_iacuc):
    item = dict(occupancy)
    iacuc = normalize_iacuc_number(item.get("iacuc", ""))
    if iacuc:
        item["iacuc"] = iacuc
    application = applications_by_iacuc.get(iacuc, {})
    for key in ("project", "pi", "owner", "funding"):
        if not item.get(key) and application.get(key):
            item[key] = application.get(key, "")

    slot_context = slot_snapshot_context(state, item.get("slotId"))
    for key, value in slot_context.items():
        if value and not item.get(key):
            item[key] = value
    profile = billing_profile_for_occupancy(item, state)
    item["roomId"] = item.get("roomId") or slot_context.get("roomId", "")
    item["rackId"] = item.get("rackId") or slot_context.get("rackId", "")
    item["species"] = item.get("species") or profile.get("species", "")
    item["billingItem"] = item.get("billingItem") or profile.get("billingItem", "")
    item["customerType"] = item.get("customerType") or profile.get("customerType", "")
    if item.get("animalCount") in (None, ""):
        item["animalCount"] = occupancy_animal_count(item, profile) if profile.get("unit") == "animal_day" else None
    return item


def slot_snapshot_context(state, slot_id):
    slot = next((item for item in state.get("slots", []) if item.get("id") == slot_id), None)
    if not slot:
        return {}
    rack = next((item for item in state.get("racks", []) if item.get("id") == slot.get("rackId")), None)
    room = next((item for item in state.get("rooms", []) if item.get("id") == (rack or {}).get("roomId")), None)
    return {
        "roomId": (room or {}).get("id", ""),
        "rackId": (rack or {}).get("id", ""),
        "roomName": (room or {}).get("name", ""),
        "rackName": (rack or {}).get("name", ""),
        "slotCode": slot.get("code", ""),
    }


def occupancy_structured_values(occupancy):
    return {
        "room_id": occupancy.get("roomId", ""),
        "rack_id": occupancy.get("rackId", ""),
        "species": occupancy.get("species", ""),
        "billing_item": occupancy.get("billingItem", ""),
        "customer_type": occupancy.get("customerType", ""),
        "animal_count": as_int(occupancy.get("animalCount")),
    }


def as_int(value):
    return int(value) if value not in (None, "") else None


def as_float(value):
    return float(value) if value not in (None, "") else None


def ensure_default_admin(conn):
    if has_any_user(conn):
        return
    now = now_iso()
    insert_user(
        conn,
        {
            "id": new_id("user"),
            "username": DEFAULT_ADMIN_USERNAME,
            "display_name": "系统管理员",
            "password_hash": hash_password(DEFAULT_ADMIN_PASSWORD),
            "role": "admin",
            "room_ids": "[]",
            "active": 1,
            "created_at": now,
            "updated_at": now,
        },
    )
    conn.commit()


def hash_password(password):
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return "pbkdf2_sha256$200000$" + base64.b64encode(salt).decode() + "$" + base64.b64encode(digest).decode()


def verify_password(password, password_hash):
    try:
        algorithm, iterations, salt_text, digest_text = password_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        salt = base64.b64decode(salt_text)
        expected = base64.b64decode(digest_text)
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(iterations))
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def create_session(conn, user_id):
    token = secrets.token_urlsafe(32)
    token_hash = hash_token(token)
    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(days=SESSION_TTL_DAYS)
    insert_session(conn, token_hash, user_id, now.isoformat(), expires_at.isoformat())
    conn.commit()
    return token, expires_at


def delete_session(conn, token):
    if not token:
        return
    delete_session_by_token_hash(conn, hash_token(token))
    conn.commit()


def hash_token(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def user_from_token(conn, token):
    if not token:
        return None
    return get_user_by_session_token_hash(conn, hash_token(token), now_iso(), sanitize_user)


def authenticate(conn, username, password):
    row = get_active_user_by_username(conn, username)
    if not row or not verify_password(password, row["password_hash"]):
        return None
    return sanitize_user(row)


def sanitize_user(row):
    if not row:
        return None
    return {
        "id": row["id"],
        "username": row["username"],
        "displayName": row["display_name"],
        "role": row["role"],
        "roomIds": json.loads(row["room_ids"] or "[]"),
        "active": bool(row["active"]),
    }


def list_users(conn):
    return list_users_repository(conn, sanitize_user)


def create_user(conn, payload):
    username = str(payload.get("username", "")).strip()
    password = str(payload.get("password", ""))
    display_name = str(payload.get("displayName", "")).strip() or username
    role = payload.get("role", "room_admin")
    room_ids = payload.get("roomIds", [])
    if not username or not password:
        raise ValueError("用户名和密码不能为空")
    if role not in ("admin", "room_admin"):
        raise ValueError("角色只能是 admin 或 room_admin")
    if not isinstance(room_ids, list):
        raise ValueError("roomIds 必须是数组")

    now = now_iso()
    user_id = new_id("user")
    insert_user(
        conn,
        {
            "id": user_id,
            "username": username,
            "display_name": display_name,
            "password_hash": hash_password(password),
            "role": role,
            "room_ids": dump_json(room_ids),
            "active": 1,
            "created_at": now,
            "updated_at": now,
        },
    )
    conn.commit()
    row = get_user_by_id(conn, user_id)
    return sanitize_user(row)


def update_user(conn, actor, user_id, payload):
    if user_id == actor["id"]:
        raise PermissionError("不能在账号管理中修改当前登录账号")

    row = get_user_by_id(conn, user_id)
    if not row:
        raise LookupError("账号不存在")

    username = str(payload.get("username", row["username"])).strip()
    display_name = str(payload.get("displayName", row["display_name"])).strip() or username
    password = str(payload.get("password", ""))
    role = payload.get("role", row["role"])
    room_ids = payload.get("roomIds", json.loads(row["room_ids"] or "[]"))

    if not username:
        raise ValueError("用户名不能为空")
    if role not in ("admin", "room_admin"):
        raise ValueError("角色只能是 admin 或 room_admin")
    if not isinstance(room_ids, list):
        raise ValueError("roomIds 必须是数组")

    now = now_iso()
    if password:
        update_user_with_password(conn, user_id, username, display_name, hash_password(password), role, dump_json(room_ids), now)
    else:
        update_user_without_password(conn, user_id, username, display_name, role, dump_json(room_ids), now)
    delete_sessions_by_user_id(conn, user_id)
    conn.commit()
    row = get_user_by_id(conn, user_id)
    return sanitize_user(row)


def delete_user(conn, actor, user_id):
    if user_id == actor["id"]:
        raise PermissionError("不能删除当前登录账号")

    row = get_user_by_id(conn, user_id)
    if not row:
        raise LookupError("账号不存在")

    delete_sessions_by_user_id(conn, user_id)
    delete_user_by_id(conn, user_id)
    conn.commit()


def validate_state_write_permission(conn, actor, old_state, new_state):
    if actor["role"] == "admin":
        return

    allowed_rooms = set(actor.get("roomIds", []))
    if not allowed_rooms:
        raise PermissionError("当前账号没有可编辑的饲养间")

    old_rooms = {item.get("id"): item for item in old_state.get("rooms", [])}
    new_rooms = {item.get("id"): item for item in new_state.get("rooms", [])}
    if set(old_rooms) != set(new_rooms):
        raise PermissionError("房间管理员不能新增或删除饲养间")
    for room_id in changed_keys(old_rooms, new_rooms):
        if room_id not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间配置")
        old_room = old_rooms.get(room_id, {})
        new_room = new_rooms.get(room_id, {})
        allowed_room_update_keys = {"rackCount"}
        changed_fields = {key for key in set(old_room) | set(new_room) if old_room.get(key) != new_room.get(key)}
        if not changed_fields.issubset(allowed_room_update_keys):
            raise PermissionError("房间管理员不能修改饲养间基础信息")

    old_racks = {item.get("id"): item for item in old_state.get("racks", [])}
    new_racks = {item.get("id"): item for item in new_state.get("racks", [])}
    rack_rooms = rack_room_map(old_state, new_state)
    for rack_id in changed_keys(old_racks, new_racks):
        if rack_rooms.get(rack_id) not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼架配置")
    if comparable_items(old_state.get("billingRules", [])) != comparable_items(new_state.get("billingRules", [])):
        raise PermissionError("房间管理员不能修改计费规则")
    if old_state.get("baseRate") != new_state.get("baseRate"):
        raise PermissionError("房间管理员不能修改计费规则")
    if comparable_items(old_state.get("adjustments", [])) != comparable_items(new_state.get("adjustments", [])):
        raise PermissionError("房间管理员不能修改减免规则")

    old_slots = {item.get("id"): item for item in old_state.get("slots", [])}
    new_slots = {item.get("id"): item for item in new_state.get("slots", [])}
    slot_rooms = slot_room_map(new_state)
    old_slot_rooms = slot_room_map(old_state)
    for slot_id in changed_keys(old_slots, new_slots):
        room_id = slot_rooms.get(slot_id) or old_slot_rooms.get(slot_id)
        if room_id not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼位结构")
    for slot_id, new_slot in new_slots.items():
        old_slot = dict(old_slots.get(slot_id, {}))
        if not old_slot:
            continue
        comparable_old = {k: v for k, v in old_slot.items() if k != "status"}
        comparable_new = {k: v for k, v in new_slot.items() if k != "status"}
        if comparable_old != comparable_new and slot_rooms.get(slot_id) not in allowed_rooms:
            raise PermissionError("房间管理员不能修改笼位结构")
        if old_slot.get("status") != new_slot.get("status") and slot_rooms.get(slot_id) not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼位状态")

    old_occupancies = {item.get("id"): item for item in old_state.get("occupancies", [])}
    new_occupancies = {item.get("id"): item for item in new_state.get("occupancies", [])}
    changed_ids = changed_keys(old_occupancies, new_occupancies)
    for occupancy_id in changed_ids:
        item = new_occupancies.get(occupancy_id) or old_occupancies.get(occupancy_id) or {}
        if slot_rooms.get(item.get("slotId")) not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的笼位信息")

    old_tasks = {item.get("id"): item for item in old_state.get("placementTasks", [])}
    new_tasks = {item.get("id"): item for item in new_state.get("placementTasks", [])}
    for task_id in changed_keys(old_tasks, new_tasks):
        item = new_tasks.get(task_id) or old_tasks.get(task_id) or {}
        if item.get("targetRoomId") not in allowed_rooms:
            raise PermissionError("不能修改未授权饲养间的待进驻任务")


def comparable_items(items):
    return {item.get("id"): item for item in items}


def changed_keys(old_items, new_items):
    keys = set(old_items) | set(new_items)
    return {key for key in keys if old_items.get(key) != new_items.get(key)}


def slot_room_map(state):
    rack_rooms = {rack.get("id"): rack.get("roomId") for rack in state.get("racks", [])}
    return {slot.get("id"): rack_rooms.get(slot.get("rackId")) for slot in state.get("slots", [])}


def rack_room_map(*states):
    rooms = {}
    for state in states:
        for rack in state.get("racks", []):
            rooms[rack.get("id")] = rack.get("roomId")
    return rooms


def build_audit_events(actor, old_state, new_state, at):
    events = []
    slot_labels = slot_label_map(new_state)
    old_occupancies = {item.get("id"): item for item in old_state.get("occupancies", [])}
    new_occupancies = {item.get("id"): item for item in new_state.get("occupancies", [])}
    for occupancy_id in sorted(changed_keys(old_occupancies, new_occupancies)):
        old_item = old_occupancies.get(occupancy_id)
        new_item = new_occupancies.get(occupancy_id)
        item = new_item or old_item or {}
        slot_id = item.get("slotId", "")
        label = slot_labels.get(slot_id, slot_id)
        if old_item is None:
            action = "occupancy.created"
            message = f"{actor['displayName']} 新增笼位 {label} 的占用记录"
        elif new_item is None:
            action = "occupancy.deleted"
            message = f"{actor['displayName']} 删除笼位 {label} 的占用记录"
        elif new_item.get("status") == "ended" and new_item.get("endReason") == "sampled" and old_item.get("status") != "ended":
            action = "occupancy.sampled"
            message = f"{actor['displayName']} 将笼位 {label} 标记为已取材，最后计费日期 {new_item.get('endDate', '')}"
        elif new_item.get("status") == "ended" and new_item.get("endReason") == "cleared" and old_item.get("status") != "ended":
            action = "occupancy.cleared"
            message = f"{actor['displayName']} 将笼位 {label} 设为空"
        else:
            action = "occupancy.updated"
            message = f"{actor['displayName']} 更新笼位 {label} 的占用信息"
        events.append(audit_event(actor, action, "occupancy", occupancy_id, message, [slot_id], at, old_item, new_item))

    old_rooms = {item.get("id"): item for item in old_state.get("rooms", [])}
    new_rooms = {item.get("id"): item for item in new_state.get("rooms", [])}
    for room_id in sorted(changed_keys(old_rooms, new_rooms)):
        old_item = old_rooms.get(room_id)
        new_item = new_rooms.get(room_id)
        action = "room.updated"
        if old_item is None:
            action = "room.created"
        elif new_item is None:
            action = "room.deleted"
        name = (new_item or old_item or {}).get("name", room_id)
        message = f"{actor['displayName']} {action_label(action)}饲养间 {name}"
        events.append(audit_event(actor, action, "room", room_id, message, [], at, old_item, new_item))

    old_batches = {item.get("id"): item for item in old_state.get("intakeBatches", [])}
    new_batches = {item.get("id"): item for item in new_state.get("intakeBatches", [])}
    for batch_id in sorted(changed_keys(old_batches, new_batches)):
        old_item = old_batches.get(batch_id)
        new_item = new_batches.get(batch_id)
        action = "intake_batch.updated"
        if old_item is None:
            action = "intake_batch.created"
        elif new_item is None:
            action = "intake_batch.deleted"
        label = (new_item or old_item or {}).get("batchNo", batch_id)
        message = f"{actor['displayName']} {action_label(action)}待接收批次 {label}"
        events.append(audit_event(actor, action, "intake_batch", batch_id, message, [], at, old_item, new_item))

    old_tasks = {item.get("id"): item for item in old_state.get("placementTasks", [])}
    new_tasks = {item.get("id"): item for item in new_state.get("placementTasks", [])}
    for task_id in sorted(changed_keys(old_tasks, new_tasks)):
        old_item = old_tasks.get(task_id)
        new_item = new_tasks.get(task_id)
        action = "placement_task.updated"
        if old_item is None:
            action = "placement_task.created"
        elif new_item is None:
            action = "placement_task.deleted"
        label = (new_item or old_item or {}).get("batchNo", task_id)
        message = f"{actor['displayName']} {action_label(action)}待进驻任务 {label}"
        events.append(audit_event(actor, action, "placement_task", task_id, message, [], at, old_item, new_item))

    return events[:100]


def action_label(action):
    if action.endswith(".created"):
        return "新增"
    if action.endswith(".deleted"):
        return "删除"
    return "更新"


def audit_event(actor, action, entity_type, entity_id, message, slot_ids, at, before, after):
    return {
        "id": new_id("audit"),
        "actorUserId": actor["id"],
        "actorUsername": actor["username"],
        "actorDisplayName": actor["displayName"],
        "action": action,
        "entityType": entity_type,
        "entityId": entity_id,
        "message": message,
        "slotIds": slot_ids,
        "at": at,
        "before": before,
        "after": after,
    }


def write_audit_events(conn, events):
    insert_audit_events(conn, events)
    if events:
        invalidate_data_cache_prefixes("audit_events::")


def merge_audit_logs(client_logs, events):
    normalized_events = [
        {
            "id": event["id"],
            "message": event["message"],
            "at": event["at"],
            "actorUsername": event["actorUsername"],
            "actorDisplayName": event["actorDisplayName"],
            "action": event["action"],
            "slotIds": event["slotIds"],
        }
        for event in events
    ]
    seen = set()
    merged = []
    for item in normalized_events + list(client_logs):
        item_id = item.get("id")
        if item_id in seen:
            continue
        seen.add(item_id)
        merged.append(item)
    return merged[:500]


def write_perf_summary(started_at, rows_changed=0, **fields):
    return {
        "total_ms": round((time.perf_counter() - started_at) * 1000, 1),
        "rows_changed": rows_changed,
        **{key: value for key, value in fields.items() if value not in (None, "")},
    }


def slot_label_map(state):
    rack_by_id = {rack.get("id"): rack for rack in state.get("racks", [])}
    room_by_id = {room.get("id"): room for room in state.get("rooms", [])}
    labels = {}
    for slot in state.get("slots", []):
        rack = rack_by_id.get(slot.get("rackId"), {})
        room = room_by_id.get(rack.get("roomId"), {})
        rack_index = rack.get("index", "")
        rack_code = str(rack_index).zfill(2) if str(rack_index).isdigit() else str(rack_index)
        labels[slot.get("id")] = f"{room.get('name', '')}-{rack_code}-{slot.get('code', '')}".strip("-")
    return labels


def new_id(prefix):
    return f"{prefix}-{secrets.token_hex(8)}"


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def today_iso():
    return datetime.now().date().isoformat()


def format_http_date(value):
    return format_datetime(value, usegmt=True)


def read_iacuc_index():
    with connect_db() as conn:
        return read_iacuc_index_repository(conn, IACUC_INDEX_PATH, LEGACY_IACUC_INDEX_PATH)


def write_experiment_applications(conn, items, imported_at):
    replace_experiment_applications(conn, items, imported_at, application_payload)


def list_principal_identities(conn):
    identity_by_pi = {
        clean_text(item.get("pi", "")): item
        for item in read_principal_identity_payloads(conn)
        if clean_text(item.get("pi", ""))
    }
    principal_names = {clean_text(name) for name in list_distinct_principal_names(conn)}
    principal_names.update(identity_by_pi.keys())
    items = []
    for pi_name in sorted((name for name in principal_names if name), key=lambda value: value.lower()):
        saved = identity_by_pi.get(pi_name, {})
        principal_type = normalize_principal_type(saved.get("principalType", BILLING_PRINCIPAL_INDEPENDENT))
        items.append(
            {
                "pi": pi_name,
                "principalType": principal_type,
                "freeCageAllowance": free_cages_for_principal_type(principal_type),
                "updatedAt": saved.get("updatedAt", ""),
            }
        )
    return items


def read_principal_identities():
    cached = cache_get("principal_identities")
    if cached is not None:
        return cached
    with connect_db() as conn:
        return cache_set("principal_identities", list_principal_identities(conn))


def save_principal_identity(conn, payload, actor, pi_name):
    pi_name = clean_text(pi_name or payload.get("pi", ""))
    if not pi_name:
        raise ValueError("项目负责人不能为空")
    principal_type = normalize_principal_type(payload.get("principalType", ""))
    now = now_iso()
    item = {
        "pi": pi_name,
        "principalType": principal_type,
        "freeCageAllowance": free_cages_for_principal_type(principal_type),
        "updatedAt": now,
    }
    upsert_principal_identity(conn, pi_name, principal_type, now, dump_json(item))
    event = audit_event(
        actor,
        "principal_identity.updated",
        "principal_identity",
        pi_name,
        f"{actor['displayName']} 更新项目负责人 {pi_name} 身份为 {principal_type_label(principal_type)}",
        [],
        now,
        None,
        item,
    )
    write_audit_events(conn, [event])
    invalidate_data_cache("principal_identities")
    invalidate_data_cache_prefixes("quantity_sheets::", "billing_workflows::")
    return item, merge_audit_logs([], [event])


def application_payload(item, imported_at):
    raw_iacuc = clean_text(item.get("rawIacuc", "") or item.get("iacuc", ""))
    normalized = {
        "id": clean_text(item.get("id", "")),
        "iacuc": clean_text(item.get("iacuc", "") or raw_iacuc),
        "rawIacuc": raw_iacuc,
        "project": clean_text(item.get("project", "")),
        "pi": clean_text(item.get("pi", "")),
        "owner": clean_text(item.get("owner", "")),
        "funding": clean_text(item.get("funding", "")),
        "importedAt": imported_at,
    }
    return normalized


def save_iacuc_index_file(items):
    save_iacuc_index_file_repository(IACUC_INDEX_PATH, items)


def system_update_status():
    current = current_revision()
    current_version = normalize_release_version(app_version())
    if not CAGELEDGER_UPDATE_CHECK_ENABLED:
        return {
            "repository": CAGELEDGER_REPOSITORY_URL,
            "repositoryUrl": CAGELEDGER_REPOSITORY_URL,
            "branch": CAGELEDGER_BRANCH,
            "appVersion": app_version(),
            "current": current or None,
            "currentShort": short_revision(current),
            "currentVersion": current_version or None,
            "latest": None,
            "latestShort": "",
            "latestVersion": None,
            "latestUrl": None,
            "latestMessage": None,
            "latestDate": None,
            "updateAvailable": None,
            "checkedAt": now_iso(),
            "disabled": True,
        }

    latest = latest_remote_release()
    latest_sha = latest.get("sha") or ""
    latest_version = normalize_release_version(latest.get("version"))
    update_available = None
    if current_version and latest_version:
        update_available = compare_release_versions(current_version, latest_version) < 0

    return {
        "repository": CAGELEDGER_REPOSITORY_URL,
        "repositoryUrl": CAGELEDGER_REPOSITORY_URL,
        "branch": CAGELEDGER_BRANCH,
        "appVersion": app_version(),
        "current": current or None,
        "currentShort": short_revision(current),
        "currentVersion": current_version or None,
        "latest": latest_sha or None,
        "latestShort": short_revision(latest_sha),
        "latestVersion": latest_version or None,
        "latestUrl": latest.get("url"),
        "latestMessage": latest.get("message"),
        "latestDate": latest.get("date"),
        "updateAvailable": update_available,
        "checkedAt": now_iso(),
    }


def system_info():
    return {
        "name": "CageLedger",
        "title": "CageLedger 实验动物笼位管理与计费系统",
        "description": "实验动物笼位管理与计费系统",
        "version": app_version(),
        "organization": CAGELEDGER_ORGANIZATION,
        "department": CAGELEDGER_DEPARTMENT,
        "developer": CAGELEDGER_DEVELOPER,
        "contactEmail": CAGELEDGER_CONTACT_EMAIL,
        "license": CAGELEDGER_LICENSE,
        "copyright": CAGELEDGER_COPYRIGHT,
        "repository": CAGELEDGER_REPOSITORY_URL,
        "repositoryUrl": CAGELEDGER_REPOSITORY_URL,
        "branch": CAGELEDGER_BRANCH,
        "revision": current_revision() or None,
        "revisionShort": short_revision(current_revision()),
    }


def app_version():
    if CAGELEDGER_APP_VERSION:
        return CAGELEDGER_APP_VERSION
    package_path = ROOT / "package.json"
    try:
        return json.loads(package_path.read_text(encoding="utf-8")).get("version", "")
    except (OSError, json.JSONDecodeError):
        return ""


def current_revision():
    if CAGELEDGER_VERSION:
        return CAGELEDGER_VERSION
    return read_git_revision(ROOT)


def read_git_revision(root):
    git_dir = root / ".git"
    if git_dir.is_file():
        content = git_dir.read_text(encoding="utf-8", errors="replace").strip()
        if content.startswith("gitdir:"):
            git_dir = (git_dir.parent / content.split(":", 1)[1].strip()).resolve()
    if not git_dir.exists():
        return ""

    head_path = git_dir / "HEAD"
    if not head_path.exists():
        return ""
    head = head_path.read_text(encoding="utf-8", errors="replace").strip()
    if head.startswith("ref:"):
        ref = head.split(":", 1)[1].strip()
        ref_path = git_dir / ref
        if ref_path.exists():
            return ref_path.read_text(encoding="utf-8", errors="replace").strip()
        packed_refs = git_dir / "packed-refs"
        if packed_refs.exists():
            for line in packed_refs.read_text(encoding="utf-8", errors="replace").splitlines():
                if not line or line.startswith("#") or line.startswith("^"):
                    continue
                sha, _, packed_ref = line.partition(" ")
                if packed_ref == ref:
                    return sha.strip()
        return ""
    return head


def latest_remote_release():
    repository = parse_gitea_repository_url(CAGELEDGER_REPOSITORY_URL)
    url = f"{repository['baseUrl']}/api/v1/repos/{repository['owner']}/{repository['repo']}/releases/latest"
    headers = {"Accept": "application/json", "User-Agent": "CageLedger"}
    if CAGELEDGER_GITEA_TOKEN:
        headers["Authorization"] = f"token {CAGELEDGER_GITEA_TOKEN}"
    request = Request(url, headers=headers)
    try:
        with urlopen(request, timeout=8) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        if exc.code in (HTTPStatus.UNAUTHORIZED, HTTPStatus.FORBIDDEN, HTTPStatus.NOT_FOUND):
            raise ValueError("Gitea 更新检查失败：请确认仓库地址正确，并为私有仓库配置只读 token") from exc
        raise ValueError(f"Gitea 返回错误：HTTP {exc.code}") from exc
    except URLError as exc:
        raise ValueError(f"无法连接 Gitea：{exc.reason}") from exc
    except TimeoutError as exc:
        raise ValueError("连接 Gitea 超时") from exc

    tag_name = str(payload.get("tag_name") or "").strip()
    if not tag_name:
        raise ValueError("Gitea 未返回可用发布版本")

    return {
        "sha": payload.get("target_commitish", ""),
        "version": tag_name,
        "url": payload.get("html_url", ""),
        "message": first_line(payload.get("name", "") or payload.get("body", "")),
        "date": payload.get("published_at") or payload.get("created_at") or "",
    }


def parse_gitea_repository_url(value):
    parsed = urlparse(str(value or "").strip())
    path_parts = [part for part in parsed.path.strip("/").split("/") if part]
    if parsed.scheme not in ("http", "https") or not parsed.netloc or len(path_parts) < 2:
        raise ValueError("项目仓库地址无效")

    owner = path_parts[-2]
    repo = path_parts[-1]
    if repo.endswith(".git"):
        repo = repo[:-4]
    if not owner or not repo:
        raise ValueError("项目仓库地址无效")

    return {
        "baseUrl": f"{parsed.scheme}://{parsed.netloc}",
        "owner": owner,
        "repo": repo,
    }


def revisions_match(current, latest):
    current = str(current or "").strip()
    latest = str(latest or "").strip()
    return bool(current and latest and (current.startswith(latest) or latest.startswith(current)))


def normalize_release_version(value):
    return str(value or "").strip().removeprefix("v")


def release_version_key(value):
    version = normalize_release_version(value)
    match = re.fullmatch(r"(\d+)\.(\d+)\.(\d+)([A-Za-z]*)(\d*)", version)
    if not match:
        return None
    major, minor, patch, suffix, suffix_number = match.groups()
    return (
        int(major),
        int(minor),
        int(patch),
        suffix or "",
        int(suffix_number) if suffix_number else -1,
    )


def compare_release_versions(left, right):
    left_key = release_version_key(left)
    right_key = release_version_key(right)
    if left_key is None or right_key is None:
        return (normalize_release_version(left) > normalize_release_version(right)) - (
            normalize_release_version(left) < normalize_release_version(right)
        )
    return (left_key > right_key) - (left_key < right_key)


def short_revision(value):
    value = str(value or "").strip()
    return value[:7] if value else None


def first_line(value):
    return str(value or "").splitlines()[0] if value else ""


def parse_iacuc_csv(raw):
    text = decode_csv_bytes(raw)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV 文件缺少表头")

    field_by_name = {clean_text(name): name for name in reader.fieldnames if clean_text(name)}
    required = {
        "iacuc": "动物伦理编号",
        "project": "动物实验名称",
        "pi": "项目负责人",
        "owner": "实验负责人",
    }
    missing = [label for label in required.values() if label not in field_by_name]
    if missing:
        raise ValueError(f"CSV 缺少必要列：{', '.join(missing)}")
    funding_field = next(
        (
            field_by_name[label]
            for label in ("项目来源", "支撑经费", "经费来源", "课题来源", "经费项目")
            if label in field_by_name
        ),
        None,
    )

    items = []
    duplicate_count = 0
    row_count = 0
    empty_iacuc_count = 0
    seen_iacucs = set()
    for row in reader:
        row_count += 1
        raw_iacuc = clean_text(row.get(field_by_name[required["iacuc"]], ""))
        iacuc = normalize_iacuc_number(raw_iacuc)
        if not iacuc:
            empty_iacuc_count += 1
            continue
        if iacuc in seen_iacucs:
            duplicate_count += 1
        seen_iacucs.add(iacuc)
        items.append({
            "id": f"app-{row_count:06d}",
            "iacuc": raw_iacuc,
            "rawIacuc": raw_iacuc,
            "project": clean_text(row.get(field_by_name[required["project"]], "")),
            "pi": clean_text(row.get(field_by_name[required["pi"]], "")),
            "owner": clean_text(row.get(field_by_name[required["owner"]], "")),
            "funding": clean_text(row.get(funding_field, "")) if funding_field else "",
        })

    return {
        "items": items,
        "summary": {
            "rowCount": row_count,
            "count": len(items),
            "emptyIacucCount": empty_iacuc_count,
            "duplicateCount": duplicate_count,
        },
    }


def decode_csv_bytes(raw):
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV 编码无法识别，请使用 UTF-8 或 GB18030")


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def bounded_int(value, default, min_value, max_value):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return max(min_value, min(number, max_value))


def filtered_where(filter_specs, filters):
    where = []
    params = []
    for key, expression in filter_specs:
        value = clean_text(filters.get(key, ""))
        if not value:
            continue
        where.append(expression)
        params.append(value)
    return " AND ".join(where), tuple(params)


def normalize_iacuc_number(value):
    text = clean_text(value)
    text = re.sub(r"（.*?）", "", text)
    text = re.sub(r"\(.*?\)", "", text)
    return text.strip()


def is_valid_iacuc_number(value):
    return bool(re.search(r"\d", value))


def list_quantity_sheets(conn):
    return list_quantity_sheets_repository(conn)


def list_quantity_sheets_by_month_iacuc(conn, month, iacuc):
    return list_quantity_sheets_by_month_iacuc_repository(conn, month, iacuc)


def list_quantity_sheets_by_month_pi(conn, month, pi):
    return list_quantity_sheets_by_month_pi_repository(conn, month, pi)


def list_quantity_sheets_page(conn, filters):
    return list_quantity_sheets_page_repository(conn, filters, filtered_where)


def get_quantity_sheet(conn, sheet_id):
    row = get_quantity_sheet_repository(conn, sheet_id)
    if not row:
        raise LookupError("数量统计表不存在")
    return row


def get_current_billing_statement(conn, statement_id):
    return get_current_billing_statement_repository(conn, clean_text(statement_id))


def save_quantity_sheet(conn, payload, actor, sheet_id=None):
    started_at = time.perf_counter()
    now = now_iso()
    sheet = normalize_quantity_sheet(payload, sheet_id, now)
    validate_quantity_sheet_permission(actor, sheet)
    validate_quantity_sheet_animal_requirements(conn, sheet)
    exists = conn.execute("SELECT 1 FROM quantity_sheets WHERE id = ?", (sheet["id"],)).fetchone()
    db_values = quantity_sheet_db_values(sheet)
    if exists:
        update_quantity_sheet_repository(conn, sheet, db_values)
        action = "quantity_sheet.updated"
        message = f"{actor['displayName']} 更新 {sheet['iacuc']} {sheet['month']} 数量统计表"
        status = HTTPStatus.OK
    else:
        insert_quantity_sheet_repository(conn, sheet, db_values)
        action = "quantity_sheet.created"
        message = f"{actor['displayName']} 创建 {sheet['iacuc']} {sheet['month']} 数量统计表"
        status = HTTPStatus.CREATED

    transfer_events, affected_sheets = sync_quantity_sheet_transfer_rows(conn, sheet, actor, now)
    event = audit_event(actor, action, "quantity_sheet", sheet["id"], message, [], now, None, sheet)
    events = [event, *transfer_events]
    write_audit_events(conn, events)
    log_perf(
        "quantity_sheet.save",
        started_at,
        sheet_id=sheet["id"],
        affected=len(affected_sheets),
        rows=len(sheet.get("rows", [])),
    )
    perf = write_perf_summary(started_at, rows_changed=1 + len(affected_sheets), affected=len(affected_sheets), rows=len(sheet.get("rows", [])))
    return sheet, affected_sheets, merge_audit_logs([], events), status, perf


def delete_quantity_sheet(conn, actor, sheet_id):
    sheet = get_quantity_sheet(conn, sheet_id)
    validate_quantity_sheet_permission(actor, sheet)
    now = now_iso()
    delete_quantity_sheet_by_id_repository(conn, sheet_id)
    event = audit_event(
        actor,
        "quantity_sheet.deleted",
        "quantity_sheet",
        sheet_id,
        f"{actor['displayName']} 删除 {sheet.get('iacuc', '')} {sheet.get('month', '')} 数量统计表",
        [],
        now,
        sheet,
        None,
    )
    write_audit_events(conn, [event])
    return merge_audit_logs([], [event])


def quantity_sheet_db_values(sheet):
    return (
        sheet["month"],
        sheet["iacuc"],
        sheet.get("roomId", ""),
        sheet.get("roomName", ""),
        sheet.get("manager", ""),
        sheet.get("project", ""),
        sheet.get("pi", ""),
        sheet.get("owner", ""),
        sheet.get("funding", ""),
        sheet["updatedAt"],
        dump_json(sheet),
    )


def sync_quantity_sheet_transfer_rows(conn, source_sheet, actor, now):
    return sync_quantity_sheet_transfer_rows_service(conn, source_sheet, actor, now, quantity_service_deps())


def normalize_quantity_sheet(payload, sheet_id, updated_at):
    source = payload.get("sheet") if isinstance(payload, dict) and isinstance(payload.get("sheet"), dict) else payload
    if not isinstance(source, dict):
        raise ValueError("数量统计表必须是 JSON 对象")

    month = clean_text(source.get("month", ""))
    iacuc = clean_text(source.get("iacuc", ""))
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("结算月份格式应为 YYYY-MM")
    if not normalize_iacuc_number(iacuc):
        raise ValueError("IACUC 编号不能为空")

    rows = source.get("rows", [])
    if not isinstance(rows, list):
        raise ValueError("统计表明细必须是数组")

    sheet = {
        "id": clean_text(sheet_id or source.get("id") or new_id("qsheet")),
        "month": month,
        "roomId": clean_text(source.get("roomId", "")),
        "roomName": clean_text(source.get("roomName", "")),
        "manager": clean_text(source.get("manager", "")),
        "iacuc": iacuc,
        "project": clean_text(source.get("project", "")),
        "pi": clean_text(source.get("pi", "")),
        "owner": clean_text(source.get("owner", "")),
        "contact": clean_text(source.get("contact", "")),
        "funding": clean_text(source.get("funding", "")),
        "billingUnit": "animal_day" if clean_text(source.get("billingUnit", "")) == "animal_day" else "cage_day",
        "animalDetailEnabled": parse_bool(source.get("animalDetailEnabled")),
        "initialAnimalCount": as_int(source.get("initialAnimalCount")),
        "initialCageCount": as_int(source.get("initialCageCount")),
        "pageCount": max(as_int(source.get("pageCount")) or 1, 1),
        "rows": [normalize_quantity_sheet_row(row, month) for row in rows],
        "updatedAt": updated_at,
    }
    sheet["rows"] = sorted(sheet["rows"], key=lambda item: (item["date"], item["id"]))
    return sheet


def parse_bool(value):
    if isinstance(value, bool):
        return value
    return clean_text(value).lower() in ("1", "true", "yes", "on")


def read_room_payload(conn, room_id):
    room_key = clean_text(room_id)
    if not room_key:
        return None
    row = conn.execute("SELECT payload FROM rooms WHERE id = ?", (room_key,)).fetchone()
    return json.loads(row["payload"]) if row else None


def read_room_payloads_for_context(conn, room_ids=None, room_names=None):
    ids = [clean_text(item) for item in (room_ids or []) if clean_text(item)]
    names = [clean_text(item) for item in (room_names or []) if clean_text(item)]
    clauses = []
    params = []
    if ids:
        clauses.append(f"id IN ({placeholders(ids)})")
        params.extend(ids)
    if names:
        clauses.append(f"name IN ({placeholders(names)})")
        params.extend(names)
    if not clauses:
        return []
    rows = conn.execute(f"SELECT payload FROM rooms WHERE {' OR '.join(clauses)} ORDER BY rowid", tuple(params)).fetchall()
    return [json.loads(row["payload"]) for row in rows]


def validate_quantity_sheet_animal_requirements(conn, sheet):
    room = read_room_payload(conn, sheet.get("roomId"))
    profile = billing_profile_for_room(room, sheet.get("billingUnit"))
    if profile["unit"] != "animal_day":
        return
    has_animal_balance = any((row.get("animalCount") or 0) > 0 for row in sheet.get("rows", []))
    if not has_animal_balance:
        raise ValueError("该房间按只/天计费，请打开动物数量并补充结余总数")


def read_rooms_for_quantity_sheets(conn, sheets):
    return read_room_payloads_for_context(
        conn,
        room_ids=[sheet.get("roomId", "") for sheet in sheets],
        room_names=[sheet.get("roomName", "") for sheet in sheets],
    )


def normalize_quantity_sheet_row(row, month):
    if not isinstance(row, dict):
        raise ValueError("统计表明细行必须是 JSON 对象")
    date = normalize_sheet_date(row.get("date", ""), month)
    return {
        "id": clean_text(row.get("id", "")) or new_id("qrow"),
        "date": date,
        "addedCount": as_int(row.get("addedCount")),
        "addedType": clean_text(row.get("addedType", "")),
        "removedCount": as_int(row.get("removedCount")),
        "removedType": clean_text(row.get("removedType", "")),
        "transferInFromIacuc": normalize_iacuc_number(row.get("transferInFromIacuc", "")),
        "transferOutToIacuc": normalize_iacuc_number(row.get("transferOutToIacuc", "")),
        "animalCount": as_int(row.get("animalCount")),
        "cageCount": as_int(row.get("cageCount")),
        "handler": clean_text(row.get("handler", "")),
        "balanceSource": "manual" if clean_text(row.get("balanceSource", "")) == "manual" else "auto",
        "notes": clean_text(row.get("notes", "")),
        "transferSourceSheetId": clean_text(row.get("transferSourceSheetId", "")),
        "transferSourceIacuc": normalize_iacuc_number(row.get("transferSourceIacuc", "")),
        "transferMirrorContrib": {
            clean_text(key): max(as_int(value) or 0, 0)
            for key, value in (row.get("transferMirrorContrib") or {}).items()
            if clean_text(key)
        }
        if isinstance(row.get("transferMirrorContrib"), dict)
        else {},
    }


def normalize_sheet_date(value, month):
    text = clean_text(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        date = text
    elif re.fullmatch(r"\d{1,2}[./-]\d{1,2}", text):
        year = month.split("-", 1)[0]
        month_no, day = [int(part) for part in re.split(r"[./-]", text)]
        date = f"{int(year):04d}-{month_no:02d}-{day:02d}"
    elif re.fullmatch(r"\d{1,2}", text):
        date = f"{month}-{int(text):02d}"
    else:
        raise ValueError("统计表日期格式应为 YYYY-MM-DD、M.D 或当月日期")
    if not date.startswith(month + "-"):
        raise ValueError("统计表明细日期必须属于结算月份")
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("统计表明细日期无效") from exc
    return date


def validate_quantity_sheet_permission(actor, sheet):
    if actor:
        return
    raise PermissionError("请先登录")


def generate_quantity_sheet_statement(conn, sheet_id, payload, actor):
    sheet = get_quantity_sheet(conn, sheet_id)
    validate_quantity_sheet_permission(actor, sheet)
    status = clean_text(payload.get("status", "draft")) or "draft"
    persist = bool(payload.get("persist"))
    if status not in ("draft", "locked"):
        raise ValueError("结算单状态只能是 draft 或 locked")

    sheet_iacuc = normalize_iacuc_number(sheet.get("iacuc", ""))
    if not sheet_iacuc:
        raise ValueError("数量统计表缺少伦理号，无法生成按伦理号拆分的结算单")
    sheets = list_quantity_sheets_by_month_iacuc(conn, sheet["month"], sheet_iacuc)
    for item in sheets:
        validate_quantity_sheet_permission(actor, item)
    pi_name = clean_text(sheet.get("pi", ""))
    principal_type_by_pi = read_principal_type_by_pi(conn)
    principal_type = principal_type_by_pi.get(pi_name, BILLING_PRINCIPAL_INDEPENDENT)
    # IACUC 分表阶段不应用 PI 免费笼位，避免跨伦理号结算失真。
    free_cages = 0
    rooms = read_rooms_for_quantity_sheets(conn, sheets)
    lines = quantity_sheet_statement_lines(sheets, free_cages, rooms)
    generated_at = now_iso()
    iacucs = sorted({normalize_iacuc_number(item.get("iacuc", "")) for item in sheets if item.get("iacuc")})
    statement_iacuc = iacucs[0] if iacucs else sheet_iacuc
    statement = {
        "id": new_id("stmt"),
        "iacuc": statement_iacuc,
        "iacucs": iacucs,
        "month": sheet["month"],
        "project": "、".join(sorted({item.get("project", "") for item in sheets if item.get("project")})),
        "pi": pi_name,
        "owner": sheet.get("owner", ""),
        "funding": "、".join(sorted({item.get("funding", "") for item in sheets if item.get("funding")})),
        "sourceType": "quantity_sheet",
        "sourceId": sheet["id"],
        "sourceIds": [item["id"] for item in sheets],
        "sourceLabel": "数量统计表",
        "roomName": "、".join(sorted({item.get("roomName", "") for item in sheets if item.get("roomName")})),
        "manager": "、".join(sorted({item.get("manager", "") for item in sheets if item.get("manager")})),
        "billingUnit": "cage_day",
        "principalType": principal_type,
        "freeCageAllowance": free_cages,
        "tierLimit": BILLING_TIER_LIMIT,
        "baseUnitPrice": BILLING_TIER_BASE_PRICE,
        "overageUnitPrice": BILLING_TIER_OVER_PRICE,
        "totalCageDays": sum(line["cageCount"] for line in lines),
        "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
        "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
        "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
        "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
        "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
        "totalAmount": lines[-1]["cumulative"] if lines else 0,
        "status": status,
        "generatedAt": generated_at,
        "lockedAt": generated_at if status == "locked" else "",
    }
    for line in lines:
        line["statementId"] = statement["id"]

    if not persist:
        return statement, lines, []

    workflow, version, statement, lines, workflow_events = save_billing_statement_workflow(
        conn,
        statement,
        lines,
        actor,
        f"根据数量统计表生成 {statement_iacuc} {sheet['month']} 饲养费结算单",
    )
    event = audit_event(
        actor,
        "billing_statement.generated_from_quantity_sheet",
        "billing_workflow",
        workflow["id"],
        f"{actor['displayName']} 根据数量统计表生成 {statement_iacuc} {sheet['month']} 饲养费结算单",
        [],
        generated_at,
        sheet,
        {"workflow": workflow, "version": version},
    )
    write_audit_events(conn, [event])
    return statement, lines, merge_audit_logs([], [event])


def quantity_sheet_statement_lines(sheets, free_cages, rooms=None):
    if not sheets:
        return []

    room_by_id = {room.get("id"): room for room in rooms or []}
    sheet_states = []
    sheet_state_by_iacuc = {}
    month = sheets[0]["month"]
    for sheet in sheets:
        rows_by_date = {}
        for row in sheet.get("rows", []):
            rows_by_date.setdefault(row["date"], []).append(row)
        profile = billing_profile_for_room(room_by_id.get(sheet.get("roomId"), {}), sheet.get("billingUnit"))
        state = {
            "sheet": sheet,
            "rowsByDate": rows_by_date,
            "profile": profile,
            "animalCount": sheet.get("initialAnimalCount") or 0,
            "cageCount": sheet.get("initialCageCount") or 0,
        }
        sheet_states.append(state)
        iacuc = normalize_iacuc_number(sheet.get("iacuc", ""))
        if iacuc and iacuc not in sheet_state_by_iacuc:
            sheet_state_by_iacuc[iacuc] = state

    cumulative = 0
    lines = []
    for date in dates_in_month(month):
        transfer_deltas = {}
        breakdown = []
        animal_count = 0
        cage_count = 0
        charge_groups = {}
        quantity_row_ids = []
        for state in sheet_states:
            day_rows = state["rowsByDate"].get(date, [])
            for row in day_rows:
                added_count = row.get("addedCount") or 0
                removed_count = row.get("removedCount") or 0
                profile = state["profile"]
                if profile["unit"] == "animal_day":
                    if row.get("animalCount") is not None:
                        state["animalCount"] = row.get("animalCount") or 0
                    else:
                        state["animalCount"] = max(state["animalCount"] + added_count - removed_count, 0)
                    if row.get("cageCount") is not None:
                        state["cageCount"] = row.get("cageCount") or 0
                else:
                    if row.get("cageCount") is not None:
                        state["cageCount"] = row.get("cageCount") or 0
                    else:
                        state["cageCount"] = max(state["cageCount"] + added_count - removed_count, 0)
                    if row.get("animalCount") is not None:
                        state["animalCount"] = row.get("animalCount") or 0
                transfer_out_iacuc = normalize_iacuc_number(row.get("transferOutToIacuc", ""))
                if transfer_out_iacuc and removed_count > 0:
                    transfer_deltas[transfer_out_iacuc] = transfer_deltas.get(transfer_out_iacuc, 0) + removed_count
                transfer_in_from_iacuc = normalize_iacuc_number(row.get("transferInFromIacuc", ""))
                if transfer_in_from_iacuc and added_count > 0:
                    transfer_deltas[transfer_in_from_iacuc] = transfer_deltas.get(transfer_in_from_iacuc, 0) - added_count
            sheet = state["sheet"]
            quantity_row_ids.extend(row["id"] for row in day_rows)

        for iacuc, delta in transfer_deltas.items():
            target = sheet_state_by_iacuc.get(iacuc)
            if not target:
                continue
            if target["profile"]["unit"] == "animal_day":
                target["animalCount"] = max((target.get("animalCount") or 0) + delta, 0)
            else:
                target["cageCount"] = max((target.get("cageCount") or 0) + delta, 0)

        for state in sheet_states:
            sheet = state["sheet"]
            profile = state["profile"]
            animal_count += state["animalCount"]
            cage_count += state["cageCount"]
            charge_count = state["animalCount"] if profile["unit"] == "animal_day" else state["cageCount"]
            add_charge_group(charge_groups, profile, charge_count)
            if state["cageCount"] or state["animalCount"]:
                breakdown.append(
                    {
                        "iacuc": sheet.get("iacuc", ""),
                        "project": sheet.get("project", ""),
                        "animalCount": state["animalCount"],
                        "cageCount": state["cageCount"],
                        "billingItem": profile["billingItem"],
                        "billingUnit": profile["unit"],
                        "customerType": profile["customerType"],
                        "unitPrice": profile["unitPrice"],
                        "overageUnitPrice": BILLING_TIER_OVER_PRICE if profile["tiered"] else 0,
                        "tiered": bool(profile["tiered"]),
                        "freeAllowance": bool(profile["freeAllowance"]),
                    }
                )

        charges = combined_daily_charge(charge_groups, free_cages)
        cumulative += charges["amount"]
        lines.append(
            {
                "id": new_id("line"),
                "date": date,
                "animalCount": animal_count,
                "cageCount": cage_count,
                **charges,
                "cumulative": cumulative,
                "iacucBreakdown": breakdown,
                "quantitySheetRowIds": quantity_row_ids,
                "occupancyIds": [],
            }
        )
    return lines


def generate_billing_statement(conn, payload, actor):
    iacuc = normalize_iacuc_number(payload.get("iacuc", ""))
    requested_pi = clean_text(payload.get("pi", ""))
    month = clean_text(payload.get("month", ""))
    status = clean_text(payload.get("status", "draft")) or "draft"
    persist = bool(payload.get("persist"))
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("结算月份格式应为 YYYY-MM")
    if status not in ("draft", "locked"):
        raise ValueError("结算单状态只能是 draft 或 locked")
    if not iacuc:
        raise ValueError("请先选择伦理号后再生成结算单")

    occupancies = read_occupancies_for_billing(conn, month, iacuc=iacuc)
    applications_by_iacuc = read_applications_by_iacuc(conn)
    state = read_billing_state_for_occupancies(conn, occupancies)
    rooms = state["rooms"]
    dates = dates_in_month(month)
    generated_at = now_iso()
    cumulative = 0
    lines = []
    pi_name = requested_pi or pi_for_iacuc(iacuc, applications_by_iacuc, occupancies)
    if not pi_name:
        raise ValueError("项目负责人不能为空")
    principal_type_by_pi = read_principal_type_by_pi(conn)
    principal_type = principal_type_by_pi.get(pi_name, BILLING_PRINCIPAL_INDEPENDENT)
    # IACUC 分表阶段不应用 PI 免费笼位，避免跨伦理号结算失真。
    free_cages = 0
    iacucs = [iacuc]

    for date in dates:
        active_items = [
            item
            for item in occupancies
            if normalize_iacuc_number(item.get("iacuc", "")) == iacuc and occupancy_active_on_date(item, date)
        ]
        charge_groups = {}
        cage_count = 0
        animal_count = 0
        for item in active_items:
            profile = billing_profile_for_occupancy(item, state)
            if profile["unit"] == "animal_day":
                count = occupancy_animal_count(item, profile)
                animal_count += count
            else:
                count = 1
                cage_count += 1
            add_charge_group(charge_groups, profile, count)
        charges = combined_daily_charge(charge_groups, free_cages)
        amount = charges["amount"]
        cumulative += amount
        breakdown = []
        if active_items:
            for item in active_items:
                profile = billing_profile_for_occupancy(item, state)
                found = next(
                    (
                        entry
                        for entry in breakdown
                        if entry["iacuc"] == iacuc
                        and entry.get("billingItem") == profile["billingItem"]
                        and entry.get("customerType") == profile["customerType"]
                    ),
                    None,
                )
                if not found:
                    found = {
                        "iacuc": iacuc,
                        "project": statement_application_snapshot(iacuc, applications_by_iacuc, occupancies).get("project", ""),
                        "animalCount": 0,
                        "cageCount": 0,
                        "billingItem": profile["billingItem"],
                        "billingUnit": profile["unit"],
                        "customerType": profile["customerType"],
                        "unitPrice": profile["unitPrice"],
                        "overageUnitPrice": BILLING_TIER_OVER_PRICE if profile["tiered"] else 0,
                        "tiered": bool(profile["tiered"]),
                        "freeAllowance": bool(profile["freeAllowance"]),
                    }
                    breakdown.append(found)
                if profile["unit"] == "animal_day":
                    found["animalCount"] += occupancy_animal_count(item, profile)
                else:
                    found["cageCount"] += 1
        line = {
            "id": new_id("line"),
            "date": date,
            "animalCount": animal_count,
            "cageCount": cage_count,
            **charges,
            "amount": amount,
            "cumulative": cumulative,
            "iacucBreakdown": breakdown,
            "occupancyIds": [item.get("id") for item in active_items if item.get("id")],
        }
        lines.append(line)

    application = statement_application_snapshot(iacuc, applications_by_iacuc, occupancies)
    statement = {
        "id": new_id("stmt"),
        "iacuc": iacuc,
        "iacucs": iacucs,
        "month": month,
        "project": application.get("project", ""),
        "pi": pi_name,
        "owner": application.get("owner", ""),
        "funding": application.get("funding", ""),
        "sourceType": "cage_map",
        "sourceLabel": "动态笼位图",
        "billingUnit": statement_billing_unit_from_lines(lines),
        "principalType": principal_type,
        "freeCageAllowance": free_cages,
        "tierLimit": BILLING_TIER_LIMIT,
        "baseUnitPrice": BILLING_TIER_BASE_PRICE,
        "overageUnitPrice": BILLING_TIER_OVER_PRICE,
        "totalCageDays": sum(line["cageCount"] for line in lines),
        "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
        "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
        "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
        "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
        "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
        "totalAmount": cumulative,
        "status": status,
        "generatedAt": generated_at,
        "lockedAt": generated_at if status == "locked" else "",
    }
    for line in lines:
        line["statementId"] = statement["id"]

    if not persist:
        return statement, lines, []

    workflow, version, statement, lines, workflow_events = save_billing_statement_workflow(
        conn,
        statement,
        lines,
        actor,
        f"生成 {pi_name} {month} 饲养费结算单",
    )
    detail_context = occupancy_detail_context(occupancies, rooms)
    upsert_reimbursement_record_from_statement(conn, workflow, statement, lines, detail_context, "workflow")
    recalculate_reimbursement_accumulations(conn, pi_name)
    event = audit_event(
        actor,
        "billing_statement.generated",
        "billing_workflow",
        workflow["id"],
        f"{actor['displayName']} 生成 {pi_name} {month} 饲养费结算单",
        [],
        generated_at,
        None,
        {"workflow": workflow, "version": version},
    )
    write_audit_events(conn, [event])
    return statement, lines, merge_audit_logs([], [event])


def generate_billing_statement_by_pi(conn, payload, actor):
    month = clean_text(payload.get("month", ""))
    pi_name = clean_text(payload.get("pi", ""))
    status = clean_text(payload.get("status", "draft")) or "draft"
    source_type = clean_text(payload.get("sourceType", "cage_map")) or "cage_map"
    persist = bool(payload.get("persist"))
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("结算月份格式应为 YYYY-MM")
    if status not in ("draft", "locked"):
        raise ValueError("结算单状态只能是 draft 或 locked")
    if source_type not in ("cage_map", "quantity_sheet"):
        raise ValueError("sourceType 只能是 cage_map 或 quantity_sheet")
    if not pi_name:
        raise ValueError("按 PI 合表需要提供项目负责人")

    principal_type_by_pi = read_principal_type_by_pi(conn)
    principal_type = principal_type_by_pi.get(pi_name, BILLING_PRINCIPAL_INDEPENDENT)
    free_cages = billing_free_cages_for_pi(principal_type_by_pi, pi_name)
    generated_at = now_iso()
    iacucs = []

    if source_type == "cage_map":
        applications_by_iacuc = read_applications_by_iacuc(conn)
        occupancies = read_occupancies_for_billing(conn, month, pi=pi_name)
        state = read_billing_state_for_occupancies(conn, occupancies)
        rooms = state["rooms"]
        iacucs = sorted(
            {
                normalize_iacuc_number(item.get("iacuc", ""))
                for item in occupancies
                if clean_text(item.get("pi", "")) == pi_name and normalize_iacuc_number(item.get("iacuc", ""))
            }
        )
        cumulative = 0
        lines = []
        for date in dates_in_month(month):
            active_items = [
                item
                for item in occupancies
                if clean_text(item.get("pi", "")) == pi_name and occupancy_active_on_date(item, date)
            ]
            charge_groups = {}
            cage_count = 0
            animal_count = 0
            for item in active_items:
                profile = billing_profile_for_occupancy(item, state)
                if profile["unit"] == "animal_day":
                    count = occupancy_animal_count(item, profile)
                    animal_count += count
                else:
                    count = 1
                    cage_count += 1
                add_charge_group(charge_groups, profile, count)
            charges = combined_daily_charge(charge_groups, free_cages)
            cumulative += charges["amount"]
            breakdown = []
            for item in active_items:
                item_iacuc = normalize_iacuc_number(item.get("iacuc", ""))
                profile = billing_profile_for_occupancy(item, state)
                found = next(
                    (
                        entry
                        for entry in breakdown
                        if entry["iacuc"] == item_iacuc
                        and entry.get("billingItem") == profile["billingItem"]
                        and entry.get("customerType") == profile["customerType"]
                    ),
                    None,
                )
                if not found:
                    found = {
                        "iacuc": item_iacuc,
                        "project": item.get("project", ""),
                        "animalCount": 0,
                        "cageCount": 0,
                        "billingItem": profile["billingItem"],
                        "billingUnit": profile["unit"],
                        "customerType": profile["customerType"],
                        "unitPrice": profile["unitPrice"],
                        "overageUnitPrice": BILLING_TIER_OVER_PRICE if profile["tiered"] else 0,
                        "tiered": bool(profile["tiered"]),
                        "freeAllowance": bool(profile["freeAllowance"]),
                    }
                    breakdown.append(found)
                if profile["unit"] == "animal_day":
                    found["animalCount"] += occupancy_animal_count(item, profile)
                else:
                    found["cageCount"] += 1
            lines.append(
                {
                    "id": new_id("line"),
                    "date": date,
                    "animalCount": animal_count,
                    "cageCount": cage_count,
                    **charges,
                    "amount": charges["amount"],
                    "cumulative": cumulative,
                    "iacucBreakdown": breakdown,
                    "occupancyIds": [item.get("id") for item in active_items if item.get("id")],
                }
            )
        application = statement_pi_snapshot(pi_name, applications_by_iacuc, occupancies)
        statement = {
            "id": new_id("stmt"),
            "iacuc": f"pi::{pi_name}",
            "iacucs": iacucs,
            "month": month,
            "project": application.get("project", ""),
            "pi": pi_name,
            "owner": application.get("owner", ""),
            "funding": application.get("funding", ""),
            "sourceType": "pi_merged_cage_map",
            "sourceLabel": "动态笼位图（按 PI 合表）",
            "billingUnit": statement_billing_unit_from_lines(lines),
            "principalType": principal_type,
            "freeCageAllowance": free_cages,
            "tierLimit": BILLING_TIER_LIMIT,
            "baseUnitPrice": BILLING_TIER_BASE_PRICE,
            "overageUnitPrice": BILLING_TIER_OVER_PRICE,
            "totalCageDays": sum(line["cageCount"] for line in lines),
            "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
            "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
            "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
            "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
            "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
            "totalAmount": cumulative,
            "status": status,
            "generatedAt": generated_at,
            "lockedAt": generated_at if status == "locked" else "",
        }
        detail_context = occupancy_detail_context(occupancies, rooms)
    else:
        sheets = list_quantity_sheets_by_month_pi(conn, month, pi_name)
        if not sheets:
            raise ValueError("未找到该 PI 在结算月份内的数量统计表")
        for item in sheets:
            validate_quantity_sheet_permission(actor, item)
        iacucs = sorted({normalize_iacuc_number(item.get("iacuc", "")) for item in sheets if item.get("iacuc")})
        rooms = read_rooms_for_quantity_sheets(conn, sheets)
        lines = quantity_sheet_statement_lines(sheets, free_cages, rooms)
        statement = {
            "id": new_id("stmt"),
            "iacuc": f"pi::{pi_name}",
            "iacucs": iacucs,
            "month": month,
            "project": "、".join(sorted({item.get("project", "") for item in sheets if item.get("project")})),
            "pi": pi_name,
            "owner": "、".join(sorted({item.get("owner", "") for item in sheets if item.get("owner")})),
            "funding": "、".join(sorted({item.get("funding", "") for item in sheets if item.get("funding")})),
            "sourceType": "pi_merged_quantity_sheet",
            "sourceIds": [item["id"] for item in sheets],
            "sourceLabel": "数量统计表（按 PI 合表）",
            "roomName": "、".join(sorted({item.get("roomName", "") for item in sheets if item.get("roomName")})),
            "manager": "、".join(sorted({item.get("manager", "") for item in sheets if item.get("manager")})),
            "billingUnit": statement_billing_unit_from_lines(lines),
            "principalType": principal_type,
            "freeCageAllowance": free_cages,
            "tierLimit": BILLING_TIER_LIMIT,
            "baseUnitPrice": BILLING_TIER_BASE_PRICE,
            "overageUnitPrice": BILLING_TIER_OVER_PRICE,
            "totalCageDays": sum(line["cageCount"] for line in lines),
            "totalFreeCageDays": sum(line.get("freeCages", 0) for line in lines),
            "totalBillableCageDays": sum(line.get("billableCages", 0) for line in lines),
            "totalTier1CageDays": sum(line.get("tier1BillableCages", 0) for line in lines),
            "totalTier2CageDays": sum(line.get("tier2BillableCages", 0) for line in lines),
            "totalAnimalDays": sum(line.get("animalCount", 0) for line in lines),
            "totalAmount": lines[-1]["cumulative"] if lines else 0,
            "status": status,
            "generatedAt": generated_at,
            "lockedAt": generated_at if status == "locked" else "",
        }
        detail_context = quantity_sheet_detail_context(sheets, rooms)

    for line in lines:
        line["statementId"] = statement["id"]
    if not persist:
        return statement, lines, []
    workflow, version, statement, lines, workflow_events = save_billing_statement_workflow(
        conn,
        statement,
        lines,
        actor,
        f"按 PI 合表生成 {pi_name} {month} 饲养费结算单",
    )
    upsert_reimbursement_record_from_statement(conn, workflow, statement, lines, detail_context, "workflow")
    recalculate_reimbursement_accumulations(conn, pi_name)
    event = audit_event(
        actor,
        "billing_statement.generated_by_pi",
        "billing_workflow",
        workflow["id"],
        f"{actor['displayName']} 按 PI 合表生成 {pi_name} {month} 饲养费结算单",
        [],
        generated_at,
        None,
        {"workflow": workflow, "version": version},
    )
    write_audit_events(conn, [event])
    return statement, lines, merge_audit_logs([], [event])


def dates_in_month(month):
    year, month_no = [int(part) for part in month.split("-")]
    day_count = calendar.monthrange(year, month_no)[1]
    return [f"{year:04d}-{month_no:02d}-{day:02d}" for day in range(1, day_count + 1)]


def tiered_daily_charge(cage_count, free_cages):
    cage_count = max(as_int(cage_count) or 0, 0)
    free_cages = min(max(as_int(free_cages) or 0, 0), cage_count)
    tier1_cages = min(cage_count, BILLING_TIER_LIMIT)
    tier2_cages = max(cage_count - BILLING_TIER_LIMIT, 0)
    tier1_free = min(free_cages, tier1_cages)
    tier2_free = min(max(free_cages - tier1_free, 0), tier2_cages)
    tier1_billable = max(tier1_cages - tier1_free, 0)
    tier2_billable = max(tier2_cages - tier2_free, 0)
    amount = tier1_billable * BILLING_TIER_BASE_PRICE + tier2_billable * BILLING_TIER_OVER_PRICE
    return {
        "freeCages": free_cages,
        "billableCages": tier1_billable + tier2_billable,
        "tier1Cages": tier1_cages,
        "tier2Cages": tier2_cages,
        "tier1BillableCages": tier1_billable,
        "tier2BillableCages": tier2_billable,
        "unitPrice": BILLING_TIER_BASE_PRICE,
        "overageUnitPrice": BILLING_TIER_OVER_PRICE,
        "discountPercent": 0,
        "amount": amount,
    }


def normalize_billing_item(value):
    text = clean_text(value)
    return text if text in BILLING_RULES else "mouse_standard"


def normalize_customer_type(value):
    return "external" if clean_text(value) == "external" else "internal"


def normalize_billing_unit(value):
    return "animal_day" if clean_text(value) == "animal_day" else "cage_day"


def billing_item_for_species(species):
    return {
        "mouse": "mouse_standard",
        "rat": "rat_standard",
        "guinea_pig": "guinea_pig",
        "rabbit": "rabbit",
        "monkey": "monkey",
        "pig": "pig",
        "dog": "dog",
    }.get(clean_text(species), "mouse_standard")


def infer_billing_item_from_room(room=None):
    room = room or {}
    text = " ".join(
        clean_text(room.get(key, ""))
        for key in ("name", "area", "defaultSpecies", "species")
        if clean_text(room.get(key, ""))
    )
    if re.search(r"糖尿病.*大鼠|大鼠.*糖尿病", text):
        return "rat_diabetic"
    if re.search(r"糖尿病.*小鼠|小鼠.*糖尿病", text):
        return "mouse_diabetic"
    if "豚鼠" in text:
        return "guinea_pig"
    if "大鼠" in text:
        return "rat_standard"
    if "兔" in text:
        return "rabbit"
    if "猴" in text:
        return "monkey"
    if "犬" in text or "狗" in text:
        return "dog"
    if "猪" in text:
        return "pig"
    if "小鼠" in text:
        return "mouse_standard"
    return ""


def room_has_manual_billing_profile(room=None, inferred_billing_item=""):
    room = room or {}
    if room.get("billingProfileConfirmed"):
        return True
    if not room.get("billingProfileConfigured"):
        return False
    fallback_mouse = (
        clean_text(room.get("defaultBillingItem", "")) in ("", "mouse_standard")
        and clean_text(room.get("defaultSpecies", "")) in ("", "mouse")
    )
    return not (fallback_mouse and inferred_billing_item and inferred_billing_item != "mouse_standard")


def billing_profile_for_room(room=None, fallback_unit=None):
    room = room or {}
    inferred_billing_item = infer_billing_item_from_room(room)
    manual_billing = room_has_manual_billing_profile(room, inferred_billing_item)
    billing_item = normalize_billing_item(
        (room.get("defaultBillingItem") if manual_billing else inferred_billing_item)
        or room.get("defaultBillingItem")
        or billing_item_for_species(room.get("defaultSpecies", "mouse"))
    )
    if fallback_unit == "animal_day" and billing_item == "mouse_standard":
        billing_item = "guinea_pig"
    rule = BILLING_RULES.get(billing_item, BILLING_RULES["mouse_standard"])
    customer_type = normalize_customer_type(room.get("defaultCustomerType", "internal"))
    unit_price = rule["externalPrice"] if customer_type == "external" else rule["internalPrice"]
    return {
        "facility": clean_text(room.get("facility", "zhujiang")) or "zhujiang",
        "species": rule["species"],
        "billingItem": billing_item,
        "customerType": customer_type,
        "unit": rule["unit"],
        "unitPrice": unit_price,
        "tiered": bool(rule.get("tiered") and customer_type == "internal"),
        "freeAllowance": bool(rule.get("freeAllowance") and customer_type == "internal"),
        "defaultAnimalCount": max(as_int(room.get("defaultAnimalCount")) or 1, 1),
    }


def billing_profile_for_occupancy(occupancy, state):
    slot = next((item for item in state.get("slots", []) if item.get("id") == occupancy.get("slotId")), None)
    rack = next((item for item in state.get("racks", []) if item.get("id") == (slot or {}).get("rackId")), None)
    room = next((item for item in state.get("rooms", []) if item.get("id") == (rack or {}).get("roomId")), None)
    base = billing_profile_for_room(room)
    billing_item = normalize_billing_item(occupancy.get("billingItem") or base["billingItem"])
    rule = BILLING_RULES.get(billing_item, BILLING_RULES["mouse_standard"])
    customer_type = normalize_customer_type(occupancy.get("customerType") or base["customerType"])
    unit_price = rule["externalPrice"] if customer_type == "external" else rule["internalPrice"]
    return {
        **base,
        "species": rule["species"],
        "billingItem": billing_item,
        "customerType": customer_type,
        "unit": rule["unit"],
        "unitPrice": unit_price,
        "tiered": bool(rule.get("tiered") and customer_type == "internal"),
        "freeAllowance": bool(rule.get("freeAllowance") and customer_type == "internal"),
    }


def occupancy_animal_count(occupancy, profile):
    return max(as_int(occupancy.get("animalCount")) or as_int(profile.get("defaultAnimalCount")) or 1, 1)


def flat_daily_charge(count, profile):
    count = max(as_int(count) or 0, 0)
    amount = count * float(profile.get("unitPrice") or 0)
    if profile.get("unit") == "animal_day":
        return {
            "freeCages": 0,
            "billableCages": 0,
            "tier1Cages": 0,
            "tier2Cages": 0,
            "tier1BillableCages": 0,
            "tier2BillableCages": 0,
            "billableAnimals": count,
            "unitPrice": profile.get("unitPrice") or 0,
            "overageUnitPrice": 0,
            "discountPercent": 0,
            "amount": amount,
        }
    return {
        "freeCages": 0,
        "billableCages": count,
        "tier1Cages": count,
        "tier2Cages": 0,
        "tier1BillableCages": count,
        "tier2BillableCages": 0,
        "billableAnimals": 0,
        "unitPrice": profile.get("unitPrice") or 0,
        "overageUnitPrice": 0,
        "discountPercent": 0,
        "amount": amount,
    }


def add_charge_group(groups, profile, count):
    count = max(as_int(count) or 0, 0)
    if count <= 0:
        return
    key = "|".join(
        [
            profile.get("billingItem", ""),
            profile.get("customerType", ""),
            profile.get("unit", ""),
            str(profile.get("unitPrice") or 0),
        ]
    )
    if key not in groups:
        groups[key] = {"profile": profile, "count": 0}
    groups[key]["count"] += count


def combined_daily_charge(groups, free_cages):
    total = {
        "freeCages": 0,
        "billableCages": 0,
        "tier1Cages": 0,
        "tier2Cages": 0,
        "tier1BillableCages": 0,
        "tier2BillableCages": 0,
        "billableAnimals": 0,
        "unitPrice": BILLING_TIER_BASE_PRICE,
        "overageUnitPrice": BILLING_TIER_OVER_PRICE,
        "discountPercent": 0,
        "amount": 0,
    }
    remaining_free_cages = max(as_int(free_cages) or 0, 0)
    for group in groups.values():
        profile = group["profile"]
        count = group["count"]
        if profile.get("tiered"):
            allowance = remaining_free_cages if profile.get("freeAllowance") else 0
            charges = tiered_daily_charge(count, allowance)
            remaining_free_cages = max(remaining_free_cages - charges.get("freeCages", 0), 0)
        else:
            charges = flat_daily_charge(count, profile)
        for key in (
            "freeCages",
            "billableCages",
            "tier1Cages",
            "tier2Cages",
            "tier1BillableCages",
            "tier2BillableCages",
            "billableAnimals",
            "amount",
        ):
            total[key] += charges.get(key, 0)
        total["unitPrice"] = charges.get("unitPrice", total["unitPrice"])
        if charges.get("overageUnitPrice"):
            total["overageUnitPrice"] = charges.get("overageUnitPrice")
    return total


def statement_billing_unit_from_lines(lines):
    has_animals = any((line.get("animalCount") or 0) > 0 for line in lines)
    has_cages = any((line.get("cageCount") or 0) > 0 for line in lines)
    if has_animals and has_cages:
        return "mixed"
    return "animal_day" if has_animals else "cage_day"


def occupancy_active_on_date(item, date):
    if item.get("status") not in ("active", "ended"):
        return False
    if not item.get("startDate") or item.get("startDate") > date:
        return False
    if item.get("endDate") and item.get("endDate") < date:
        return False
    return True


def billing_unit_price_for(rules, date):
    for rule in rules:
        after_start = not rule.get("effectiveStart") or rule.get("effectiveStart") <= date
        before_end = not rule.get("effectiveEnd") or rule.get("effectiveEnd") >= date
        if rule.get("unit") == "cage_day" and after_start and before_end:
            return float(rule.get("price") or 0)
    return 4.5


def billing_discount_for(adjustments, iacuc, date):
    for adjustment in adjustments:
        in_range = (
            (not adjustment.get("effectiveStart") or adjustment.get("effectiveStart") <= date)
            and (not adjustment.get("effectiveEnd") or adjustment.get("effectiveEnd") >= date)
        )
        if (
            adjustment.get("targetType") == "iacuc"
            and normalize_iacuc_number(adjustment.get("targetId", "")) == iacuc
            and adjustment.get("type") == "discount"
            and in_range
        ):
            return float(adjustment.get("value") or 0)
    return 0


def billing_free_cages_for(adjustments, pi_name):
    for adjustment in adjustments:
        if (
            adjustment.get("targetType") == "pi"
            and clean_text(adjustment.get("targetId", "")) == pi_name
            and adjustment.get("type") == "free_cages"
        ):
            if adjustment.get("principalType"):
                return free_cages_for_principal_type(adjustment.get("principalType"))
            return max(as_int(adjustment.get("value")) or 0, 0)
    return FREE_CAGES_DEFAULT


def read_principal_type_by_pi(conn):
    return {
        clean_text(pi_name): normalize_principal_type(principal_type)
        for pi_name, principal_type in read_principal_type_by_pi_repository(conn).items()
        if clean_text(pi_name)
    }


def billing_free_cages_for_pi(principal_type_by_pi, pi_name):
    return free_cages_for_principal_type(principal_type_by_pi.get(clean_text(pi_name), BILLING_PRINCIPAL_INDEPENDENT))


def normalize_principal_type(value):
    return BILLING_PRINCIPAL_PI if value == BILLING_PRINCIPAL_PI else BILLING_PRINCIPAL_INDEPENDENT


def free_cages_for_principal_type(value):
    return FREE_CAGES_PI if normalize_principal_type(value) == BILLING_PRINCIPAL_PI else FREE_CAGES_INDEPENDENT


def principal_type_label(value):
    return "PI" if normalize_principal_type(value) == BILLING_PRINCIPAL_PI else "独立科研人员"


def pi_for_iacuc(iacuc, applications_by_iacuc, occupancies):
    if not iacuc:
        return ""
    application = applications_by_iacuc.get(iacuc)
    if application and application.get("pi"):
        return clean_text(application.get("pi", ""))
    for item in occupancies:
        if normalize_iacuc_number(item.get("iacuc", "")) == iacuc and item.get("pi"):
            return clean_text(item.get("pi", ""))
    return ""


def statement_application_snapshot(iacuc, applications_by_iacuc, occupancies):
    application = applications_by_iacuc.get(iacuc)
    if application:
        return application
    for item in occupancies:
        if normalize_iacuc_number(item.get("iacuc", "")) == iacuc:
            return {
                "project": item.get("project", ""),
                "pi": item.get("pi", ""),
                "owner": item.get("owner", ""),
                "funding": item.get("funding", ""),
            }
    return {}


def statement_pi_snapshot(pi_name, applications_by_iacuc, occupancies):
    projects = []
    owners = []
    fundings = []
    for application in applications_by_iacuc.values():
        if clean_text(application.get("pi", "")) == pi_name:
            projects.append(application.get("project", ""))
            owners.append(application.get("owner", ""))
            fundings.append(application.get("funding", ""))
    for item in occupancies:
        if clean_text(item.get("pi", "")) == pi_name:
            projects.append(item.get("project", ""))
            owners.append(item.get("owner", ""))
            fundings.append(item.get("funding", ""))
    return {
        "project": "、".join(sorted({value for value in projects if value})),
        "pi": pi_name,
        "owner": "、".join(sorted({value for value in owners if value})),
        "funding": "、".join(sorted({value for value in fundings if value})),
    }


def normalize_workflow_source(source_type):
    text = clean_text(source_type or "")
    return text if text else "cage_map"


def workflow_scope_for_statement(statement):
    source_type = normalize_workflow_source(statement.get("sourceType", ""))
    pi_name = clean_text(statement.get("pi", ""))
    if source_type.startswith("pi_merged_") and pi_name:
        return "pi", f"pi::{pi_name}"
    if source_type in ("cage_map", "quantity_sheet") and pi_name:
        return "pi", f"pi::{pi_name}"
    iacuc = clean_text(statement.get("iacuc", ""))
    return "iacuc", iacuc


def billing_workflow_business_key(scope_type, scope_key, month, source_type):
    return "|".join([clean_text(scope_type), clean_text(scope_key), clean_text(month), normalize_workflow_source(source_type)])


def make_statement_document_number(statement, version_no):
    source = normalize_workflow_source(statement.get("sourceType", ""))
    source_code = {
        "quantity_sheet": "QS",
        "pi_merged_quantity_sheet": "PQS",
        "pi_merged_cage_map": "PCM",
    }.get(source, "CM")
    month = re.sub(r"\D", "", clean_text(statement.get("month", "")) or "000000")
    iacuc = re.sub(r"[^A-Za-z0-9]", "", clean_text(statement.get("iacuc", "")) or "UNKNOWN").upper()
    return f"CL-{source_code}-{month}-{iacuc}-V{int(version_no):02d}"


def enrich_statement_for_workflow(
    statement,
    *,
    workflow_id,
    version_id,
    version_no,
    version_status,
    workflow_status,
    document_number,
):
    return {
        **statement,
        "id": version_id,
        "workflowId": workflow_id,
        "versionId": version_id,
        "versionNo": version_no,
        "versionStatus": version_status,
        "workflowStatus": workflow_status,
        "documentNumber": document_number,
    }


def build_version_payload(
    statement,
    workflow_id,
    version_no,
    version_status,
    workflow_status,
    generated_at,
    voided_at,
    voided_by,
    void_reason,
):
    return {
        "id": statement["id"],
        "workflowId": workflow_id,
        "versionNo": version_no,
        "versionStatus": version_status,
        "workflowStatus": workflow_status,
        "generatedAt": generated_at,
        "voidedAt": voided_at,
        "voidedBy": voided_by,
        "voidReason": void_reason,
        "documentNumber": statement.get("documentNumber", ""),
        "statement": statement,
        "summary": {
            "iacuc": statement.get("iacuc", ""),
            "iacucs": statement.get("iacucs", []),
            "month": statement.get("month", ""),
            "sourceType": statement.get("sourceType", ""),
            "pi": statement.get("pi", ""),
            "project": statement.get("project", ""),
            "owner": statement.get("owner", ""),
            "funding": statement.get("funding", ""),
            "totalAmount": statement.get("totalAmount", 0),
            "totalCageDays": statement.get("totalCageDays", 0),
            "status": statement.get("status", "draft"),
        },
    }


def build_workflow_payload(workflow_id, iacuc, month, source_type, workflow_status, current_version, latest_event_at):
    statement = current_version.get("statement", {})
    scope_type, scope_key = workflow_scope_for_statement(statement)
    timestamps = {
        "generatedAt": current_version.get("generatedAt", ""),
        "sentAt": statement.get("sentAt", ""),
        "signedReturnedAt": statement.get("signedReturnedAt", ""),
        "submittedToFinanceAt": statement.get("submittedToFinanceAt", ""),
    }
    return {
        "id": workflow_id,
        "businessKey": billing_workflow_business_key(scope_type, scope_key, month, source_type),
        "scopeType": scope_type,
        "scopeKey": scope_key,
        "iacuc": iacuc,
        "iacucs": statement.get("iacucs", []),
        "month": month,
        "sourceType": source_type,
        "workflowStatus": workflow_status,
        "currentVersionId": current_version.get("id", ""),
        "currentVersionNo": current_version.get("versionNo", 0),
        "currentVersion": current_version,
        "latestEventAt": latest_event_at,
        "pi": statement.get("pi", ""),
        "project": statement.get("project", ""),
        "owner": statement.get("owner", ""),
        "funding": statement.get("funding", ""),
        "totalAmount": statement.get("totalAmount", 0),
        "totalCageDays": statement.get("totalCageDays", 0),
        **timestamps,
    }


def build_workflow_event_payload(event_id, workflow_id, version_id, event_type, from_status, to_status, actor, at, channel, note):
    return {
        "id": event_id,
        "workflowId": workflow_id,
        "versionId": version_id,
        "eventType": event_type,
        "fromStatus": from_status,
        "toStatus": to_status,
        "actor": {
            "id": actor.get("id", ""),
            "username": actor.get("username", ""),
            "displayName": actor.get("displayName", ""),
        },
        "channel": channel,
        "note": note,
        "at": at,
    }


def get_billing_workflow_by_key(conn, business_key):
    return get_billing_workflow_by_key_repository(conn, business_key)


def get_billing_workflow(conn, workflow_id):
    return get_billing_workflow_repository(conn, workflow_id)


def get_billing_workflow_detail(conn, workflow_id):
    return get_billing_workflow_detail_repository(conn, workflow_id)


def get_billing_version(conn, version_id):
    return get_billing_version_repository(conn, version_id)


def list_billing_workflows(conn):
    return list_billing_workflows_repository(conn)


def list_billing_workflows_page(conn, filters):
    return list_billing_workflows_page_repository(conn, filters, clean_text, WORKFLOW_STATUS_FINANCE)


def list_billing_workflow_lines(conn, workflow_id, version_id=""):
    workflow = get_billing_workflow(conn, workflow_id)
    if not workflow:
        raise LookupError("结算流程不存在")
    selected_version_id = clean_text(version_id) or clean_text(workflow.get("currentVersionId", ""))
    if not selected_version_id:
        return {"workflowId": workflow_id, "versionId": "", "lines": []}
    return {
        "workflowId": workflow_id,
        "versionId": selected_version_id,
        "lines": list_billing_statement_line_summaries_for_version_repository(conn, selected_version_id),
    }


def delete_billing_workflow(conn, workflow_id):
    workflow = get_billing_workflow(conn, workflow_id)
    if not workflow:
        raise LookupError("结算流程不存在")
    delete_billing_workflow_tree_repository(conn, workflow_id)
    return workflow


def list_billing_workflow_versions(conn, workflow_id):
    return list_billing_workflow_versions_repository(conn, workflow_id)


def billing_workflow_detail_item(workflow):
    return billing_workflow_detail_item_repository(workflow)


def list_billing_workflow_events(conn, workflow_id):
    return list_billing_workflow_events_repository(conn, workflow_id)


def list_billing_statement_lines_for_version(conn, version_id):
    return list_billing_statement_lines_for_version_repository(conn, version_id)


def list_current_billing_statements(conn):
    return list_current_billing_statements_repository(conn)


def list_reimbursement_records_page(conn, filters):
    return list_reimbursement_records_page_repository(conn, filters, clean_text)


def get_reimbursement_record(conn, record_id):
    return get_reimbursement_record_repository(conn, record_id)


def get_reimbursement_record_by_key(conn, business_key):
    return get_reimbursement_record_by_key_repository(conn, business_key)


def get_reimbursement_record_by_workflow_id(conn, workflow_id):
    return get_reimbursement_record_by_workflow_id_repository(conn, workflow_id)


def list_reimbursement_records_for_pi(conn, pi_name):
    return list_reimbursement_records_for_pi_repository(conn, pi_name)


def list_reimbursement_record_summaries_for_pi(conn, pi_name):
    return list_reimbursement_record_summaries_for_pi_repository(conn, pi_name)


def upsert_reimbursement_record(conn, payload):
    existing_by_key = get_reimbursement_record_by_key(conn, payload.get("businessKey", ""))
    if existing_by_key and existing_by_key.get("id"):
        payload["id"] = existing_by_key["id"]
    else:
        row = conn.execute("SELECT payload FROM reimbursement_records WHERE id = ?", (payload.get("id", ""),)).fetchone()
        if row:
            existing_by_id = json.loads(row["payload"])
            payload["id"] = existing_by_id.get("id", payload.get("id", ""))
            if not payload.get("businessKey") and existing_by_id.get("businessKey"):
                payload["businessKey"] = existing_by_id["businessKey"]
    upsert_reimbursement_record_repository(conn, payload)


def delete_reimbursement_record(conn, record_id):
    delete_reimbursement_record_repository(conn, record_id)


def reimbursement_status_label(value):
    return {
        REIMBURSEMENT_STATUS_PENDING: "待提交",
        REIMBURSEMENT_STATUS_REIMBURSING: "报销中",
        REIMBURSEMENT_STATUS_COMPLETED: "已完成",
    }.get(normalize_reimbursement_status(value), "待提交")


def reimbursement_record_id(month, pi_name):
    raw = f"{clean_text(month)}|{clean_text(pi_name)}"
    slug = re.sub(r"[^A-Za-z0-9]+", "-", clean_text(month)).strip("-") or "record"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]
    return f"rrc-{slug}-{digest}"


def quantity_sheet_detail_context(sheets, rooms):
    room_by_id = {room.get("id"): room for room in rooms}
    context = {}
    for sheet in sheets:
        iacuc = normalize_iacuc_number(sheet.get("iacuc", ""))
        if not iacuc:
            continue
        room = room_by_id.get(sheet.get("roomId")) or next((item for item in rooms if clean_text(item.get("name")) == clean_text(sheet.get("roomName"))), {})
        profile = billing_profile_for_room(room or {})
        current = context.get(iacuc) or {
            "facility": profile.get("facility", ""),
            "funding": "",
            "species": room.get("defaultSpecies", ""),
            "project": "",
            "owner": "",
            "roomNames": set(),
        }
        if sheet.get("funding"):
            current["funding"] = join_distinct_text(current.get("funding", ""), sheet.get("funding", ""))
        if sheet.get("project"):
            current["project"] = join_distinct_text(current.get("project", ""), sheet.get("project", ""))
        if sheet.get("owner"):
            current["owner"] = join_distinct_text(current.get("owner", ""), sheet.get("owner", ""))
        if room.get("defaultSpecies"):
            current["species"] = join_distinct_text(current.get("species", ""), room.get("defaultSpecies", ""))
        if sheet.get("roomName"):
            current["roomNames"].add(sheet.get("roomName"))
        context[iacuc] = current
    return finalize_reimbursement_detail_context(context)


def occupancy_detail_context(occupancies, rooms):
    room_by_id = {room.get("id"): room for room in rooms}
    context = {}
    for item in occupancies:
        iacuc = normalize_iacuc_number(item.get("iacuc", ""))
        if not iacuc:
            continue
        room = room_by_id.get(item.get("roomId"), {})
        profile = billing_profile_for_room(room or {})
        current = context.get(iacuc) or {
            "facility": profile.get("facility", ""),
            "funding": "",
            "species": "",
            "project": "",
            "owner": "",
            "roomNames": set(),
        }
        current["funding"] = join_distinct_text(current.get("funding", ""), item.get("funding", ""))
        current["species"] = join_distinct_text(current.get("species", ""), item.get("species", ""))
        current["project"] = join_distinct_text(current.get("project", ""), item.get("project", ""))
        current["owner"] = join_distinct_text(current.get("owner", ""), item.get("owner", ""))
        if item.get("roomName"):
            current["roomNames"].add(item.get("roomName"))
        context[iacuc] = current
    return finalize_reimbursement_detail_context(context)


def finalize_reimbursement_detail_context(context):
    finalized = {}
    for iacuc, item in context.items():
        finalized[iacuc] = {
            **item,
            "roomNames": sorted(item.get("roomNames", set())),
        }
    return finalized


def join_distinct_text(current, value):
    values = [part for part in [clean_text(current), clean_text(value)] if part]
    if not values:
        return ""
    return "、".join(sorted(set(values)))


def reimbursement_detail_context_from_workflow(conn, workflow, statement):
    detail_context = {}
    iacucs = [normalize_iacuc_number(value) for value in statement.get("iacucs", []) if normalize_iacuc_number(value)]
    applications_by_iacuc = read_applications_by_iacuc(conn)
    if statement.get("sourceType") == "pi_merged_quantity_sheet":
        sheets = list_quantity_sheets_by_month_pi(conn, statement.get("month"), clean_text(statement.get("pi")))
        rooms = read_room_payloads_for_context(
            conn,
            room_ids=[item.get("roomId", "") for item in sheets],
            room_names=[item.get("roomName", "") for item in sheets],
        )
        detail_context = quantity_sheet_detail_context(sheets, rooms)
    elif statement.get("sourceType") == "quantity_sheet":
        sheets = list_quantity_sheets_by_month_iacuc(conn, statement.get("month"), normalize_iacuc_number(statement.get("iacuc")))
        rooms = read_room_payloads_for_context(
            conn,
            room_ids=[item.get("roomId", "") for item in sheets],
            room_names=[item.get("roomName", "") for item in sheets],
        )
        detail_context = quantity_sheet_detail_context(sheets, rooms)
    else:
        occupancies = read_occupancies_for_billing(
            conn,
            statement.get("month", ""),
            iacuc="" if clean_text(statement.get("sourceType", "")).startswith("pi_merged_") else statement.get("iacuc", ""),
            pi=statement.get("pi", "") if clean_text(statement.get("pi", "")) else "",
        )
        rooms = read_room_payloads_for_context(conn, room_ids=[item.get("roomId", "") for item in occupancies])
        detail_context = occupancy_detail_context(occupancies, rooms)
    for iacuc in iacucs:
        snapshot = statement_application_snapshot(iacuc, applications_by_iacuc, [])
        current = detail_context.get(iacuc, {"roomNames": []})
        detail_context[iacuc] = {
            **current,
            "funding": current.get("funding") or snapshot.get("funding", ""),
            "project": current.get("project") or snapshot.get("project", ""),
            "owner": current.get("owner") or snapshot.get("owner", ""),
        }
    return detail_context


def build_reimbursement_record_payload(existing, workflow, statement, lines, detail_context_by_iacuc, source):
    summary = summarize_statement(statement, lines, detail_context_by_iacuc, BILLING_TIER_LIMIT)
    month = clean_text(statement.get("month", ""))
    pi_name = clean_text(statement.get("pi", ""))
    business_key = reimbursement_business_key(month, pi_name)
    current_month_amount = coerce_reimbursement_money(statement.get("totalAmount", 0))
    support_amount = coerce_reimbursement_money(summary.get("supportAmount", 0))
    payable_amount = coerce_reimbursement_money(summary.get("payableAmount", current_month_amount - support_amount))
    paid_amount = coerce_reimbursement_money(existing.get("paidAmount", 0) if existing else 0)
    reimbursement_status = normalize_reimbursement_status(existing.get("reimbursementStatus") if existing else "")
    payload = {
        "id": existing.get("id") if existing else reimbursement_record_id(month, pi_name),
        "businessKey": business_key,
        "month": month,
        "pi": pi_name,
        "workflowId": workflow.get("id", "") if workflow else existing.get("workflowId", "") if existing else "",
        "workflowStatus": workflow.get("workflowStatus", "") if workflow else existing.get("workflowStatus", "") if existing else "",
        "reimbursementStatus": reimbursement_status,
        "currentMonthAmount": current_month_amount,
        "supportAmount": support_amount,
        "payableAmount": payable_amount,
        "paidAmount": paid_amount,
        "unpaidAmount": coerce_reimbursement_money(max(payable_amount - paid_amount, 0)),
        "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0) if existing else 0),
        "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0) if existing else 0),
        "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0) if existing else 0),
        "fundBookNo": clean_text(existing.get("fundBookNo", "") if existing else ""),
        "reimbursementFormNo": clean_text(existing.get("reimbursementFormNo", "") if existing else ""),
        "approvedBudget": existing.get("approvedBudget", "") if existing else "",
        "notes": clean_text(existing.get("notes", "") if existing else ""),
        "completedAt": clean_text(existing.get("completedAt", "") if existing else ""),
        "source": source,
        "latestEventAt": workflow.get("latestEventAt", "") if workflow else statement.get("generatedAt", ""),
        "updatedAt": now_iso(),
        "details": summary.get("details", []),
        "iacucs": [detail.get("iacuc", "") for detail in summary.get("details", []) if detail.get("iacuc")],
        "statementVersionId": statement.get("versionId", "") or statement.get("id", ""),
        "documentNumber": statement.get("documentNumber", ""),
        "billingUnit": statement.get("billingUnit", ""),
        "project": statement.get("project", ""),
        "owner": statement.get("owner", ""),
        "funding": statement.get("funding", ""),
    }
    if payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED and payload["paidAmount"] + 1e-9 < payload["payableAmount"]:
        payload["reimbursementStatus"] = REIMBURSEMENT_STATUS_REIMBURSING if (payload["fundBookNo"] or payload["reimbursementFormNo"]) else REIMBURSEMENT_STATUS_PENDING
        payload["completedAt"] = ""
    return payload


def upsert_reimbursement_record_from_statement(conn, workflow, statement, lines, detail_context_by_iacuc, source="workflow"):
    business_key = reimbursement_business_key(statement.get("month", ""), statement.get("pi", ""))
    existing = get_reimbursement_record_by_key(conn, business_key) or {}
    payload = build_reimbursement_record_payload(existing, workflow, statement, lines, detail_context_by_iacuc, source)
    upsert_reimbursement_record(conn, payload)
    return payload


def recalculate_reimbursement_accumulations(conn, pi_name):
    records = list_reimbursement_records_for_pi(conn, clean_text(pi_name))
    records = sorted(records, key=lambda item: (clean_text(item.get("month", "")), clean_text(item.get("latestEventAt", "")), clean_text(item.get("id", ""))))
    accumulated_payable = 0.0
    accumulated_paid = 0.0
    for record in records:
        payable_amount = coerce_reimbursement_money(record.get("payableAmount", 0))
        paid_amount = coerce_reimbursement_money(record.get("paidAmount", 0))
        record["unpaidAmount"] = coerce_reimbursement_money(max(payable_amount - paid_amount, 0))
        accumulated_payable += payable_amount
        accumulated_paid += paid_amount
        record["accumulatedPayable"] = coerce_reimbursement_money(accumulated_payable)
        record["accumulatedPaid"] = coerce_reimbursement_money(accumulated_paid)
        record["accumulatedUnpaid"] = coerce_reimbursement_money(accumulated_payable - accumulated_paid)
        record["updatedAt"] = now_iso()
        upsert_reimbursement_record(conn, record)


def recalculate_all_reimbursement_accumulations(conn):
    rows = conn.execute("SELECT DISTINCT pi FROM reimbursement_records WHERE TRIM(COALESCE(pi, '')) != ''").fetchall()
    for row in rows:
        recalculate_reimbursement_accumulations(conn, row["pi"])


def ensure_excel_import_supported():
    if not openpyxl:
        raise ValueError("当前运行环境缺少 Excel 导入依赖 openpyxl")


def month_key(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    if isinstance(value, (int, float)) and openpyxl_from_excel:
        try:
            parsed = openpyxl_from_excel(value)
        except Exception:
            parsed = None
        if isinstance(parsed, datetime):
            return parsed.strftime("%Y-%m")
        if isinstance(parsed, date):
            return parsed.strftime("%Y-%m")
    text = clean_text(value)
    if not text:
        return ""
    compact = re.search(r"(?<!\d)(20\d{2})(\d{2})(?!\d)", text)
    if compact:
        return f"{compact.group(1)}-{compact.group(2)}"
    dotted = re.search(r"(20\d{2})[./-](\d{1,2})", text)
    if dotted:
        return f"{dotted.group(1)}-{int(dotted.group(2)):02d}"
    chinese = re.search(r"(20\d{2})年\s*(\d{1,2})月", text)
    if chinese:
        return f"{chinese.group(1)}-{int(chinese.group(2)):02d}"
    return ""


def split_multiline_values(value):
    text = str(value or "").replace("\r", "\n")
    parts = [clean_text(part) for part in re.split(r"[\n、/]", text)]
    seen = set()
    values = []
    for part in parts:
        if not part or part in seen:
            continue
        seen.add(part)
        values.append(part)
    return values


def distinct_text_list(values):
    seen = set()
    result = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def find_header_row(sheet, required_labels, max_scan_rows=12):
    normalized_required = [clean_text(label).lower() for label in required_labels]
    for row_index in range(1, min(sheet.max_row, max_scan_rows) + 1):
        labels = [clean_text(sheet.cell(row_index, column).value).lower() for column in range(1, sheet.max_column + 1)]
        if all(any(required in label for label in labels if label) for required in normalized_required):
            return row_index, labels
    raise ValueError(f"工作表 {sheet.title} 缺少必需表头")


def find_header_column(labels, *patterns):
    candidates = [clean_text(pattern).lower() for pattern in patterns]
    for index, label in enumerate(labels, start=1):
        if not label:
            continue
        for candidate in candidates:
            if candidate and candidate in label:
                return index
    return 0


def monthly_sheet_month(sheet, header_row):
    for row_index in range(1, min(header_row, 3) + 1):
        for column_index in range(1, min(sheet.max_column, 8) + 1):
            key = month_key(sheet.cell(row_index, column_index).value)
            if key:
                return key
    key = month_key(sheet.title)
    if key:
        return key
    raise ValueError(f"无法识别工作表 {sheet.title} 的月份")


def next_imported_record_id(prefix="reim-import"):
    return f"{prefix}-{secrets.token_hex(8)}"


def build_monthly_import_groups(workbook):
    groups = {}
    for sheet in workbook.worksheets:
        try:
            header_row, labels = find_header_row(sheet, ["项目负责人", "IACUC", "应缴纳"])
        except ValueError:
            continue
        month = monthly_sheet_month(sheet, header_row)
        pi_col = find_header_column(labels, "项目负责人")
        facility_col = find_header_column(labels, "设施")
        iacuc_col = find_header_column(labels, "iacuc")
        funding_col = find_header_column(labels, "伦理对应的经费")
        species_col = find_header_column(labels, "动物品系")
        current_amount_col = find_header_column(labels, "产生的饲养费")
        support_amount_col = find_header_column(labels, "单位支持")
        payable_amount_col = find_header_column(labels, "应缴纳")
        fund_book_col = find_header_column(labels, "经费本号")
        reimbursement_form_col = find_header_column(labels, "单号")
        notes_col = find_header_column(labels, "备注")
        room_col = find_header_column(labels, "房间号")
        start_col = find_header_column(labels, "实验开始")
        end_col = find_header_column(labels, "实验结束")
        if not (pi_col and iacuc_col and current_amount_col and payable_amount_col):
            continue
        empty_rows = 0
        for row_index in range(header_row + 1, min(sheet.max_row, 1200) + 1):
            pi_name = clean_text(sheet.cell(row_index, pi_col).value)
            iacuc = normalize_iacuc_number(sheet.cell(row_index, iacuc_col).value)
            amount = coerce_reimbursement_money(sheet.cell(row_index, current_amount_col).value)
            if not pi_name and not iacuc and amount <= 0:
                empty_rows += 1
                if empty_rows >= 20:
                    break
                continue
            empty_rows = 0
            if not pi_name:
                continue
            business_key = reimbursement_business_key(month, pi_name)
            entry = groups.setdefault(
                business_key,
                {
                    "month": month,
                    "pi": pi_name,
                    "source": "imported",
                    "workflowId": "",
                    "workflowStatus": "",
                    "details": [],
                    "currentMonthAmount": 0.0,
                    "supportAmount": 0.0,
                    "payableAmount": 0.0,
                    "fundBookNos": [],
                    "reimbursementFormNos": [],
                    "notes": [],
                    "species": [],
                    "funding": [],
                },
            )
            support_amount = coerce_reimbursement_money(sheet.cell(row_index, support_amount_col).value) if support_amount_col else 0.0
            payable_amount = coerce_reimbursement_money(sheet.cell(row_index, payable_amount_col).value)
            note_parts = [clean_text(sheet.cell(row_index, notes_col).value)] if notes_col else []
            start_month = month_key(sheet.cell(row_index, start_col).value) if start_col else ""
            end_month = month_key(sheet.cell(row_index, end_col).value) if end_col else ""
            if start_month:
                note_parts.append(f"实验开始：{start_month}")
            if end_month:
                note_parts.append(f"实验结束：{end_month}")
            funding = clean_text(sheet.cell(row_index, funding_col).value) if funding_col else ""
            species = clean_text(sheet.cell(row_index, species_col).value) if species_col else ""
            rooms = split_multiline_values(sheet.cell(row_index, room_col).value) if room_col else []
            facility = clean_text(sheet.cell(row_index, facility_col).value) if facility_col else ""
            entry["currentMonthAmount"] += amount
            entry["supportAmount"] += support_amount
            entry["payableAmount"] += payable_amount
            entry["fundBookNos"].append(clean_text(sheet.cell(row_index, fund_book_col).value) if fund_book_col else "")
            entry["reimbursementFormNos"].append(clean_text(sheet.cell(row_index, reimbursement_form_col).value) if reimbursement_form_col else "")
            entry["notes"].extend([part for part in note_parts if part])
            if species:
                entry["species"].append(species)
            if funding:
                entry["funding"].append(funding)
            entry["details"].append(
                {
                    "iacuc": iacuc,
                    "facility": facility,
                    "funding": funding,
                    "species": species,
                    "project": "",
                    "owner": "",
                    "amount": coerce_reimbursement_money(amount),
                    "supportAmount": coerce_reimbursement_money(support_amount),
                    "payableAmount": coerce_reimbursement_money(payable_amount),
                    "roomNames": rooms,
                    "statementVersionId": "",
                }
            )
    return groups


def import_monthly_reimbursement_workbook(conn, file_body, actor):
    ensure_excel_import_supported()
    workbook = openpyxl.load_workbook(io.BytesIO(file_body), data_only=True)
    groups = build_monthly_import_groups(workbook)
    if not groups:
        raise ValueError("未识别到可导入的月度汇总数据")
    saved = []
    audits = []
    imported_months = set()
    imported_pis = set()
    for entry in groups.values():
        existing = get_reimbursement_record_by_key(conn, reimbursement_business_key(entry["month"], entry["pi"])) or {}
        fund_book_no = "；".join(distinct_text_list(entry["fundBookNos"]))
        reimbursement_form_no = "；".join(distinct_text_list(entry["reimbursementFormNos"]))
        notes = "；".join(distinct_text_list(entry["notes"]))
        payload = {
            "id": existing.get("id") or next_imported_record_id(),
            "businessKey": reimbursement_business_key(entry["month"], entry["pi"]),
            "month": entry["month"],
            "pi": entry["pi"],
            "workflowId": existing.get("workflowId", ""),
            "workflowStatus": existing.get("workflowStatus", ""),
            "reimbursementStatus": normalize_reimbursement_status(
                existing.get("reimbursementStatus") or infer_import_status(fund_book_no, reimbursement_form_no, notes)
            ),
            "currentMonthAmount": coerce_reimbursement_money(entry["currentMonthAmount"]),
            "supportAmount": coerce_reimbursement_money(entry["supportAmount"]),
            "payableAmount": coerce_reimbursement_money(entry["payableAmount"]),
            "paidAmount": coerce_reimbursement_money(existing.get("paidAmount", 0)),
            "unpaidAmount": 0,
            "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0)),
            "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0)),
            "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0)),
            "fundBookNo": fund_book_no or clean_text(existing.get("fundBookNo", "")),
            "reimbursementFormNo": reimbursement_form_no or clean_text(existing.get("reimbursementFormNo", "")),
            "approvedBudget": existing.get("approvedBudget", ""),
            "notes": notes or clean_text(existing.get("notes", "")),
            "completedAt": clean_text(existing.get("completedAt", "")),
            "source": "imported",
            "latestEventAt": now_iso(),
            "updatedAt": now_iso(),
            "details": entry["details"],
            "iacucs": distinct_text_list(detail.get("iacuc", "") for detail in entry["details"]),
            "statementVersionId": existing.get("statementVersionId", ""),
            "documentNumber": existing.get("documentNumber", ""),
            "billingUnit": existing.get("billingUnit", ""),
            "project": clean_text(existing.get("project", "")),
            "owner": clean_text(existing.get("owner", "")),
            "funding": "；".join(distinct_text_list(entry["funding"])) or clean_text(existing.get("funding", "")),
        }
        payload["unpaidAmount"] = coerce_reimbursement_money(max(payload["payableAmount"] - payload["paidAmount"], 0))
        if payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED and payload["paidAmount"] + 1e-9 < payload["payableAmount"]:
            payload["reimbursementStatus"] = infer_import_status(payload["fundBookNo"], payload["reimbursementFormNo"], payload["notes"])
            payload["completedAt"] = ""
        if payload["reimbursementStatus"] == REIMBURSEMENT_STATUS_COMPLETED and not payload["completedAt"]:
            payload["completedAt"] = now_iso()
        upsert_reimbursement_record(conn, payload)
        saved.append(payload)
        imported_months.add(payload["month"])
        imported_pis.add(payload["pi"])
    for pi_name in imported_pis:
        recalculate_reimbursement_accumulations(conn, pi_name)
    event_time = now_iso()
    audit = audit_event(
        actor,
        "reimbursement.import_monthly",
        "reimbursement_record",
        f"monthly:{len(saved)}",
        f"{actor['displayName']} 导入月度报销台账 {len(saved)} 条",
        [],
        event_time,
        None,
        {"months": sorted(imported_months), "count": len(saved)},
    )
    write_audit_events(conn, [audit])
    audits.append(audit)
    invalidate_data_cache_prefixes("reimbursement_records::", "billing_workflows::")
    return {"items": saved, "auditLogs": merge_audit_logs([], audits), "count": len(saved), "months": sorted(imported_months)}


def arrears_summary_columns(sheet):
    month_columns = []
    max_columns = min(sheet.max_column, 120)
    for column_index in range(6, max_columns + 1):
        label = sheet.cell(1, column_index).value
        month = month_key(label)
        if month:
            month_columns.append((column_index, month))
    return month_columns


def import_arrears_reimbursement_workbook(conn, file_body, actor):
    ensure_excel_import_supported()
    workbook = openpyxl.load_workbook(io.BytesIO(file_body), data_only=True)
    if not workbook.worksheets:
        raise ValueError("欠缴工作簿为空")
    sheet = workbook.worksheets[0]
    month_columns = arrears_summary_columns(sheet)
    if not month_columns:
        raise ValueError("未识别到欠缴月份列")
    saved = []
    imported_pis = set()
    empty_rows = 0
    for row_index in range(2, 1200):
        pi_name = clean_text(sheet.cell(row_index, 2).value)
        total_amount = coerce_reimbursement_money(sheet.cell(row_index, 5).value)
        if not pi_name and total_amount <= 0:
            empty_rows += 1
            if empty_rows >= 20:
                break
            continue
        empty_rows = 0
        if not pi_name:
            continue
        species = clean_text(sheet.cell(row_index, 3).value)
        funding = clean_text(sheet.cell(row_index, 4).value)
        for column_index, month in month_columns:
            monthly_unpaid = coerce_reimbursement_money(sheet.cell(row_index, column_index).value)
            if monthly_unpaid <= 0:
                continue
            business_key = reimbursement_business_key(month, pi_name)
            existing = get_reimbursement_record_by_key(conn, business_key) or {}
            if existing and existing.get("source") != "imported":
                continue
            details = existing.get("details") or [
                {
                    "iacuc": "",
                    "facility": "",
                    "funding": funding,
                    "species": species,
                    "project": "",
                    "owner": "",
                    "amount": monthly_unpaid,
                    "supportAmount": 0,
                    "payableAmount": monthly_unpaid,
                    "roomNames": [],
                    "statementVersionId": "",
                }
            ]
            payload = {
                "id": existing.get("id") or next_imported_record_id("reim-arrears"),
                "businessKey": business_key,
                "month": month,
                "pi": pi_name,
                "workflowId": existing.get("workflowId", ""),
                "workflowStatus": existing.get("workflowStatus", ""),
                "reimbursementStatus": normalize_reimbursement_status(existing.get("reimbursementStatus") or REIMBURSEMENT_STATUS_PENDING),
                "currentMonthAmount": coerce_reimbursement_money(existing.get("currentMonthAmount", monthly_unpaid) or monthly_unpaid),
                "supportAmount": coerce_reimbursement_money(existing.get("supportAmount", 0)),
                "payableAmount": coerce_reimbursement_money(existing.get("payableAmount", monthly_unpaid) or monthly_unpaid),
                "paidAmount": coerce_reimbursement_money(existing.get("paidAmount", 0)),
                "unpaidAmount": 0,
                "accumulatedPayable": coerce_reimbursement_money(existing.get("accumulatedPayable", 0)),
                "accumulatedPaid": coerce_reimbursement_money(existing.get("accumulatedPaid", 0)),
                "accumulatedUnpaid": coerce_reimbursement_money(existing.get("accumulatedUnpaid", 0)),
                "fundBookNo": clean_text(existing.get("fundBookNo", "")),
                "reimbursementFormNo": clean_text(existing.get("reimbursementFormNo", "")),
                "approvedBudget": existing.get("approvedBudget", ""),
                "notes": join_distinct_text(existing.get("notes", ""), f"欠缴汇算导入：累计欠缴 {total_amount:.2f} 元"),
                "completedAt": clean_text(existing.get("completedAt", "")),
                "source": "imported",
                "latestEventAt": now_iso(),
                "updatedAt": now_iso(),
                "details": details,
                "iacucs": distinct_text_list(detail.get("iacuc", "") for detail in details),
                "statementVersionId": existing.get("statementVersionId", ""),
                "documentNumber": existing.get("documentNumber", ""),
                "billingUnit": existing.get("billingUnit", ""),
                "project": clean_text(existing.get("project", "")),
                "owner": clean_text(existing.get("owner", "")),
                "funding": funding or clean_text(existing.get("funding", "")),
            }
            payload["unpaidAmount"] = coerce_reimbursement_money(max(payload["payableAmount"] - payload["paidAmount"], 0))
            upsert_reimbursement_record(conn, payload)
            saved.append(payload)
            imported_pis.add(pi_name)
    if not saved:
        raise ValueError("未识别到可导入的欠缴记录")
    for pi_name in imported_pis:
        recalculate_reimbursement_accumulations(conn, pi_name)
    event_time = now_iso()
    audit = audit_event(
        actor,
        "reimbursement.import_arrears",
        "reimbursement_record",
        f"arrears:{len(saved)}",
        f"{actor['displayName']} 导入欠缴汇算 {len(saved)} 条",
        [],
        event_time,
        None,
        {"count": len(saved)},
    )
    write_audit_events(conn, [audit])
    invalidate_data_cache_prefixes("reimbursement_records::", "billing_workflows::")
    return {"items": saved, "auditLogs": merge_audit_logs([], [audit]), "count": len(saved)}


def reimbursement_detail_payload(conn, record):
    record_id = clean_text(record.get("id", ""))
    workflow_id = clean_text(record.get("workflowId", ""))
    cache_key_value = cache_key(
        "reimbursement_records::detail",
        record_id=record_id,
        workflow_id=workflow_id,
        updated_at=clean_text(record.get("updatedAt", "")),
        latest_event_at=clean_text(record.get("latestEventAt", "")),
    )
    cached = cache_get(cache_key_value)
    if cached is not None:
        return cached
    workflow = get_billing_workflow_detail(conn, record.get("workflowId", "")) if record.get("workflowId") else None
    workflow_versions = list_billing_workflow_versions(conn, workflow["id"]) if workflow else []
    workflow_events = list_billing_workflow_events(conn, workflow["id"]) if workflow else []
    history = list_reimbursement_record_summaries_for_pi(conn, record.get("pi", ""))
    payload = {
        "item": record,
        "workflow": workflow if workflow else None,
        "workflowVersions": workflow_versions,
        "workflowEvents": workflow_events,
        "history": history,
    }
    return cache_set(cache_key_value, payload)

def save_billing_statement_workflow(conn, statement, lines, actor, note=""):
    result = save_billing_statement_workflow_service(conn, statement, lines, actor, note, billing_workflow_service_deps())
    invalidate_data_cache_prefixes("billing_workflows::", "billing_statements::", "reimbursement_records::")
    return result


def insert_billing_workflow(conn, payload):
    insert_billing_workflow_repository(
        conn,
        payload,
        payload.get("businessKey", billing_workflow_business_key(payload.get("scopeType", ""), payload.get("scopeKey", ""), payload.get("month", ""), payload.get("sourceType", ""))),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
        as_int(payload.get("currentVersionNo")) or 0,
    )


def update_billing_workflow(conn, payload):
    update_billing_workflow_repository(
        conn,
        payload,
        payload.get("businessKey", billing_workflow_business_key(payload.get("scopeType", ""), payload.get("scopeKey", ""), payload.get("month", ""), payload.get("sourceType", ""))),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
        as_int(payload.get("currentVersionNo")) or 0,
    )


def insert_billing_version(conn, payload):
    insert_billing_version_repository(
        conn,
        payload,
        as_int(payload.get("versionNo")) or 1,
        payload.get("versionStatus", VERSION_STATUS_ACTIVE),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
    )


def update_billing_version(conn, payload):
    update_billing_version_repository(
        conn,
        payload,
        as_int(payload.get("versionNo")) or 1,
        payload.get("versionStatus", VERSION_STATUS_ACTIVE),
        payload.get("workflowStatus", WORKFLOW_STATUS_GENERATED),
    )


def replace_version_lines(conn, version_id, lines):
    replace_billing_statement_version_lines_repository(conn, version_id, lines)


def insert_billing_workflow_event(conn, payload):
    insert_billing_workflow_event_repository(conn, payload)


def update_workflow_status(conn, workflow_id, next_status, actor, note=""):
    return update_workflow_status_service(conn, workflow_id, next_status, actor, note, billing_workflow_service_deps())


def parse_multipart_upload(content_type, raw):
    if "multipart/form-data" not in content_type:
        raise ValueError("请使用 multipart/form-data 上传文件")
    boundary = multipart_boundary(content_type)
    delimiter = b"--" + boundary
    for part in raw.split(delimiter):
        part = part.strip(b"\r\n")
        if not part or part == b"--" or b"\r\n\r\n" not in part:
            continue
        header_blob, body = part.split(b"\r\n\r\n", 1)
        headers = header_blob.decode("utf-8", errors="replace")
        disposition = next((line for line in headers.split("\r\n") if line.lower().startswith("content-disposition:")), "")
        if 'name="file"' not in disposition:
            continue
        filename = multipart_filename(disposition)
        return filename, body.rstrip(b"\r\n")
    raise ValueError("没有找到上传字段 file")


def multipart_boundary(content_type):
    for segment in content_type.split(";"):
        segment = segment.strip()
        if segment.startswith("boundary="):
            value = segment.split("=", 1)[1].strip()
            if value.startswith('"') and value.endswith('"'):
                value = value[1:-1]
            return value.encode("utf-8")
    raise ValueError("上传请求缺少 multipart boundary")


def multipart_filename(disposition):
    match = re.search(r'filename="([^"]*)"', disposition)
    return match.group(1) if match else ""


class CageLedgerHandler(SimpleHTTPRequestHandler):
    server_version = "CageLedger/0.2"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def end_headers(self):
        add_default_headers(self)
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/api/health":
            self.send_json({"ok": True, "database": str(DB_PATH), "system": system_info()})
            return
        if path.startswith("/api/public/cage-card/"):
            qr_id = unquote(path.rsplit("/", 1)[-1])
            try:
                with connect_db() as conn:
                    self.send_json(public_cage_card_payload(conn, qr_id))
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            return
        if path == "/api/system/info":
            self.send_json(system_info())
            return
        if path == "/api/auth/me":
            user = self.current_user()
            if not user:
                self.send_json({"user": None}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_json({"user": user})
            return
        if path == "/api/state":
            user = self.require_user()
            if not user:
                return
            payload = read_state()
            self.send_json({**payload, "state": filter_state_for_actor(payload.get("state"), user)})
            return
        if path == "/api/bootstrap":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                query = parse_qs(urlparse(self.path).query)
                scope = clean_text(query.get("scope", ["summary"])[0]).lower() or "summary"
                room_id = clean_text(query.get("roomId", [""])[0])
                self.send_json(read_bootstrap_state(conn, user, scope, room_id))
            return
        if path == "/api/iacuc-index":
            if not self.require_user():
                return
            self.send_json(read_iacuc_index())
            return
        if path == "/api/iacuc-index/status":
            if not self.require_user():
                return
            payload = read_iacuc_index()
            self.send_json({key: payload[key] for key in ("count", "updatedAt", "source")})
            return
        if path == "/api/system/update-check":
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                self.send_json(system_update_status())
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_GATEWAY)
            return
        if path == "/api/users":
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            with connect_db() as conn:
                self.send_json({"users": list_users(conn)})
            return
        if path == "/api/quantity-sheets":
            if not self.require_user():
                return
            with connect_db() as conn:
                self.send_json(list_quantity_sheets_page(conn, self.list_filters()))
            return
        if path == "/api/quantity-sheet-rooms":
            if not self.require_user():
                return
            with connect_db() as conn:
                rooms = read_payloads(conn, "rooms", "rowid")
            self.send_json({"items": rooms})
            return
        sheet_id = self.quantity_sheet_route(path)
        if sheet_id:
            user = self.require_user()
            if not user:
                return
            try:
                with connect_db() as conn:
                    sheet = get_quantity_sheet(conn, sheet_id)
                    validate_quantity_sheet_permission(user, sheet)
                    self.send_json({"item": sheet})
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            return
        if path == "/api/principal-identities":
            user = self.require_user()
            if not user:
                return
            self.send_json({"items": read_principal_identities()})
            return
        if path == "/api/infrastructure/occupancies":
            user = self.require_user()
            if not user:
                return
            query = parse_qs(urlparse(self.path).query)
            filters = {
                "month": clean_text(query.get("month", [""])[0]),
                "iacuc": clean_text(query.get("iacuc", [""])[0]),
                "pi": clean_text(query.get("pi", [""])[0]),
            }
            try:
                with connect_db() as conn:
                    self.send_json(read_billing_occupancies(conn, user, filters))
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        if path == "/api/billing-workflows":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                self.send_json(list_billing_workflows_page(conn, self.list_filters()))
            return
        if path == "/api/reimbursement-records":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                self.send_json(list_reimbursement_records_page(conn, self.list_filters()))
            return
        workflow_id = self.billing_workflow_lines_route(path)
        if workflow_id:
            user = self.require_user()
            if not user:
                return
            query = parse_qs(urlparse(self.path).query)
            version_id = clean_text(query.get("versionId", [""])[0])
            try:
                with connect_db() as conn:
                    self.send_json(list_billing_workflow_lines(conn, workflow_id, version_id))
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            return
        workflow_id = self.billing_workflow_route(path)
        if workflow_id:
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                workflow = get_billing_workflow_detail(conn, workflow_id)
                if not workflow:
                    self.send_json({"error": "结算流程不存在"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json(
                    {
                        "workflow": workflow,
                        "versions": list_billing_workflow_versions(conn, workflow_id),
                        "events": list_billing_workflow_events(conn, workflow_id),
                    }
                )
            return
        reimbursement_id = self.reimbursement_record_route(path)
        if reimbursement_id:
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                record = get_reimbursement_record(conn, reimbursement_id)
                if not record:
                    self.send_json({"error": "报销台账不存在"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json(reimbursement_detail_payload(conn, record))
            return
        if path == "/api/billing-statements":
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                self.send_json({"items": list_current_billing_statements(conn)})
            return
        statement_id = self.billing_statement_route(path)
        if statement_id:
            user = self.require_user()
            if not user:
                return
            with connect_db() as conn:
                statement = get_current_billing_statement(conn, statement_id)
                if not statement:
                    self.send_json({"error": "结算单不存在"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json({"item": statement})
            return
        if path in ENTITY_ENDPOINTS:
            user = self.require_user()
            if not user:
                return
            self.send_entity_list(ENTITY_ENDPOINTS[path], user)
            return
        endpoint, item_id = self.entity_route(path)
        if endpoint and item_id:
            user = self.require_user()
            if not user:
                return
            collection = WRITABLE_ENTITY_ENDPOINTS[endpoint]["collection"]
            with connect_db() as conn:
                state = filter_state_for_actor(read_cached_state(conn), user)
            item = next((entry for entry in state.get(collection, []) if entry.get("id") == item_id), None)
            if not item:
                self.send_json({"error": "记录不存在"}, HTTPStatus.NOT_FOUND)
                return
            self.send_json({"item": item})
            return
        if path.startswith("/scan/cage-card/") or path.startswith("/c/"):
            self.send_spa_index()
            return
        super().do_GET()

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/auth/login":
            self.handle_login()
            return
        if path == "/api/auth/logout":
            self.handle_logout()
            return
        if path == "/api/iacuc-index/upload":
            self.handle_iacuc_upload()
            return
        if path == "/api/billing-statements/generate":
            self.handle_billing_statement_generate()
            return
        if path == "/api/billing-statements/generate-by-pi":
            self.handle_billing_statement_generate_by_pi()
            return
        if path == "/api/billing-workflows/advance":
            self.handle_billing_workflow_advance()
            return
        if path == "/api/reimbursement-records/import-monthly":
            self.handle_reimbursement_monthly_import()
            return
        if path == "/api/reimbursement-records/import-arrears":
            self.handle_reimbursement_arrears_import()
            return
        if path == "/api/infrastructure":
            self.handle_infrastructure_write()
            return
        sheet_id = self.quantity_sheet_generate_route(path)
        if sheet_id:
            self.handle_quantity_sheet_statement_generate(sheet_id)
            return
        if path == "/api/quantity-sheets":
            self.handle_quantity_sheet_save(None)
            return
        batch_id = self.intake_batch_confirm_route(path)
        if batch_id:
            self.handle_intake_batch_confirm(batch_id)
            return
        task_id = self.placement_task_action_route(path, "reserve")
        if task_id:
            self.handle_placement_task_reserve(task_id)
            return
        task_id = self.placement_task_action_route(path, "move-in")
        if task_id:
            self.handle_placement_task_move_in(task_id)
            return
        task_id = self.placement_task_action_route(path, "reassign-room")
        if task_id:
            self.handle_placement_task_reassign_room(task_id)
            return
        if path == "/api/users":
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                body = self.read_json_body()
                with connect_db() as conn:
                    created = create_user(conn, body)
                self.send_json({"user": created}, HTTPStatus.CREATED)
            except sqlite3.IntegrityError:
                self.send_json({"error": "用户名已存在"}, HTTPStatus.CONFLICT)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        if path in WRITABLE_ENTITY_ENDPOINTS:
            self.handle_entity_write("POST", path, None)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_PUT(self):
        path = urlparse(self.path).path
        if path == "/api/state":
            user = self.require_user()
            if not user:
                return

            try:
                body = self.read_json_body()
                state = body.get("state")
                if not isinstance(state, dict):
                    self.send_error(HTTPStatus.BAD_REQUEST, "Request body must contain a state object")
                    return
                self.send_json(write_state(state, user))
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            except ValueError as exc:
                self.send_error(HTTPStatus.BAD_REQUEST, str(exc))
            return

        user_id = self.user_route(path)
        if user_id:
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                body = self.read_json_body()
                with connect_db() as conn:
                    updated = update_user(conn, user, user_id, body)
                self.send_json({"user": updated})
            except sqlite3.IntegrityError:
                self.send_json({"error": "用户名已存在"}, HTTPStatus.CONFLICT)
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return

        endpoint, item_id = self.entity_route(path)
        if endpoint and item_id:
            self.handle_entity_write("PUT", endpoint, item_id)
            return
        principal_name = self.principal_identity_route(path)
        if principal_name:
            self.handle_principal_identity_save(principal_name)
            return
        sheet_id = self.quantity_sheet_route(path)
        if sheet_id:
            self.handle_quantity_sheet_save(sheet_id)
            return
        reimbursement_id = self.reimbursement_record_route(path)
        if reimbursement_id:
            self.handle_reimbursement_record_update(reimbursement_id)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_DELETE(self):
        path = urlparse(self.path).path
        user_id = self.user_route(path)
        if user_id:
            user = self.require_user()
            if not user:
                return
            if user["role"] != "admin":
                self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
                return
            try:
                with connect_db() as conn:
                    delete_user(conn, user, user_id)
                self.send_json({"ok": True})
            except LookupError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            return

        endpoint, item_id = self.entity_route(path)
        if endpoint and item_id:
            self.handle_entity_write("DELETE", endpoint, item_id)
            return
        sheet_id = self.quantity_sheet_route(path)
        if sheet_id:
            self.handle_quantity_sheet_delete(sheet_id)
            return
        workflow_id = self.billing_workflow_route(path)
        if workflow_id:
            self.handle_billing_workflow_delete(workflow_id)
            return
        reimbursement_id = self.reimbursement_record_route(path)
        if reimbursement_id:
            self.handle_reimbursement_record_delete(reimbursement_id)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def read_json_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValueError("Missing request body")
        if length > MAX_BODY_BYTES:
            raise ValueError("Request body is too large")
        raw = self.rfile.read(length)
        try:
            return json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError as exc:
            raise ValueError("Invalid JSON body") from exc

    def read_raw_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValueError("Missing request body")
        if length > MAX_BODY_BYTES:
            raise ValueError("Request body is too large")
        return self.rfile.read(length)

    def read_optional_json_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        if length > MAX_BODY_BYTES:
            raise ValueError("Request body is too large")
        raw = self.rfile.read(length)
        try:
            return json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError as exc:
            raise ValueError("Invalid JSON body") from exc

    def send_json(self, payload, status=HTTPStatus.OK):
        send_json_response(self, payload, status)

    def send_spa_index(self):
        body = (ROOT / "index.html").read_text(encoding="utf-8")
        if "<base " not in body:
            body = body.replace("<head>", '<head>\n    <base href="/" />', 1)
        body_bytes = body.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body_bytes)))
        self.end_headers()
        self.wfile.write(body_bytes)

    def handle_login(self):
        try:
            body = self.read_json_body()
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        username = str(body.get("username", "")).strip()
        password = str(body.get("password", ""))
        with connect_db() as conn:
            user = authenticate(conn, username, password)
            if not user:
                self.send_json({"error": "用户名或密码错误"}, HTTPStatus.UNAUTHORIZED)
                return
            token, expires_at = create_session(conn, user["id"])
            now = now_iso()
            event = audit_event(
                user,
                "auth.login",
                "session",
                user["id"],
                f"{user['displayName']} 登录系统",
                [],
                now,
                None,
                {
                    "username": user["username"],
                    "role": user["role"],
                    "clientIp": self.client_address[0] if self.client_address else "",
                    "userAgent": self.headers.get("User-Agent", ""),
                },
            )
            write_audit_events(conn, [event])
            conn.commit()
        self.send_response(HTTPStatus.OK)
        body_bytes = json.dumps({"user": user}, ensure_ascii=False).encode("utf-8")
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body_bytes)))
        self.send_header("Cache-Control", "no-store")
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE}={token}; Path=/; HttpOnly; SameSite=Lax; Expires={format_http_date(expires_at)}",
        )
        self.end_headers()
        self.wfile.write(body_bytes)

    def handle_logout(self):
        token = self.session_token()
        with connect_db() as conn:
            delete_session(conn, token)
        self.send_response(HTTPStatus.OK)
        body = b'{"ok": true}'
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")
        self.end_headers()
        self.wfile.write(body)

    def handle_iacuc_upload(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return

        try:
            raw = self.read_raw_body()
            filename, file_body = parse_multipart_upload(self.headers.get("Content-Type", ""), raw)
            if filename and not filename.lower().endswith(".csv"):
                raise ValueError("目前只支持上传 CSV 文件")
            parsed = parse_iacuc_csv(file_body)
            now = now_iso()
            file_items = [application_payload(item, now) for item in parsed["items"]]
            save_iacuc_index_file(file_items)
            event = audit_event(
                user,
                "iacuc_index.uploaded",
                "iacuc_index",
                "iacuc-index",
                f"{user['displayName']} 上传 IACUC 索引 {len(parsed['items'])} 条",
                [],
                now,
                None,
                {"filename": filename, **parsed["summary"]},
            )
            with connect_db() as conn:
                write_experiment_applications(conn, parsed["items"], now)
                write_audit_events(conn, [event])
                conn.commit()
            self.send_json(
                {
                    "ok": True,
                    "filename": filename,
                    "updatedAt": now,
                    **parsed["summary"],
                    "items": file_items,
                    "auditLogs": merge_audit_logs([], [event]),
                }
            )
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_statement_generate(self):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        try:
            with connect_db() as conn:
                statement, lines, audit_logs = generate_billing_statement(conn, body, user)
                workflow = get_billing_workflow(conn, statement.get("workflowId", "")) if statement.get("workflowId") else None
                conn.commit()
            self.send_json({"statement": statement, "lines": lines, "workflow": workflow, "auditLogs": audit_logs}, HTTPStatus.CREATED)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_statement_generate_by_pi(self):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        try:
            with connect_db() as conn:
                statement, lines, audit_logs = generate_billing_statement_by_pi(conn, body, user)
                workflow = get_billing_workflow(conn, statement.get("workflowId", "")) if statement.get("workflowId") else None
                reimbursement = get_reimbursement_record_by_workflow_id(conn, statement.get("workflowId", "")) if statement.get("workflowId") else None
                conn.commit()
            self.send_json(
                {
                    "statement": statement,
                    "lines": lines,
                    "workflow": workflow,
                    "reimbursementItem": reimbursement_record_list_item(reimbursement) if reimbursement else None,
                    "auditLogs": audit_logs,
                },
                HTTPStatus.CREATED,
            )
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_workflow_advance(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            body = self.read_json_body()
            workflow_id = clean_text(body.get("workflowId", ""))
            to_status = clean_text(body.get("toStatus", ""))
            note = clean_text(body.get("note", ""))
            with connect_db() as conn:
                workflow, version, event = update_workflow_status(conn, workflow_id, to_status, user, note)
                reimbursement = get_reimbursement_record_by_workflow_id(conn, workflow_id)
                if reimbursement:
                    reimbursement["workflowStatus"] = workflow.get("workflowStatus", "")
                    reimbursement["latestEventAt"] = workflow.get("latestEventAt", "") or event.get("at", "")
                    reimbursement["updatedAt"] = now_iso()
                    upsert_reimbursement_record(conn, reimbursement)
                audit = audit_event(
                    user,
                    f"billing_workflow.{to_status}",
                    "billing_workflow",
                    workflow_id,
                    f"{user['displayName']} 更新 {workflow.get('iacuc', '')} {workflow.get('month', '')} 结算流程状态",
                    [],
                    event["at"],
                    None,
                    {"workflow": workflow, "version": version, "event": event},
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("billing_workflows::", "billing_statements::", "reimbursement_records::")
            self.send_json(
                {
                    "workflow": workflow,
                    "event": event,
                    "reimbursementItem": reimbursement_record_list_item(reimbursement) if reimbursement else None,
                    "auditLogs": merge_audit_logs([], [audit]),
                }
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_billing_workflow_delete(self, workflow_id):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            with connect_db() as conn:
                workflow = delete_billing_workflow(conn, workflow_id)
                reimbursement = get_reimbursement_record_by_workflow_id(conn, workflow_id)
                deleted_reimbursement_id = ""
                if reimbursement:
                    if reimbursement_has_manual_entry(reimbursement) or reimbursement.get("source") == "imported":
                        reimbursement["workflowId"] = ""
                        reimbursement["workflowStatus"] = "workflow_deleted"
                        reimbursement["notes"] = join_distinct_text(reimbursement.get("notes", ""), "原流程已删除")
                        reimbursement["latestEventAt"] = now_iso()
                        reimbursement["updatedAt"] = now_iso()
                        upsert_reimbursement_record(conn, reimbursement)
                    else:
                        deleted_reimbursement_id = reimbursement.get("id", "")
                        delete_reimbursement_record(conn, reimbursement["id"])
                at = now_iso()
                audit = audit_event(
                    user,
                    "billing_workflow.deleted",
                    "billing_workflow",
                    workflow_id,
                    f"{user['displayName']} 删除 {workflow.get('pi') or workflow.get('iacuc', '')} {workflow.get('month', '')} 结算流程",
                    [],
                    at,
                    workflow,
                    None,
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("billing_workflows::", "billing_statements::", "reimbursement_records::")
            self.send_json(
                {
                    "ok": True,
                    "workflow": workflow,
                    "reimbursementItem": reimbursement_record_list_item(reimbursement) if reimbursement and not deleted_reimbursement_id else None,
                    "deletedReimbursementId": deleted_reimbursement_id,
                    "auditLogs": merge_audit_logs([], [audit]),
                }
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)

    def handle_reimbursement_monthly_import(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            raw = self.read_raw_body()
            filename, file_body = parse_multipart_upload(self.headers.get("Content-Type", ""), raw)
            if filename and not filename.lower().endswith(".xlsx"):
                raise ValueError("请上传月汇总 Excel 文件")
            with connect_db() as conn:
                payload = import_monthly_reimbursement_workbook(conn, file_body, user)
                conn.commit()
            self.send_json({**payload, "filename": filename}, HTTPStatus.CREATED)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_reimbursement_arrears_import(self):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            raw = self.read_raw_body()
            filename, file_body = parse_multipart_upload(self.headers.get("Content-Type", ""), raw)
            if filename and not filename.lower().endswith(".xlsx"):
                raise ValueError("请上传欠缴汇算 Excel 文件")
            with connect_db() as conn:
                payload = import_arrears_reimbursement_workbook(conn, file_body, user)
                conn.commit()
            self.send_json({**payload, "filename": filename}, HTTPStatus.CREATED)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_reimbursement_record_update(self, record_id):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            patch = self.read_json_body()
            with connect_db() as conn:
                existing = get_reimbursement_record(conn, record_id)
                if not existing:
                    raise LookupError("报销台账不存在")
                updated = merge_reimbursement_edit(existing, patch)
                updated["latestEventAt"] = now_iso()
                updated["updatedAt"] = now_iso()
                upsert_reimbursement_record(conn, updated)
                recalculate_reimbursement_accumulations(conn, updated.get("pi", ""))
                refreshed = get_reimbursement_record(conn, record_id) or updated
                audit = audit_event(
                    user,
                    "reimbursement.updated",
                    "reimbursement_record",
                    record_id,
                    f"{user['displayName']} 更新 {refreshed.get('pi', '')} {refreshed.get('month', '')} 报销台账",
                    [],
                    refreshed.get("updatedAt", now_iso()),
                    existing,
                    refreshed,
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("reimbursement_records::")
            with connect_db() as conn:
                detail = reimbursement_detail_payload(conn, get_reimbursement_record(conn, record_id) or refreshed)
            self.send_json({**detail, "auditLogs": merge_audit_logs([], [audit])})
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_reimbursement_record_delete(self, record_id):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            with connect_db() as conn:
                record = get_reimbursement_record(conn, record_id)
                if not record:
                    raise LookupError("报销台账不存在")
                pi_name = record.get("pi", "")
                delete_reimbursement_record(conn, record_id)
                recalculate_reimbursement_accumulations(conn, pi_name)
                audit = audit_event(
                    user,
                    "reimbursement.deleted",
                    "reimbursement_record",
                    record_id,
                    f"{user['displayName']} 删除 {record.get('pi', '')} {record.get('month', '')} 报销台账",
                    [],
                    now_iso(),
                    record,
                    None,
                )
                write_audit_events(conn, [audit])
                conn.commit()
            invalidate_data_cache_prefixes("reimbursement_records::")
            self.send_json({"ok": True, "item": record, "auditLogs": merge_audit_logs([], [audit])})
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)

    def current_user(self):
        with connect_db() as conn:
            return user_from_token(conn, self.session_token())

    def require_user(self):
        user = self.current_user()
        if not user:
            self.send_json({"error": "请先登录"}, HTTPStatus.UNAUTHORIZED)
            return None
        return user

    def session_token(self):
        cookie = self.headers.get("Cookie", "")
        for part in cookie.split(";"):
            if "=" not in part:
                continue
            key, value = part.strip().split("=", 1)
            if key == SESSION_COOKIE:
                return value
        return ""

    def list_filters(self, default_limit=10000, max_limit=10000):
        query = parse_qs(urlparse(self.path).query)
        value = lambda key: query.get(key, [""])[0]
        return {
            "limit": bounded_int(value("limit"), default_limit, 1, max_limit),
            "offset": bounded_int(value("offset"), 0, 0, 1_000_000),
            "status": value("status"),
            "month": value("month"),
            "iacuc": value("iacuc"),
            "pi": value("pi"),
            "roomId": value("roomId"),
            "roomName": value("roomName"),
            "sourceType": value("sourceType"),
            "entityType": value("entityType"),
            "action": value("action"),
            "onlyUnpaid": value("onlyUnpaid"),
        }

    def send_entity_list(self, table, actor):
        with connect_db() as conn:
            if table == "audit_events":
                filters = self.list_filters(default_limit=500, max_limit=1000)
                payload = list_audit_events_page(conn, filters, filtered_where)
                items = payload["items"]
                page = payload["page"]
            elif table == "intake_batches":
                filters = self.list_filters()
                payload = list_intake_batches_page(conn, filters, filtered_where, ENTITY_ORDER_BY)
                items = payload["items"]
                page = payload["page"]
            elif table == "placement_tasks":
                filters = self.list_filters()
                if actor.get("role") != "admin":
                    allowed_rooms = [clean_text(item) for item in actor.get("roomIds", []) if clean_text(item)]
                    requested_room_id = clean_text(filters.get("roomId", ""))
                    if requested_room_id and requested_room_id not in allowed_rooms:
                        filters["roomIds"] = []
                        filters["roomId"] = ""
                    elif not requested_room_id:
                        filters["roomIds"] = allowed_rooms
                payload = list_placement_tasks_page(conn, filters, ENTITY_ORDER_BY, clean_text)
                items = payload["items"]
                page = payload["page"]
            else:
                rows = conn.execute(f"SELECT payload FROM {table} ORDER BY {ENTITY_ORDER_BY.get(table, 'rowid')}").fetchall()
                items = [json.loads(row["payload"]) for row in rows]
                page = None
        collection = {
            "rooms": "rooms",
            "racks": "racks",
            "cage_slots": "slots",
            "occupancies": "occupancies",
            "placement_tasks": "placementTasks",
        }.get(table)
        if collection:
            items = filter_entity_payloads_for_actor(collection, items, actor)
        response = {"items": items}
        if page:
            response["page"] = page
        self.send_json(response)

    def entity_route(self, path):
        for endpoint in WRITABLE_ENTITY_ENDPOINTS:
            prefix = endpoint + "/"
            if path.startswith(prefix):
                item_id = unquote(path[len(prefix) :])
                if "/" not in item_id and item_id:
                    return endpoint, item_id
        return None, None

    def user_route(self, path):
        prefix = "/api/users/"
        if not path.startswith(prefix):
            return None
        user_id = unquote(path[len(prefix) :])
        if "/" in user_id or not user_id:
            return None
        return user_id

    def quantity_sheet_route(self, path):
        prefix = "/api/quantity-sheets/"
        if not path.startswith(prefix):
            return None
        sheet_id = unquote(path[len(prefix) :])
        if "/" in sheet_id or not sheet_id:
            return None
        return sheet_id

    def billing_statement_route(self, path):
        prefix = "/api/billing-statements/"
        if not path.startswith(prefix):
            return None
        statement_id = unquote(path[len(prefix) :])
        if "/" in statement_id or not statement_id:
            return None
        return statement_id

    def principal_identity_route(self, path):
        prefix = "/api/principal-identities/"
        if not path.startswith(prefix):
            return None
        pi_name = unquote(path[len(prefix) :])
        if "/" in pi_name or not pi_name:
            return None
        return pi_name

    def billing_workflow_route(self, path):
        prefix = "/api/billing-workflows/"
        if not path.startswith(prefix):
            return None
        workflow_id = unquote(path[len(prefix) :])
        if "/" in workflow_id or not workflow_id:
            return None
        return workflow_id

    def billing_workflow_lines_route(self, path):
        prefix = "/api/billing-workflows/"
        suffix = "/lines"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        workflow_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in workflow_id or not workflow_id:
            return None
        return workflow_id

    def reimbursement_record_route(self, path):
        prefix = "/api/reimbursement-records/"
        if not path.startswith(prefix):
            return None
        record_id = unquote(path[len(prefix) :])
        if "/" in record_id or not record_id:
            return None
        return record_id

    def intake_batch_confirm_route(self, path):
        prefix = "/api/intake-batches/"
        suffix = "/confirm-receipt"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        batch_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in batch_id or not batch_id:
            return None
        return batch_id

    def placement_task_action_route(self, path, action):
        prefix = "/api/placement-tasks/"
        suffix = f"/{action}"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        task_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in task_id or not task_id:
            return None
        return task_id

    def quantity_sheet_generate_route(self, path):
        prefix = "/api/quantity-sheets/"
        suffix = "/generate-statement"
        if not path.startswith(prefix) or not path.endswith(suffix):
            return None
        sheet_id = unquote(path[len(prefix) : -len(suffix)])
        if "/" in sheet_id or not sheet_id:
            return None
        return sheet_id

    def handle_entity_write(self, method, endpoint, item_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_optional_json_body() if method == "DELETE" else self.read_json_body()
            payload, status = write_entity_state(endpoint, method, item_id, body, user)
            self.send_json(payload, status)
        except sqlite3.IntegrityError:
            self.send_json({"error": "实体 id 已存在"}, HTTPStatus.CONFLICT)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_infrastructure_write(self):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            payload = write_infrastructure_state(body, user)
            self.send_json(payload, HTTPStatus.CREATED)
        except sqlite3.IntegrityError:
            self.send_json({"error": "实体 id 已存在"}, HTTPStatus.CONFLICT)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_quantity_sheet_save(self, sheet_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            with connect_db() as conn:
                sheet, affected_sheets, audit_logs, status, perf = save_quantity_sheet(conn, body, user, sheet_id)
                conn.commit()
            invalidate_data_cache("principal_identities")
            invalidate_data_cache_prefixes("quantity_sheets::", "billing_workflows::")
            self.send_json({"item": sheet, "affectedItems": affected_sheets, "auditLogs": audit_logs, "perf": perf}, status)
        except sqlite3.IntegrityError:
            self.send_json({"error": "数量统计表 id 已存在"}, HTTPStatus.CONFLICT)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_principal_identity_save(self, pi_name):
        user = self.require_user()
        if not user:
            return
        if user["role"] != "admin":
            self.send_json({"error": "需要管理员权限"}, HTTPStatus.FORBIDDEN)
            return
        try:
            body = self.read_json_body()
            with connect_db() as conn:
                item, audit_logs = save_principal_identity(conn, body, user, pi_name)
                conn.commit()
            self.send_json({"item": item, "auditLogs": audit_logs})
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_quantity_sheet_delete(self, sheet_id):
        user = self.require_user()
        if not user:
            return
        try:
            with connect_db() as conn:
                audit_logs = delete_quantity_sheet(conn, user, sheet_id)
                conn.commit()
            invalidate_data_cache("principal_identities")
            invalidate_data_cache_prefixes("quantity_sheets::", "billing_workflows::")
            self.send_json({"ok": True, "auditLogs": audit_logs})
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)

    def handle_quantity_sheet_statement_generate(self, sheet_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_optional_json_body()
            with connect_db() as conn:
                statement, lines, audit_logs = generate_quantity_sheet_statement(conn, sheet_id, body, user)
                conn.commit()
            self.send_json({"statement": statement, "lines": lines, "auditLogs": audit_logs}, HTTPStatus.CREATED)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_intake_batch_confirm(self, batch_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(persist_intake_receipt_confirmation(batch_id, body, user), HTTPStatus.CREATED)
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_placement_task_reserve(self, task_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(
                persist_placement_action(
                    task_id,
                    user,
                    lambda state: reserve_placement_task(state, task_id, clean_text(body.get("slotId")), user),
                )
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_placement_task_move_in(self, task_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(
                persist_placement_action(
                    task_id,
                    user,
                    lambda state: move_in_placement_task(state, task_id, body.get("actualMoveInDate"), user),
                )
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_placement_task_reassign_room(self, task_id):
        user = self.require_user()
        if not user:
            return
        try:
            body = self.read_json_body()
            self.send_json(
                persist_placement_action(
                    task_id,
                    user,
                    lambda state: reassign_placement_task_room(state, task_id, clean_text(body.get("roomId")), user),
                )
            )
        except LookupError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)


configure_database(initialize_schema)


def main():
    ensure_database_ready()
    server = ThreadingHTTPServer((HOST, PORT), CageLedgerHandler)
    print(f"CageLedger server listening on http://{HOST}:{PORT}")
    print(f"SQLite database: {DB_PATH}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
